import sys
import logging
import openpyxl
from functools import partial
from openpyxl.styles import Font
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from PyQt5 import uic, QtWidgets, QtCore
from PyQt5.QtGui import QStandardItemModel, QKeySequence
from PyQt5.QtWidgets import QDialog, QMessageBox, QMenu, QInputDialog, QShortcut
from PyQt5.QtCore import Qt, QTimer
from datetime import datetime
from commonmd import *
#for non_ui version-------------------------
#from consumableinclosing_ui import Ui_ConsumableInClosingDialog

# Closing of the Receipt Product
class ConsumableInClosingDialog(QDialog, SubWindowBase):
#for non_ui version-------------------------
#class ConsumableInClosingDialog(QDialog, Ui_ConsumableInClosingDialog, SubWindowBase):
    
    def __init__(self, current_username = None, current_datetime = None):
        super().__init__()

        self.conn, self.cursor = connect_to_database2()

        uic.loadUi("consumableinclosing.ui", self)
        #for non_ui version-------------------------
        #self.setupUi(self)

        # Initialize current_username and current_datetime directly
        self.current_username, self.current_datetime = initialize_username_and_datetime(current_username, current_datetime)

        # Add window flags
        self.setWindowFlags(self.windowFlags() | Qt.WindowMinimizeButtonHint | Qt.WindowMaximizeButtonHint | Qt.WindowCloseButtonHint)  

        # Enable automatic deletion on close
        self.setAttribute(QtCore.Qt.WA_DeleteOnClose)  

        # Create tv_consumableinclosing and QSortFilterProxyModel
        self.model = QStandardItemModel()
        self.proxy_model = NumericStringSortModel(self.model)
        self.proxy_model.setSourceModel(self.model)        
        
        # Define the column types
        column_types = ["numeric", "", "", "numeric", "numeric", "numeric", "numeric", "", "", "", "", ]

        # Set the custom delegate for the specific column
        delegate = NumericDelegate(column_types, self.tv_consumableinclosing)
        self.tv_consumableinclosing.setItemDelegate(delegate)
        self.tv_consumableinclosing.setModel(self.proxy_model)
       
        # Enable sorting
        self.tv_consumableinclosing.setSortingEnabled(True)

        # Enable alternating row colors
        self.tv_consumableinclosing.setAlternatingRowColors(True)  
        
        # Hide the first index column
        self.tv_consumableinclosing.verticalHeader().setVisible(False)

        # While selecting row in tv_consumableinclosing, each cell values to displayed to designated widgets
        #self.tv_consumableinclosing.clicked.connect(self.show_selected_data)      

        # Fill combobox items when the application starts
        self.get_combobox_contents()

        # Initial display of data
        self.make_data()
        self.conn_button_to_method()
        self.conn_signal_to_slot()

        # Set tab order for input widgets
        self.set_tab_order()
        
        # Initiate CTRL+C, CTRL+V and ENTER
        self.setup_shortcuts()

        self.date_clock_setup()

    # Setup current date time 
    def date_clock_setup(self):
        self.lbl_consumableinclosing_date_time.setText("Initializing...")
        # Set up a QTimer to update the datetime label every second
        self.update_timer = QTimer(self)
        self.update_timer.timeout.connect(self.update_currentdatetime)
        self.update_timer.start(1000)  # Update every 1000 milliseconds (1 second)
        
        current_date = datetime.now()
        current_year = current_date.year
        current_month = current_date.month

        self.cb_consumableinclosing_year.setCurrentText(str(current_year))
        self.cb_consumableinclosing_month.setCurrentText(str(current_month))

        cur_year = int(self.cb_consumableinclosing_year.currentText())
        cur_month = int(self.cb_consumableinclosing_month.currentText())

        first_day = get_first_day_of_month(cur_year, cur_month)
        last_day = get_last_day_of_month(cur_year, cur_month)

        self.entry_consumableinclosing_firstday.setText(first_day.strftime("%Y/%m/%d"))  # Convert date to string
        self.entry_consumableinclosing_lastday.setText(last_day.strftime("%Y/%m/%d"))  # Convert date to string

    # Display current date time 
    def update_currentdatetime(self):
        now = datetime.now()
        curr_date = now.strftime("%Y/%m/%d")
        curr_time = now.strftime("%H:%M:%S")
        ddt = f"{curr_date} {curr_time}"
        self.lbl_consumableinclosing_date_time.setText(ddt)

    # Create a shortcut for CTRL+C/CTRL+V/ Return key
    def setup_shortcuts(self):
        self.copy_shortcut = QShortcut(QKeySequence.Copy, self.tv_consumableinclosing, partial(self.copy_cells, self.tv_consumableinclosing))
        self.paste_shortcut = QShortcut(QKeySequence.Paste, self.tv_consumableinclosing, partial(self.paste_cells, self.tv_consumableinclosing))
        self.return_shortcut = QShortcut(Qt.Key_Return, self.tv_consumableinclosing, partial(self.handle_return_key, self.tv_consumableinclosing))

    # Call process_key_event and pass the event and your QTableWidget instance
    def keyPressEvent(self, event):
        tv_widget = self.tv_consumableinclosing
        self.process_key_event(event, tv_widget)

    # Pass combobox info and sql to next method
    def get_combobox_contents(self):
        self.insert_combobox_initiate(self.cb_consumableinclosing_cname, "SELECT DISTINCT cname FROM customer where type01='con'")

    # Initiate Combo_Box 
    def insert_combobox_initiate(self, combo_box, sql_query):
        self.combobox_initializing(combo_box, sql_query) 
        self.cb_consumableinclosing_cname.setCurrentIndex(0) 

    # Connect the button to the method
    def conn_button_to_method(self):
        self.pb_consumableinclosing_show.clicked.connect(self.make_data)
        self.pb_consumableinclosing_execute.clicked.connect(self.execute_closing)
        self.pb_consumableinclosing_status.clicked.connect(self.change_status)
        self.pb_consumableinclosing_export.clicked.connect(self.export_table)
        self.pb_consumableinclosing_amount.clicked.connect(self.calculate_sum)
        self.pb_consumableinclosing_clear.clicked.connect(self.clear_inputs)

    # Connect signal to method    
    def conn_signal_to_slot(self):
        self.cb_consumableinclosing_year.activated.connect(self.closing_receipt_year_month_changed)
        self.cb_consumableinclosing_month.activated.connect(self.closing_receipt_year_month_changed)
        self.cb_consumableinclosing_cname.activated.connect(self.customer_description_changed)

    # tab order for product window
    def set_tab_order(self):
        widgets = [self.pb_consumableinclosing_show, self.entry_consumableinclosing_ccode, self.cb_consumableinclosing_cname,
            self.cb_consumableinclosing_year, self.cb_consumableinclosing_month, self.entry_consumableinclosing_firstday,
            self.entry_consumableinclosing_lastday, self.entry_consumableinclosing_supval, self.entry_consumableinclosing_vat,
            self.entry_consumableinclosing_ttlval, self.entry_consumableinclosing_remark, self.pb_consumableinclosing_status,
            self.pb_consumableinclosing_execute, self.pb_consumableinclosing_export]

        for i in range(len(widgets) - 1):
            self.setTabOrder(widgets[i], widgets[i + 1])

    # To reduce duplications
    def common_query_statement(self):
        tv_widget = self.tv_consumableinclosing
        
        query = "Select id, cname, pdescription, qty, unitprice, payment, 0 as totalpaymemt, trx_date, status, sdescription, remark"
        query = query + " From vw_aplist_02" 
        query = query + " Where 1=0"

        self.cursor.execute(query)
        column_info = self.cursor.description
        column_names = [col[0] for col in column_info]       

        sql_query = "Select "
        sql_query = sql_query + " id, cname, pdescription, qty, unitprice, payment, Clng(payment*1.1) as totalpayment, trx_date, status, sdescription, remark"
        sql_query = sql_query + " From vw_aplist_02" 
        #sql_query = sql_query + " Where status <> '9'"
        sql_query = sql_query + " order by id desc"
        column_widths = [80, 100, 100, 100, 100, 100]

        return sql_query, tv_widget, column_info, column_names, column_widths

    # show receiptproduct table data inside of the MDI
    def make_data(self):
        sql_query, tv_widget, column_info, column_names, column_widths = self.common_query_statement() 
        self.populate_dialog(self.cursor, sql_query, tv_widget, column_info, column_names,column_widths)

    # Monthly Closing
    def execute_closing(self):

        self.entry_consumableinclosing_supval.setText("")
        self.entry_consumableinclosing_vat.setText("")
        self.entry_consumableinclosing_ttlval.setText("")
        
        cus_code = self.entry_consumableinclosing_ccode.text()
        sdate = self.entry_consumableinclosing_firstday.text()
        edate = self.entry_consumableinclosing_lastday.text()

        if len(cus_code) > 0:
            query = "Select "
            query = query + " id, cname, pdescription, qty, unitprice, payment, Clng(payment*1.1) as totalpayment, trx_date, status, sdescription, remark"
            query = query + " From vw_aplist_02" 
            query = query + " Where "
            query = query + " ccode = " + cus_code
            query = query + " And (trx_date >= #" + sdate + "# and trx_date <= #" + edate + "#)"      # MS Access에서 날짜는 # #로 감싸줘야 함
            query = query + " order by id desc"
        else:
            query = "Select "
            query = query + " id, cname, pdescription, qty, unitprice, payment, Clng(payment*1.1) as totalpayment, trx_date, status, sdescription, remark"
            query = query + " From vw_aplist_02" 
            query = query + " Where "
            query = query + " trx_date >= #" + sdate + "# and trx_date <= #" + edate + "#"      # MS Access에서 날짜는 # #로 감싸줘야 함
            query = query + " order by id desc"            

        sql_query, tv_widget, column_info, column_names, column_widths = self.common_query_statement() 
        self.populate_dialog(self.cursor, query, tv_widget, column_info, column_names,column_widths)

    # Status change
    def change_status(self):
        confirm_dialog = self.show_closing_confirmation_dialog()

        if confirm_dialog == QMessageBox.Yes:
        
            sdate = self.entry_consumableinclosing_firstday.text()
            edate = self.entry_consumableinclosing_lastday.text()
            ccode = self.entry_consumableinclosing_ccode.text()
            cname = self.cb_consumableinclosing_cname.currentText()

            if len(ccode)>0:
                self.cursor.execute('''UPDATE aplist SET status='9' WHERE ccode=? and (trx_date >= ? and trx_date <= ?)''' , (ccode, sdate, edate))
            else:
                self.cursor.execute('''UPDATE aplist SET status='9' WHERE trx_date >= ? and trx_date <= ?''' , (sdate, edate))
            QMessageBox.about(self, "검색 조건 확인", f"거래처명: {cname} \n마감시작일: {sdate} \n마감종료일: {edate} \n\n위 조건으로 검색을 수행합니다!")
            
            self.conn.commit()
            self.show_update_success_message()
            self.refresh_data()         
        else:
            self.show_cancel_message("데이터 변경 취소")

    # Export data to Excel sheet                
    def export_table(self):
        output_subfolder = "closing"                        # set the output subfoler name
        tv_widget = self.tv_consumableinclosing             # set the name of table widget
        filetext = "매입마감"
        sheet_name = "data"                                 # set the excel sheet name
        filename = prefix_get_file_name(sheet_name, filetext, output_subfolder)                # get the file name / export_to_excel 밑에 두면 get_file_name을 두 번 실행시켜서 파일의 번호가 달라지므로 미리 만들어서 filename에 담아둠.
        numeric_columns=[0, 3, 4 ,5, 6]                     # set the numerical column index
        prefix_export_to_excel(output_subfolder, tv_widget, sheet_name, filetext, numeric_columns)

        if filename:
            full_file_path = os.path.join(output_subfolder, filename)
            self.excel_formatting(sheet_name, full_file_path)
            self.excel_sum_result(full_file_path)
            QMessageBox.about(self, "파일 생성 완료", f"엑셀 파일이 {full_file_path}에 \n생성 되었습니다!") 
        else:
            QMessageBox.about(self, "코드 확인", "파일을 찾지 못했습니다!") 

    def excel_formatting(self, sheet_name, full_file_path):
        wb = load_workbook(full_file_path)
        sheet_name = sheet_name
        ws = wb[sheet_name]
        
        # Find the last row in column K
        last_row = ws.max_row + 1

        # Handling Integer Column
        # Apply number format to columns D:G
        for col_num in range(4, 8):
            col_letter = openpyxl.utils.get_column_letter(col_num)
            for row in ws.iter_rows(min_row=2, max_row=last_row, min_col=col_num, max_col=col_num):
                for cell in row:
                    cell.number_format = '#,##0'

        # Handling column which has the string date type
        # Set formula from K2:K16 as formulas
        for row_num in range(2, last_row+1):
            date_str = ws[f'H{row_num}'].value
            if date_str is not None:
                # Adjust the format to match your Excel date format
                date_obj = datetime.strptime(date_str, "%Y-%m-%d") 
                timestamp = int(date_obj.timestamp())
                excel_date = (date_obj - datetime(1899, 12, 30)).days
                ws[f'K{row_num}'].value = excel_date
            else:
                # Handle the case when the cell is empty or doesn't contain a valid date string
                ws[f'K{row_num}'].value = None

        # Copy the values from range K2:K16
        copied_values = []
        for row_num in range(2, last_row+1):
            copied_values.append([ws[f'K{row_num}'].value])

        # Paste the copied values to column H, starting from H2 
        for row_num, value in enumerate(copied_values, start=2):
            ws[f'H{row_num}'].value = value[0]

        # Apply date number format to the cells in column H
        for row_num in range(2, last_row+1):
            ws[f'H{row_num}'].number_format = 'yyyy-mm-dd'

        # Delete column K
        ws.delete_cols(11)

        column_widths = [5, 10, 15, 10, 12, 12, 15, 12, 10, 10] # set the width of each column
        title_font = Font(bold=True, name="Arial", size=10)     # set the font for the first row
        arial_font = Font(name="Arial", size=10)                # set the forn from the second row to max row

        set_column_widths(ws, column_widths)    # reset column widths

        ws.freeze_panes = 'D2'              # freeze panes D2 cell
        ws.auto_filter.ref = ws.dimensions  # apply auto filter
        ws.sheet_view.showGridLines = False # remove gridlines
            
        set_font(ws, [1], range(1, len(column_widths) + 1), title_font)                 # first row font setting
        set_font(ws, range(2, last_row), range(1, len(column_widths) + 1), arial_font)  # from second row to last row font setting
        
        wb.save(full_file_path)

    # Add to the result of summmation of the column
    def excel_sum_result(self, full_file_path):
        wb = load_workbook(full_file_path)

        # Delete the default "Sheet" sheet if it exists
        if "Sheet" in wb.sheetnames:
            sheet_to_delete = wb["Sheet"]
            wb.remove(sheet_to_delete)
            
        # Create the 'closing' sheet and set it as the active sheet
        sheet = wb.create_sheet('closing')
        sheet_name = 'closing'
        ws = wb[sheet_name]

        sdate = self.entry_consumableinclosing_firstday.text()
        edate = self.entry_consumableinclosing_lastday.text()

        sql_query = '''
                SELECT pcode, pdescription, um, SUM(qty) AS TotalQty, SUM(CLng(payment*1.1)) AS TotalPayment
                FROM vw_aplist_02
                WHERE trx_date BETWEEN #{}# AND #{}#
                GROUP BY pcode, pdescription, um
                ORDER BY pcode'''.format(sdate, edate)

        self.cursor.execute(sql_query)
        results = self.cursor.fetchall()

        # Write column headers
        column_headers = ["pcode", "pdescription", "um", "TotalQty", "TotalPayment"]
        for col_num, header in enumerate(column_headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header)

        # Write data to the worksheet
        for row_num, row_data in enumerate(results, 2):
            for col_num, cell_value in enumerate(row_data, 1):
                cell = sheet.cell(row=row_num, column=col_num, value=cell_value)            

        # Add summation functions to the last rows
        last_row = len(results) + 2  # Adjust for headers and 1-based indexing
        sheet.cell(row=last_row, column=1, value="Total").font 
        sheet.cell(row=last_row, column=5, value="=SUM(E2:E{})".format(last_row - 1))
        
        # Make the summation row bold
        bold_only_font = Font(name='Calibri', size=11, bold=True)
        for col in [1, 5]:
            cell = sheet.cell(row=last_row, column=col)
            cell.font = bold_only_font
        
        # Select columns B and C & Apply the "#,##0" number format to the selected columns
        columns_range = sheet['D:E']

        for column in columns_range:
            for cell in column:
                cell.number_format = '#,##0'

        #----------------------------------------------------------------------------------------------
        column_widths = [10, 15, 10, 12, 20]                    # set the width of each column
        title_font = Font(bold=True, name="Arial", size=10)     # set the font for the first row
        arial_font = Font(name="Arial", size=10)                # set the forn from the second row to max row

        set_column_widths(ws, column_widths)    # reset column widths

        ws.freeze_panes = 'D2'                  # freeze panes D2 cell
        ws.auto_filter.ref = ws.dimensions      # apply auto filter
        ws.sheet_view.showGridLines = False     # remove gridlines
            
        set_font(ws, [1], range(1, len(column_widths) + 1), title_font)                 # first row font setting
        set_font(ws, range(2, last_row), range(1, len(column_widths) + 1), arial_font)  # from second row to last row font setting
        #----------------------------------------------------------------------------------------------

        wb.save(full_file_path)

    # Calculate the sum of values in a specific column
    def calculate_sum(self):
        
        self.entry_consumableinclosing_supval.clear()
        self.entry_consumableinclosing_vat.clear()
        self.entry_consumableinclosing_ttlval.clear()

        COLUMN = 5
        r = 0
        for i in range(self.tv_consumableinclosing.model().rowCount()):
            item = self.tv_consumableinclosing.model().item(i, COLUMN)
            if item and item.data(Qt.DisplayRole):
                value = int(item.data(Qt.DisplayRole))
                r += value
        #print(f"Sum = {r}")
       
        self.entry_consumableinclosing_supval.setText(str(r))
        vat = int(r * 0.1)
        self.entry_consumableinclosing_vat.setText(str(vat))
        ttlval = int(r * 1.1)
        self.entry_consumableinclosing_ttlval.setText(str(ttlval))


    # Combo box year activated and connect
    def closing_receipt_year_month_changed(self):
        self.entry_consumableinclosing_firstday.clear()
        self.entry_consumableinclosing_lastday.clear()
        
        selected_year = self.cb_consumableinclosing_year.currentText()
        selected_month = self.cb_consumableinclosing_month.currentText()

        cur_year = int(selected_year)
        cur_month = int(selected_month)

        first_day = get_first_day_of_month(cur_year, cur_month)
        last_day = get_last_day_of_month(cur_year, cur_month)

        self.entry_consumableinclosing_firstday.setText(first_day.strftime("%Y/%m/%d"))  # Convert date to string
        self.entry_consumableinclosing_lastday.setText(last_day.strftime("%Y/%m/%d"))  # Convert date to string

            
    # customer code editing finished and connect
    def customer_description_changed(self):
        self.entry_consumableinclosing_ccode.clear()
        selected_item = self.cb_consumableinclosing_cname.currentText()

        if selected_item:
            query = f"SELECT DISTINCT ccode From customer WHERE cname ='{selected_item}'"
            line_edit_widgets = [self.entry_consumableinclosing_ccode]
        
            # Check if any line edit widgets are provided
            if line_edit_widgets:
                self.lineEdit_contents(line_edit_widgets, query)
            else:
                pass

    # Clear Inputs
    def clear_inputs(self):
        for line_edit in self.findChildren(QtWidgets.QLineEdit):
            line_edit.clear()
        self.cb_consumableinclosing_cname.setCurrentIndex(0)

    # Clear all entry and combo boxes
    def refresh_data(self):
        self.make_data()

if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    dialog = ConsumableInClosingDialog()
    dialog.show()
    sys.exit(app.exec())