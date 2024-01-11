import sys
import logging
from functools import partial
from openpyxl.styles import Font
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from PyQt5 import uic, QtWidgets, QtCore
from PyQt5.QtGui import QStandardItemModel, QKeySequence
from PyQt5.QtWidgets import QDialog, QMessageBox, QMenu, QInputDialog, QShortcut
from PyQt5.QtCore import Qt
from datetime import datetime
from datetime import timedelta
from datetime import date
from commonmd import *
from cal import CalendarView
#for non_ui version-------------------------
#from cost_ui import Ui_CostDialog

# Cost table contents -----------------------------------------------------
class CostDialog(QDialog, SubWindowBase):
#for non_ui version-------------------------
#class CostDialog(QDialog, Ui_CostDialog, SubWindowBase): 
    
    def __init__(self, current_username = None, current_datetime = None):
        super().__init__()

        self.conn, self.cursor = connect_to_database1()

        # Load ui file
        uic.loadUi("cost.ui", self)
        #for non_ui version-------------------------
        #self.setupUi(self)

        # Add window flags
        self.setWindowFlags(self.windowFlags() | Qt.WindowMinimizeButtonHint | Qt.WindowMaximizeButtonHint | Qt.WindowCloseButtonHint)  

        # Initialize current_username and current_datetime directly
        self.current_username, self.current_datetime = initialize_username_and_datetime(current_username, current_datetime)

        # Enable automatic deletion on close
        self.setAttribute(QtCore.Qt.WA_DeleteOnClose)  
        
        # Create tv_cost and QSortFilterProxyModel
        self.model = QStandardItemModel()
        self.proxy_model = NumericStringSortModel(self.model)
        self.proxy_model.setSourceModel(self.model)        
        
        # Define the column types
        column_types = ["numeric", "numeric", "numeric", "", "", "", "", "", "", "", "", ""]
    
        # Set the custom delegate for the specific column
        delegate = NumericDelegate(column_types, self.tv_cost)
        self.tv_cost.setItemDelegate(delegate)
        self.tv_cost.setModel(self.proxy_model)
       
        # Enable sorting
        self.tv_cost.setSortingEnabled(True)

        # Enable alternating row colors
        self.tv_cost.setAlternatingRowColors(True)  
        
        # Hide the first index column
        self.tv_cost.verticalHeader().setVisible(False)

        # While selecting row in tv_cost, each cell values to displayed to designated widgets
        self.tv_cost.clicked.connect(self.show_selected_data)

        # Fill combobox items when the application starts
        self.get_combobox_contents()

        # Initial Display of data
        self.make_data()
        self.connect_btn_method()
        self.conn_signal_to_slot()

        # Set tab order for input widgets
        self.set_tab_order()

        # Initiate CTRL+C, CTRL+V and ENTER
        self.setup_shortcuts()
        
        # Create context menus ----------------------------------------------------------------------------------
        self.context_menu1 = self.create_context_menu(self.entry_cost_efffrom)
        self.context_menu2 = self.create_context_menu(self.entry_cost_effthru)
        self.context_menu3 = self.create_context_menu(self.entry_costchange_efffrom)
        self.context_menu4 = self.create_context_menu(self.entry_costchange_effthru)

        self.entry_cost_efffrom.setContextMenuPolicy(Qt.CustomContextMenu)
        self.entry_cost_efffrom.customContextMenuRequested.connect(self.show_context_menu1)

        self.entry_cost_effthru.setContextMenuPolicy(Qt.CustomContextMenu)
        self.entry_cost_effthru.customContextMenuRequested.connect(self.show_context_menu2)

        self.entry_costchange_efffrom.setContextMenuPolicy(Qt.CustomContextMenu)
        self.entry_costchange_efffrom.customContextMenuRequested.connect(self.show_context_menu3)

        self.entry_costchange_effthru.setContextMenuPolicy(Qt.CustomContextMenu)
        self.entry_costchange_effthru.customContextMenuRequested.connect(self.show_context_menu4)
        #-----------------------------------------------------------------------------------------------------------------

        self.entry_stylesheet_as_is()
        self.hide_costchange_widget()

        # Make log file
        self.make_logfiles("access_cost.log")

    # Create a shortcut for CTRL+C/CTRL+V/ Return key
    def setup_shortcuts(self):
        self.copy_shortcut = QShortcut(QKeySequence.Copy, self.tv_cost, partial(self.copy_cells, self.tv_cost))
        self.paste_shortcut = QShortcut(QKeySequence.Paste, self.tv_cost, partial(self.paste_cells, self.tv_cost))
        self.return_shortcut = QShortcut(Qt.Key_Return, self.tv_cost, partial(self.handle_return_key, self.tv_cost))

    # Change styles for the selected widgets 1
    def entry_stylesheet_as_is(self):
        self.entry_costchange_cost.setStyleSheet('color:black;background:rgb(255,255,255)')
        self.entry_costchange_efffrom.setStyleSheet('color:black;background:rgb(255,255,255)')
        self.entry_costchange_effthru.setStyleSheet('color:black;background:rgb(255,255,255)')
        self.entry_costchange_remark.setStyleSheet('color:black;background:rgb(255,255,255)')

    # Change styles for the selected widgets 2
    def entry_stylesheet_to_be(self):
        self.entry_costchange_cost.setStyleSheet('color:white;background:rgb(255,0,0)')
        self.entry_costchange_efffrom.setStyleSheet('color:white;background:rgb(255,0,0)')
        self.entry_costchange_effthru.setStyleSheet('color:white;background:rgb(255,0,0)')
        self.entry_costchange_remark.setStyleSheet('color:white;background:rgb(255,0,0)')

    # Show widgets for the cost change parts 
    def show_costchange_widget(self):
        ddt, endofdate = self.display_eff_date()
        self.pb_costchange_update.setVisible(True)
        self.entry_costchange_cost.setReadOnly(False)
        self.entry_costchange_efffrom.setReadOnly(False)
        self.entry_costchange_effthru.setReadOnly(False)
        self.entry_costchange_remark.setReadOnly(False)
        self.entry_costchange_effthru.setText(endofdate)
        self.entry_stylesheet_to_be()

    # Hide widgets for the cost change parts 
    def hide_costchange_widget(self):
        self.pb_costchange_update.setVisible(False)
        self.entry_costchange_cost.setReadOnly(True)
        self.entry_costchange_efffrom.setReadOnly(True)
        self.entry_costchange_effthru.setReadOnly(True)
        self.entry_costchange_remark.setReadOnly(True)        
        self.entry_stylesheet_as_is()

    # Mouse Right click, show "달력보기" menu ---------------------------------------------------------------------
    def create_context_menu(self, target_lineedit):
        context_menu = QMenu()
        custom_action = context_menu.addAction("달력보기")
        custom_action.triggered.connect(lambda: self.show_calendar(target_lineedit))
        return context_menu

    def show_context_menu1(self, pos):
        self.context_menu1.exec_(self.entry_cost_efffrom.mapToGlobal(pos))
    def show_context_menu2(self, pos):
        self.context_menu2.exec_(self.entry_cost_effthru.mapToGlobal(pos))
    def show_context_menu3(self, pos):
        self.context_menu3.exec_(self.entry_costchange_efffrom.mapToGlobal(pos))
    def show_context_menu4(self, pos):
        self.context_menu4.exec_(self.entry_costchange_effthru.mapToGlobal(pos))

    # Show Calendar
    def show_calendar(self, target_lineedit):
        calendar_dialog = CalendarView()
        calendar_dialog.selected_date_changed.connect(lambda date: self.set_selected_date(date, target_lineedit))
        calendar_dialog.exec()

    # Show selected date to the select Qlineedit
    def set_selected_date(self, date, target_lineedit):
        if target_lineedit == self.entry_cost_efffrom:
            target_lineedit.setText(date)
        elif target_lineedit == self.entry_cost_effthru:
            target_lineedit.setText(date)
        elif target_lineedit == self.entry_costchange_efffrom:
            target_lineedit.setText(date)
        elif target_lineedit == self.entry_costchange_effthru:
            target_lineedit.setText(date)     

    # Call process_key_event and pass the event and your QTableWidget instance
    def keyPressEvent(self, event):
        tv_widget = self.tv_cost
        self.process_key_event(event, tv_widget)

    # Display current date only
    def display_eff_date(self):
        now = datetime.now()
        curr_date = now.strftime("%Y/%m/%d")
        ddt = f"{curr_date}"         
        endofdate = "2050/12/31"

        return ddt, endofdate

    # Pass combobox info and sql to next method
    def get_combobox_contents(self):
        self.insert_combobox_initiate(self.cb_cost_class1, "SELECT DISTINCT class1 FROM product ORDER BY class1")
        self.insert_combobox_initiate(self.cb_cost_class2, "SELECT DISTINCT class2 FROM product ORDER BY class2")
        self.insert_combobox_initiate(self.cb_cost_class3, "SELECT DISTINCT class3 FROM cost ORDER BY class3")
        self.insert_combobox_initiate(self.cb_cost_class4, "SELECT DISTINCT class4 FROM cost ORDER BY class4")                

    # Initiate Combo_Box 
    def insert_combobox_initiate(self, combo_box, sql_query):    
        self.combobox_initializing(combo_box, sql_query)
        self.lbl_cost_id.setText("")
        self.entry_cost_code1.setText("")
        self.entry_cost_code2.setText("")
        self.cb_cost_class1.setCurrentIndex(0) 
        self.cb_cost_class2.setCurrentIndex(0)
        self.cb_cost_class3.setCurrentIndex(0)
        self.cb_cost_class4.setCurrentIndex(0)

    # Connect button to method
    def connect_btn_method(self):
        self.pb_cost_show.clicked.connect(self.make_data)
        self.pb_cost_search.clicked.connect(self.search_data)
        self.pb_cost_currentprice.clicked.connect(self.current_cost)
        self.pb_cost_close.clicked.connect(self.close_dialog)
        self.pb_cost_clear.clicked.connect(self.clear_data)

        self.pb_cost_insert.clicked.connect(self.tb_insert)
        self.pb_cost_update.clicked.connect(self.tb_update)
        self.pb_cost_delete.clicked.connect(self.tb_delete)
        self.pb_cost_excel_export.clicked.connect(self.export_table)
        self.pb_costchange_update.clicked.connect(self.reflect_cost_change)

    # Connect signal to method    
    def conn_signal_to_slot(self):
        self.cb_cost_class1.activated.connect(self.cost_class1_changed)
        self.cb_cost_class2.activated.connect(self.cost_class2_changed)
        self.cb_cost_class3.activated.connect(self.cost_class3_changed)
        self.cb_cost_class4.activated.connect(self.cost_class4_changed)
        self.entry_costchange_efffrom.editingFinished.connect(self.costchange_efffrom_change)

    # tab order for cost window
    def set_tab_order(self):
        widgets = [self.pb_cost_show, self.entry_cost_code1, self.entry_cost_code2,
            self.cb_cost_class1, self.cb_cost_class2, self.cb_cost_class3,
            self.cb_cost_class4, self.entry_cost_um, self.entry_cost_cost,
            self.entry_cost_efffrom, self.entry_cost_effthru, self.entry_cost_remark, 
            self.pb_cost_currentprice, self.pb_cost_search, self.pb_cost_clear, 
            self.pb_cost_insert, self.pb_cost_update, self.pb_cost_delete,
            self.pb_cost_close,self.entry_costchange_cost, self.entry_costchange_efffrom,
            self.entry_costchange_effthru, self.entry_costchange_remark, self.pb_costchange_update]
        
        for i in range(len(widgets) - 1):
            self.setTabOrder(widgets[i], widgets[i + 1])

    # To reduce duplications
    def common_query_statement(self):
        tv_widget = self.tv_cost
        self.cursor.execute("SELECT id, pcode, costcode, class1, class2, class3, class4, um, cost, efffrom, effthru, remark FROM vw_cost where 1=0")
        column_info = self.cursor.description
        column_names = [col[0] for col in column_info]    

        sql_query = "SELECT id, pcode, costcode, class1, class2, class3, class4, um, cost, efffrom, effthru, remark FROM vw_cost"
        column_widths = [80, 100, 100, 100, 100, 100]

        return sql_query, tv_widget, column_info, column_names, column_widths
    
    # show cost table data inside of the MDI
    def make_data(self):
        sql_query, tv_widget, column_info, column_names, column_widths = self.common_query_statement() 
        self.populate_dialog(self.cursor, sql_query, tv_widget, column_info, column_names,column_widths)

    # Get the value of other variables
    def get_cost_input(self):
        pcode = int(self.entry_cost_code1.text())
        costcode = int(self.entry_cost_code2.text())
        class1 = str(self.cb_cost_class1.currentText())
        class2 = str(self.cb_cost_class2.currentText())
        class3 = str(self.cb_cost_class3.currentText())
        class4 = str(self.cb_cost_class4.currentText())
        um = str(self.entry_cost_um.text())

        input_val = self.entry_cost_cost.text()
        if input_val.replace(".", "").replace("-", "").isdigit():
            cost = float(input_val)
        else:
            cost = 0

        #cost = float(self.entry_cost_cost.text())

        #date_format_pattern = r'\d{4}/\d{2}/\d{2}'
        #srtdt = str(self.entry_cost_efffrom.text())
        #efffrom = srtdt if re.match(date_format_pattern, srtdt) else "2023/01/01"
        #enddt = str(self.entry_cost_effthru.text())
        #effthru = enddt if re.match(date_format_pattern, srtdt) else "2030/12/31"        
        
        efffrom = str(self.entry_cost_efffrom.text())
        effthru = str(self.entry_cost_effthru.text())
        remark = str(self.entry_cost_remark.text())

        return pcode, costcode, class1, class2, class3, class4, um, cost, efffrom, effthru, remark

    # Make Common values set
    def common_values_set(self):
        username = self.current_username
        user_id = self.userID_gen(username)
        formatted_datetime = self.dt_time_info()

        return username, user_id, formatted_datetime

    # insert new cost data to MySQL table
    def tb_insert(self):
        currentid = self.lbl_cost_id.text()
        if not currentid:
            confirm_dialog = self.show_insert_confirmation_dialog()
            
            if confirm_dialog == QMessageBox.Yes:

                idx = self.max_row_id("cost")                                                                  # Get the max id 
                username, user_id, formatted_datetime = self.common_values_set()
                pcode, costcode, class1, class2, class3, class4, um, cost, efffrom, effthru, remark = self.get_cost_input()           # Get the value of other variables

                if (idx>0 and pcode>0 and abs(costcode)>0 and abs(cost)>0) and all(len(var) > 0 for var in (class3, class4, um, efffrom, effthru)):
                    self.cursor.execute('''INSERT INTO cost (id, pcode, costcode, class3, class4, um, cost, efffrom, effthru, trxdate, userid, remark) 
                                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)'''
                                , (idx, pcode, costcode, class3, class4, um, cost, efffrom, effthru, formatted_datetime, user_id, remark))
                    self.conn.commit()

                    self.show_insert_success_message()
                    self.refresh_data() 
                    logging.info(f"User {username} inserted {idx} row to the cost table.")
                else:
                    self.show_missing_message("입력 이상")
                    return
            else:
                self.show_cancel_message("데이터 추가 취소")
                return
        else:
            QMessageBox.information(self, "Input Error", "기존 id가 선택된 상태에서는 신규 입력이 불가합니다!")
            return

    # revise the values in the selected row
    def tb_update(self):
        
        id_text = self.lbl_cost_id.text()
        if id_text.strip(): #id_text가 공백이 아닌 경우를 확인

            conA = '''단가 오류 수정 - 현재 데이터 수정, 추가 행을 만들지 않음!'''
            conB = '''단가 갱신 - 현재 행의 종료일 변경 + 변경된 단가 추가 행을 삽입!'''
            
            conditions = [conA, conB]
            condition, okPressed = QInputDialog.getItem(self, "입력 조건 선택", "어떤 작업을 진행하시겠습니까?:", conditions, 0, False)

            if okPressed:
                if condition == conA:
                    self.fix_typo()     
                elif condition == conB:
                    self.show_costchange_widget()
                else:
                    return

        else:
            QMessageBox.about(self, "수정 조건 확인", "선택된 행이 없습니다!")
            return
        
    # Fix typing error in the cost
    def fix_typo(self):

        confirm_dialog = self.show_update_confirmation_dialog()

        if confirm_dialog == QMessageBox.Yes:
            
            id_text = self.lbl_cost_id.text()
            idx = int(id_text)

            username, user_id, formatted_datetime = self.common_values_set()    
            pcode, costcode, class1, class2, class3, class4, um, cost, efffrom, effthru, remark = self.get_cost_input()
            
            fix_efffrom = str(self.entry_cost_efffrom.text())
            fix_effthru = str(self.entry_cost_effthru.text())

            if (idx>0 and pcode>0 and abs(costcode)>=0 and abs(cost)>=0) and all(len(var) > 0 for var in (class3, class4, um, efffrom, effthru)):
                self.cursor.execute('''UPDATE cost SET 
                            pcode=?, costcode=?, class3=?, class4=?, um=?, cost=?, efffrom=?, effthru=?, trxdate=?, userid=?, remark=? WHERE id=?'''
                            , (pcode, costcode, class3, class4, um, cost, fix_efffrom, fix_effthru, formatted_datetime, user_id, remark, idx))
                self.conn.commit()
                self.show_update_success_message()
                self.refresh_data()  
                logging.info(f"User {username} updated row number {idx} in the cost table.")
            else:
                self.show_missing_message("입력 이상")
                return
        else:
            self.show_cancel_message("데이터 변경 취소")
            return

    # Changed cost values
    def changed_cost_value(self):
        chgcost = float(self.entry_costchange_cost.text())
        chgefffrom = str(self.entry_costchange_efffrom.text())
        chgeffthru = str(self.entry_costchange_effthru.text())
        chgremark = str(self.entry_costchange_remark.text())

        return chgcost, chgefffrom, chgeffthru, chgremark

    # Reflect changed cost values 1. change cost effthru date for the current record 2. add newly added cost record for the cost table.
    def reflect_cost_change(self):
        idx = int(self.max_row_id("cost"))
        username, user_id, formatted_datetime = self.common_values_set()
        pcode, costcode, class1, class2, class3, class4, um, cost, efffrom, effthru, remark = self.get_cost_input()
        chgcost, chgefffrom, chgeffthru, chgremark = self.changed_cost_value()

        if (idx>0 and abs(chgcost)>=0) and all(len(var) > 0 for var in (chgefffrom, chgeffthru)):
            #기존 단가 ID의 유효종료일을 변경유효시작일 -1 일로 수정
            org_id = str(self.lbl_cost_id.text())
            effthru1 = str(self.entry_cost_effthru.text()) #다시 불러와야 함....
            self.cursor.execute('''UPDATE cost SET 
                            pcode=?, costcode=?, class3=?, class4=?, um=?, cost=?, efffrom=?, effthru=?, trxdate=?, userid=?, remark=? WHERE id=?'''
                            , (pcode, costcode, class3, class4, um, cost, efffrom, effthru1, formatted_datetime, user_id, remark, org_id))
            self.conn.commit()            
            
            #변경된 단가 삽입
            self.cursor.execute('''INSERT INTO cost (id, pcode, costcode, class3, class4, um, cost, efffrom, effthru, trxdate, userid, remark) 
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)'''
                            , (idx, pcode, costcode, class3, class4, um, chgcost, chgefffrom, chgeffthru, formatted_datetime, user_id, chgremark))
            
            self.conn.commit()
                       
            self.show_update_success_message()
            self.refresh_data()
            logging.info(f"User {username} updated row number {idx} in the cost table.")
            self.hide_costchange_widget()
        else:
            self.show_missing_message("입력 이상")
            return

    # delete row according to id selected
    def tb_delete(self):
        confirm_dialog = self.show_delete_confirmation_dialog()

        if confirm_dialog == QMessageBox.Yes:
            idx = self.lbl_cost_id.text()
            username, user_id, formatted_datetime = self.common_values_set()            
            self.cursor.execute("DELETE FROM cost WHERE id=?", (idx,))
            self.conn.commit()
            self.show_delete_success_message()
            self.refresh_data()  
            logging.info(f"User {username} deleted {idx} row to the cost table.")            
        else:
            self.show_cancel_message("데이터 삭제 취소")

    # Search data
    def search_data(self):
        class1 = self.cb_cost_class1.currentText()
        class2 = self.cb_cost_class2.currentText()
        class3 = self.cb_cost_class3.currentText()
        class4 = self.cb_cost_class4.currentText()
        org_code = self.entry_cost_code1.text()

        conditions = {'v01': (class1, "class1 like '%{}%'"),
                    'v02': (class2, "class2 like '%{}%'"),
                    'v03': (class3, "class3 like '%{}%'"),
                    'v04': (class4, "class4 like '%{}%'"),
                    'v05': (org_code, "pcode like '%{}%'"),}
        
        selected_conditions = []

        for key, (value, condition_format) in conditions.items():
            if len(value) > 0:
                selected_conditions.append(condition_format.format(value))

        if not selected_conditions:
            QMessageBox.about(self, "검색 조건 확인", "검색 조건이 비어 있습니다!")
            return

        # Join the selected conditions to form the SQL query
        query = f"SELECT id, pcode, costcode, class1, class2, class3, class4, um, cost, efffrom, effthru, remark FROM vw_cost WHERE {' AND '.join(selected_conditions)}"

        QMessageBox.about(self, "검색 조건 확인", f"대분류: {class1} \n 중분류: {class2}\n 소분류:{class3}\n 상세분류:{class4}\n 코드:{org_code} \n\n위 조건으로 검색을 수행합니다!")
        
        sql_query, tv_widget, column_info, column_names, column_widths = self.common_query_statement() 
        self.populate_dialog(self.cursor, query, tv_widget, column_info, column_names,column_widths)

    # Current Price Only
    def current_cost_statement(self):

        today = date.today()
        sdate = today.strftime('%Y-%m-%d')

        class1 = self.cb_cost_class1.currentText()
        class2 = self.cb_cost_class2.currentText()
        class3 = self.cb_cost_class3.currentText()
        class4 = self.cb_cost_class4.currentText()
        org_code = self.entry_cost_code1.text()

        conditions = {
            'v01': (class1, "class1 like '%{}%'"),
            'v02': (class2, "class2 like '%{}%'"),
            'v03': (class3, "class3 like '%{}%'"),
            'v04': (class4, "class4 like '%{}%'"),
            'v05': (org_code, "pcode like '%{}%'"),}
        
        selected_conditions = []

        for key, (value, condition_format) in conditions.items():
            if len(value) > 0:
                selected_conditions.append(condition_format.format(value))
                query = f"SELECT id, pcode, costcode, class1, class2, class3, class4, um, cost, efffrom, effthru, remark FROM vw_cost WHERE {' AND '.join(selected_conditions)}"
                query += f" AND (efffrom <= #{sdate}# AND effthru >= #{sdate}#)"  

        if not selected_conditions:
            QMessageBox.about(self, "검색 조건 확인", "검색 조건이 비어있어 전체 품목의 현재 단가를 출력합니다!")
            query = f"SELECT id, pcode, costcode, class1, class2, class3, class4, um, cost, efffrom, effthru, remark From vw_cost WHERE" 
            query += f"(efffrom <= #{sdate}# AND effthru >= #{sdate}#)"  

        column_widths1 = [80, 100, 100, 100, 100, 100]        
        sql_query, tv_widget, column_info, column_names, column_widths = self.common_query_statement() 

        return query, tv_widget, column_info, column_names, column_widths1       

    def current_cost(self):
        query, tv_widget, column_info, column_names,column_widths = self.current_cost_statement() 
        self.populate_dialog(self.cursor, query, tv_widget, column_info, column_names,column_widths)

    # Combobox index changed
    def cost_class1_changed(self):
        self.entry_cost_code1.clear()
        pass

    def cost_class2_changed(self):
        self.entry_cost_code1.clear()
        selected_item1 = self.cb_cost_class1.currentText()
        selected_item2 = self.cb_cost_class2.currentText()

        if selected_item1 and selected_item2:
            query = f"SELECT DISTINCT pcode From vw_cost WHERE class1 = '{selected_item1}' and class2 ='{selected_item2}'"
            line_edit_widgets = [self.entry_cost_code1]
        
            # Check if any line edit widgets are provided
            if line_edit_widgets:
                self.lineEdit_contents(line_edit_widgets, query)
            else:
                pass

    def cost_class3_changed(self):
        self.entry_cost_code2.clear()
        pass

    def cost_class4_changed(self):
        self.entry_cost_code2.clear()
        selected_item3 = self.cb_cost_class3.currentText()
        selected_item4 = self.cb_cost_class4.currentText()

        if selected_item3 and selected_item4:
            query = f"SELECT DISTINCT costcode From vw_cost WHERE class3 ='{selected_item3}' and class4 = '{selected_item4}'"
            line_edit_widgets = [self.entry_cost_code2]
        
            # Check if any line edit widgets are provided
            if line_edit_widgets:
                self.lineEdit_contents(line_edit_widgets, query)
            else:
                pass

    # while cost change efffrom date, change cost effthru date
    def costchange_efffrom_change(self):
        chg_date_str = self.entry_costchange_efffrom.text()

        try:
            chg_date = parse_date(chg_date_str)                 # 날짜 문자열을 날짜 객체로 변환
            org_date = chg_date - timedelta(days=1)             # 날짜에서 1일을 빼줍니다.
            org_date_str = org_date.strftime('%Y-%m-%d')        # 결과를 문자열로 변환
            self.entry_cost_effthru.setText(org_date_str)       # 변경된 cost effthru 날짜를 표시
        
        except ValueError as e:
            QMessageBox.critical(self, "날짜 형식 오류", str(e)) # 날짜 형식이 잘못된 경우 사용자에게 알림

    # Export data to Excel sheet                
    def export_table(self):
        output_subfolder = "data_list"          # set the output subfoler name
        tv_widget = self.tv_cost                # set the name of table widget
        sheet_name = "cost"                     # set the excel sheet name
        filename = get_file_name(output_subfolder, sheet_name)    # get the file name / export_to_excel 밑에 두면 get_file_name을 두 번 실행시켜서 파일의 번호가 달라지므로 미리 만들어서 filename에 담아둠.
        numeric_columns=[0, 1, 2, 4, 8]                  # set the numerical column index
        export_to_excel(output_subfolder, tv_widget, sheet_name, numeric_columns)
               
        if filename:
            full_file_path = os.path.join(output_subfolder, filename)
            self.excel_formatting(sheet_name, full_file_path)
        else:
            QMessageBox.about(self, "코드 확인", "파일을 찾지 못했습니다!")    

    # Excel sheet formatting
    def excel_formatting(self, sheet_name, full_file_path):
        wb = load_workbook(full_file_path)
        sheet_name = sheet_name
        ws = wb[sheet_name]
        last_row = ws.max_row + 1

        column_widths = [8, 10, 12, 10, 10, 8, 10, 8, 10, 12, 12, 25]        # set the width of each column
        title_font = Font(bold=True, name="Arial", size=10)     # set the font for the first row
        arial_font = Font(name="Arial", size=10)                # set the forn from the second row to max row

        set_column_widths(ws, column_widths)    # reset column widths

        ws.freeze_panes = 'D2'              # freeze panes D2 cell
        ws.auto_filter.ref = ws.dimensions  # apply auto filter
        ws.sheet_view.showGridLines = False # remove gridlines
            
        set_font(ws, [1], range(1, len(column_widths) + 1), title_font)                 # first row font setting
        set_font(ws, range(2, last_row), range(1, len(column_widths) + 1), arial_font)  # from second row to last row font setting
        
        wb.save(full_file_path)
        QMessageBox.about(self, "파일 생성 완료", f"data_list folder에 \n엑셀 파일이 {full_file_path}로 \n생성 되었습니다!")
        
    # clear input field entry
    def clear_data(self):
        self.lbl_cost_id.setText("")
        for line_edit in self.findChildren(QtWidgets.QLineEdit):
            line_edit.clear()

        self.display_eff_date()            
        self.cb_cost_class1.setCurrentIndex(0)
        self.cb_cost_class2.setCurrentIndex(0)
        self.cb_cost_class3.setCurrentIndex(0)
        self.cb_cost_class4.setCurrentIndex(0)

    # table widget cell double click
    def show_selected_data(self, item):
        # Get the row index of the clicked item
        row_index = item.row()

        # Initialize a list to store the cell values
        cell_values = []

        # Loop through the columns and retrieve the text from each cell
        for column_index in range(12):  # 12 columns
            cell_text = self.tv_cost.model().item(row_index, column_index).text()
            cell_values.append(cell_text)                                      

        # Populate the input widgets with the data from the selected row
        self.lbl_cost_id.setText(cell_values[0])
        self.entry_cost_code1.setText(cell_values[1])
        self.entry_cost_code2.setText(cell_values[2])
        self.cb_cost_class1.setCurrentText(cell_values[3])
        self.cb_cost_class2.setCurrentText(cell_values[4])
        self.cb_cost_class3.setCurrentText(cell_values[5])
        self.cb_cost_class4.setCurrentText(cell_values[6])                        
        self.entry_cost_um.setText(cell_values[7])
        self.entry_cost_cost.setText(cell_values[8])
        self.entry_cost_efffrom.setText(cell_values[9])
        self.entry_cost_effthru.setText(cell_values[10])
        self.entry_cost_remark.setText(cell_values[11])
            
    def refresh_data(self):
        self.clear_data()
        self.make_data()

if __name__ == "__main__":

    log_subfolder = "logs"
    os.makedirs(log_subfolder, exist_ok=True)
    log_file_path = os.path.join(log_subfolder, "access_cost.log")

    logging.basicConfig(
        filename=log_file_path,  
        level=logging.INFO,    
        format="%(asctime)s [%(levelname)s] - %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    app = QtWidgets.QApplication(sys.argv)
    dialog = CostDialog()
    dialog.show()
    sys.exit(app.exec())