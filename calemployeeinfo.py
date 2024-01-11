import sys
import logging
import math
from functools import partial
from openpyxl.styles import Font
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from PyQt5 import uic, QtWidgets, QtCore
from PyQt5.QtGui import QStandardItemModel, QKeySequence
from PyQt5.QtWidgets import QMessageBox, QDialog, QShortcut, QMenu
from PyQt5.QtCore import Qt
from commonmd import *

# Calendar Master table contents -----------------------------------------------------
class CalEmployeeInfoDialog(QDialog, SubWindowBase):

    def __init__(self, current_username = None, current_datetime = None):
        super().__init__()

        self.conn, self.cursor = connect_to_database3()    
        uic.loadUi("calemployeeinfo.ui", self)

        # Add window flags
        self.setWindowFlags(self.windowFlags() | Qt.WindowMinimizeButtonHint | Qt.WindowMaximizeButtonHint | Qt.WindowCloseButtonHint)  

        # Initialize current_username and current_datetime directly
        self.current_username, self.current_datetime = initialize_username_and_datetime(current_username, current_datetime)

        # Enable automatic deletion on close
        self.setAttribute(QtCore.Qt.WA_DeleteOnClose)  

        # Create tv_pensionrate and QSortFilterProxyModel
        self.model = QStandardItemModel()
        self.proxy_model = NumericStringSortModel(self.model)
        self.proxy_model.setSourceModel(self.model)

        # Define the column types
        column_types = ["numeric", "numeric", "", "numeric", "", "", "numeric", "", "", "", ""] 

        delegate = NumericDelegate(column_types, self.tv_employeeinfo)
        self.tv_employeeinfo.setItemDelegate(delegate)
        self.tv_employeeinfo.setModel(self.proxy_model)

        # Enable Sorting
        self.tv_employeeinfo.setSortingEnabled(True)

        # Enable alternating row colors
        self.tv_employeeinfo.setAlternatingRowColors(True)  

        # Hide the first index column
        self.tv_employeeinfo.verticalHeader().setVisible(False)

        # While selecting row in tv_widget, each cell values to displayed to designated widgets
        self.tv_employeeinfo.clicked.connect(self.show_selected_employeeinfo_data)

        # Fill combobox items when the application starts
        self.get_combobox_contents()

        # Populate the data
        self.make_data() 
        self.connect_btn_method()
        self.conn_signal_to_slot()

        # Set tab order for input widgets
        self.set_tab_order()

        # Initiate CTRL+C, CTRL+V and ENTER
        self.setup_shortcuts()

        # Make log file
        self.make_logfiles("access_CalEmployeeInfo.log")        

    # Create a shortcut for CTRL+C/CTRL+V/ Return key
    def setup_shortcuts(self):
        self.copy_shortcut = QShortcut(QKeySequence.Copy, self.tv_employeeinfo, partial(self.copy_cells, self.tv_employeeinfo))
        self.paste_shortcut = QShortcut(QKeySequence.Paste, self.tv_employeeinfo, partial(self.paste_cells, self.tv_employeeinfo))
        self.return_shortcut = QShortcut(Qt.Key_Return, self.tv_employeeinfo, partial(self.handle_return_key, self.tv_employeeinfo))

    # Call process_key_event and pass the event and your QTableWidget instance
    def keyPressEvent(self, event):
        tv_widget = self.tv_employeeinfo
        self.process_key_event(event, tv_widget)

    # Pass combobox info and sql to next method
    def get_combobox_contents(self):
        self.insert_combobox_initiate(self.cb_employeeinfo_cname, "SELECT DISTINCT cname FROM company ORDER BY cname")
        self.insert_combobox_initiate(self.cb_employeeinfo_ename, "SELECT DISTINCT ename FROM employee ORDER BY ename")
        self.insert_combobox_initiate(self.cb_employeeinfo_familyno, "SELECT DISTINCT HC FROM taxtable ORDER BY HC")
        self.insert_combobox_initiate(self.cb_employeeinfo_regpension, "SELECT DISTINCT regpension FROM vw_employeeinfo ORDER BY regpension")

    # Initiate Combo_Box 
    def insert_combobox_initiate(self, combo_box, sql_query):
        self.combobox_initializing(combo_box, sql_query) 
        self.cb_employeeinfo_cname.setCurrentIndex(0) 
        self.cb_employeeinfo_ename.setCurrentIndex(0) 
        self.cb_employeeinfo_familyno.setCurrentIndex(0) 
        self.cb_employeeinfo_regpension.setCurrentIndex(0) 

    # Connect button to method
    def connect_btn_method(self):
        self.pb_employeeinfo_show.clicked.connect(self.make_data)
        self.pb_employeeinfo_clear.clicked.connect(self.clear_data)
        self.pb_employeeinfo_excel_export.clicked.connect(self.export_data)
        self.pb_employeeinfo_close.clicked.connect(self.close_dialog)
        self.pb_employeeinfo_search.clicked.connect(self.search_data)

        self.pb_employeeinfo_insert.clicked.connect(self.tv_insert)
        self.pb_employeeinfo_update.clicked.connect(self.tv_update)
        self.pb_employeeinfo_delete.clicked.connect(self.tv_delete)
      
    # Connect signal to method    
    def conn_signal_to_slot(self):
        #pass
        #self.entry_employeeinfo_.editingFinished.connect()
        self.cb_employeeinfo_ename.activated.connect(self.employeeinfo_ename_changed)
        self.cb_employeeinfo_cname.activated.connect(self.cb_employeeinfo_cname_changed)

    # Base Salary change
    def base_salary_changed(self):
        self.entry_pension_itvalue.clear()
        self.entry_pension_rtvalue.clear()

        selected_text = self.entry_pension_taxedsalary.text()
        selected_item = self.cb_pension_dedhc.currentText()

        if selected_item and selected_text:
            query = f"SELECT * FROM taxtable WHERE GE <= {selected_text} AND HC = {selected_item} ORDER BY id DESC"
            try:
                self.cursor.execute(query)

                result = self.cursor.fetchone()
                if result:
                    item01 = str(result[4])
                    self.entry_pension_itvalue.setText(item01)
            except Exception as e:
                print(f"Error executing SQL query: {e}")

        itv = math.floor((float(self.entry_pension_itvalue.text())*0.1)/10) * 10
        self.entry_pension_rtvalue.setText(str(itv))
        
        # Calculate National Pension-------------------------------------------------------------------
        # 1. Calculate the age
        mybirthday = self.entry_pension_birthdate.text()
        # Convert the birthdate string to a datetime object
        birthdate = datetime.strptime(mybirthday, "%Y/%m/%d")
        # Get the current date
        today = datetime.now()
        # Calculate the age
        age = today.year - birthdate.year - ((today.month, today.day) < (birthdate.month, birthdate.day))
        # Check if the person is 60 years old or older
        is_60_or_older = age >= 60

        # 2. Calculate National Pension Value and Display
        if is_60_or_older:
            self.entry_pension_npvalue.setText("0")
        else:
            np = math.floor(float(self.entry_pension_taxedsalary.text()) * float(self.entry_pension_nprate.text())/10)*10
            self.entry_pension_npvalue.setText(str(np))

        # Calculate Health Insurance-------------------------------------------------------------------
        hi = math.floor(float(self.entry_pension_taxedsalary.text()) * float(self.entry_pension_hcrate.text())/10)*10
        self.entry_pension_hcvalue.setText(str(hi))

        # 3. Calculate Long Term Care-------------------------------------------------------------------
        ltc = math.floor(hi * float(self.entry_pension_ltcrate.text()) / 10) * 10
        self.entry_pension_ltcvalue.setText(str(ltc))

        # 4. Calculate Long Term Care-------------------------------------------------------------------
        ei = math.floor(float(self.entry_pension_taxedsalary.text()) * float(self.entry_pension_eirate.text())/10)*10
        self.entry_pension_eivalue.setText(str(ei))

    # tab order for calmaster window
    def set_tab_order(self):
        widgets = [self.cb_employeeinfo_cname, self.cb_employeeinfo_ename, self.entry_employeeinfo_regid, 
            self.cb_employeeinfo_familyno, self.cb_employeeinfo_regpension, self.entry_employeeinfo_joindt, 
            self.entry_employeeinfo_exitdt, self.entry_employeeinfo_remark, ]
        
        for i in range(len(widgets) - 1):
            self.setTabOrder(widgets[i], widgets[i + 1])

    def query_statement(self):
        tv_widget = self.tv_employeeinfo
        
        self.cursor.execute("SELECT id, ccode, cname, ecode, ename, regid, nofamily, regpension, joindt, exitdt, remark FROM vw_employeeinfo WHERE 1=0")
        column_info = self.cursor.description
        column_names = [col[0] for col in column_info]

        sql_query = "Select id, ccode, cname, ecode, ename, regid, nofamily, regpension, joindt, exitdt, remark from vw_employeeinfo order By ename"
        column_widths = [80, 100, 120, 100, 100, 150, 80, 80, 100, 100, 150]

        return sql_query, tv_widget, column_info, column_names, column_widths

    # To reduce duplications
    def make_data(self):
        sql_query, tv_widget, column_info, column_names, column_widths = self.query_statement() 
        self.populate_dialog(self.cursor, sql_query, tv_widget, column_info, column_names,column_widths)

    # Get the value of other variables
    def get_employeeinfo_input(self):
        ccode = int(self.entry_employeeinfo_ccode.text())
        ecode = int(self.entry_employeeinfo_ecode.text())
        regid = str(self.entry_employeeinfo_regid.text())
        nofamily = int(self.cb_employeeinfo_familyno.currentText())
        regpension = str(self.cb_employeeinfo_regpension.currentText())
        joindt = str(self.entry_employeeinfo_joindt.text())
        exitdt = str(self.entry_employeeinfo_exitdt.text())
        remark = str(self.entry_employeeinfo_remark.text())

        return ccode, ecode, regid, nofamily, regpension, joindt, exitdt, remark

    # Make Common values set
    def common_values_set(self):
        username = self.current_username
        user_id = self.userID_gen(username)
        formatted_datetime = self.dt_time_info()
        return username, user_id, formatted_datetime

    # insert new employeeinfo data to MySQL table
    def tv_insert(self):
        currentid = self.lbl_employeeinfo_id.text()
        if not currentid:
            confirm_dialog = self.show_insert_confirmation_dialog()
            
            if confirm_dialog == QMessageBox.Yes:

                idx = self.max_row_id("employeeinfo")                                                                  # Get the max id 
                username, user_id, formatted_datetime = self.common_values_set()
                ccode, ecode, regid, nofamily, regpension, joindt, exitdt, remark = self.get_employeeinfo_input()           # Get the value of other variables

                if (idx>0 and ccode>0 and ecode>0 and nofamily>0) and all(len(var) > 0 for var in (regid, regpension, joindt,)):
                    self.cursor.execute('''INSERT INTO employeeinfo (id, ccode, ecode, regid, nofamily, regpension, joindt, exitdt, trxdate, userid, remark) 
                                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)'''
                                , (idx, ccode, ecode, regid, nofamily, regpension, joindt, exitdt, formatted_datetime, user_id, remark))
                    self.conn.commit()

                    self.show_insert_success_message()
                    self.refresh_data() 
                    logging.info(f"User {username} inserted {idx} row to the employeeinfo table.")
                else:
                    self.show_missing_message("입력 이상")
                    return
            else:
                self.show_cancel_message("데이터 추가 취소")
                return
        else:
            QMessageBox.information(self, "Input Error", "기존 id가 선택된 상태에서는 신규 입력이 불가합니다!")
            return

    # revise the values in the selected row
    def tv_update(self):
        confirm_dialog = self.show_update_confirmation_dialog()

        if confirm_dialog == QMessageBox.Yes:
            idx = int(self.lbl_employeeinfo_id.text())
            username, user_id, formatted_datetime = self.common_values_set()
            ccode, ecode, regid, nofamily, regpension, joindt, exitdt, remark = self.get_employeeinfo_input() 

            if (idx>0 and ccode>0 and ecode>0 and nofamily>0) and all(len(var) > 0 for var in (regid, regpension, joindt,)):
                self.cursor.execute('''UPDATE employeeinfo SET ccode=?, ecode=?, regid=?, nofamily=?, regpension=?, joindt=?, exitdt=?, trxdate=?, userid=?, remark=? WHERE id=?'''
                            , (ccode, ecode, regid, nofamily, regpension, joindt, exitdt, formatted_datetime, user_id, remark, idx))
                self.conn.commit()
                self.show_update_success_message()
                self.refresh_data()
                logging.info(f"User {username} updated row number {idx} in the employeeinfo table.")
            else:
                self.show_missing_message("입력 이상")
                return
        else:
            self.show_cancel_message("데이터 변경 취소")
            return

    # delete row according to id selected
    def tv_delete(self):
        confirm_dialog = self.show_delete_confirmation_dialog()

        if confirm_dialog == QMessageBox.Yes:
            idx = self.lbl_employeeinfo_id.text()
            username, user_id, formatted_datetime = self.common_values_set()
            self.cursor.execute("DELETE FROM employeeinfo WHERE id=?", (idx,))
            self.conn.commit()
            self.show_delete_success_message()
            self.refresh_data()
            logging.info(f"User {username} deleted {idx} row to the employeeinfo table.")
        else:
            self.show_cancel_message("데이터 삭제 취소")
            return

    # clear input field entry
    def clear_data(self):
        for line_edit in self.findChildren(QtWidgets.QLineEdit):
            line_edit.clear()
        self.lbl_employeeinfo_id.setText("")
        self.cb_employeeinfo_cname.setCurrentIndex(0) 
        self.cb_employeeinfo_ename.setCurrentIndex(0) 
        self.cb_employeeinfo_familyno.setCurrentIndex(0) 
        self.cb_employeeinfo_regpension.setCurrentIndex(0) 

    # Combobox employeeinfo index changed
    def employeeinfo_ename_changed(self):
        self.entry_employeeinfo_ecode.clear()
        self.entry_employeeinfo_regid.clear()
        selected_item = self.cb_employeeinfo_ename.currentText()

        if selected_item:
            
            qry = f"SELECT COUNT(*) FROM vw_employeeinfo WHERE ename ='{selected_item}'"
            cnt = self.cursor.execute(qry).fetchone()[0]

            if cnt != 0:
                query = f"SELECT DISTINCT ecode, regid FROM vw_employeeinfo WHERE ename ='{selected_item}'"
            else:
                query = f"SELECT ecode, regid FROM residenceno WHERE ename ='{selected_item}'"

        line_edit_widgets = [self.entry_employeeinfo_ecode, self.entry_employeeinfo_regid]
        self.lineEdit_contents(line_edit_widgets, query)


    def cb_employeeinfo_cname_changed(self):
        self.entry_employeeinfo_ccode.clear()

        selected_item = self.cb_employeeinfo_cname.currentText()

        if selected_item:
            query = f"SELECT DISTINCT ccode From company WHERE cname ='{selected_item}'"
            line_edit_widgets = [self.entry_employeeinfo_ccode, ]
        
            # Check if any line edit widgets are provided
            if line_edit_widgets:
                self.lineEdit_contents(line_edit_widgets, query)
            else:
                pass
 
     # Search data
    def search_data(self):
        cname = self.cb_employeeinfo_cname.currentText()
        ename = self.cb_employeeinfo_ename.currentText()
        regp = self.cb_employeeinfo_regpension.currentText()

        conditions = {
                    'v01': (cname, "cname like '%{}%'"),
                    'v02': (ename, "ename like '%{}%'"),
                    'v03': (regp, "regpension like '%{}%'"),}
        
        selected_conditions = []

        for key, (value, condition_format) in conditions.items():
            if len(value) > 0:
                selected_conditions.append(condition_format.format(value))

        if not selected_conditions:
            QMessageBox.about(self, "검색 조건 확인", "검색 조건이 비어 있습니다!")
            return

        # Join the selected conditions to form the SQL query
        query = f"SELECT id, ccode, cname, ecode, ename, regid, nofamily, regpension, joindt, exitdt, remark from vw_employeeinfo WHERE {' AND '.join(selected_conditions)} ORDER BY ename"

        QMessageBox.about(self, "검색 조건 확인", f"거래처명: {cname}\n 직원이름: {ename}\n 4대보험가입여부: {regp} \n\n위 조건으로 검색을 수행합니다!")
        
        sql_query, tv_widget, column_info, column_names, column_widths = self.query_statement()
        self.populate_dialog(self.cursor, query, tv_widget, column_info, column_names, column_widths)


    # table view click
    def show_selected_employeeinfo_data(self, item):
        # Get the row index of the clicked item
        row_index = item.row()

        # Initialize a list to store the cell values
        cell_values = []

        # Loop through the columns and retrieve the text from each cell
        for column_index in range(11): 
            cell_text = self.tv_employeeinfo.model().item(row_index, column_index).text()
            cell_values.append(cell_text)
            
        # Populate the input widgets with the data from the selected row
        self.lbl_employeeinfo_id.setText(cell_values[0])
        self.entry_employeeinfo_ccode.setText(cell_values[1])
        self.cb_employeeinfo_cname.setCurrentText(cell_values[2])
        self.entry_employeeinfo_ecode.setText(cell_values[3])
        self.cb_employeeinfo_ename.setCurrentText(cell_values[4])
        self.entry_employeeinfo_regid.setText(cell_values[5])
        self.cb_employeeinfo_familyno.setCurrentText(cell_values[6])
        self.cb_employeeinfo_regpension.setCurrentText(cell_values[7])
        self.entry_employeeinfo_joindt.setText(cell_values[8])
        self.entry_employeeinfo_exitdt.setText(cell_values[9])
        self.entry_employeeinfo_remark.setText(cell_values[10])



    # 선택된 각 위젯의 내용을 엑셀로 내보내기
    def export_data(self):
        output_subfolder = "data_list"                # set the output subfoler name
        sheet_name = "employeeinfo"                   # set the excel sheet name
        filename = get_file_name(output_subfolder, sheet_name)    # get the file name

        # Ensure the subfolder exists; create it if it doesn't
        os.makedirs(output_subfolder, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        if filename:
            full_file_path = os.path.join(output_subfolder, filename)
    
            # 각 위젯에서 내용 가져와서 엑셀에 쓰기
            data_to_export = [
                (self.label.text(), self.lbl_employeeinfo_id.text()),
                (self.label_3.text(), self.entry_employeeinfo_ccode.text()),
                (self.label_4.text(), self.cb_employeeinfo_cname.currentText()),
                (self.label_10.text(), self.entry_employeeinfo_ecode.text()),
                (self.label_8.text(), self.cb_employeeinfo_ename.currentText()),
                (self.label_9.text(), self.entry_employeeinfo_regid.text()),
                (self.label_11.text(), self.cb_employeeinfo_familyno.currentText()),
                (self.label_5.text(), self.cb_employeeinfo_regpension.currentText()),
                (self.label_12.text(), self.entry_employeeinfo_joindt.text()),
                (self.label_13.text(), self.entry_employeeinfo_exitdt.text()),
                (self.label_7.text(), self.entry_employeeinfo_remark.text()),
                #(self..text() , self..text() ),
                #(self..text() , self..text() ),
                #(self..text() , self..text() ),
                #(self..text() , self..text() ),
                ]

            #for i, value in enumerate(data_to_export, start=1):
            #    ws.cell(row=i, column=1, value=value)
            
            for i, (label, value) in enumerate(data_to_export, start=1):
                ws.cell(row=i, column=1, value=label)
                ws.cell(row=i, column=2, value=value)
            
            # 엑셀 파일 저장
            wb.save(full_file_path)
        
            self.excel_formatting(sheet_name, full_file_path)

    # Excel sheet formatting
    def excel_formatting(self, sheet_name, full_file_path):
        wb = load_workbook(full_file_path)
        sheet_name = sheet_name
        ws = wb[sheet_name]
        last_row = ws.max_row + 2

        # Insert headers at A1 and B1
        ws.insert_rows(1)
        ws['A1'] = '구분'
        ws['B1'] = '내용'

        column_widths = [20, 15]                                # set the width of each column
        title_font = Font(bold=True, name="Arial", size=10)     # set the font for the first row
        arial_font = Font(name="Arial", size=10)                # set the forn from the second row to max row

        set_column_widths(ws, column_widths)    # reset column widths

        ws.freeze_panes = 'B2'              # freeze panes D2 cell
        ws.auto_filter.ref = ws.dimensions  # apply auto filter
        ws.sheet_view.showGridLines = False # remove gridlines
            
        set_font(ws, [1], range(1, len(column_widths) + 1), title_font)                 # first row font setting
        set_font(ws, range(2, last_row), range(1, len(column_widths) + 1), arial_font)  # from second row to last row font setting

        wb.save(full_file_path)
        QMessageBox.about(self, "파일 생성 완료", f"data_list folder에 \n엑셀 파일이 {full_file_path}로 \n생성 되었습니다!")  

    # Refresh data
    def refresh_data(self):
        self.clear_data()
        self.make_data()       

if __name__ == "__main__":
    
    log_subfolder = "logs"
    os.makedirs(log_subfolder, exist_ok=True)
    log_file_path = os.path.join(log_subfolder, "access_CalEmployeeInfo.log")

    logging.basicConfig(
        filename=log_file_path,  
        level=logging.INFO,    
        format="%(asctime)s [%(levelname)s] - %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
    app = QtWidgets.QApplication(sys.argv)
    dialog = CalEmployeeInfoDialog()
    dialog.show()
    sys.exit(app.exec())
