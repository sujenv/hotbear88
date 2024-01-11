import os
import re
import pyodbc
import calendar
import logging
import inspect
from datetime import datetime, timedelta, date
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle
from PyQt5 import QtWidgets
from PyQt5.QtGui import QColor, QFont, QStandardItem, QStandardItemModel, QKeySequence
from PyQt5.QtCore import Qt, QSortFilterProxyModel, QVariant
from PyQt5.QtWidgets import QMessageBox, QApplication, QTableWidgetItem, QMessageBox, QStyledItemDelegate, QMdiSubWindow

# Set the working directory to the folder where your script is located
script_folder = os.path.dirname(os.path.abspath(__file__))
os.chdir(script_folder)

# Relative path to the 'dbs' folder
relative_dbs_folder = 'dbs'

# build the connection to MS Access database
db_driver = '{Microsoft Access Driver (*.mdb, *.accdb)}'

# for the first db connection
def connect_to_database1():
    filename = 'araps.accdb'
    db_path = os.path.join(relative_dbs_folder, filename)
    conn = pyodbc.connect(rf'DRIVER={db_driver};' rf'DBQ={db_path};')
    cursor = conn.cursor()
    return conn, cursor

# for the second db connection
def connect_to_database2():
    filename = 'consumables.accdb'
    db_path = os.path.join(relative_dbs_folder, filename)
    password = 'tnwp153700!'
    conn = pyodbc.connect(rf'DRIVER={db_driver};' rf'DBQ={db_path};' rf'PWD={password};')
    cursor = conn.cursor()    
       
    return conn, cursor

# for third db connection
def connect_to_database3():
    filename = 'payroll.accdb'
    db_path = os.path.join(relative_dbs_folder, filename)
    conn = pyodbc.connect(rf'DRIVER={db_driver};' rf'DBQ={db_path};')
    cursor = conn.cursor()
    return conn, cursor

# for third db connection
def connect_to_database4():
    filename = 'aptdb.accdb'
    db_path = os.path.join(relative_dbs_folder, filename)
    conn = pyodbc.connect(rf'DRIVER={db_driver};' rf'DBQ={db_path};')
    cursor = conn.cursor()
    return conn, cursor

# Set first day of month 
def get_first_day_of_month(year, month):
    return date(year, month, 1)

def get_last_day_of_month(year, month):
    _, last_day = calendar.monthrange(year, month)
    return date(year, month, last_day)  

class NumericDelegate(QStyledItemDelegate):
    def __init__(self, column_type, parent=None):
        super().__init__(parent)
        self.column_type = column_type

    def initStyleOption(self, option, index):
        super().initStyleOption(option, index)
        # Access the column_type attribute here
        if self.column_type[index.column()] == "numeric":
            value = index.data(Qt.DisplayRole)
            if isinstance(value, (int, float)):
                # Format the numeric value without scientific notation
                #option.text = f'{value:.0f}'
                option.text = '{:.0f}'.format(value)

# NumericStringSortModel class for custom sorting
class NumericStringSortModel(QSortFilterProxyModel):
    def lessThan(self, left, right):
        left_data = left.data().strip()
        right_data = right.data().strip()
    
        # Compare numeric and string values
        try:
            left_data_numeric = float(left_data)
            right_data_numeric = float(right_data)
            return left_data_numeric < right_data_numeric
        except ValueError:
            # If conversion to float fails, compare as strings
            return left_data < right_data
        
class SubWindowBase:

    # populate the dialong as sub window of the main window
    def make_dialog(self, dialog_class):
        
        # Instantiate the specific dialog
        dialog = dialog_class

        sub_window = QMdiSubWindow()
        sub_window.setWidget(dialog)

        self.mdi_area.addSubWindow(sub_window)
        sub_window.show()
        
        # added to avoid duplication of sub window
        self.open_sub_windows.append(sub_window)
    # Display current date only
    def display_trx_date(self):
        now = datetime.now()
        curr_date = now.strftime("%Y/%m/%d")
        ddt = f"{curr_date}"         
        return ddt

    # For First DB Get the user_id thru execution of SQL
    def userID_gen(self, username):
        self.cursor.execute('''SELECT ecode FROM employee WHERE ename = ?''', (username,))
        result = self.cursor.fetchone()
        uid = result[0] if result is not None else None
        return uid

    # Check maximum id number to create new row number
    def max_row_id(self, tbname):
        self.cursor.execute(f"SELECT MAX(id) FROM {tbname}")
        row = self.cursor.fetchone()
        max_id = row[0]
        
        # Check if the table is empty
        if max_id is None:
            idx = 1
        else:
            # Increment the maximum ID to get the next available ID
            idx = max_id + 1

        return idx
    
    # Get the maximum column number from the table
    def max_col_id(self, tbname):
        self.cursor.execute(f"SELECT * FROM {tbname} WHERE 1=0")
        column_info = self.cursor.description
        max_col = len(column_info)
        col_count = max_col
    
        return col_count
    
    # Get the transaction date time info
    def dt_time_info(self):
        # Update self.current_datetime to the current time
        self.current_datetime = datetime.now()
        cdatetime_str = self.current_datetime

        if isinstance(cdatetime_str, str):
            # Define the regular expression pattern to match Korean characters
            korean_pattern = r'(\d+)년 (\d+)월 (\d+)일 (\S+) (\S+) (\d+:\d+:\d+)'

            match = re.match(korean_pattern, cdatetime_str)

            if match:
                year, month, day, day_of_week, am_pm, time = match.groups()
                hour, minute, second = map(int, time.split(':'))
                
                if am_pm == '오후':
                    # Convert PM hours to 24-hour format
                    hour = (hour % 12) + 12
                
                formatted_datetime = datetime(int(year), int(month), int(day), hour, minute, second)
                fdt = formatted_datetime.strftime('%Y-%m-%d %p %I:%M:%S')
            else:
                fdt = "Datetime format not recognized"
                
        elif isinstance(cdatetime_str, datetime):
            # If it's already a datetime object, keep it as is
            cdatetime = cdatetime_str
            fdt = datetime.strftime(cdatetime, '%Y-%m-%d %p %I:%M:%S')

        return fdt
    
    # Combobox initializing
    def combobox_initializing(self, combo_box, sql_query, params=None):
        combo_box.addItem("")  # Add a blank item as the first option
        if params:
            self.cursor.execute(sql_query, params)
        else:
            self.cursor.execute(sql_query)
        for row in self.cursor.fetchall():
            combo_box.addItem(str(row[0]))
    
    # Set combobox item index as zero
    def clear_combobox_selections(self, combo_box):
        for combo_box in self.findChildren(QtWidgets.QComboBox):
            if combo_box.currentIndex() == 0 and combo_box.currentText() != "":
                combo_box.insertItem(0, "")  # Add an empty string at the beginning
            combo_box.setCurrentIndex(0)

    # For Multiple QLineEdit contents display
    def lineEdit_contents(self, line_edit_widgets, sql_query):
        num_widgets = len(line_edit_widgets)

        for index, line_edit in enumerate(line_edit_widgets):
            #line_edit.clear()
            self.cursor.execute(sql_query)
            result = self.cursor.fetchone()

            if result:
                if num_widgets == 1:
                    item01 = str(result[0])
                    line_edit.setText(item01)
                elif num_widgets == 2:
                    item01 = str(result[0])
                    item02 = str(result[1])
                    if index == 0:
                        line_edit.setText(item01)
                    elif index == 1:
                        line_edit.setText(item02)
                elif num_widgets == 3:
                    item01 = str(result[0])
                    item02 = str(result[1])
                    item03 = str(result[2])
                    if index == 0:
                        line_edit.setText(item01)
                    elif index == 1:
                        line_edit.setText(item02)
                    elif index == 2:
                        line_edit.setText(item03)
                elif num_widgets == 4:
                    item01 = str(result[0])
                    item02 = str(result[1])
                    item03 = str(result[2])
                    item04 = str(result[3])
                    if index == 0:
                        line_edit.setText(item01)
                    elif index == 1:
                        line_edit.setText(item02)
                    elif index == 2:
                        line_edit.setText(item03)
                    elif index == 3:
                        line_edit.setText(item04)
            else:
                line_edit.setText("")

    # For Multiple combo box contents display
    def insert_combobox_contents_changed(self, combobox_widgets, sql_query):
        num_widgets = len(combobox_widgets)
        for index, combo_box in enumerate(combobox_widgets):
            
            combo_box.clear() # Clear existing items
            combo_box.addItem("")  # Add a blank item as the first option

            self.cursor.execute(sql_query)
            items = self.cursor.fetchall()

            if items:
                if num_widgets == 1:
                    combo_box.addItems([str(item[0]) for item in items]) 
                    combo_box.setCurrentText(str(items[0][0]))
                elif num_widgets == 2:
                    if index == 0:
                        combo_box.addItems([str(item[0]) for item in items])
                        combo_box.setCurrentText(str(items[0][0]))
                    elif index == 1:
                        combo_box.addItems([str(item[1]) for item in items])
                        combo_box.setCurrentText(str(items[0][1]))
            else:
                print(num_widgets)  

    # for initializing first db table data
    def populate_dialog(self, cursor, sql_query, tv_widget, column_info, column_names, column_widths):
        cursor.execute(sql_query)
        rows = cursor.fetchall()
        columns = [column[0] for column in cursor.description] # Get the column names from the database cursor

        self.populate_dialog_common(tv_widget, column_info, column_names,column_widths, columns, rows)

    def populate_dialog_common(self, tv_widget, column_info, column_names,column_widths, columns, rows):
        # Create a new instance of QStandardItemModel
        new_model = QStandardItemModel()

        # Set column headers
        new_model.setHorizontalHeaderLabels(column_names)

        for row_index, row_data in enumerate(rows):
            row_items = []
            for col_index, col_data in enumerate(row_data):
                item = QStandardItem()

                # Check if column type is numeric
                if column_info[col_index] == "numeric":
                    try:
                        numeric_value = float(col_data)
                        item.setData(numeric_value, Qt.DisplayRole)
                    except ValueError:
                        # If conversion to float fails, set the original string value
                        item.setData(col_data, Qt.DisplayRole)

                else:
                    # Handle other types (e.g., datetime, empty string, etc.)
                    if col_data is None:
                        item.setData("", Qt.DisplayRole)
                    elif isinstance(col_data, datetime):
                        formatted_date = col_data.strftime("%Y-%m-%d")
                        item.setData(formatted_date, Qt.DisplayRole)
                    else:
                        item.setData(col_data, Qt.DisplayRole)

                item.setTextAlignment(Qt.AlignCenter)
                row_items.append(item)

            # Set the background color of each cell to gray for alternating rows
            if row_index % 2 == 0:
                for item in row_items:
                    item.setBackground(QColor(255, 255, 255))

            # Append each row to the model
            new_model.appendRow(row_items)
        
        # Set the new model to the table view
        tv_widget.setModel(new_model)

        # Resize columns to fit their content
        #self.ui.tv_calmaster.resizeColumnsToContents()

        # Set column widths for specific columns (modify the sizes as needed)
        for col_index, width in enumerate(column_widths):
            tv_widget.setColumnWidth(col_index, width)

        # Set column header background color and font
        for col_index, col_name in enumerate(column_names):
            header_item = QStandardItem(col_name)

            # Set the header data on the proxy model
            self.proxy_model.setHeaderData(col_index, Qt.Horizontal, header_item)

        # Directly set the header properties on the table view
        header_view = tv_widget.horizontalHeader()

        # Set the background color for the header of each column individually
        tv_widget.horizontalHeader().setStyleSheet("QHeaderView::section { background-color: rgb(220, 220, 220); }")

        # Set the font for the header
        header_font = QFont()
        header_font.setBold(True)
        header_view.setFont(header_font)    

    # to inform the close of the dialong to user
    def close_dialog(self):
        confirm_dialog = QMessageBox.question(
            self, "Confirm Close", "Are you sure you want to close this operation?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        
        if confirm_dialog == QMessageBox.Yes:
            # Delay for a short moment (optional) to allow the user to see the message
            parent = self.parentWidget()
            if parent is not None:
                parent.close()
            else:
                # Handle the case where the parent widget is None (e.g., no parent)
                self.close()            
        else:
            pass

    def process_key_event(self, event, tv_widget):
        try:
            if event.matches(QKeySequence.Copy):
                self.copy_cells(tv_widget)
            elif event.matches(QKeySequence.Paste):
            #if event.matches(QKeySequence.Paste):
                self.paste_cells(tv_widget)
            elif event.key() in [Qt.Key_Return, Qt.Key_Enter]:
                self.handle_return_key(tv_widget)
            else:
                # Handle other key events or remove this block if not need
                pass
        
        except IndexError as e:
            self.handle_error(f"IndexError: {e}")
        except Exception as e:
            self.handle_error(f"An unexpected error occurred: {e}")

    def copy_cells(self, tv_widget):
        selected_indexes = sorted(tv_widget.selectedIndexes(), key=lambda index: index.row())

        # Prepare a list to store the copied values
        copied_data = []

        # Track the current row to detect when we move to a new row
        current_row = selected_indexes[0].row()

        for index in selected_indexes:
            if index.row() != current_row:
                # Move to a new row, add a newline character to separate rows
                copied_data.append('\n')
                current_row = index.row()

            value = tv_widget.model().itemFromIndex(index).text()
            copied_data.append(value + '\t')  # Separate cells with a tab

        # Join the values into a single string
        clipboard_data = ''.join(copied_data)

        # Set the clipboard data
        clipboard = QApplication.clipboard()
        clipboard.setText(clipboard_data)

        return copied_data

    def paste_cells(self, tv_widget):
        #print("Pasting cells...")
        clipboard = QApplication.clipboard()
        mime_data = clipboard.mimeData()

        if mime_data.hasText():
            clipboard_data = mime_data.text()
            #print("Clipboard Data:", repr(clipboard_data))  # Add this line for debugging
            lines = clipboard_data.split('\n')
            
            for i, line in enumerate(lines):
                if not line.strip('\n'):
                    continue  # Skip empty lines

                columns = line.split('\t')
                for j, text in enumerate(columns):
                    item = QStandardItem(text)
                    self.model.setItem(tv_widget.currentRow() + i, tv_widget.currentColumn() + j, item)
        
    def handle_return_key(self, tv_widget):
        selection_model = tv_widget.selectionModel()
        current_indexes = selection_model.selectedIndexes()

        if not current_indexes:
            # If no cells are selected, do nothing
            return

        current_row = current_indexes[0].row()
        current_column = current_indexes[0].column()
        next_row = current_row + 1

        if next_row < tv_widget.model().rowCount():
            tv_widget.setCurrentIndex(tv_widget.model().index(next_row, current_column))
        else:
            # If you want to wrap around to the first row when reaching the last row
            tv_widget.setCurrentIndex(tv_widget.model().index(0, current_column))

    def handle_error(self, message):
        QMessageBox.critical(self, "Error", message)

    def show_insert_confirmation_dialog(self):
        return QMessageBox.question(
            self, "새로운 데이터 추가 확인", "새로운 데이터를 추가하시겠습니까?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No)

    def show_insert_success_message(self):
        QMessageBox.about(self, "새로운 데이터 등록 완료", "새로 추가한 데이터가 등록 완료되었습니다.")    

    def show_update_confirmation_dialog(self):
        return QMessageBox.question(
            self, "기존 데이터 수정", "입력하신 내용으로 기존 데이터를 수정하시겠습니까?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No)

    def show_update_success_message(self):
        QMessageBox.about(self, "데이터 변경 완료", "선택하신 데이터가 변경 완료되었습니다.")    

    def show_delete_confirmation_dialog(self):
        return QMessageBox.question(
            self, "삭제 확인", "선택하신 내용을 삭제하시겠습니까?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No)

    def show_delete_success_message(self):
        QMessageBox.about(self, "데이터 삭제 완료", "선택하신 데이터가 삭제되었습니다.")    

    def show_closing_confirmation_dialog(self):
        return QMessageBox.question(self, "마감 확인", "Are you sure you want to finalize closing work?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
    
    def show_missing_message(self, message):
        QMessageBox.about(self, message, "빠진 데이터가 없는지 확인해 주십시오!")

    def show_missing_message_update(self, message):
        QMessageBox.about(self, message, "데이터를 먼저 선택해 주세요! \n")

    def show_cancel_message(self, message):
        QMessageBox.about(self, message, "데이터 변경 작업이 취소되었습니다.")

    # Make log file
    def make_logfiles(self, log_file_name):
        log_subfolder = "logs"
        os.makedirs(log_subfolder, exist_ok=True)
        log_file_path = os.path.join(log_subfolder, log_file_name)

        logging.basicConfig(
            filename=log_file_path,
            level=logging.INFO,
            format="%(asctime)s [%(levelname)s] - %(message)s",
            datefmt="%Y-%m-%d %H:%M:%S",
        )        

    # DB update
    # Move employee and customer table contents from araps to consumables ----------------
    def connect_to_db(self, db_path, pwd=None):
        conn_str = f'DRIVER={{Microsoft Access Driver (*.mdb, *.accdb)}};DBQ={db_path};'
        if pwd:
            conn_str += f'PWD={pwd};'
        return pyodbc.connect(conn_str)

    def close_connection(self, cursor, connection):
        cursor.close()
        connection.close()

    def move_data(self, source_cursor, target_cursor, source_table, target_table, columns=None, conditions=None):
        try:
            # Delete existing data in the target table
            target_cursor.execute(f'DELETE FROM {target_table}')

            # Define columns to select
            columns_str = '*' if columns is None else ', '.join(columns)

            # Retrieve data from the source table
            query = f'SELECT {columns_str} FROM {source_table}'
            if conditions:
                query += f' WHERE {conditions}'
            source_cursor.execute(query)
            rows = source_cursor.fetchall()

            # Insert data into the target table
            for row in rows:
                placeholders = ', '.join(['?' for _ in row])
                target_cursor.execute(f'INSERT INTO {target_table} VALUES ({placeholders})', row)

            # Commit the transaction
            target_cursor.commit()

        except Exception as e:
            print(f"Error moving data: {e}")

    def ini_move_data(self):
        self.relative_dbs_folder = 'dbs'
        self.a_db_path = os.path.join(self.relative_dbs_folder, 'araps.accdb')
        self.b_db_path = os.path.join(self.relative_dbs_folder, 'consumables.accdb')
        self.c_db_path = os.path.join(self.relative_dbs_folder, 'aptdb.accdb')
        self.d_db_path = os.path.join(self.relative_dbs_folder, 'payroll.accdb')
        
        try:
            with self.connect_to_db(self.a_db_path) as conn_a, \
                    self.connect_to_db(self.b_db_path, pwd='tnwp153700!') as conn_b, \
                    self.connect_to_db(self.c_db_path) as conn_c, \
                    self.connect_to_db(self.d_db_path) as conn_d:
                
                cursor_a = conn_a.cursor()
                cursor_b = conn_b.cursor()
                cursor_c = conn_c.cursor()
                cursor_d = conn_d.cursor()

                # Move data from customer table in a_db to customer table in b_db
                self.move_data(cursor_a, cursor_b, 'customer', 'customer')
                # Move data from employee table in a_db to employee table in b_db
                self.move_data(cursor_a, cursor_b, 'employee', 'employee')

                # Move data from employee table in a_db to employee table in c_db
                self.move_data(cursor_a, cursor_c, 'employee', 'employee')
                # Move data from customer table in a_db to customer table in c_db
                self.move_data(cursor_a, cursor_c, 'customer', 'customer', columns=['id', 'ccode', 'cname'], conditions="remark like '%헌옷%'")                

                # Move data from employee table in a_db to employee table in d_db
                self.move_data(cursor_a, cursor_d, 'customer', 'company', columns=['id', 'ccode', 'cname'], conditions="type01='s'")
                self.move_data(cursor_a, cursor_d, 'employee', 'employee', columns=['id', 'ecode', 'ename', 'class1', 'class2', 'area', 'remark'])
                self.move_data(cursor_a, cursor_d, 'vw_regno', 'residenceno', columns=['id', 'ecode', 'ename', 'regid'])
                self.move_data(cursor_a, cursor_d, 'vw_emp_car', 'employeecar', columns=['id', 'cpno1', 'cpno2', 'ecode', 'class1', 'efffrom', 'effthru'])

                # Move data from aptdb table in c_db to payroll table in d_db
                self.move_data(cursor_c, cursor_d, 'vw_apt_master', 'aptmaster', columns=['id', 'acode', 'adesc'])

        except Exception as e:
            print(f"Error during data movement: {e}")
            print(f"Error occurred in line {inspect.currentframe().f_lineno}")

        finally:
            # Close connections
            self.close_connection(cursor_a, conn_a)
            self.close_connection(cursor_b, conn_b)
            self.close_connection(cursor_c, conn_c)
            self.close_connection(cursor_d, conn_d)

#----------------------------------------------------------------------------------------
# Belows are top-level functions that can be accessed from anywhere within the module where it's defined. 
# It's part of the module's global scope.

# Initialize current_username and current_datetime
def initialize_username_and_datetime(username, current_datetime):
    if username is None:
        username = "test_user"
    
    if current_datetime is None:
        current_datetime = datetime.now()

    return username, current_datetime

# Excel Formatting
def set_column_widths(ws, column_widths):
    for col, width in enumerate(column_widths, 1):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = width

def set_font(ws, rows, columns, font):
    for row in rows:
        for col in columns:
            cell = ws.cell(row=row, column=col)
            cell.font = font

# Excel file build
def setup_workbook_and_worksheet(sheet_name):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    return wb, ws

def write_header(ws, tv_widget):
    model = tv_widget.model()
    if model is not None:
        header_labels = [get_header_label(col, tv_widget) for col in range(model.columnCount())]
        ws.append(header_labels)
    else:
        print("No model set for the QTableView.")

def write_data_rows(wb, ws, tv_widget, numeric_columns):
    model = tv_widget.model()
    numeric_style = NamedStyle(name='Numeric', number_format='0')    
    
    if model is not None:
        for row in range(model.rowCount()):
            row_data = get_row_data(row, tv_widget)
            ws.append(row_data)

            # Format numeric columns as numbers
            if numeric_columns:
                for col, value in enumerate(row_data):
                    if col in numeric_columns:
                        cell = ws.cell(row=row + 2, column=col + 1)  # Adding 2 to row to account for header
                        try:
                            numeric_value = float(value)
                            cell.value = numeric_value
                            cell.number_format = '0'
                        except ValueError:
                            # If conversion to float fails, set the original string value
                            cell.value = value
            else:
                # If numeric_columns is not specified, set all columns as they are
                for col, value in enumerate(row_data):
                    cell = ws.cell(row=row + 2, column=col + 1)
                    cell.value = value
        # Add the numeric_style to the workbook if provided
        if wb:
            wb.add_named_style(numeric_style)
    else:
        print("No model set for the QTableView.")

def get_header_label(col, tv_widget):
    header_model = tv_widget.horizontalHeader().model()
    if header_model is not None:
        header_item = header_model.headerData(col, Qt.Horizontal)
        return str(header_item) if header_item else ""
    else:
        return ""   

def get_row_data(row, tv_widget):
    model = tv_widget.model()
    row_data = []
    if model is not None:
        for col in range(model.columnCount()):
            item = model.item(row, col)
            if isinstance(item, QTableWidgetItem):
                row_data.append(get_item_value(item))
            else:
                row_data.append(item.text() if item else "")
    return row_data

def get_item_value(item):
    original_value = item.original_value if hasattr(item, 'original_value') else item.text()
    if isinstance(original_value, (int, float)):
        return original_value
    elif isinstance(original_value, str) and is_date(original_value):
        return original_value
    else:
        return original_value

def is_date(value):
    try:
        datetime.strptime(value, "%Y%m%d")  # Adjust the format to match your date format
        return True
    except ValueError:
        return False
    
def get_file_name(output_subfolder, base_name):
    current_datetime = datetime.now()
    formatted_date = current_datetime.strftime("%Y%m%d")
    original_date = base_name + "_" + formatted_date
    default_file_name = original_date.replace("/", "")
    extension = ".xlsx"
    counter = 1

    while True:
        candidate_file_name = f"{default_file_name}_{counter:02d}{extension}"
        full_candidate_path = os.path.join(output_subfolder, candidate_file_name)
        #print(f"Checking: {full_candidate_path}")

        if not os.path.exists(full_candidate_path):
            #print(f"File does not exist, returning: {candidate_file_name}")
            return candidate_file_name
        
        #print(f"File exists, trying next: {candidate_file_name}")
        counter += 1

def export_to_excel(output_subfolder, tv_widget, sheet_name, numeric_columns):
    wb, ws = setup_workbook_and_worksheet(sheet_name)
    write_header(ws, tv_widget)
    write_data_rows(wb, ws, tv_widget, numeric_columns)
    filename = get_file_name(output_subfolder, sheet_name)

    # Ensure the subfolder exists; create it if it doesn't
    os.makedirs(output_subfolder, exist_ok=True)

    if filename:
        full_file_path = os.path.join(output_subfolder, filename)
        wb.active = wb[sheet_name]
        wb.save(full_file_path)

def prefix_get_file_name(base_name, filetext, output_subfolder):
    current_datetime = datetime.now()
    formatted_date = current_datetime.strftime("%Y%m%d")
    original_date = filetext + "_" + base_name + "_" + formatted_date
    default_file_name = original_date.replace("/", "")
    extension = ".xlsx"
    counter = 1

    while True:
        candidate_file_name = f"{default_file_name}_{counter:02d}{extension}"
        full_candidate_path = os.path.join(output_subfolder, candidate_file_name)
        #print(f"Checking: {full_candidate_path}")

        if not os.path.exists(full_candidate_path):
            #print(f"File does not exist, returning: {candidate_file_name}")
            return candidate_file_name
        
        #print(f"File exists, trying next: {candidate_file_name}")
        counter += 1

def prefix_export_to_excel(output_subfolder, tv_widget, sheet_name, filetext, numeric_columns):
    
    wb, ws = setup_workbook_and_worksheet(sheet_name)
    write_header(ws, tv_widget)
    write_data_rows(wb, ws, tv_widget, numeric_columns)
    filename = prefix_get_file_name(sheet_name, filetext, output_subfolder)

    # Ensure the subfolder exists; create it if it doesn't
    os.makedirs(output_subfolder, exist_ok=True)

    if filename:
        full_file_path = os.path.join(output_subfolder, filename)
        wb.active = wb[sheet_name]
        wb.save(full_file_path)

# Clear the contents of each Widget
def clear_widget_data(widget):
    if isinstance(widget, QtWidgets.QLineEdit):
        widget.clear()
    elif isinstance(widget, QtWidgets.QComboBox):
        widget.setCurrentIndex(0)
    elif isinstance(widget, QtWidgets.QWidget):
        for child in widget.findChildren(QtWidgets.QWidget):
            clear_widget_data(child)

# recognize the date format as regular expression
def parse_date(chg_date_str):
    
    match = re.match(r'(\d{4})[/-](\d{2})[/-](\d{2})', chg_date_str)
    if match:
        year, month, day = map(int, match.groups())
        return datetime(year, month, day)
    else:
        raise ValueError("날짜 형식이 잘못되었습니다. 올바른 형식은 'YYYY/MM/DD' 또는 'YYYY-MM-DD'입니다.")
    
# Display current year/month/date info
def disply_date_info():
    now = datetime.now()
    curr_date = now.strftime("%Y/%m/%d")
    ddt = f"{curr_date}" 

    # Calculate the date one year later
    one_year_later = now + timedelta(days=365)
    ddt_1 = one_year_later.strftime("%Y/%m/%d")

    return ddt, ddt_1

# 주민번호에서 생년월일 추출
def extract_birthdate(registration_number):    
    # 주민번호에서 "-" 제거
    registration_number = registration_number.replace("-", "")
    
    # 주민번호에서 생년월일 추출    
    year = int(registration_number[:2])
    month = int(registration_number[2:4])
    day = int(registration_number[4:6])

    # 8번째 자리에 따라 처리
    seventh_digit = int(registration_number[6])
    if seventh_digit in [1, 2]:
        # 19세기인 경우
        year += 1900
    elif seventh_digit in [3, 4]:
        # 20세기인 경우
        year += 2000

    return year, month, day

def format_birthdate(registration_number):
    year, month, day = extract_birthdate(registration_number)
    return "{:04d}/{:02d}/{:02d}".format(year, month, day)

