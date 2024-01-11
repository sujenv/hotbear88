import sys
import logging
import math
from functools import partial
from openpyxl.styles import Font
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from PyQt5 import uic, QtWidgets, QtCore
from PyQt5.QtGui import QStandardItemModel, QKeySequence
from PyQt5.QtWidgets import QMessageBox, QDialog, QShortcut, QMenu
from PyQt5.QtCore import Qt
from commonmd import *
from cal import CalendarView

# Calendar Master table contents -----------------------------------------------------
class CalWorkingHrInhOtDialog(QDialog, SubWindowBase):

    def __init__(self, current_username = None, current_datetime = None):
        super().__init__()

        self.conn, self.cursor = connect_to_database3()    
        uic.loadUi("calwkhrinhot.ui", self)

        # Add window flags
        self.setWindowFlags(self.windowFlags() | Qt.WindowMinimizeButtonHint | Qt.WindowMaximizeButtonHint | Qt.WindowCloseButtonHint)  

        # Initialize current_username and current_datetime directly
        self.current_username, self.current_datetime = initialize_username_and_datetime(current_username, current_datetime)

        # Enable automatic deletion on close
        self.setAttribute(QtCore.Qt.WA_DeleteOnClose)  

        # Create tv_workinghrreg and QSortFilterProxyModel
        self.model = QStandardItemModel()
        self.proxy_model = NumericStringSortModel(self.model)
        self.proxy_model.setSourceModel(self.model)

        # Define the column types
        column_types = ["", "numeric", "", "", "", ""] 

        # Set the custom delegate for the specific column
        delegate = NumericDelegate(column_types, self.tv_workinghrreg)
        self.tv_workinghrreg.setItemDelegate(delegate)
        self.tv_workinghrreg.setModel(self.proxy_model)

        # Enable Sorting
        self.tv_workinghrreg.setSortingEnabled(True)

        # Enable alternating row colors
        self.tv_workinghrreg.setAlternatingRowColors(True)  
        
        # Hide the first index column
        self.tv_workinghrreg.verticalHeader().setVisible(False)

        # While selecting row in tv_workinghrreg, each cell values to displayed to designated widgets
        #self.tv_workinghrreg.clicked.connect(self.show_selected_data)

        # Fill combobox items when the application starts
        self.get_combobox_contents()

        # Populate the data
        self.make_data() 
        self.connect_btn_method()
        self.conn_signal_to_slot()

        # Set tab order for input widgets
        self.set_tab_order()

        # Initiate CTRL+C, CTRL+V and ENTER
        self.setup_shortcuts()

        # Automatically input current date
        self.display_currentdate()

        # Base value setting
        self.cb_workinghrreg_wk_wkd.setCurrentIndex(6)
        self.entry_workinghrreg_whr.setText("8")

        # Create context menus ----------------------------------------------------------------------------------
        self.context_menu1 = self.create_context_menu(self.entry_workinghrreg_startdt)
        self.context_menu2 = self.create_context_menu(self.entry_workinghrreg_enddt)

        self.entry_workinghrreg_startdt.setContextMenuPolicy(Qt.CustomContextMenu)
        self.entry_workinghrreg_startdt.customContextMenuRequested.connect(self.show_context_menu1)

        self.entry_workinghrreg_enddt.setContextMenuPolicy(Qt.CustomContextMenu)
        self.entry_workinghrreg_enddt.customContextMenuRequested.connect(self.show_context_menu2)

        # Make log file
        self.make_logfiles("access_CalWorkingHrOuthOt.log")        

    # Create a shortcut for CTRL+C/CTRL+V/ Return key
    def setup_shortcuts(self):
        self.copy_shortcut = QShortcut(QKeySequence.Copy, self.tv_workinghrreg, partial(self.copy_cells, self.tv_workinghrreg))
        self.paste_shortcut = QShortcut(QKeySequence.Paste, self.tv_workinghrreg, partial(self.paste_cells, self.tv_workinghrreg))
        self.return_shortcut = QShortcut(Qt.Key_Return, self.tv_workinghrreg, partial(self.handle_return_key, self.tv_workinghrreg))

    # Mouse Right click, show "달력보기" menu ---------------------------------------------------------------------
    def create_context_menu(self, target_lineedit):
        context_menu = QMenu()
        custom_action = context_menu.addAction("달력보기")
        custom_action.triggered.connect(lambda: self.show_calendar(target_lineedit))
        return context_menu

    def show_context_menu1(self, pos):
        self.context_menu1.exec_(self.entry_workinghrreg_startdt.mapToGlobal(pos))
    def show_context_menu2(self, pos):
        self.context_menu2.exec_(self.entry_workinghrreg_enddt.mapToGlobal(pos))

    # Show Calendar
    def show_calendar(self, target_lineedit):
        calendar_dialog = CalendarView()
        calendar_dialog.selected_date_changed.connect(lambda date: self.set_selected_date(date, target_lineedit))
        calendar_dialog.exec()

    # Show selected date to the select Qlineedit
    def set_selected_date(self, date, target_lineedit):
        if target_lineedit == self.entry_workinghrreg_startdt:
            target_lineedit.setText(date)
        elif target_lineedit == self.entry_workinghrreg_enddt:
            target_lineedit.setText(date)

    # Display current date and endofdate
    def display_eff_date(self):
        now = datetime.now()
        curr_date = now.strftime("%Y/%m/%d")
        ddt = f"{curr_date}"         
        endofdate = "2050/12/31"
        return ddt, endofdate
    
    # Display current date only
    def display_currentdate(self):
        ddt, ddt_1 = disply_date_info()
        self.entry_workinghrreg_startdt.setText(ddt)
        self.entry_workinghrreg_enddt.setText(ddt_1)

    # Call process_key_event and pass the event and your QTableWidget instance
    def keyPressEvent(self, event):
        tv_widget = self.tv_workinghrreg    
        self.process_key_event(event, tv_widget)

    # Pass combobox info and sql to next method
    def get_combobox_contents(self):
        self.insert_combobox_initiate(self.cb_workinghrreg_wk_wkd, "SELECT DISTINCT da FROM calmaster Where da < 7 ORDER BY da")

    # Initiate Combo_Box 
    def insert_combobox_initiate(self, combo_box, sql_query):
        self.combobox_initializing(combo_box, sql_query) 
        self.cb_workinghrreg_wk_wkd.setCurrentIndex(0) 

    # Connect button to method
    def connect_btn_method(self):
        self.pb_workinghrreg_search.clicked.connect(self.calc_data)
        self.pb_workinghrreg_show.clicked.connect(self.make_data)
        self.pb_workinghrreg_clear.clicked.connect(self.clear_data)
        self.pb_workinghrreg_excel_export.clicked.connect(self.export_data)
        self.pb_workinghrreg_close.clicked.connect(self.close_dialog)

    # Connect signal to method    
    def conn_signal_to_slot(self):
        self.entry_workinghrreg_startdt.editingFinished.connect(self.end_date_change)

    # Effective Thru Date Autoset
    def end_date_change(self):
        curr_date = self.entry_workinghrreg_startdt.text()
        dt01 = datetime.strptime(curr_date, "%Y/%m/%d")
        end_date = dt01 + timedelta(days=365)
        dt_end = end_date.strftime("%Y/%m/%d")

        self.entry_workinghrreg_enddt.setText(dt_end)
        self.entry_workinghrreg_year_whrbot.setText("")
        self.entry_workinghrreg_year_whrsot.setText("")
        self.entry_workinghrreg_mth_whrbot.setText("")
        self.entry_workinghrreg_mth_whrsot.setText("")
        self.entry_workinghrreg_mth_whrbothr.setText("")
        self.entry_workinghrreg_mth_whrsothr.setText("")


    # tab order for calmaster window
    def set_tab_order(self):
        widgets = [self.pb_workinghrreg_show, 
            self.entry_workinghrreg_startdt, self.entry_workinghrreg_enddt, self.pb_workinghrreg_search,
            self.cb_workinghrreg_wk_wkd, self.entry_workinghrreg_whr]
        
        for i in range(len(widgets) - 1):
            self.setTabOrder(widgets[i], widgets[i + 1])

    # To reduce duplications
    def common_query_statement(self):
        tv_widget = self.tv_workinghrreg
        
        self.cursor.execute("SELECT * FROM vw_inh WHERE 1=0")
        column_info = self.cursor.description
        column_names = [col[0] for col in column_info]

        sql_query = "Select * from vw_inh order By caldt"
        column_widths = [120, 80, 80, 80, 120, 150]

        return sql_query, tv_widget, column_info, column_names, column_widths

    # To reduce duplications
    def make_data(self):
        sql_query, tv_widget, column_info, column_names, column_widths = self.common_query_statement() 
        self.populate_dialog(self.cursor, sql_query, tv_widget, column_info, column_names,column_widths)

    # Make Common values set
    def common_values_set(self):
        username = self.current_username
        user_id = self.userID_gen(username)
        formatted_datetime = self.dt_time_info()
        return username, user_id, formatted_datetime

    # clear input field entry
    def clear_data(self):
        for line_edit in self.findChildren(QtWidgets.QLineEdit):
            line_edit.clear()
        
        self.cb_workinghrreg_wk_wkd.setCurrentIndex(0)
        self.cb_workinghrreg_wk_wkd.setCurrentIndex(6)
        self.entry_workinghrreg_whr.setText("8")

        self.display_currentdate()

    # Search data
    def calc_data(self):
        startdt = self.entry_workinghrreg_startdt.text()
        enddt = self.entry_workinghrreg_enddt.text()
        condition = ''' ('SOT', 'BOT')'''
        
        conditions = {'v01': (startdt, "caldt >= #{}#"), 'v02': (enddt, "caldt <= #{}#"), 'v03': (condition, "inh in {}")}
        
        selected_conditions = []

        for key, (value, condition_format) in conditions.items():
            if len(value) > 0:
                selected_conditions.append(condition_format.format(value))

        if not selected_conditions:
            QMessageBox.about(self, "검색 조건 확인", "검색 조건이 비어 있습니다!")
            return

        # Join the selected conditions to form the SQL query
        query = f"SELECT * FROM vw_inh WHERE {' AND '.join(selected_conditions)} ORDER BY caldt"

        QMessageBox.about(self, "검색 조건 확인", f"계약시작일: {startdt} \n계약종료일: {enddt} \n\n조건으로 내근직 현장 특근 시간을 계산합니다!")

        sql_query, tv_widget, column_info, column_names, column_widths = self.common_query_statement() 
        self.populate_dialog(self.cursor, query, tv_widget, column_info, column_names, column_widths)

        self.calculate_days()

    # Calculate the sum of values in a specific column
    def calculate_days(self):
        self.entry_workinghrreg_year_whrbot.clear()
        self.entry_workinghrreg_year_whrsot.clear()
        self.entry_workinghrreg_mth_whrbot.clear()
        self.entry_workinghrreg_mth_whrsot.clear()

        COLUMN = 1 # index 기준으로 count
        r_bot = 0  # Sum for "BOT"
        r_sot = 0  # Sum for "SOT"

        model = self.tv_workinghrreg.model()  # Get the model associated with the table view

        for i in range(model.rowCount()):
            # Assuming column index 2 is used for "BOT" and "SOT"
            item_bot_sot = model.item(i, 2)
            if item_bot_sot and item_bot_sot.data(Qt.DisplayRole):
                value = float(model.item(i, COLUMN).data(Qt.DisplayRole))

                # Check conditions and accumulate sums accordingly
                if item_bot_sot.data(Qt.DisplayRole) == "BOT":
                    r_bot += value
                elif item_bot_sot.data(Qt.DisplayRole) == "SOT":
                    r_sot += value
        
        self.entry_workinghrreg_year_whrbot.setText(str(r_bot))
        self.entry_workinghrreg_year_whrsot.setText(str(r_sot))

        m_bot = round(365/12/7,3)
        m_sot = round(r_sot/12,3)
        self.entry_workinghrreg_mth_whrbot.setText(str(m_bot))
        self.entry_workinghrreg_mth_whrsot.setText(str(m_sot))

        workdays = float(self.cb_workinghrreg_wk_wkd.currentText())
        workhrs = float(self.entry_workinghrreg_whr.text())
        
        # 월평균기준OT시간
        baseOThrs = float(self.entry_workinghrreg_mth_whrbot.text()) * workhrs
        self.entry_workinghrreg_mth_whrbothr.setText(str(baseOThrs))

        # 월평균변동OT시간
        variOthrs = float(self.entry_workinghrreg_mth_whrsot.text()) * workhrs
        self.entry_workinghrreg_mth_whrsothr.setText(str(variOthrs))

    # 선택된 각 위젯의 내용을 엑셀로 내보내기
    def export_data(self):
        output_subfolder = "data_list"          # set the output subfoler name
        sheet_name = "inh_ot"                   # set the excel sheet name
        filename = get_file_name(output_subfolder, sheet_name)    # get the file name

        # Ensure the subfolder exists; create it if it doesn't
        os.makedirs(output_subfolder, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        if filename:
            full_file_path = os.path.join(output_subfolder, filename)
    
            # 각 위젯에서 내용 가져와서 엑셀에 쓰기
            data_to_export = [
                (self.label_2.text(), self.entry_workinghrreg_startdt.text()),
                (self.label_11.text(), self.entry_workinghrreg_enddt.text()),
                (self.label_16.text(), self.entry_workinghrreg_year_whrbot.text()),
                (self.label_28.text(), self.entry_workinghrreg_mth_whrbot.text()),
                (self.label_27.text(), self.entry_workinghrreg_year_whrsot.text()),
                (self.label_29.text(), self.entry_workinghrreg_mth_whrsot.text()),
                (self.label_15.text(), self.cb_workinghrreg_wk_wkd.currentText()),
                (self.label_5.text(), self.entry_workinghrreg_whr.text()),
                (self.label_31.text(), self.entry_workinghrreg_mth_whrbothr.text()),
                (self.label_32.text(), self.entry_workinghrreg_mth_whrsothr.text()),

                ]

            #for i, value in enumerate(data_to_export, start=1):
            #    ws.cell(row=i, column=1, value=value)
            
            for i, (label, value) in enumerate(data_to_export, start=1):
                ws.cell(row=i, column=1, value=label)
                ws.cell(row=i, column=2, value=value)
            
            # 엑셀 파일 저장
            wb.save(full_file_path)
        
            self.excel_formatting(sheet_name, full_file_path)

    # Excel sheet formatting
    def excel_formatting(self, sheet_name, full_file_path):
        wb = load_workbook(full_file_path)
        sheet_name = sheet_name
        ws = wb[sheet_name]
        last_row = ws.max_row + 2

        # Insert headers at A1 and B1
        ws.insert_rows(1)
        ws['A1'] = '구분'
        ws['B1'] = '내용'

        column_widths = [20, 15]                                # set the width of each column
        title_font = Font(bold=True, name="Arial", size=10)     # set the font for the first row
        arial_font = Font(name="Arial", size=10)                # set the forn from the second row to max row

        set_column_widths(ws, column_widths)    # reset column widths

        ws.freeze_panes = 'B2'              # freeze panes D2 cell
        ws.auto_filter.ref = ws.dimensions  # apply auto filter
        ws.sheet_view.showGridLines = False # remove gridlines
            
        set_font(ws, [1], range(1, len(column_widths) + 1), title_font)                 # first row font setting
        set_font(ws, range(2, last_row), range(1, len(column_widths) + 1), arial_font)  # from second row to last row font setting

        wb.save(full_file_path)
        QMessageBox.about(self, "파일 생성 완료", f"data_list folder에 \n엑셀 파일이 {full_file_path}로 \n생성 되었습니다!")  

    # Refresh data
    def refresh_data(self):
        self.clear_data()
        self.make_data()       

if __name__ == "__main__":
    
    log_subfolder = "logs"
    os.makedirs(log_subfolder, exist_ok=True)
    log_file_path = os.path.join(log_subfolder, "access_CalWorkingHrOuthOt.log")

    logging.basicConfig(
        filename=log_file_path,  
        level=logging.INFO,    
        format="%(asctime)s [%(levelname)s] - %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
    app = QtWidgets.QApplication(sys.argv)
    dialog = CalWorkingHrInhOtDialog()
    dialog.show()
    sys.exit(app.exec())
