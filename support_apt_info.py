import sys
import logging
from functools import partial
from openpyxl.styles import Font
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from PyQt5 import uic, QtWidgets, QtCore
from PyQt5.QtGui import QStandardItemModel, QKeySequence
from PyQt5.QtWidgets import QDialog, QMessageBox, QShortcut, QMenu, QInputDialog
from PyQt5.QtCore import Qt
from datetime import datetime
from cal import CalendarView
from commonmd import *
#for non_ui version-------------------------
#from support_apt_info_ui import Ui_SupportAptInfoDialog

#table contents -----------------------------------------------------
class SupportAptInfoDialog(QDialog, SubWindowBase):
#for non_ui version-------------------------
#class SupportAptInfoDialog(QDialog, Ui_SupportAptInfoDialog, SubWindowBase):
     
    def __init__(self, current_username = None, current_datetime = None):
        super().__init__()    
  
        self.conn, self.cursor = connect_to_database3()

        # load ui file
        uic.loadUi("support_apt_info.ui", self)
        #for non_ui version-------------------------
        #self.setupUi(self)

        # Add window flags
        self.setWindowFlags(self.windowFlags() | Qt.WindowMinimizeButtonHint | Qt.WindowMaximizeButtonHint | Qt.WindowCloseButtonHint)  

        # Initialize current_username and current_datetime directly
        self.current_username, self.current_datetime = initialize_username_and_datetime(current_username, current_datetime)

        # Enable automatic deletion on close
        self.setAttribute(QtCore.Qt.WA_DeleteOnClose)  

        # Create tv_supportaptinfo and QSortFilterProxyModel
        self.model = QStandardItemModel()
        self.proxy_model = NumericStringSortModel(self.model)
        self.proxy_model.setSourceModel(self.model)

        # Define the column types
        column_types = ["numeric", "numeric", "", "", "", "", "numeric", "", "", "", ]

        # Set the custom delegate for the specific column
        delegate = NumericDelegate(column_types,self.tv_supportaptinfo)
        self.tv_supportaptinfo.setItemDelegate(delegate)
        self.tv_supportaptinfo.setModel(self.proxy_model)

        # Enable sorting
        self.tv_supportaptinfo.setSortingEnabled(True)

        # Enable alternating row colors
        self.tv_supportaptinfo.setAlternatingRowColors(True)  
        
        # Hide the first index column
        self.tv_supportaptinfo.verticalHeader().setVisible(False)

        # While selecting row in tv_supportaptinfo, each cell values to displayed to designated widgets
        self.tv_supportaptinfo.clicked.connect(self.show_selected_data)

        # Fill combobox items when the application starts
        self.get_combobox_contents()

        # Initiate display of data
        self.make_data()
        self.conn_button_to_method()
        self.connect_signal_slot()

        # Set tab order for input widgets
        self.set_tab_order()

        # Initiate CTRL+C, CTRL+V and ENTER
        self.setup_shortcuts()

        # Create context menus ----------------------------------------------------------------------------------
        self.context_menu1 = self.create_context_menu(self.entry_supportaptinfo_efffrom)
        self.context_menu2 = self.create_context_menu(self.entry_supportaptinfo_effthru)
        self.context_menu3 = self.create_context_menu(self.entry_supportaptinfo_cefffrom)
        self.context_menu4 = self.create_context_menu(self.entry_supportaptinfo_ceffthru)

        self.entry_supportaptinfo_efffrom.setContextMenuPolicy(Qt.CustomContextMenu)
        self.entry_supportaptinfo_efffrom.customContextMenuRequested.connect(self.show_context_menu1)
        
        self.entry_supportaptinfo_effthru.setContextMenuPolicy(Qt.CustomContextMenu)
        self.entry_supportaptinfo_effthru.customContextMenuRequested.connect(self.show_context_menu2)

        self.entry_supportaptinfo_cefffrom.setContextMenuPolicy(Qt.CustomContextMenu)
        self.entry_supportaptinfo_cefffrom.customContextMenuRequested.connect(self.show_context_menu3)

        self.entry_supportaptinfo_ceffthru.setContextMenuPolicy(Qt.CustomContextMenu)
        self.entry_supportaptinfo_ceffthru.customContextMenuRequested.connect(self.show_context_menu4)

        self.entry_stylesheet_as_is()
        self.hide_change_widget()

        # Make log file
        self.make_logfiles("access_SupportAptInfoDialog.log")

    # Create a shortcut for CTRL+C/CTRL+V/ Return key
    def setup_shortcuts(self):
        self.copy_shortcut = QShortcut(QKeySequence.Copy, self.tv_supportaptinfo, partial(self.copy_cells, self.tv_supportaptinfo))
        self.paste_shortcut = QShortcut(QKeySequence.Paste, self.tv_supportaptinfo, partial(self.paste_cells, self.tv_supportaptinfo))
        self.return_shortcut = QShortcut(Qt.Key_Return, self.tv_supportaptinfo, partial(self.handle_return_key, self.tv_supportaptinfo))

    # Change styles for the selected widgets 1
    def entry_stylesheet_as_is(self):
        self.entry_supportaptinfo_cecode.setStyleSheet('color:white;background:rgb(0,0,0)')
        self.cb_supportaptinfo_cename.setStyleSheet('color:white;background:rgb(0,0,0)')
        self.entry_supportaptinfo_csval.setStyleSheet('color:white;background:rgb(0,0,0)')
        self.entry_supportaptinfo_cdescr.setStyleSheet('color:white;background:rgb(0,0,0)')
        self.entry_supportaptinfo_cefffrom.setStyleSheet('color:white;background:rgb(0,0,0)')
        self.entry_supportaptinfo_ceffthru.setStyleSheet('color:white;background:rgb(0,0,0)')
        self.entry_supportaptinfo_cremark.setStyleSheet('color:white;background:rgb(0,0,0)')

    # Change styles for the selected widgets 2
    def entry_stylesheet_to_be(self):
        self.entry_supportaptinfo_cecode.setStyleSheet('color:black;background:rgb(255,255,0)')
        self.cb_supportaptinfo_cename.setStyleSheet('color:black;background:rgb(255,255,0)')
        self.entry_supportaptinfo_csval.setStyleSheet('color:black;background:rgb(255,255,0)')
        self.entry_supportaptinfo_cdescr.setStyleSheet('color:black;background:rgb(255,255,0)')
        self.entry_supportaptinfo_cefffrom.setStyleSheet('color:black;background:rgb(255,255,0)')
        self.entry_supportaptinfo_ceffthru.setStyleSheet('color:black;background:rgb(255,255,0)')
        self.entry_supportaptinfo_cremark.setStyleSheet('color:black;background:rgb(255,255,0)')

    # Show widgets for the cost change parts 
    def show_change_widget(self):
        self.entry_supportaptinfo_cecode.setReadOnly(False)
        self.cb_supportaptinfo_cename.setEnabled(True)
        self.entry_supportaptinfo_csval.setReadOnly(False)
        self.entry_supportaptinfo_cdescr.setReadOnly(False)
        self.entry_supportaptinfo_cefffrom.setReadOnly(False)
        self.entry_supportaptinfo_ceffthru.setReadOnly(False)
        self.entry_supportaptinfo_cremark.setReadOnly(False)
        self.entry_stylesheet_to_be()

    # Hide widgets for the cost change parts 
    def hide_change_widget(self):
        self.entry_supportaptinfo_cecode.setReadOnly(True)
        self.cb_supportaptinfo_cename.setEnabled(False)
        self.entry_supportaptinfo_csval.setReadOnly(True)
        self.entry_supportaptinfo_cdescr.setReadOnly(True)
        self.entry_supportaptinfo_cefffrom.setReadOnly(True)
        self.entry_supportaptinfo_ceffthru.setReadOnly(True)
        self.entry_supportaptinfo_cremark.setReadOnly(True)
        self.entry_stylesheet_as_is()


    # Mouse Right click, show "달력보기" menu ---------------------------------------------------------------------
    def create_context_menu(self, target_lineedit):
        context_menu = QMenu()
        custom_action = context_menu.addAction("달력보기")
        custom_action.triggered.connect(lambda: self.show_calendar(target_lineedit))
        return context_menu

    def show_context_menu1(self, pos):
        self.context_menu1.exec_(self.entry_supportaptinfo_efffrom.mapToGlobal(pos))
    def show_context_menu2(self, pos):
        self.context_menu2.exec_(self.entry_supportaptinfo_effthru.mapToGlobal(pos))
    def show_context_menu3(self, pos):
        self.context_menu3.exec_(self.entry_supportaptinfo_cefffrom.mapToGlobal(pos))
    def show_context_menu4(self, pos):
        self.context_menu4.exec_(self.entry_supportaptinfo_ceffthru.mapToGlobal(pos))

    # Show Calendar
    def show_calendar(self, target_lineedit):
        calendar_dialog = CalendarView()
        calendar_dialog.selected_date_changed.connect(lambda date: self.set_selected_date(date, target_lineedit))
        calendar_dialog.exec()

    # Show selected date to the select Qlineedit
    def set_selected_date(self, date, target_lineedit):
        if target_lineedit == self.entry_supportaptinfo_efffrom:
            target_lineedit.setText(date)
        elif target_lineedit == self.entry_supportaptinfo_effthru:
            target_lineedit.setText(date)
        elif target_lineedit == self.entry_supportaptinfo_cefffrom:
            target_lineedit.setText(date)
        elif target_lineedit == self.entry_supportaptinfo_ceffthru:
            target_lineedit.setText(date)

    # Call process_key_event and pass the event and your QTableWidget instance
    def keyPressEvent(self, event):
        tv_widget = self.tv_supportaptinfo
        self.process_key_event(event, tv_widget)

    # Display end of date only
    def display_eff_date(self):
        endofdate = "2050/12/31"

        return endofdate
    
    # Pass combobox info and sql to next method
    def get_combobox_contents(self):
        self.insert_combobox_initiate(self.cb_supportaptinfo_ename, "SELECT DISTINCT ename FROM employee ORDER BY ename")
        self.insert_combobox_initiate(self.cb_supportaptinfo_aname, "SELECT DISTINCT aname FROM aptmaster ORDER BY aname")
        self.insert_combobox_initiate(self.cb_supportaptinfo_cename, "SELECT DISTINCT ename FROM employee ORDER BY ename")

    # Initiate Combo_Box 
    def insert_combobox_initiate(self, combo_box, sql_query):
        self.combobox_initializing(combo_box, sql_query) 
        self.lbl_supportaptinfo_id.setText("")
        self.entry_supportaptinfo_acode.setText("")
        self.entry_supportaptinfo_ecode.setText("")
        self.cb_supportaptinfo_ename.setCurrentIndex(0) 
        self.cb_supportaptinfo_aname.setCurrentIndex(0)         

    # Connect button to method
    def conn_button_to_method(self):
        self.pb_supportaptinfo_show.clicked.connect(self.make_data)
        self.pb_supportaptinfo_search.clicked.connect(self.search_data)        
        self.pb_supportaptinfo_close.clicked.connect(self.close_dialog)
        self.pb_supportaptinfo_clear.clicked.connect(self.clear_data)

        self.pb_supportaptinfo_insert.clicked.connect(self.tb_insert)
        self.pb_supportaptinfo_update.clicked.connect(self.SelectionMessageBox)
        self.pb_supportaptinfo_delete.clicked.connect(self.tb_delete)
        self.pb_supportaptinfo_xlexport.clicked.connect(self.export_table)
        self.pb_supportaptinfo_change.clicked.connect(self.reflect_change)
        
    # Connect Signal to Slot
    def connect_signal_slot(self):
        self.cb_supportaptinfo_ename.activated.connect(self.cb_supportaptinfo_ename_changed)
        self.cb_supportaptinfo_aname.activated.connect(self.cb_supportaptinfo_aname_changed)
        self.cb_supportaptinfo_cename.activated.connect(self.cb_supportaptinfo_cename_changed)
        self.entry_supportaptinfo_cefffrom.editingFinished.connect(self.ceffrom_changed)

    # tab order for employee window
    def set_tab_order(self):
       
        widgets = [self.pb_supportaptinfo_show, self.entry_supportaptinfo_ecode, self.cb_supportaptinfo_ename,
            self.entry_supportaptinfo_acode, self.entry_supportaptinfo_efffrom, self.entry_supportaptinfo_descr,
            self.entry_supportaptinfo_sval, self.entry_supportaptinfo_efffrom, self.entry_supportaptinfo_remark, 
            
            self.pb_supportaptinfo_search, self.pb_supportaptinfo_clear, self.pb_supportaptinfo_insert, 
            self.pb_supportaptinfo_update, self.pb_supportaptinfo_delete, self.pb_supportaptinfo_close]
        
        for i in range(len(widgets) - 1):
            self.setTabOrder(widgets[i], widgets[i + 1])

    # To reduce duplications
    def common_query_statement(self):
        tv_widget = self.tv_supportaptinfo
        self.cursor.execute("SELECT * FROM vw_supportapt WHERE 1=0")
        column_info = self.cursor.description
        column_names = [col[0] for col in column_info]

        sql_query = "Select * From vw_supportapt"
        column_widths = [80, 100, 100, 100, 250, 200, 100, 100, 100, 150]

        return sql_query, tv_widget, column_info, column_names, column_widths 

    # show employee table data inside of the MDI
    def make_data(self):
        sql_query, tv_widget, column_info, column_names, column_widths = self.common_query_statement() 
        self.populate_dialog(self.cursor, sql_query, tv_widget, column_info, column_names,column_widths)

    # Get the value of other variables
    def get_basic_input(self):
        ecode = int(self.entry_supportaptinfo_ecode.text())
        acode = str(self.entry_supportaptinfo_acode.text())
        descr = str(self.entry_supportaptinfo_descr.text())
        sval = float(self.entry_supportaptinfo_sval.text())        
        efffrom = str(self.entry_supportaptinfo_efffrom.text())
        effthru = str(self.entry_supportaptinfo_effthru.text())
        remark = str(self.entry_supportaptinfo_remark.text())

        return ecode, acode, descr, sval, efffrom, effthru, remark

    # Make Common values set
    def common_values_set(self):
        username = self.current_username
        user_id = self.userID_gen(username)
        formatted_datetime = self.dt_time_info()

        return username, user_id, formatted_datetime    

    # insert new employee data to MySQL table
    def tb_insert(self):
        currentid = self.lbl_supportaptinfo_id.text()
        if not currentid:
            confirm_dialog = self.show_insert_confirmation_dialog()
            
            if confirm_dialog == QMessageBox.Yes:

                idx = self.max_row_id("supportapt")
                username, user_id, formatted_datetime = self.common_values_set()
                ecode, acode, descr, sval, efffrom, effthru, remark = self.get_basic_input()  
                
                if (idx>0 and ecode and acode and sval) and all(len(var) > 0 for var in (descr, efffrom, effthru)):
                
                    self.cursor.execute('''INSERT INTO supportapt (id, ecode, acode, descr, sval, efffrom, effthru, trxdate, userid, remark) 
                                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)'''
                                , (idx, ecode, acode, descr, sval, efffrom, effthru, formatted_datetime, user_id, remark))
                    self.conn.commit()
                    self.show_insert_success_message()
                    self.refresh_data() 
                    logging.info(f"User {username} inserted id number {idx}, at the supportapt table.")
                else:
                    self.show_missing_message("입력 이상")
                    return
            else:
                self.show_cancel_message("데이터 추가 취소")
                return
        else:
            QMessageBox.information(self, "Input Error", "기존 id가 선택된 상태에서는 신규 입력이 불가합니다!")
            return

    # update 조건에 따라 분기 할 것
    def SelectionMessageBox(self):

        # If there's no selection
        if len(self.lbl_supportaptinfo_id.text()) == 0:
            self.show_missing_message_update("입력 확인")

        # In case of row selection

        conA = '''내용 중 오류 수정 - 현재 행을 수정, 추가 행을 만들지 않음!'''
        conB = '''내용의 변경 또는 갱신 - 현재 행의 종료일 변경, 변경된 내용으로 추가 행을 만듦!'''
        
        conditions = [conA, conB]
        condition, okPressed = QInputDialog.getItem(self, "입력 조건 선택", "어떤 작업을 진행하시겠습니까?:", conditions, 0, False)

        if okPressed:
            if condition == conA:
                self.fix_typo()     
            elif condition == conB:
                self.show_change_widget()
            else:
                return
            
    # revise the values in the selected row
    def fix_typo(self):

        confirm_dialog = self.show_update_confirmation_dialog()

        if confirm_dialog == QMessageBox.Yes:
            idx = int(self.lbl_supportaptinfo_id.text())
            username, user_id, formatted_datetime = self.common_values_set()         
            ecode, acode, descr, sval, efffrom, effthru, remark = self.get_basic_input()  
                
            if (idx>0 and ecode and acode and sval) and all(len(var) > 0 for var in (descr, efffrom, effthru)):
                self.cursor.execute('''UPDATE supportapt SET ecode=?, acode=?, descr=?, sval=?, efffrom=?, effthru=?, trxdate=?, userid=?, remark=? WHERE id=?'''
                            , (ecode, acode, descr, sval, efffrom, effthru, formatted_datetime, user_id, remark, idx))
                self.conn.commit()
                self.show_update_success_message()
                self.refresh_data()  
                logging.info(f"User {username} updated row number {idx} in the supportapt table.")            
            else:
                self.show_missing_message("입력 이상")
                return
        else:
            self.show_cancel_message("데이터 변경 취소")
            return

    # Get the Changed Infos
    def get_changed_info(self):
        cecode = int(self.entry_supportaptinfo_cecode.text())
        cacode = str(self.entry_supportaptinfo_acode.text())
        cdescr = str(self.entry_supportaptinfo_cdescr.text())
        csval = float(self.entry_supportaptinfo_csval.text())
        cefffrom = str(self.entry_supportaptinfo_cefffrom.text())
        ceffthru = str(self.entry_supportaptinfo_ceffthru.text())
        cremark = str(self.entry_supportaptinfo_cremark.text())
        
        return cecode, cacode, cdescr, csval, cefffrom, ceffthru, cremark

    # Bank Account Change and Insert
    def reflect_change(self):

        username, user_id, formatted_datetime = self.common_values_set()
        cecode, cacode, cdescr, csval, cefffrom, ceffthru, cremark  = self.get_changed_info()

        idx = int(self.max_row_id("gpai"))        
                
        if (idx>0 and cecode and csval) and all(len(var) > 0 for var in (cacode, cdescr, cefffrom, ceffthru)):

            org_id = str(self.lbl_supportaptinfo_id.text())
            effthru1 = str(self.entry_supportaptinfo_effthru.text()) # 다시 불러와야 함..
            # 기존 id의 유효종료일을 변경유효시작일 -1 일로 수정
            self.cursor.execute('''UPDATE supportapt SET effthru=? WHERE id=?''', (effthru1, org_id))
            
            #변경된 내용을 신규로 추가
            self.cursor.execute('''INSERT INTO supportapt (id, ecode, acode, descr, sval, efffrom, effthru, trxdate, userid, remark) 
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)'''
                        , (idx, cecode, cacode, cdescr, csval, cefffrom, ceffthru, formatted_datetime, user_id, cremark))
            
            self.conn.commit()
            self.show_update_success_message()
            self.refresh_data()
            logging.info(f"User {username} updated row number {idx} in the support apt table.")
            
        else:
            self.show_missing_message("입력 이상")
            return

        self.entry_stylesheet_as_is()
        self.hide_change_widget()

    # delete row according to id selected
    def tb_delete(self):
        confirm_dialog = self.show_delete_confirmation_dialog()

        if confirm_dialog == QMessageBox.Yes:
            idx = self.lbl_supportaptinfo_id.text()
            username, user_id, formatted_datetime = self.common_values_set()

            self.cursor.execute("DELETE FROM supportapt WHERE id=?", (idx,))
            self.conn.commit()
            self.show_delete_success_message()
            self.refresh_data()  
            logging.info(f"User {username} deleted id number {idx}, at the supportapt table.")       
        else:
            self.show_cancel_message("데이터 삭제 취소")

    # Search data
    def search_data(self):
      
        ename = self.cb_supportaptinfo_ename.currentText()
        aname = self.cb_supportaptinfo_aname.currentText()
        
        conditions = {'v01': (ename, "ename like '%{}%'"),
                      'v02': (aname, "aname like '%{}%'")
                      }
        
        selected_conditions = []

        for key, (value, condition_format) in conditions.items():
            if len(value) > 0:
                selected_conditions.append(condition_format.format(value))

        if not selected_conditions:
            QMessageBox.about(self, "검색 조건 확인", "검색 조건이 비어 있습니다!")
            return

        # Join the selected conditions to form the SQL query
        query = f"SELECT * FROM vw_supportapt WHERE {' AND '.join(selected_conditions)} ORDER BY ename, aname"

        QMessageBox.about(self, "검색 조건 확인", f"직원명:{ename} \n아파트명:{ename} \n\n위 조건으로 검색을 수행합니다!")
        
        sql_query, tv_widget, column_info, column_names, column_widths = self.common_query_statement()
        self.populate_dialog(self.cursor, query, tv_widget, column_info, column_names,column_widths)


    # Employee Name Index Changed
    def cb_supportaptinfo_ename_changed(self):
        self.entry_supportaptinfo_ecode.clear()
        selected_item = self.cb_supportaptinfo_ename.currentText()

        if selected_item:
            query = f"SELECT DISTINCT ecode From employee WHERE ename ='{selected_item}'"
            line_edit_widgets = [self.entry_supportaptinfo_ecode]
        
            # Check if any line edit widgets are provided
            if line_edit_widgets:
                self.lineEdit_contents(line_edit_widgets, query)
            else:
                pass

    # Apt Name Index Changed
    def cb_supportaptinfo_aname_changed(self):
        self.entry_supportaptinfo_acode.clear()
        selected_item = self.cb_supportaptinfo_aname.currentText()

        if selected_item:
            query = f"SELECT DISTINCT acode From aptmaster WHERE aname ='{selected_item}'"
            line_edit_widgets = [self.entry_supportaptinfo_acode]
        
            # Check if any line edit widgets are provided
            if line_edit_widgets:
                self.lineEdit_contents(line_edit_widgets, query)
            else:
                pass

    # CEmployee Name Index Changed
    def cb_supportaptinfo_cename_changed(self):
        self.entry_supportaptinfo_cecode.clear()
        selected_item = self.cb_supportaptinfo_cename.currentText()

        if selected_item:
            query = f"SELECT DISTINCT ecode From employee WHERE ename ='{selected_item}'"
            line_edit_widgets = [self.entry_supportaptinfo_cecode]
        
            # Check if any line edit widgets are provided
            if line_edit_widgets:
                self.lineEdit_contents(line_edit_widgets, query)
            else:
                pass

    # Ceffective From Date Changed
    def ceffrom_changed(self):
        chg_date_str = self.entry_supportaptinfo_cefffrom.text()
        try:
            chg_date = parse_date(chg_date_str)                             # 날짜 문자열을 날짜 객체로 변환
            org_date = chg_date - timedelta(days=1)                         # 날짜에서 1일을 빼줍니다.
            org_date_str = org_date.strftime('%Y-%m-%d')                    # 결과를 문자열로 변환
            self.entry_supportaptinfo_effthru.setText(org_date_str)                   # 변경된 effthru 날짜를 표시
        
            endofdate = self.display_eff_date()
            self.entry_supportaptinfo_ceffthru.setText(endofdate)

        except ValueError as e:
            QMessageBox.critical(self, "날짜 형식 오류", str(e))             # 날짜 형식이 잘못된 경우 사용자에게 알림



    # Export data to Excel sheet                
    def export_table(self):
        output_subfolder = "data_list"              # set the output subfoler name
        table_widget = self.tv_supportaptinfo             # set the name of table widget
        sheet_name = "support_apt"                 # set the excel sheet name
        filename = get_file_name(output_subfolder, sheet_name)    # get the file name / export_to_excel 밑에 두면 get_file_name을 두 번 실행시켜서 파일의 번호가 달라지므로 미리 만들어서 filename에 담아둠.
        numeric_columns=[0, 1]                      # set the numerical column index
        export_to_excel(output_subfolder, table_widget, sheet_name, numeric_columns)
               
        if filename:
            full_file_path = os.path.join(output_subfolder, filename)
            self.excel_formatting(sheet_name, full_file_path)
        else:
            QMessageBox.about(self, "코드 확인", "파일을 찾지 못했습니다!")    

    # Excel sheet formatting
    def excel_formatting(self, sheet_name, full_file_path):
        wb = load_workbook(full_file_path)
        sheet_name = sheet_name
        ws = wb[sheet_name]
        last_row = ws.max_row + 1

        column_widths = [8, 14, 10, 10, 10, 8, 10, 10, 16, 20, 10, 10, 20]  # set the width of each column
        title_font = Font(bold=True, name="Arial", size=10)                 # set the font for the first row
        arial_font = Font(name="Arial", size=10)                            # set the forn from the second row to max row

        set_column_widths(ws, column_widths)        # reset column widths

        ws.freeze_panes = 'D2'                      # freeze panes D2 cell
        ws.auto_filter.ref = ws.dimensions          # apply auto filter
        ws.sheet_view.showGridLines = False         # remove gridlines
            
        set_font(ws, [1], range(1, len(column_widths) + 1), title_font)                 # first row font setting
        set_font(ws, range(2, last_row), range(1, len(column_widths) + 1), arial_font)  # from second row to last row font setting
        
        wb.save(full_file_path)
        QMessageBox.about(self, "파일 생성 완료", f"data_list folder에 \n엑셀 파일이 {full_file_path}로 \n생성 되었습니다!")   

    # clear input field entry
    def clear_data(self):
        self.lbl_supportaptinfo_id.setText("")
        for line_edit in self.findChildren(QtWidgets.QLineEdit):
            line_edit.clear()

    # table widget cell double click
    def show_selected_data(self, item):
        # Get the row index of the clicked item
        row_index = item.row()

        # Initialize a list to store the cell values
        cell_values = []

        # Loop through the columns and retrieve the text from each cell
        for column_index in range(10):  # 10 columns
            cell_text = self.tv_supportaptinfo.model().item(row_index, column_index).text()
            cell_values.append(cell_text)

        # Populate the input widgets with the data from the selected row
        self.lbl_supportaptinfo_id.setText(cell_values[0])
        self.entry_supportaptinfo_ecode.setText(cell_values[1])
        self.cb_supportaptinfo_ename.setCurrentText(cell_values[2])
        self.entry_supportaptinfo_acode.setText(cell_values[3])
        self.cb_supportaptinfo_aname.setCurrentText(cell_values[4])
        self.entry_supportaptinfo_descr.setText(cell_values[5])
        self.entry_supportaptinfo_sval.setText(cell_values[6])
        self.entry_supportaptinfo_efffrom.setText(cell_values[7])
        self.entry_supportaptinfo_effthru.setText(cell_values[8])
        self.entry_supportaptinfo_remark.setText(cell_values[9])
    
    def refresh_data(self):
        self.clear_data()
        self.make_data()


if __name__ == "__main__":

    log_subfolder = "logs"
    os.makedirs(log_subfolder, exist_ok=True)
    log_file_path = os.path.join(log_subfolder, "access_SupportAptInfoDialog.log")

    logging.basicConfig(
        filename=log_file_path,  
        level=logging.INFO,    
        format="%(asctime)s [%(levelname)s] - %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
        
    app = QtWidgets.QApplication(sys.argv)
    dialog = SupportAptInfoDialog()
    dialog.show()
    sys.exit(app.exec())