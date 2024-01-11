import sys
import logging
import openpyxl
from functools import partial
from openpyxl.styles import Font
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from PyQt5 import uic, QtWidgets, QtCore
from PyQt5.QtGui import QStandardItemModel, QKeySequence
from PyQt5.QtWidgets import QDialog, QMessageBox, QMenu, QInputDialog, QShortcut
from PyQt5.QtCore import Qt, QTimer
from datetime import datetime
from commonmd import *
#for non_ui version-------------------------
#from consumableoutclosing_ui import Ui_ConsumableOutClosing

# Sales Product table contents -----------------------------------------------------
class ConsumableOutClosingDialog(QDialog, SubWindowBase):
#for non_ui version-------------------------
#class ConsumableOutClosing(QDialog, Ui_ConsumableOutClosing, SubWindowBase):
     
    def __init__(self, current_username = None, current_datetime = None):
        super().__init__()

        self.conn, self.cursor = connect_to_database2()

        uic.loadUi("consumableoutclosing.ui", self)
        #for non_ui version-------------------------
        #self.setupUi(self)

        # Add window flags
        self.setWindowFlags(self.windowFlags() | Qt.WindowMinimizeButtonHint | Qt.WindowMaximizeButtonHint | Qt.WindowCloseButtonHint)  

        # Initialize current_username and current_datetime directly
        self.current_username, self.current_datetime = initialize_username_and_datetime(current_username, current_datetime)

        # Enable automatic deletion on close
        self.setAttribute(QtCore.Qt.WA_DeleteOnClose)  

        # Create tv_consumableoutclosing and QSortFilterProxyModel
        self.model = QStandardItemModel()
        self.proxy_model = NumericStringSortModel(self.model)
        self.proxy_model.setSourceModel(self.model)        
        
        # Define the column types
        column_types = ["numeric", "", "", "numeric", "numeric", "numeric", "numeric", "", "", "", "", ]

        # Set the custom delegate for the specific column
        delegate = NumericDelegate(column_types, self.tv_consumableoutclosing)
        self.tv_consumableoutclosing.setItemDelegate(delegate)
        self.tv_consumableoutclosing.setModel(self.proxy_model)
       
        # Enable sorting
        self.tv_consumableoutclosing.setSortingEnabled(True)

        # Enable alternating row colors
        self.tv_consumableoutclosing.setAlternatingRowColors(True)  
        
        # Hide the first index column
        self.tv_consumableoutclosing.verticalHeader().setVisible(False)

        # While selecting row in tv_consumableoutclosing, each cell values to displayed to designated widgets
        #self.tv_consumableoutclosing.clicked.connect(self.show_selected_data)

        # Initial display of data
        self.make_data()
        self.conn_button_to_method()
        self.conn_signal_to_slot()

        # Initiate CTRL+C, CTRL+V and ENTER
        self.setup_shortcuts()

        self.date_clock_setup()

    # Setup current date time 
    def date_clock_setup(self):
        self.lbl_consumableoutclosing_date_time.setText("Initializing...")

        # Set up a QTimer to update the datetime label every second
        self.update_timer = QTimer(self)
        self.update_timer.timeout.connect(self.update_currentdatetime)
        self.update_timer.start(1000)  # Update every 1000 milliseconds (1 second)
        
        current_date = datetime.now()
        current_year = current_date.year
        current_month = current_date.month

        self.cb_consumableoutclosing_year.setCurrentText(str(current_year))
        self.cb_consumableoutclosing_month.setCurrentText(str(current_month))

        cur_year = int(self.cb_consumableoutclosing_year.currentText())
        cur_month = int(self.cb_consumableoutclosing_month.currentText())

        first_day = get_first_day_of_month(cur_year, cur_month)
        last_day = get_last_day_of_month(cur_year, cur_month)

        self.entry_consumableoutclosing_firstday.setText(first_day.strftime("%Y/%m/%d"))  # Convert date to string
        self.entry_consumableoutclosing_lastday.setText(last_day.strftime("%Y/%m/%d"))  # Convert date to string

    # Display current date time 
    def update_currentdatetime(self):
        now = datetime.now()
        curr_date = now.strftime("%Y/%m/%d")
        curr_time = now.strftime("%H:%M:%S")
        ddt = f"{curr_date} {curr_time}"
        self.lbl_consumableoutclosing_date_time.setText(ddt)

    # Create a shortcut for CTRL+C/CTRL+V/ Return key
    def setup_shortcuts(self):
        self.copy_shortcut = QShortcut(QKeySequence.Copy, self.tv_consumableoutclosing, partial(self.copy_cells, self.tv_consumableoutclosing))
        self.paste_shortcut = QShortcut(QKeySequence.Paste, self.tv_consumableoutclosing, partial(self.paste_cells, self.tv_consumableoutclosing))
        self.return_shortcut = QShortcut(Qt.Key_Return, self.tv_consumableoutclosing, partial(self.handle_return_key, self.tv_consumableoutclosing))

    # Call process_key_event and pass the event and your QTableWidget instance
    def keyPressEvent(self, event):
        tv_widget = self.tv_consumableoutclosing
        self.process_key_event(event, tv_widget)

    # Connect the button to the method
    def conn_button_to_method(self):
        self.pb_consumableoutclosing_show.clicked.connect(self.make_data)
        self.pb_consumableoutclosing_execute.clicked.connect(self.execute_closing)
        self.pb_consumableoutclosing_status.clicked.connect(self.change_status)
        self.pb_consumableoutclosing_export.clicked.connect(self.export_to_excel)

    # Connect signal to method    
    def conn_signal_to_slot(self):
        self.cb_consumableoutclosing_year.activated.connect(self.closing_sales_year_month_changed)
        self.cb_consumableoutclosing_month.activated.connect(self.closing_sales_year_month_changed)

    # To reduce duplications
    def common_query_statement(self):
        tv_widget = self.tv_consumableoutclosing

        query = "SELECT id, ename, pdescription, qty, salesprice, receipt as sup_price, trx_date, scode as status, sdescription, remark"
        query = query + " FROM vw_arlist_02"
        query = query + " WHERE 1=0"
        
        self.cursor.execute(query)
        column_info = self.cursor.description
        column_names = [col[0] for col in column_info]       

        sql_query = "Select "
        sql_query = sql_query + " id, ename, pdescription, qty, salesprice, receipt as sup_price, trx_date, scode as status, sdescription, remark"
        sql_query = sql_query + " From vw_arlist_02" 
        sql_query = sql_query + " order by id desc"
        column_widths = [80, 100, 100, 100, 100, 100]

        return sql_query, tv_widget, column_info, column_names, column_widths

    # show salesproduct table data inside of the MDI
    def make_data(self):
        sql_query, tv_widget, column_info, column_names, column_widths = self.common_query_statement() 
        self.populate_dialog(self.cursor, sql_query, tv_widget, column_info, column_names,column_widths)

    # Monthly Closing
    def execute_closing(self):

        sdate = self.entry_consumableoutclosing_firstday.text()
        edate = self.entry_consumableoutclosing_lastday.text()

        query = "Select "
        query = query + " id, ename, pdescription, qty, salesprice, receipt as sup_price, trx_date, scode as status, sdescription, remark"
        query = query + " From vw_arlist_02" 
        query = query + " Where "
        query = query + " trx_date >= #" + sdate + "# and trx_date <= #" + edate + "#"      # MS Access에서 날짜는 # #로 감싸줘야 함
        query = query + " order by id desc"

        column_widths = [80, 100, 100, 100, 100, 100]

        sql_query, tv_widget, column_info, column_names, column_widths = self.common_query_statement() 
        self.populate_dialog(self.cursor, query, tv_widget, column_info, column_names,column_widths)

    # Status change
    def change_status(self):
        confirm_dialog = self.show_closing_confirmation_dialog()

        if confirm_dialog == QMessageBox.Yes:
        
            sdate = self.entry_consumableoutclosing_firstday.text()
            edate = self.entry_consumableoutclosing_lastday.text()

            self.cursor.execute('''UPDATE arlist SET status='9' WHERE trx_date >= ? and trx_date <= ?''' , (sdate, edate))
            self.conn.commit()
            self.show_update_success_message()
            self.refresh_data()  
        else:
            self.show_cancel_message("데이터 변경 취소")


    # Combo box year activated and connect
    def closing_sales_year_month_changed(self):
        self.entry_consumableoutclosing_firstday.clear()
        self.entry_consumableoutclosing_lastday.clear()
        
        selected_year = self.cb_consumableoutclosing_year.currentText()
        selected_month = self.cb_consumableoutclosing_month.currentText()

        cur_year = int(selected_year)
        cur_month = int(selected_month)

        first_day = get_first_day_of_month(cur_year, cur_month)
        last_day = get_last_day_of_month(cur_year, cur_month)

        self.entry_consumableoutclosing_firstday.setText(first_day.strftime("%Y/%m/%d"))  # Convert date to string
        self.entry_consumableoutclosing_lastday.setText(last_day.strftime("%Y/%m/%d"))  # Convert date to string

    # Export data to Excel sheet   
    def export_to_excel(self):
        output_subfolder = "closing"                        # set the output subfoler name
        tv_widget = self.tv_consumableoutclosing            # set the name of table widget
        filetext = "매출마감"
        sheet_name = "data"                                 # set the excel sheet name
        filename = prefix_get_file_name(sheet_name, filetext, output_subfolder)     # get the file name / export_to_excel 밑에 두면 get_file_name을 두 번 실행시켜서 파일의 번호가 달라지므로 미리 만들어서 filename에 담아둠.
        numeric_columns=[0, 3, 4 ,5, 6]                     # set the numerical column index
        prefix_export_to_excel(output_subfolder, tv_widget, sheet_name, filetext, numeric_columns)

        if filename:
            full_file_path = os.path.join(output_subfolder, filename)
            self.excel_formatting(sheet_name, full_file_path)
            self.excel_sum_result(full_file_path)
            self.closing_by_ename(full_file_path) 
            QMessageBox.about(self, "파일 생성 완료", f"엑셀 파일이 {full_file_path}에 \n생성 되었습니다!") 
        else:
            QMessageBox.about(self, "코드 확인", "파일을 찾지 못했습니다!") 
        
    # Excel sheet formatting
    def excel_formatting(self, sheet_name, full_file_path):
        wb = load_workbook(full_file_path)
        sheet_name = sheet_name
        ws = wb[sheet_name]
        
        # Find the last row in column K
        last_row = ws.max_row + 1

        # Handling Integer Column
        # Apply number format to columns D:G
        for col_num in range(4, 7):
            col_letter = openpyxl.utils.get_column_letter(col_num)
            for row in ws.iter_rows(min_row=2, max_row=last_row, min_col=col_num, max_col=col_num):
                for cell in row:
                    cell.number_format = '#,##0'

        # Handling column which has the string date type
        # Set formula from K2:K16 as formulas
        for row_num in range(2, last_row+1):
            date_str = ws[f'G{row_num}'].value  # column letter need to change
            if date_str is not None:
                # Adjust the format to match your Excel date format
                date_obj = datetime.strptime(date_str, "%Y-%m-%d") 
                timestamp = int(date_obj.timestamp())
                excel_date = (date_obj - datetime(1899, 12, 30)).days
                ws[f'K{row_num}'].value = excel_date
            else:
                # Handle the case when the cell is empty or doesn't contain a valid date string
                ws[f'K{row_num}'].value = None

        # Copy the values from range K2:K16
        copied_values = []
        for row_num in range(2, last_row+1):
            copied_values.append([ws[f'K{row_num}'].value])

        # Paste the copied values to column G, starting from G2 
        for row_num, value in enumerate(copied_values, start=2):
            ws[f'G{row_num}'].value = value[0]  # column letter need to change and matched

        # Apply date number format to the cells in column G
        for row_num in range(2, last_row+1):
            ws[f'G{row_num}'].number_format = 'yyyy-mm-dd' # column letter need to change and matched

        # Delete column K
        ws.delete_cols(11)

        column_widths = [5, 10, 15, 8, 12, 12, 15, 10, 10, 10] # set the width of each column
        title_font = Font(bold=True, name="Arial", size=10)     # set the font for the first row
        arial_font = Font(name="Arial", size=10)                # set the forn from the second row to max row

        set_column_widths(ws, column_widths)    # reset column widths

        ws.freeze_panes = 'D2'              # freeze panes D2 cell
        ws.auto_filter.ref = ws.dimensions  # apply auto filter
        ws.sheet_view.showGridLines = False # remove gridlines
            
        set_font(ws, [1], range(1, len(column_widths) + 1), title_font)                 # first row font setting
        set_font(ws, range(2, last_row), range(1, len(column_widths) + 1), arial_font)  # from second row to last row font setting
        
        wb.save(full_file_path)

    # Add to the result of summmation of the column
    def excel_sum_result(self, full_file_path):
        wb = load_workbook(full_file_path)

        # Delete the default "Sheet" sheet if it exists
        if "Sheet" in wb.sheetnames:
            sheet_to_delete = wb["Sheet"]
            wb.remove(sheet_to_delete)
            
        # Create the 'closing' sheet and set it as the active sheet
        sheet = wb.create_sheet('closing')
        sheet_name = 'closing'
        ws = wb[sheet_name]

        sdate = self.entry_consumableoutclosing_firstday.text()
        edate = self.entry_consumableoutclosing_lastday.text()

        sql_query = '''
            SELECT ename, pdescription, um, SUM(qty) AS TotalQty, SUM(CLng(receipt)) AS TotalPayment
            FROM vw_arlist_02
            WHERE trx_date BETWEEN #{}# AND #{}# and ecode <> 20301003
            GROUP BY ename, pdescription, um
            ORDER BY ename'''.format(sdate, edate)

        self.cursor.execute(sql_query)
        results = self.cursor.fetchall()
            
        # Write column headers
        column_headers = ["ename", "pdescription", "um", "TotalQty", "TotalPayment"]
        for col_num, header in enumerate(column_headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header)

        # Write data to the worksheet
        for row_num, row_data in enumerate(results, 2):
            for col_num, cell_value in enumerate(row_data, 1):
                cell = sheet.cell(row=row_num, column=col_num, value=cell_value)            

        # Add summation functions to the last rows
        bold_only_font = Font(name='Calibri', size=11, bold=True)
        last_row = len(results) + 2  # Adjust for headers and 1-based indexing
        sheet.cell(row=last_row, column=1, value="Total").font 
        sheet.cell(row=last_row, column=5, value="=SUM(E2:E{})".format(last_row - 1))

        # Make the summation row bold
        bold_only_font = Font(name='Calibri', size=11, bold=True)
        for col in [1, 5]:
            cell = sheet.cell(row=last_row, column=col)
            cell.font = bold_only_font
        
        # Select columns B and C & Apply the "#,##0" number format to the selected columns
        columns_range = sheet['D:E']

        for column in columns_range:
            for cell in column:
                cell.number_format = '#,##0'
        
        #----------------------------------------------------------------------------------------------
        column_widths = [10, 15, 10, 12, 20]                    # set the width of each column
        title_font = Font(bold=True, name="Arial", size=10)     # set the font for the first row
        arial_font = Font(name="Arial", size=10)                # set the forn from the second row to max row

        set_column_widths(ws, column_widths)    # reset column widths

        ws.freeze_panes = 'D2'                  # freeze panes D2 cell
        ws.auto_filter.ref = ws.dimensions      # apply auto filter
        ws.sheet_view.showGridLines = False     # remove gridlines
            
        set_font(ws, [1], range(1, len(column_widths) + 1), title_font)                 # first row font setting
        set_font(ws, range(2, last_row), range(1, len(column_widths) + 1), arial_font)  # from second row to last row font setting
        #----------------------------------------------------------------------------------------------

        wb.save(full_file_path)

    # Personal sales results
    def closing_by_ename(self, full_file_path):
        # Create the 'closing_by_ename' sheet and set it as the active sheet
        wb = load_workbook(full_file_path)
        sheet_name = 'closing_by_ename'

        if sheet_name in wb.sheetnames:
            wb[sheet_name]  # If the sheet already exists, set it as the active sheet
        else:
            sheet = wb.create_sheet(sheet_name)
            wb.active = sheet  # Set the newly created sheet as the active sheet

        sdate = self.entry_consumableoutclosing_firstday.text()
        edate = self.entry_consumableoutclosing_lastday.text()

        sql_query = '''
            SELECT ename, SUM(CLng(receipt)) AS TotalPayment
            FROM vw_arlist_02
            WHERE trx_date BETWEEN #{}# AND #{}# and ecode <> 20301003
            GROUP BY ename
            ORDER BY ename'''.format(sdate, edate)

        self.cursor.execute(sql_query)
        results = self.cursor.fetchall()
        
        # Write column headers
        column_headers = ["이름", "금액"]
        for col_num, header in enumerate(column_headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header)

        # Write data to the worksheet
        for row_num, row_data in enumerate(results, 2):
            for col_num, cell_value in enumerate(row_data, 1):
                cell = sheet.cell(row=row_num, column=col_num, value=cell_value)            
            
        # Add summation functions to the last rows
        last_row = len(results) + 2  # Adjust for headers and 1-based indexing
        sheet.cell(row=last_row, column=1, value="Total").font 
        sheet.cell(row=last_row, column=2, value="=SUM(B2:B{})".format(last_row - 1))

        # Make the summation row bold
        bold_only_font = Font(name='Calibri', size=11, bold=True)
        for col in [1, 2]:
            cell = sheet.cell(row=last_row, column=col)
            cell.font = bold_only_font

        # Select columns B Apply the "#,##0" number format to the selected columns
        for row in sheet.iter_rows(min_row=2, max_row=len(results) + 2, min_col=2, max_col=2):
            for cell in row:
                cell.number_format = '#,##0'

        ws = wb[sheet_name]
        #----------------------------------------------------------------------------------------------
        column_widths = [10, 12]                                # set the width of each column
        title_font = Font(bold=True, name="Arial", size=10)     # set the font for the first row
        arial_font = Font(name="Arial", size=10)                # set the forn from the second row to max row

        set_column_widths(ws, column_widths)    # reset column widths

        ws.freeze_panes = 'B2'                  # freeze panes D2 cell
        ws.auto_filter.ref = ws.dimensions      # apply auto filter
        ws.sheet_view.showGridLines = False     # remove gridlines
            
        set_font(ws, [1], range(1, len(column_widths) + 1), title_font)                 # first row font setting
        set_font(ws, range(2, last_row), range(1, len(column_widths) + 1), arial_font)  # from second row to last row font setting
        #----------------------------------------------------------------------------------------------

        wb.save(full_file_path)

    # Clear all entry and combo boxes
    def refresh_data(self):
        self.make_data()


if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    dialog = ConsumableOutClosingDialog()
    dialog.show()
    sys.exit(app.exec())