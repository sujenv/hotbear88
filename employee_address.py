import sys
import logging
from functools import partial
from openpyxl.styles import Font
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from PyQt5 import uic, QtWidgets, QtCore
from PyQt5.QtGui import QStandardItemModel, QKeySequence
from PyQt5.QtWidgets import QDialog, QMessageBox, QShortcut
from PyQt5.QtCore import Qt
from datetime import datetime
from commonmd import *
#for non_ui version-------------------------
#from employeeaddress_ui import Ui_EmployeeAddressDialog

# employee address table contents -----------------------------------------------------
class EmployeeAddressDialog(QDialog, SubWindowBase):
#for non_ui version-------------------------
#class EmployeeAddressDialog(QDialog, Ui_EmployeeAddressDialog, SubWindowBase):
     
    def __init__(self, current_username = None, current_datetime = None):
        super().__init__()    
  
        self.conn, self.cursor = connect_to_database1()

        # load ui file
        uic.loadUi("employee_address.ui", self)
        #for non_ui version-------------------------
        #self.setupUi(self)

        # Add window flags
        self.setWindowFlags(self.windowFlags() | Qt.WindowMinimizeButtonHint | Qt.WindowMaximizeButtonHint | Qt.WindowCloseButtonHint)  

        # Initialize current_username and current_datetime directly
        self.current_username, self.current_datetime = initialize_username_and_datetime(current_username, current_datetime)

        # Enable automatic deletion on close
        self.setAttribute(QtCore.Qt.WA_DeleteOnClose)  

        # Create tv_address and QSortFilterProxyModel
        self.model = QStandardItemModel()
        self.proxy_model = NumericStringSortModel(self.model)
        self.proxy_model.setSourceModel(self.model)

        # Define the column types
        column_types = ["numeric", "numeric", "", "", "", "", "", ""]

        # Set the custom delegate for the specific column
        delegate = NumericDelegate(column_types,self.tv_address)
        self.tv_address.setItemDelegate(delegate)
        self.tv_address.setModel(self.proxy_model)

        # Enable sorting
        self.tv_address.setSortingEnabled(True)

        # Enable alternating row colors
        self.tv_address.setAlternatingRowColors(True)  
        
        # Hide the first index column
        self.tv_address.verticalHeader().setVisible(False)

        # While selecting row in tv_address, each cell values to displayed to designated widgets
        self.tv_address.clicked.connect(self.show_selected_data)

        # Fill combobox items when the application starts
        self.get_combobox_contents()

        # Initiate display of data
        self.make_data()
        self.conn_button_to_method()
        self.connect_signal_slot()

        # Set tab order for input widgets
        self.set_tab_order()

        # Initiate CTRL+C, CTRL+V and ENTER
        self.setup_shortcuts()

        # Make log file
        self.make_logfiles("access_EmployeeAddress.log")

    # Create a shortcut for CTRL+C/CTRL+V/ Return key
    def setup_shortcuts(self):
        self.copy_shortcut = QShortcut(QKeySequence.Copy, self.tv_address, partial(self.copy_cells, self.tv_address))
        self.paste_shortcut = QShortcut(QKeySequence.Paste, self.tv_address, partial(self.paste_cells, self.tv_address))
        self.return_shortcut = QShortcut(Qt.Key_Return, self.tv_address, partial(self.handle_return_key, self.tv_address))

    # Call process_key_event and pass the event and your QTableWidget instance
    def keyPressEvent(self, event):
        tv_widget = self.tv_address
        self.process_key_event(event, tv_widget)

    # Pass combobox info and sql to next method
    def get_combobox_contents(self):
        self.insert_combobox_initiate(self.cb_address_ename, "SELECT DISTINCT ename FROM employee ORDER BY ename")
        self.insert_combobox_initiate(self.cb_address_tname, "SELECT DISTINCT tname FROM employeetype ORDER BY tname")
        self.insert_combobox_initiate(self.cb_address_class2, "SELECT DISTINCT class2 FROM employee ORDER BY class2")

    # Initiate Combo_Box 
    def insert_combobox_initiate(self, combo_box, sql_query):
        self.combobox_initializing(combo_box, sql_query) 
        self.lbl_address_id.setText("")
        self.entry_address_code.setText("")
        self.entry_address_class1.setText("")
        self.cb_address_ename.setCurrentIndex(0) 
        self.cb_address_tname.setCurrentIndex(0) 
        self.cb_address_class2.setCurrentIndex(0) 


    # Connect button to method
    def conn_button_to_method(self):
        self.pb_address_show.clicked.connect(self.make_data)
        self.pb_address_search.clicked.connect(self.search_data)        
        self.pb_address_close.clicked.connect(self.close_dialog)
        self.pb_address_clear.clicked.connect(self.clear_data)

        self.pb_address_insert.clicked.connect(self.tb_insert)
        self.pb_address_update.clicked.connect(self.tb_update)
        self.pb_address_delete.clicked.connect(self.tb_delete)
        self.pb_address_excel_export.clicked.connect(self.export_table)

    # Connect Signal to Slot
    def connect_signal_slot(self):
        self.cb_address_ename.activated.connect(self.cb_address_ename_changed)
        self.cb_address_tname.activated.connect(self.cb_regno_type_changed)

    # tab order for employee window
    def set_tab_order(self):
       
        widgets = [self.pb_address_show, self.entry_address_code, self.cb_address_ename,
            self.entry_address_class1, self.cb_address_tname, self.cb_address_class2, 
            self.entry_address_no, self.entry_address_remark, 
            self.pb_address_search, self.pb_address_clear, self.pb_address_insert, 
            self.pb_address_update, self.pb_address_delete, self.pb_address_close]
        
        for i in range(len(widgets) - 1):
            self.setTabOrder(widgets[i], widgets[i + 1])

    # To reduce duplications
    def common_query_statement(self):
        tv_widget = self.tv_address
        self.cursor.execute("SELECT * FROM vw_empaddress WHERE 1=0")
        column_info = self.cursor.description
        column_names = [col[0] for col in column_info]

        sql_query = "Select * From vw_empaddress Order By id"
        column_widths = [80, 100, 100, 50, 60, 50, 300, 100]

        return sql_query, tv_widget, column_info, column_names, column_widths 

    # show employee table data inside of the MDI
    def make_data(self):
        sql_query, tv_widget, column_info, column_names, column_widths = self.common_query_statement() 
        self.populate_dialog(self.cursor, sql_query, tv_widget, column_info, column_names,column_widths)

    # Get the value of other variables
    def get_regno_input(self):
        ecode = int(self.entry_address_code.text())
        addressno = str(self.entry_address_no.text())
        remark = str(self.entry_address_remark.text())

        return ecode, addressno, remark

    # Make Common values set
    def common_values_set(self):
        username = self.current_username
        user_id = self.userID_gen(username)
        formatted_datetime = self.dt_time_info()

        return username, user_id, formatted_datetime    

    # insert new employee data to MySQL table
    def tb_insert(self):
        currentid = self.lbl_address_id.text()
        if not currentid:
            confirm_dialog = self.show_insert_confirmation_dialog()
            
            if confirm_dialog == QMessageBox.Yes:

                idx = self.max_row_id("employeeaddress")
                username, user_id, formatted_datetime = self.common_values_set()
                ecode, addressno, remark = self.get_regno_input()  
                
                if (idx>0 and ecode>0) and all(len(var) > 0 for var in (addressno)):
                
                    self.cursor.execute('''INSERT INTO employeeaddress (id, ecode, address, remark) 
                                VALUES (?, ?, ?, ?)'''
                                , (idx, ecode, addressno, remark))
                    self.conn.commit()
                    self.show_insert_success_message()
                    self.refresh_data() 
                    logging.info(f"User {username} inserted id number {idx}, at the address no table.")
                else:
                    self.show_missing_message("입력 이상")
                    return
            else:
                self.show_cancel_message("데이터 추가 취소")
                return
        else:
            QMessageBox.information(self, "Input Error", "기존 id가 선택된 상태에서는 신규 입력이 불가합니다!")
            return
        
    # revise the values in the selected row
    def tb_update(self):

        confirm_dialog = self.show_update_confirmation_dialog()

        if confirm_dialog == QMessageBox.Yes:
            idx = int(self.lbl_address_id.text())
            username, user_id, formatted_datetime = self.common_values_set()         
            ecode, addressno, remark = self.get_regno_input()
            
            if (idx>0 and ecode>0) and all(len(var) > 0 for var in (addressno)):
                self.cursor.execute('''UPDATE employeeaddress SET ecode=?, address=?, remark=? WHERE id=?'''
                            , (ecode, addressno, remark, idx))
                self.conn.commit()
                self.show_update_success_message()
                self.refresh_data()  
                logging.info(f"User {username} updated row number {idx} in the address no table.")            
            else:
                self.show_missing_message("입력 이상")
                return
        else:
            self.show_cancel_message("데이터 변경 취소")
            return
        
    # delete row according to id selected
    def tb_delete(self):
        confirm_dialog = self.show_delete_confirmation_dialog()

        if confirm_dialog == QMessageBox.Yes:
            idx = self.lbl_address_id.text()
            username, user_id, formatted_datetime = self.common_values_set()

            self.cursor.execute("DELETE FROM employeeaddress WHERE id=?", (idx,))
            self.conn.commit()
            self.show_delete_success_message()
            self.refresh_data()  
            logging.info(f"User {username} deleted id number {idx}, at the address no table.")       
        else:
            self.show_cancel_message("데이터 삭제 취소")

    # Search data
    def search_data(self):
        ename = self.cb_address_ename.currentText()
        tname = self.cb_address_tname.currentText()
        class2 = self.cb_address_class2.currentText()
        
        conditions = {'v01': (ename, "ename like '%{}%'"),
                      'v02': (tname, "tname like '%{}%'"),
                      'v03': (class2, "class2 like '%{}%'"),
                      }
        
        selected_conditions = []

        for key, (value, condition_format) in conditions.items():
            if len(value) > 0:
                selected_conditions.append(condition_format.format(value))

        if not selected_conditions:
            QMessageBox.about(self, "검색 조건 확인", "검색 조건이 비어 있습니다!")
            return

        # Join the selected conditions to form the SQL query
        query = f"SELECT * FROM vw_empaddress WHERE {' AND '.join(selected_conditions)} ORDER BY ename"

        QMessageBox.about(self, "검색 조건 확인", f"직원명: {ename}\n 구분1:{tname}\n 구분2:{class2} \n\n위 조건으로 검색을 수행합니다!")
        
        sql_query, tv_widget, column_info, column_names, column_widths = self.common_query_statement()
        self.populate_dialog(self.cursor, query, tv_widget, column_info, column_names,column_widths)

    # Combobox apt type index changed
    def cb_address_ename_changed(self):
        self.entry_address_code.clear()
        selected_item = self.cb_address_ename.currentText()

        if selected_item:
            query = f"SELECT DISTINCT ecode From employee WHERE ename ='{selected_item}'"
            line_edit_widgets = [self.entry_address_code]
        
            # Check if any line edit widgets are provided
            if line_edit_widgets:
                self.lineEdit_contents(line_edit_widgets, query)
            else:
                pass

    def cb_regno_type_changed(self):
        self.entry_address_class1.clear()
        selected_item = self.cb_address_tname.currentText()

        if selected_item:
            query = f"SELECT DISTINCT tcode From employeetype WHERE tname ='{selected_item}'"
            line_edit_widgets = [self.entry_address_class1]
        
            # Check if any line edit widgets are provided
            if line_edit_widgets:
                self.lineEdit_contents(line_edit_widgets, query)
            else:
                pass

    # Export data to Excel sheet                
    def export_table(self):
        output_subfolder = "data_list"          # set the output subfoler name
        table_widget = self.tv_address         # set the name of table widget
        sheet_name = "phone_no"                 # set the excel sheet name
        filename = get_file_name(output_subfolder, sheet_name)    # get the file name / export_to_excel 밑에 두면 get_file_name을 두 번 실행시켜서 파일의 번호가 달라지므로 미리 만들어서 filename에 담아둠.
        numeric_columns=[0, 1]                     # set the numerical column index
        export_to_excel(output_subfolder, table_widget, sheet_name, numeric_columns)
               
        if filename:
            full_file_path = os.path.join(output_subfolder, filename)
            self.excel_formatting(sheet_name, full_file_path)
        else:
            QMessageBox.about(self, "코드 확인", "파일을 찾지 못했습니다!")    

    # Excel sheet formatting
    def excel_formatting(self, sheet_name, full_file_path):
        wb = load_workbook(full_file_path)
        sheet_name = sheet_name
        ws = wb[sheet_name]
        last_row = ws.max_row + 1

        column_widths = [6, 10, 12, 8, 10, 8, 15, 20]                # set the width of each column
        title_font = Font(bold=True, name="Arial", size=10)     # set the font for the first row
        arial_font = Font(name="Arial", size=10)                # set the forn from the second row to max row

        set_column_widths(ws, column_widths)    # reset column widths

        ws.freeze_panes = 'D2'              # freeze panes D2 cell
        ws.auto_filter.ref = ws.dimensions  # apply auto filter
        ws.sheet_view.showGridLines = False # remove gridlines
            
        set_font(ws, [1], range(1, len(column_widths) + 1), title_font)                 # first row font setting
        set_font(ws, range(2, last_row), range(1, len(column_widths) + 1), arial_font)  # from second row to last row font setting
        
        wb.save(full_file_path)
        QMessageBox.about(self, "파일 생성 완료", f"data_list folder에 \n엑셀 파일이 {full_file_path}로 \n생성 되었습니다!")   

    # clear input field entry
    def clear_data(self):
        self.lbl_address_id.setText("")
        for line_edit in self.findChildren(QtWidgets.QLineEdit):
            line_edit.clear()

    # table widget cell double click
    def show_selected_data(self, item):
        # Get the row index of the clicked item
        row_index = item.row()

        # Initialize a list to store the cell values
        cell_values = []

        # Loop through the columns and retrieve the text from each cell
        for column_index in range(8):  # 8 columns
            cell_text = self.tv_address.model().item(row_index, column_index).text()
            cell_values.append(cell_text)

        # Populate the input widgets with the data from the selected row
        self.lbl_address_id.setText(cell_values[0])
        self.entry_address_code.setText(cell_values[1])
        self.cb_address_ename.setCurrentText(cell_values[2])
        self.entry_address_class1.setText(cell_values[3])
        self.cb_address_tname.setCurrentText(cell_values[4])
        self.cb_address_class2.setCurrentText(cell_values[5])
        self.entry_address_no.setText(cell_values[6])
        self.entry_address_remark.setText(cell_values[7])
    
    def refresh_data(self):
        self.clear_data()
        self.make_data()


if __name__ == "__main__":

    log_subfolder = "logs"
    os.makedirs(log_subfolder, exist_ok=True)
    log_file_path = os.path.join(log_subfolder, "access_EmployeeAddress.log")

    logging.basicConfig(
        filename=log_file_path,  
        level=logging.INFO,    
        format="%(asctime)s [%(levelname)s] - %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
        
    app = QtWidgets.QApplication(sys.argv)
    dialog = EmployeeAddressDialog()
    dialog.show()
    sys.exit(app.exec())