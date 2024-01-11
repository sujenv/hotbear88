import sys
import logging
from functools import partial
from openpyxl.styles import Font
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from PyQt5 import uic, QtWidgets, QtCore
from PyQt5.QtGui import QStandardItemModel, QKeySequence
from PyQt5.QtWidgets import QDialog, QMessageBox, QShortcut
from PyQt5.QtCore import Qt
from datetime import datetime
from commonmd import *
#for non_ui version-------------------------
#from customer_address_ui import Ui_CustomerAddressDialog

# Customer Address table contents -----------------------------------------------------
class CustomerAddressDialog(QDialog, SubWindowBase):
#for non_ui version-------------------------
#class CustomerAddressDialog(QDialog, Ui_CustomerAddressDialog, SubWindowBase):
     
    def __init__(self, current_username = None, current_datetime = None):
        super().__init__()    
  
        self.conn, self.cursor = connect_to_database1()

        # load ui file
        uic.loadUi("customer_address.ui", self)
        #for non_ui version-------------------------
        #self.setupUi(self)

        # Add window flags
        self.setWindowFlags(self.windowFlags() | Qt.WindowMinimizeButtonHint | Qt.WindowMaximizeButtonHint | Qt.WindowCloseButtonHint)  

        # Initialize current_username and current_datetime directly
        self.current_username, self.current_datetime = initialize_username_and_datetime(current_username, current_datetime)

        # Enable automatic deletion on close
        self.setAttribute(QtCore.Qt.WA_DeleteOnClose)  

        # Create tv_caddress and QSortFilterProxyModel
        self.model = QStandardItemModel()
        self.proxy_model = NumericStringSortModel(self.model)
        self.proxy_model.setSourceModel(self.model)

        # Define the column types
        column_types = ["numeric", "numeric", "", "", "", "", "",]

        # Set the custom delegate for the specific column
        delegate = NumericDelegate(column_types,self.tv_caddress)
        self.tv_caddress.setItemDelegate(delegate)
        self.tv_caddress.setModel(self.proxy_model)

        # Enable sorting
        self.tv_caddress.setSortingEnabled(True)

        # Enable alternating row colors
        self.tv_caddress.setAlternatingRowColors(True)  
        
        # Hide the first index column
        self.tv_caddress.verticalHeader().setVisible(False)

        # While selecting row in tv_caddress, each cell values to displayed to designated widgets
        self.tv_caddress.clicked.connect(self.show_selected_data)

        # Fill combobox items when the application starts
        self.get_combobox_contents()

        # Initiate display of data
        self.make_data()
        self.conn_button_to_method()
        self.connect_signal_slot()

        # Set tab order for input widgets
        self.set_tab_order()

        # Initiate CTRL+C, CTRL+V and ENTER
        self.setup_shortcuts()

        # Make log file
        self.make_logfiles("access_CustomerAddress.log")

    # Create a shortcut for CTRL+C/CTRL+V/ Return key
    def setup_shortcuts(self):
        self.copy_shortcut = QShortcut(QKeySequence.Copy, self.tv_caddress, partial(self.copy_cells, self.tv_caddress))
        self.paste_shortcut = QShortcut(QKeySequence.Paste, self.tv_caddress, partial(self.paste_cells, self.tv_caddress))
        self.return_shortcut = QShortcut(Qt.Key_Return, self.tv_caddress, partial(self.handle_return_key, self.tv_caddress))

    # Call process_key_event and pass the event and your QTableWidget instance
    def keyPressEvent(self, event):
        tv_widget = self.tv_caddress
        self.process_key_event(event, tv_widget)

    # Pass combobox info and sql to next method
    def get_combobox_contents(self):
        self.insert_combobox_initiate(self.cb_caddress_cname, "SELECT DISTINCT cname FROM customer ORDER BY cname")
        self.insert_combobox_initiate(self.cb_caddress_tname, "SELECT DISTINCT type01 FROM customer ORDER BY type01")
        self.insert_combobox_initiate(self.cb_caddress_active, "SELECT DISTINCT active FROM customer ORDER BY active")

    # Initiate Combo_Box 
    def insert_combobox_initiate(self, combo_box, sql_query):
        self.combobox_initializing(combo_box, sql_query) 
        self.lbl_caddress_id.setText("")
        self.entry_caddress_code.setText("")
        self.cb_caddress_cname.setCurrentIndex(0) 
        self.cb_caddress_tname.setCurrentIndex(0) 
        self.cb_caddress_active.setCurrentIndex(0) 

    # Connect button to method
    def conn_button_to_method(self):
        self.pb_caddress_show.clicked.connect(self.make_data)
        self.pb_caddress_search.clicked.connect(self.search_data)        
        self.pb_caddress_close.clicked.connect(self.close_dialog)
        self.pb_caddress_clear.clicked.connect(self.clear_data)

        self.pb_caddress_insert.clicked.connect(self.tb_insert)
        self.pb_caddress_update.clicked.connect(self.tb_update)
        self.pb_caddress_delete.clicked.connect(self.tb_delete)
        self.pb_caddress_excel_export.clicked.connect(self.export_table)

    # Connect Signal to Slot
    def connect_signal_slot(self):
        self.cb_caddress_cname.activated.connect(self.cb_regno_cname_changed)

    # tab order for employee window
    def set_tab_order(self):
       
        widgets = [self.pb_caddress_show, self.entry_caddress_code, self.cb_caddress_cname,
            self.cb_caddress_tname, self.cb_caddress_active, self.entry_caddress_address, 
            self.entry_caddress_remark, 
            self.pb_caddress_search, self.pb_caddress_clear, self.pb_caddress_insert, 
            self.pb_caddress_update, self.pb_caddress_delete, self.pb_caddress_close]
        
        for i in range(len(widgets) - 1):
            self.setTabOrder(widgets[i], widgets[i + 1])

    # To reduce duplications
    def common_query_statement(self):
        tv_widget = self.tv_caddress
        self.cursor.execute("SELECT * FROM vw_customeraddress WHERE 1=0")
        column_info = self.cursor.description
        column_names = [col[0] for col in column_info]

        sql_query = "Select * From vw_customeraddress Order By id"
        column_widths = [100, 100, 150, 200, 60, 60, 150]

        return sql_query, tv_widget, column_info, column_names, column_widths 

    # show employee table data inside of the MDI
    def make_data(self):
        sql_query, tv_widget, column_info, column_names, column_widths = self.common_query_statement() 
        self.populate_dialog(self.cursor, sql_query, tv_widget, column_info, column_names,column_widths)

    # Get the value of other variables
    def get_regno_input(self):
        ccode = int(self.entry_caddress_code.text())
        address = str(self.entry_caddress_address.text())
        remark = str(self.entry_caddress_remark.text())

        return ccode, address, remark

    # Make Common values set
    def common_values_set(self):
        username = self.current_username
        user_id = self.userID_gen(username)
        formatted_datetime = self.dt_time_info()

        return username, user_id, formatted_datetime    

    # insert new employee data to MySQL table
    def tb_insert(self):
        currentid = self.lbl_caddress_id.text()
        if not currentid:
            confirm_dialog = self.show_insert_confirmation_dialog()
            
            if confirm_dialog == QMessageBox.Yes:

                idx = self.max_row_id("customeraddress")
                username, user_id, formatted_datetime = self.common_values_set()
                ccode, address, remark = self.get_regno_input()  
                
                if (idx>0 and ccode>0) and all(len(var) > 0 for var in (address)):
                
                    self.cursor.execute('''INSERT INTO customeraddress (id, ccode, address, remark) 
                                VALUES (?, ?, ?, ?)'''
                                , (idx, ccode, address, remark))
                    self.conn.commit()
                    self.show_insert_success_message()
                    self.refresh_data() 
                    logging.info(f"User {username} inserted id number {idx}, at the customer address table.")
                else:
                    self.show_missing_message("입력 이상")
                    return
            else:
                self.show_cancel_message("데이터 추가 취소")
                return
        else:
            QMessageBox.information(self, "Input Error", "기존 id가 선택된 상태에서는 신규 입력이 불가합니다!")
            return
        
    # revise the values in the selected row
    def tb_update(self):

        confirm_dialog = self.show_update_confirmation_dialog()

        if confirm_dialog == QMessageBox.Yes:
            idx = int(self.lbl_caddress_id.text())
            username, user_id, formatted_datetime = self.common_values_set()         
            ccode, address, remark = self.get_regno_input()
            
            if (idx>0 and ccode>0) and all(len(var) > 0 for var in (address)):
                self.cursor.execute('''UPDATE customeraddress SET ccode=?, address=?, remark=? WHERE id=?'''
                            , (ccode, address, remark, idx))
                self.conn.commit()
                self.show_update_success_message()
                self.refresh_data()  
                logging.info(f"User {username} updated row number {idx} in the customer address table.")            
            else:
                self.show_missing_message("입력 이상")
                return
        else:
            self.show_cancel_message("데이터 변경 취소")
            return
        
    # delete row according to id selected
    def tb_delete(self):
        confirm_dialog = self.show_delete_confirmation_dialog()

        if confirm_dialog == QMessageBox.Yes:
            idx = self.lbl_caddress_id.text()
            username, user_id, formatted_datetime = self.common_values_set()

            self.cursor.execute("DELETE FROM customeraddress WHERE id=?", (idx,))
            self.conn.commit()
            self.show_delete_success_message()
            self.refresh_data()  
            logging.info(f"User {username} deleted id number {idx}, at the customer address table.")       
        else:
            self.show_cancel_message("데이터 삭제 취소")

    # Search data
    def search_data(self):
        cname = self.cb_caddress_cname.currentText()
        tname = self.cb_caddress_tname.currentText()
        active = self.cb_caddress_active.currentText()
        address = self.entry_caddress_address.text()
        
        conditions = {'v01': (cname, "cname like '%{}%'"),
                      'v02': (tname, "type01 = '{}'"),
                      'v03': (active, "active = '{}'"),
                      'v04': (address, "address like '%{}%'"),
                      }
        
        selected_conditions = []

        for key, (value, condition_format) in conditions.items():
            if len(value) > 0:
                selected_conditions.append(condition_format.format(value))

        if not selected_conditions:
            QMessageBox.about(self, "검색 조건 확인", "검색 조건이 비어 있습니다!")
            return

        # Join the selected conditions to form the SQL query
        query = f"SELECT * FROM vw_customeraddress WHERE {' AND '.join(selected_conditions)} ORDER BY cname"

        QMessageBox.about(self, "검색 조건 확인", f"업체명: {cname} \n구분:{tname} \n현행:{active} \n주소: {address} \n\n위 조건으로 검색을 수행합니다!")
        
        sql_query, tv_widget, column_info, column_names, column_widths = self.common_query_statement()
        self.populate_dialog(self.cursor, query, tv_widget, column_info, column_names,column_widths)

    # Combobox apt type index changed
    def cb_regno_cname_changed(self):
        self.entry_caddress_code.clear()
        selected_item = self.cb_caddress_cname.currentText()

        if selected_item:
            query = f"SELECT DISTINCT ccode From customer WHERE cname ='{selected_item}'"
            line_edit_widgets = [self.entry_caddress_code]
        
            # Check if any line edit widgets are provided
            if line_edit_widgets:
                self.lineEdit_contents(line_edit_widgets, query)
            else:
                pass

    # Export data to Excel sheet                
    def export_table(self):
        output_subfolder = "data_list"          # set the output subfoler name
        table_widget = self.tv_caddress           # set the name of table widget
        sheet_name = "customeraddress"          # set the excel sheet name
        filename = get_file_name(output_subfolder, sheet_name)    # get the file name / export_to_excel 밑에 두면 get_file_name을 두 번 실행시켜서 파일의 번호가 달라지므로 미리 만들어서 filename에 담아둠.
        numeric_columns=[0, 1]                     # set the numerical column index
        export_to_excel(output_subfolder, table_widget, sheet_name, numeric_columns)
               
        if filename:
            full_file_path = os.path.join(output_subfolder, filename)
            self.excel_formatting(sheet_name, full_file_path)
        else:
            QMessageBox.about(self, "코드 확인", "파일을 찾지 못했습니다!")    

    # Excel sheet formatting
    def excel_formatting(self, sheet_name, full_file_path):
        wb = load_workbook(full_file_path)
        sheet_name = sheet_name
        ws = wb[sheet_name]
        last_row = ws.max_row + 1

        column_widths = [6, 10, 12, 8, 10, 8, 15, 10, 20]                # set the width of each column
        title_font = Font(bold=True, name="Arial", size=10)     # set the font for the first row
        arial_font = Font(name="Arial", size=10)                # set the forn from the second row to max row

        set_column_widths(ws, column_widths)    # reset column widths

        ws.freeze_panes = 'D2'              # freeze panes D2 cell
        ws.auto_filter.ref = ws.dimensions  # apply auto filter
        ws.sheet_view.showGridLines = False # remove gridlines
            
        set_font(ws, [1], range(1, len(column_widths) + 1), title_font)                 # first row font setting
        set_font(ws, range(2, last_row), range(1, len(column_widths) + 1), arial_font)  # from second row to last row font setting
        
        wb.save(full_file_path)
        QMessageBox.about(self, "파일 생성 완료", f"data_list folder에 \n엑셀 파일이 {full_file_path}로 \n생성 되었습니다!")   

    # clear input field entry
    def clear_data(self):
        self.lbl_caddress_id.setText("")
        for line_edit in self.findChildren(QtWidgets.QLineEdit):
            line_edit.clear()

    # table widget cell double click
    def show_selected_data(self, item):
        # Get the row index of the clicked item
        row_index = item.row()

        # Initialize a list to store the cell values
        cell_values = []

        # Loop through the columns and retrieve the text from each cell
        for column_index in range(7):  # 7 columns
            cell_text = self.tv_caddress.model().item(row_index, column_index).text()
            cell_values.append(cell_text)

        # Populate the input widgets with the data from the selected row
        self.lbl_caddress_id.setText(cell_values[0])
        self.entry_caddress_code.setText(cell_values[1])
        self.cb_caddress_cname.setCurrentText(cell_values[2])
        self.entry_caddress_address.setText(cell_values[3])
        self.cb_caddress_tname.setCurrentText(cell_values[4])
        self.cb_caddress_active.setCurrentText(cell_values[5])
        self.entry_caddress_remark.setText(cell_values[6])
    
    def refresh_data(self):
        self.clear_data()
        self.make_data()


if __name__ == "__main__":

    log_subfolder = "logs"
    os.makedirs(log_subfolder, exist_ok=True)
    log_file_path = os.path.join(log_subfolder, "access_CustomerAddress.log")

    logging.basicConfig(
        filename=log_file_path,  
        level=logging.INFO,    
        format="%(asctime)s [%(levelname)s] - %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
        
    app = QtWidgets.QApplication(sys.argv)
    dialog = CustomerAddressDialog()
    dialog.show()
    sys.exit(app.exec())