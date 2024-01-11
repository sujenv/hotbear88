import sys
import logging
from functools import partial
from openpyxl.styles import Font
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from PyQt5 import uic, QtWidgets, QtCore
from PyQt5.QtGui import QStandardItemModel, QKeySequence
from PyQt5.QtWidgets import QDialog, QMessageBox, QMenu, QInputDialog, QShortcut
from PyQt5.QtCore import Qt
from datetime import datetime
from datetime import timedelta
from datetime import date
from commonmd import *
from cal import CalendarView
#for non_ui version-------------------------
#from salesprice_ui import Ui_SalesPriceDialog

# Sales Item List-----------------------------------------------------------------
class SalesPriceDialog(QDialog, SubWindowBase):
#for non_ui version-------------------------
#class SalesPriceDialog(QDialog, Ui_SalesPriceDialog, SubWindowBase):

    def __init__(self, current_username = None, current_datetime = None):
        super().__init__()

        self.conn, self.cursor = connect_to_database1()

        # Load ui file 
        uic.loadUi("salesprice.ui", self)
        #for non_ui version-------------------------
        #self.setupUi(self)

        # Add window flags
        self.setWindowFlags(self.windowFlags() | Qt.WindowMinimizeButtonHint | Qt.WindowMaximizeButtonHint | Qt.WindowCloseButtonHint)  

        # Initialize current_username and current_datetime directly
        self.current_username, self.current_datetime = initialize_username_and_datetime(current_username, current_datetime)

        # Enable automatic deletion on close
        self.setAttribute(QtCore.Qt.WA_DeleteOnClose)  

        # Create tv_salesprice and QSortFilterProxyModel
        self.model = QStandardItemModel()
        self.proxy_model = NumericStringSortModel(self.model)
        self.proxy_model.setSourceModel(self.model)        
        
        # Define the column types
        column_types = ["numeric", "numeric", "", "numeric", "", "numeric", "", "numeric", "", "", "", ""]

        # Set the custom delegate for the specific column
        delegate = NumericDelegate(column_types, self.tv_salesprice)
        self.tv_salesprice.setItemDelegate(delegate)
        self.tv_salesprice.setModel(self.proxy_model)
       
        # Enable sorting
        self.tv_salesprice.setSortingEnabled(True)

        # Enable alternating row colors
        self.tv_salesprice.setAlternatingRowColors(True)  
        
        # Hide the first index column
        self.tv_salesprice.verticalHeader().setVisible(False)

        # While selecting row in tv_salesprice, each cell values to displayed to designated widgets
        self.tv_salesprice.clicked.connect(self.show_selected_data)


        # Fill combobox items when the application starts
        self.get_combobox_contents()

        # Initiate display of data
        self.make_data()
        self.conn_button_to_method()
        self.connect_signal_slot()

        # Set tab order for input widgets
        self.set_tab_order()

        # Initiate CTRL+C, CTRL+V and ENTER
        self.setup_shortcuts()

        # Create context menus ----------------------------------------------------------------------------------
        self.context_menu1 = self.create_context_menu(self.entry_salesprice_efffrom)
        self.context_menu2 = self.create_context_menu(self.entry_salesprice_effthru)
        self.context_menu3 = self.create_context_menu(self.entry_salespricechange_efffrom)
        self.context_menu4 = self.create_context_menu(self.entry_salespricechange_effthru)

        self.entry_salesprice_efffrom.setContextMenuPolicy(Qt.CustomContextMenu)
        self.entry_salesprice_efffrom.customContextMenuRequested.connect(self.show_context_menu1)

        self.entry_salesprice_effthru.setContextMenuPolicy(Qt.CustomContextMenu)
        self.entry_salesprice_effthru.customContextMenuRequested.connect(self.show_context_menu2)

        self.entry_salespricechange_efffrom.setContextMenuPolicy(Qt.CustomContextMenu)
        self.entry_salespricechange_efffrom.customContextMenuRequested.connect(self.show_context_menu3)

        self.entry_salespricechange_effthru.setContextMenuPolicy(Qt.CustomContextMenu)
        self.entry_salespricechange_effthru.customContextMenuRequested.connect(self.show_context_menu4)
        #-----------------------------------------------------------------------------------------------------------------

        self.entry_stylesheet_as_is()
        self.hide_salespricechange_widget()

        # Make log file
        self.make_logfiles("access_salesprice.log")

    # Create a shortcut for CTRL+C/CTRL+V/ Return key
    def setup_shortcuts(self):
        self.copy_shortcut = QShortcut(QKeySequence.Copy, self.tv_salesprice, partial(self.copy_cells, self.tv_salesprice))
        self.paste_shortcut = QShortcut(QKeySequence.Paste, self.tv_salesprice, partial(self.paste_cells, self.tv_salesprice))
        self.return_shortcut = QShortcut(Qt.Key_Return, self.tv_salesprice, partial(self.handle_return_key, self.tv_salesprice))

    # Change styles for the selected widgets 1
    def entry_stylesheet_as_is(self):
        self.entry_salespricechange_price.setStyleSheet('color:black;background:rgb(255,255,255)')
        self.entry_salespricechange_efffrom.setStyleSheet('color:black;background:rgb(255,255,255)')
        self.entry_salespricechange_effthru.setStyleSheet('color:black;background:rgb(255,255,255)')
        self.entry_salespricechange_remark.setStyleSheet('color:black;background:rgb(255,255,255)')

    # Change styles for the selected widgets 2
    def entry_stylesheet_to_be(self):
        self.entry_salespricechange_price.setStyleSheet('color:white;background:rgb(255,0,0)')
        self.entry_salespricechange_efffrom.setStyleSheet('color:white;background:rgb(255,0,0)')
        self.entry_salespricechange_effthru.setStyleSheet('color:white;background:rgb(255,0,0)')
        self.entry_salespricechange_remark.setStyleSheet('color:white;background:rgb(255,0,0)')

    # Show widgets for the salesprice change parts 
    def show_salespricechange_widget(self):
        ddt, endofdate = self.display_eff_date()
        self.pb_salespricechange_update.setVisible(True)
        self.entry_salespricechange_price.setReadOnly(False)
        self.entry_salespricechange_efffrom.setReadOnly(False)
        self.entry_salespricechange_effthru.setReadOnly(False)
        self.entry_salespricechange_remark.setReadOnly(False)
        self.entry_salespricechange_effthru.setText(endofdate)
        self.entry_stylesheet_to_be()

    # Hide widgets for the salesprice change parts 
    def hide_salespricechange_widget(self):
        self.pb_salespricechange_update.setVisible(False)
        self.entry_salespricechange_price.setReadOnly(True)
        self.entry_salespricechange_efffrom.setReadOnly(True)
        self.entry_salespricechange_effthru.setReadOnly(True)
        self.entry_salespricechange_remark.setReadOnly(True)        
        self.entry_stylesheet_as_is()

    # Mouse Right click, show "달력보기" menu ---------------------------------------------------------------------
    def create_context_menu(self, target_lineedit):
        context_menu = QMenu()
        custom_action = context_menu.addAction("달력보기")
        custom_action.triggered.connect(lambda: self.show_calendar(target_lineedit))
        return context_menu

    def show_context_menu1(self, pos):
        self.context_menu1.exec_(self.entry_salesprice_efffrom.mapToGlobal(pos))

    def show_context_menu2(self, pos):
        self.context_menu2.exec_(self.entry_salesprice_effthru.mapToGlobal(pos))

    def show_context_menu3(self, pos):
        self.context_menu3.exec_(self.entry_salespricechange_efffrom.mapToGlobal(pos))

    def show_context_menu4(self, pos):
        self.context_menu4.exec_(self.entry_salespricechange_effthru.mapToGlobal(pos))

    # Show Calendar
    def show_calendar(self, target_lineedit):
        calendar_dialog = CalendarView()
        calendar_dialog.selected_date_changed.connect(lambda date: self.set_selected_date(date, target_lineedit))
        calendar_dialog.exec()

    # Show selected date to the select Qlineedit
    def set_selected_date(self, date, target_lineedit):
        if target_lineedit == self.entry_salesprice_efffrom:
            target_lineedit.setText(date)
        elif target_lineedit == self.entry_salesprice_effthru:
            target_lineedit.setText(date)
        elif target_lineedit == self.entry_salespricechange_efffrom:
            target_lineedit.setText(date)
        elif target_lineedit == self.entry_salespricechange_effthru:
            target_lineedit.setText(date)     

    # Call process_key_event and pass the event and your QTableWidget instance
    def keyPressEvent(self, event):
        tv_widget = self.tv_salesprice
        self.process_key_event(event, tv_widget)

    # Display current date & end of date
    def display_eff_date(self):
        now = datetime.now()
        curr_date = now.strftime("%Y/%m/%d")
        ddt = f"{curr_date}"         
        endofdate = "2050/12/31"

        return ddt, endofdate

    # Pass combobox info and sql to next method
    def get_combobox_contents(self):
        self.insert_combobox_initiate(self.cb_salesprice_cname, "SELECT DISTINCT cname FROM vw_customer ORDER BY cname")
        self.insert_combobox_initiate(self.cb_salesprice_iname, "SELECT DISTINCT iname FROM salesitem ORDER BY iname")
        self.insert_combobox_initiate(self.cb_salesprice_dname, "SELECT DISTINCT dname FROM delivery ORDER BY dname")

    # Initiate Combo_Box 
    def insert_combobox_initiate(self, combo_box, sql_query):
        self.combobox_initializing(combo_box, sql_query) 
        self.lbl_salesprice_id.setText("")
        self.entry_salesprice_ccode.setText("")
        self.cb_salesprice_cname.setCurrentIndex(0) 
        self.cb_salesprice_iname.setCurrentIndex(0) 
        self.cb_salesprice_dname.setCurrentIndex(0) 

    # Connect the button to the method
    def conn_button_to_method(self):
        self.pb_salesprice_show.clicked.connect(self.make_data)
        self.pb_salesprice_search.clicked.connect(self.search_data) 
        self.pb_salesprice_currentprice.clicked.connect(self.current_price)
        self.pb_salesprice_close.clicked.connect(self.close_dialog)
        self.pb_salesprice_clear.clicked.connect(self.clear_data)
        self.pb_salesprice_insert.clicked.connect(self.tb_insert)
        self.pb_salesprice_update.clicked.connect(self.tb_update)
        self.pb_salesprice_delete.clicked.connect(self.tb_delete)
        self.pb_salesprice_excel_export.clicked.connect(self.export_table)
        self.pb_salespricechange_update.clicked.connect(self.reflect_salesprice_change)

    # Connect Signal to Slot
    def connect_signal_slot(self):
        self.cb_salesprice_cname.activated.connect(self.cb_salesprice_cname_changed)
        self.cb_salesprice_iname.activated.connect(self.cb_salesprice_iname_changed)
        self.cb_salesprice_dname.activated.connect(self.cb_salesprice_dname_changed)
        
        self.entry_salespricechange_efffrom.editingFinished.connect(self.salespricechange_efffrom_change)
                
    # Tab order for sub window
    def set_tab_order(self):
        widgets = [self.pb_salesprice_show, self.entry_salesprice_ccode, self.cb_salesprice_cname,
            self.entry_salesprice_icode, self.cb_salesprice_iname, self.entry_salesprice_dcode,
            self.cb_salesprice_dname, self.entry_salesprice_price, self.entry_salesprice_um,
            self.entry_salesprice_efffrom, self.entry_salesprice_effthru, self.entry_salesprice_remark, 
            self.pb_salesprice_search, self.pb_salesprice_currentprice, self.pb_salesprice_clear, 
            self.pb_salesprice_insert, self.pb_salesprice_update, self.pb_salesprice_delete, 
            self.pb_salesprice_close, self.entry_salespricechange_price, self.entry_salespricechange_efffrom,
            self.entry_salespricechange_effthru, self.entry_salespricechange_remark, self.pb_salespricechange_update]

        for i in range(len(widgets) - 1):
            self.setTabOrder(widgets[i], widgets[i + 1])

    # To reduce duplications
    def common_query_statement(self):
        tv_widget = self.tv_salesprice
        self.cursor.execute("Select id, ccode, cname, icode, iname, dcode, dname, price, um, efffrom, effthru, remark From vw_unit_price Where 1=0")
        column_info = self.cursor.description
        column_names = [col[0] for col in column_info]  
        
        sql_query = "Select id, ccode, cname, icode, iname, dcode, dname, price, um, efffrom, effthru, remark From vw_unit_price order by id"
        column_widths = [60, 80, 150, 80]

        return sql_query, tv_widget, column_info, column_names, column_widths

    # show sales price table data inside of the MDI
    def make_data(self):
        sql_query, tv_widget, column_info, column_names, column_widths = self.common_query_statement() 
        self.populate_dialog(self.cursor, sql_query, tv_widget, column_info, column_names,column_widths)

    # Get the value of other variables
    def get_salesprice_input(self):
        ccode = int(self.entry_salesprice_ccode.text())
        icode = int(self.entry_salesprice_icode.text())
        dcode = int(self.entry_salesprice_dcode.text())
        price = float(self.entry_salesprice_price.text())
        efffrom = str(self.entry_salesprice_efffrom.text())
        effthru = str(self.entry_salesprice_effthru.text())
        remark = str(self.entry_salesprice_remark.text())
        return ccode, icode, dcode, price, efffrom, effthru, remark

    # Make Common values set
    def common_values_set(self):
        username = self.current_username
        user_id = self.userID_gen(username)
        formatted_datetime = self.dt_time_info()
        return username, user_id, formatted_datetime
    
    # insert new salesprice data to MySQL table
    def tb_insert(self):
        currentid = self.lbl_salesprice_id.text()
        if not currentid:
            confirm_dialog = self.show_insert_confirmation_dialog()
            if confirm_dialog == QMessageBox.Yes:
                idx = self.max_row_id("salesprice")
                username, user_id, formatted_datetime = self.common_values_set()
                ccode, icode, dcode, price, efffrom, effthru, remark = self.get_salesprice_input() 

                if (idx>0 and ccode>0 and icode>0 and dcode>0 and abs(price)>=0) and all(len(var) > 0 for var in (efffrom, effthru)):
                    self.cursor.execute('''INSERT INTO salesprice (id, ccode, icode, dcode, price, efffrom, effthru, trxdate, userid, remark) 
                                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)'''
                                , (idx, ccode, icode, dcode, price, efffrom, effthru, formatted_datetime, user_id, remark))
                    self.conn.commit()
                    self.show_insert_success_message()
                    self.refresh_data() 
                    logging.info(f"User {username} inserted {idx} row to the salesprice table.")
                else:
                    self.show_missing_message("입력 이상")
                    return
            else:
                self.show_cancel_message("데이터 추가 취소")
                return    
        else:
            QMessageBox.information(self, "Input Error", "기존 id가 선택된 상태에서는 신규 입력이 불가합니다!")
            return
        
    # update values in the selected row
    def tb_update(self):

        id_text = self.lbl_salesprice_id.text()
        if id_text.strip(): #id_text가 공백이 아닌 경우를 확인

            conA = '''단가 오류 수정 - 현재 데이터 수정, 추가 행을 만들지 않음!'''
            conB = '''단가 갱신 - 현재 데이터의 유효종료일 변경 및 신규 단가 추가!'''
            
            conditions = [conA, conB]
            condition, okPressed = QInputDialog.getItem(self, "입력 조건 선택", "어떤 작업을 진행하시겠습니까?:", conditions, 0, False)

            if okPressed:
                if condition == conA:
                    self.fix_typo()     
                elif condition == conB:
                    self.show_salespricechange_widget()
                else:
                    return

        else:
            QMessageBox.about(self, "수정 조건 확인", "선택된 행이 없습니다!")
            return

    # Fix typing error in the salesprice
    def fix_typo(self):

        confirm_dialog = self.show_update_confirmation_dialog()

        if confirm_dialog == QMessageBox.Yes:
            id_text = self.lbl_salesprice_id.text()
            idx = int(id_text)
            
            username, user_id, formatted_datetime = self.common_values_set()
            ccode, icode, dcode, price, efffrom, effthru, remark = self.get_salesprice_input() 

            #print(ccode, icode, dcode, price, efffrom, effthru, formatted_datetime, user_id, remark, idx)

            if (idx>0 and ccode>0 and icode>0 and dcode>0 and abs(price)>=0) and all(len(var) > 0 for var in (efffrom, effthru)):
                self.cursor.execute('''UPDATE salesprice SET 
                            ccode=?, icode=?, dcode=?, price=?, efffrom=?, effthru=?, trxdate=?, userid=?, remark=? WHERE id=?'''
                            , (ccode, icode, dcode, price, efffrom, effthru, formatted_datetime, user_id, remark, idx))
                self.conn.commit()
                self.show_update_success_message()
                self.refresh_data()
                logging.info(f"User {username} updated row number {idx} in the saleprice table.")
            else:
                self.show_missing_message("입력 이상")
                return
        else:
            self.show_cancel_message("데이터 변경 취소")
            return

    # Changed sales price values
    def changed_salesprice_value(self):
        chgprice = float(self.entry_salespricechange_price.text())
        chgefffrom = str(self.entry_salespricechange_efffrom.text())
        chgeffthru = str(self.entry_salespricechange_effthru.text())
        chgremark = str(self.entry_salespricechange_remark.text())

        return chgprice, chgefffrom, chgeffthru, chgremark

    # Reflect changed salesprice values 1. change salesprice effthru date for the current record 2. add newly added salesprice record for the salesprice table.
    def reflect_salesprice_change(self):
        
        idx = int(self.max_row_id("salesprice"))
        username, user_id, formatted_datetime = self.common_values_set()
        ccode, icode, dcode, price, efffrom, effthru, remark = self.get_salesprice_input() 
        chgprice, chgefffrom, chgeffthru, chgremark = self.changed_salesprice_value()

        if (idx>0 and abs(chgprice)>=0) and all(len(var) > 0 for var in (chgefffrom, chgeffthru)):
            
            #기존 가격 ID의 유효종료일을 변경유효시작일 -1 일로 수정
            org_id = str(self.lbl_salesprice_id.text())
            effthru1 = str(self.entry_salesprice_effthru.text()) #다시 불러와야 함....
            self.cursor.execute('''UPDATE salesprice SET
                            ccode=?, icode=?, dcode=?, price=?, efffrom=?, effthru=?, trxdate=?, userid=?, remark=? WHERE id=?'''
                            , (ccode, icode, dcode, price, efffrom, effthru1, formatted_datetime, user_id, remark, org_id))
            self.conn.commit()            
            
            #변경된 단가 삽입
            self.cursor.execute('''INSERT INTO salesprice (id, ccode, icode, dcode, price, efffrom, effthru, trxdate, userid, remark) 
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)'''
                            , (idx, ccode, icode, dcode, chgprice, chgefffrom, chgeffthru, formatted_datetime, user_id, chgremark))
            
            self.conn.commit()
                       
            self.show_update_success_message()
            self.refresh_data()
            logging.info(f"User {username} updated row number {idx} in the salesprice table.")
            self.hide_salespricechange_widget()
        else:
            self.show_missing_message("입력 이상")
            return

    # delete row according to id selected
    def tb_delete(self):
        confirm_dialog = self.show_delete_confirmation_dialog()

        if confirm_dialog == QMessageBox.Yes:
            idx = self.lbl_salesprice_id.text()
            username, user_id, formatted_datetime = self.common_values_set()
            self.cursor.execute("DELETE FROM salesprice WHERE id=?", (idx,))
            self.conn.commit()
            self.show_delete_success_message()
            self.refresh_data()
            logging.info(f"User {username} deleted {idx} row to the salesprice table.")
        else:
            self.show_cancel_message("데이터 삭제 취소")
            return

    # Search data
    def search_data(self):
        cname = self.cb_salesprice_cname.currentText()
        iname = self.cb_salesprice_iname.currentText()
        dname = self.cb_salesprice_dname.currentText()

        conditions = {'v01': (cname, "cname like '%{}%'"),
                    'v02': (iname, "iname like '%{}%'"),
                    'v03': (dname, "dname like '%{}%'"),}
        
        selected_conditions = []

        for key, (value, condition_format) in conditions.items():
            if len(value) > 0:
                selected_conditions.append(condition_format.format(value))

        if not selected_conditions:
            QMessageBox.about(self, "검색 조건 확인", "검색 조건이 비어 있습니다!")
            return

        # Join the selected conditions to form the SQL query
        query = f"SELECT id, ccode, cname, icode, iname, dcode, dname, price, um, efffrom, effthru, remark FROM vw_unit_price WHERE {' AND '.join(selected_conditions)}"

        QMessageBox.about(self, "검색 조건 확인", f"거래처명: {cname} \n 품목명: {iname}\n 배송방법:{dname} \n\n위 조건으로 검색을 수행합니다!")
        
        sql_query, tv_widget, column_info, column_names, column_widths = self.common_query_statement() 
        self.populate_dialog(self.cursor, query, tv_widget, column_info, column_names,column_widths)

    # Current Price Only
    def current_price_statement(self):

        today = date.today()
        sdate = today.strftime('%Y-%m-%d')

        cname = self.cb_salesprice_cname.currentText()
        iname = self.cb_salesprice_iname.currentText()
        dname = self.cb_salesprice_dname.currentText()

        conditions = {
            'v01': (cname, "cname like '%{}%'"),
            'v02': (iname, "iname like '%{}%'"),
            'v03': (dname, "dname like '%{}%'"),}
        
        selected_conditions = []

        for key, (value, condition_format) in conditions.items():
            if len(value) > 0:
                selected_conditions.append(condition_format.format(value))
                query = f"Select id, ccode, cname, icode, iname, dcode, dname, price, um, efffrom, effthru, remark From vw_unit_price WHERE {' AND '.join(selected_conditions)}" 
                query += f" AND (efffrom <= #{sdate}# AND effthru >= #{sdate}#)"  

        if not selected_conditions:
            QMessageBox.about(self, "검색 조건 확인", "검색 조건이 비어있어 전체 품목의 현재 단가를 출력합니다!")
            query = f"Select id, ccode, cname, icode, iname, dcode, dname, price, um, efffrom, effthru, remark From vw_unit_price WHERE" 
            query += f"(efffrom <= #{sdate}# AND effthru >= #{sdate}#)"  

        column_widths1 = [60, 80, 150, 80]
        sql_query, tv_widget, column_info, column_names, column_widths = self.common_query_statement() 

        return query, tv_widget, column_info, column_names, column_widths1 
        
    def current_price(self):
        query, tv_widget, column_info, column_names,column_widths = self.current_price_statement() 
        self.populate_dialog(self.cursor, query, tv_widget, column_info, column_names,column_widths)

    # Combobox index changed
    def cb_salesprice_cname_changed(self):
        self.entry_salesprice_ccode.clear()
        selected_item = self.cb_salesprice_cname.currentText()

        if selected_item:
            query = f"SELECT DISTINCT ccode From vw_customer WHERE cname ='{selected_item}'"
            line_edit_widgets = [self.entry_salesprice_ccode]
        
            # Check if any line edit widgets are provided
            if line_edit_widgets:
                self.lineEdit_contents(line_edit_widgets, query)
            else:
                pass

    # Combobox index changed
    def cb_salesprice_iname_changed(self):
        self.entry_salesprice_icode.clear()
        selected_item = self.cb_salesprice_iname.currentText()

        if selected_item:
            query = f"SELECT DISTINCT icode From vw_unit_price WHERE iname ='{selected_item}'"
            line_edit_widgets = [self.entry_salesprice_icode]
        
            # Check if any line edit widgets are provided
            if line_edit_widgets:
                self.lineEdit_contents(line_edit_widgets, query)
            else:
                pass

    # Combobox index changed
    def cb_salesprice_dname_changed(self):
        self.entry_salesprice_dcode.clear()
        selected_item = self.cb_salesprice_dname.currentText()

        if selected_item:
            query = f"SELECT DISTINCT dcode From vw_unit_price WHERE dname ='{selected_item}'"
            line_edit_widgets = [self.entry_salesprice_dcode]
        
            # Check if any line edit widgets are provided
            if line_edit_widgets:
                self.lineEdit_contents(line_edit_widgets, query)
            else:
                pass

    # while salesprice change efffrom date, change salesprice effthru date
    def salespricechange_efffrom_change(self):
        chg_date_str = self.entry_salespricechange_efffrom.text()

        try:
            chg_date = parse_date(chg_date_str)                     # change date string to date object
            org_date = chg_date - timedelta(days=1)                 # deduct -1 from the chg_date
            org_date_str = org_date.strftime('%Y-%m-%d')            # return the result as string
            self.entry_salesprice_effthru.setText(org_date_str)     # show converted effthru date
        
        except ValueError as e:
            QMessageBox.critical(self, "날짜 형식 오류", str(e))     # alert to the user in case of wring date format

    # Export data to Excel sheet                
    def export_table(self):
        output_subfolder = "data_list"          # set the output subfoler name
        table_widget = self.tv_salesprice       # set the name of table widget
        sheet_name = "salesprice"               # set the excel sheet name
        filename = get_file_name(output_subfolder, sheet_name)    # get the file name / export_to_excel 밑에 두면 get_file_name을 두 번 실행시켜서 파일의 번호가 달라지므로 미리 만들어서 filename에 담아둠.
        numeric_columns=[0, 1, 3, 5, 7]                  # set the numerical column index
        export_to_excel(output_subfolder, table_widget, sheet_name, numeric_columns)
               
        if filename:
            full_file_path = os.path.join(output_subfolder, filename)
            self.excel_formatting(sheet_name, full_file_path)
        else:
            QMessageBox.about(self, "코드 확인", "파일을 찾지 못했습니다!")    

    # Excel sheet formatting
    def excel_formatting(self, sheet_name, full_file_path):
        wb = load_workbook(full_file_path)
        sheet_name = sheet_name
        ws = wb[sheet_name]
        last_row = ws.max_row + 1

        column_widths = [8, 10, 12, 10, 10, 8, 10, 8, 10, 12, 12, 25]        # set the width of each column
        title_font = Font(bold=True, name="Arial", size=10)     # set the font for the first row
        arial_font = Font(name="Arial", size=10)                # set the forn from the second row to max row

        set_column_widths(ws, column_widths)    # reset column widths

        ws.freeze_panes = 'F2'              # freeze panes F2 cell
        ws.auto_filter.ref = ws.dimensions  # apply auto filter
        ws.sheet_view.showGridLines = False # remove gridlines
            
        set_font(ws, [1], range(1, len(column_widths) + 1), title_font)                 # first row font setting
        set_font(ws, range(2, last_row), range(1, len(column_widths) + 1), arial_font)  # from second row to last row font setting
        
        wb.save(full_file_path)
        QMessageBox.about(self, "파일 생성 완료", f"data_list folder에 \n엑셀 파일이 {full_file_path}로 \n생성 되었습니다!")     
        

    # clear input field entry
    def clear_data(self):
        self.lbl_salesprice_id.setText("")
        clear_widget_data(self)

    # table widget cell double click
    def show_selected_data(self, item):
        # Get the row index of the clicked item
        row_index = item.row()

        # Initialize a list to store the cell values
        cell_values = []

        # Loop through the columns and retrieve the text from each cell
        for column_index in range(12):  # 12 columns
            cell_text = self.tv_salesprice.model().item(row_index, column_index).text()
            cell_values.append(cell_text)

        # Populate the input widgets with the data from the selected row
        self.lbl_salesprice_id.setText(cell_values[0])
        self.entry_salesprice_ccode.setText(cell_values[1])
        self.cb_salesprice_cname.setCurrentText(cell_values[2])
        self.entry_salesprice_icode.setText(cell_values[3])
        self.cb_salesprice_iname.setCurrentText(cell_values[4])
        self.entry_salesprice_dcode.setText(cell_values[5])
        self.cb_salesprice_dname.setCurrentText(cell_values[6])
        self.entry_salesprice_price.setText(cell_values[7])
        self.entry_salesprice_um.setText(cell_values[8])        
        self.entry_salesprice_efffrom.setText(cell_values[9])
        self.entry_salesprice_effthru.setText(cell_values[10])
        self.entry_salesprice_remark.setText(cell_values[11])

    def refresh_data(self):
        self.clear_data()
        self.make_data()


if __name__ == "__main__":

    log_subfolder = "logs"
    os.makedirs(log_subfolder, exist_ok=True)
    log_file_path = os.path.join(log_subfolder, "access_salesprice.log")

    logging.basicConfig(
        filename=log_file_path,  
        level=logging.INFO,    
        format="%(asctime)s [%(levelname)s] - %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    app = QtWidgets.QApplication(sys.argv)
    dialog = SalesPriceDialog()
    dialog.show()
    sys.exit(app.exec())