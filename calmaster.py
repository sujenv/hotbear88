import sys
import logging
from functools import partial
from openpyxl.styles import Font
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from PyQt5 import uic, QtWidgets, QtCore
from PyQt5.QtGui import QStandardItemModel, QKeySequence
from PyQt5.QtWidgets import QMessageBox, QDialog, QShortcut
from PyQt5.QtCore import Qt
from commonmd import *

# Calendar Master table contents -----------------------------------------------------
class CalMasterDialog(QDialog, SubWindowBase):

    def __init__(self, current_username = None, current_datetime = None):
        super().__init__()

        self.conn, self.cursor = connect_to_database3()    
        uic.loadUi("calmaster.ui", self)

        # Add window flags
        self.setWindowFlags(self.windowFlags() | Qt.WindowMinimizeButtonHint | Qt.WindowMaximizeButtonHint | Qt.WindowCloseButtonHint)  

        # Initialize current_username and current_datetime directly
        self.current_username, self.current_datetime = initialize_username_and_datetime(current_username, current_datetime)

        # Enable automatic deletion on close
        self.setAttribute(QtCore.Qt.WA_DeleteOnClose)  

        # Create tv_calmaster and QSortFilterProxyModel
        self.model = QStandardItemModel()
        self.proxy_model = NumericStringSortModel(self.model)
        self.proxy_model.setSourceModel(self.model)

        # Define the column types
        column_types = ["numeric", "", "numeric", "numeric", "numeric", "", "numeric", "", "", "", ""] 

        # Set the custom delegate for the specific column
        delegate = NumericDelegate(column_types, self.tv_calmaster)
        self.tv_calmaster.setItemDelegate(delegate)
        self.tv_calmaster.setModel(self.proxy_model)

        # Enable Sorting
        self.tv_calmaster.setSortingEnabled(True)

        # Enable alternating row colors
        self.tv_calmaster.setAlternatingRowColors(True)  
        
        # Hide the first index column
        self.tv_calmaster.verticalHeader().setVisible(False)

        # While selecting row in tv_calmaster, each cell values to displayed to designated widgets
        self.tv_calmaster.clicked.connect(self.show_selected_data)

        # Fill combobox items when the application starts
        self.get_combobox_contents()

        # Populate the data
        self.make_data() 
        self.connect_btn_method()
        self.conn_signal_to_slot()

        # Set tab order for input widgets
        self.set_tab_order()

        # Initiate CTRL+C, CTRL+V and ENTER
        self.setup_shortcuts()

        # Make log file
        self.make_logfiles("access_calmaster.log")        

    # Create a shortcut for CTRL+C/CTRL+V/ Return key
    def setup_shortcuts(self):
        self.copy_shortcut = QShortcut(QKeySequence.Copy, self.tv_calmaster, partial(self.copy_cells, self.tv_calmaster))
        self.paste_shortcut = QShortcut(QKeySequence.Paste, self.tv_calmaster, partial(self.paste_cells, self.tv_calmaster))
        self.return_shortcut = QShortcut(Qt.Key_Return, self.tv_calmaster, partial(self.handle_return_key, self.tv_calmaster))

    # Call process_key_event and pass the event and your QTableWidget instance
    def keyPressEvent(self, event):
        tv_widget = self.tv_calmaster    
        self.process_key_event(event, tv_widget)

    # Pass combobox info and sql to next method
    def get_combobox_contents(self):
            self.initialize_comboboxes(self.cb_calmaster_yr, self.cb_calmaster_mth, self.cb_calmaster_da,
            self.cb_calmaster_dy, self.cb_calmaster_inh, self.cb_calmaster_outh, self.cb_calmaster_nmh,)

    # Initiate Combo_Box 
    def initialize_comboboxes(self, *combo_boxes):
        for combo_box in combo_boxes:
            field_name = combo_box.objectName().replace("cb_calmaster_", "")
            sql_query = f"SELECT DISTINCT {field_name} FROM calmaster ORDER BY {field_name}"
            self.combobox_initializing(combo_box, sql_query)
            self.clear_comboboxes(combo_box)

    # Clear Combo_Box contents
    def clear_comboboxes(self, combo_box):
        self.lbl_calmaster_id.setText("")
        self.entry_calmaster_caldt.clear()
        self.clear_combobox_selections(combo_box)

    # Connect button to method
    def connect_btn_method(self):
        self.pb_calmaster_show.clicked.connect(self.make_data)
        self.pb_calmaster_clear.clicked.connect(self.clear_data)
        self.pb_calmaster_search.clicked.connect(self.search_data)
        self.pb_calmaster_cancel.clicked.connect(self.close_dialog)
        self.pb_calmaster_check_days.clicked.connect(self.calculate_days)        
        self.pb_calmaster_excel_export.clicked.connect(self.export_table)

    # Connect signal to method    
    def conn_signal_to_slot(self):
        pass

    # tab order for calmaster window
    def set_tab_order(self):
        widgets = [self.pb_calmaster_show, self.entry_calmaster_caldt, self.cb_calmaster_yr,
            self.cb_calmaster_mth, self.cb_calmaster_da, self.cb_calmaster_dy, 
            self.entry_calmaster_val, self.cb_calmaster_inh, self.cb_calmaster_outh, 
            self.cb_calmaster_nmh, self.entry_calmaster_remark, self.pb_calmaster_search, 
            self.pb_calmaster_clear, self.pb_calmaster_insert, self.pb_calmaster_update, 
            self.pb_calmaster_delete, self.pb_calmaster_cancel, self.pb_calmaster_check_days, 
            self.pb_calmaster_excel_export]
        
        for i in range(len(widgets) - 1):
            self.setTabOrder(widgets[i], widgets[i + 1])

    # To reduce duplications
    def common_query_statement(self):
        tv_widget = self.tv_calmaster
        
        self.cursor.execute("SELECT * FROM calmaster WHERE 1=0")
        column_info = self.cursor.description
        column_names = [col[0] for col in column_info]

        sql_query = "Select * from calmaster order By id"
        column_widths = [80, 100, 80, 50, 50, 50, 50, 50, 50, 150, 150]

        return sql_query, tv_widget, column_info, column_names, column_widths

    # To reduce duplications
    def make_data(self):
        sql_query, tv_widget, column_info, column_names, column_widths = self.common_query_statement() 
        self.populate_dialog(self.cursor, sql_query, tv_widget, column_info, column_names,column_widths)

    # Get the value of other variables
    def get_calmaster_input(self):
        caldt = str(self.entry_calmaster_caldt.text())
        yr = int(self.cb_calmaster_yr.currentText())
        mth = int(self.cb_calmaster_mth.currentText())
        da = int(self.cb_calmaster_da.currentText())
        dy = str(self.cb_calmaster_dy.currentText())
        val = 1
        inh = str(self.cb_calmaster_inh.currentText())
        outh = str(self.cb_calmaster_outh.currentText())
        nmh = str(self.cb_calmaster_nmh.currentText())
        remark = str(self.entry_calmaster_remark.text())

        return caldt, yr, mth, da, dy, val, inh, outh, nmh, remark

    # Make Common values set
    def common_values_set(self):
        username = self.current_username
        user_id = self.userID_gen(username)
        formatted_datetime = self.dt_time_info()
        return username, user_id, formatted_datetime


    # clear input field entry
    def clear_data(self):
        self.lbl_calmaster_id.setText("")
        for line_edit in self.findChildren(QtWidgets.QLineEdit):
            line_edit.clear()

    # table widget cell double click
    def show_selected_data(self, item):
        # Get the row index of the clicked item
        row_index = item.row()

        # Initialize a list to store the cell values
        cell_values = []

        # Loop through the columns and retrieve the text from each cell
        for column_index in range(11):  
            cell_text = self.tv_calmaster.model().item(row_index, column_index).text()
            cell_values.append(cell_text)

        # Populate the input widgets with the data from the selected row
        self.lbl_calmaster_id.setText(cell_values[0])
        self.entry_calmaster_caldt.setText(cell_values[1])
        self.cb_calmaster_yr.setCurrentText(cell_values[2])
        self.cb_calmaster_mth.setCurrentText(cell_values[3])
        self.cb_calmaster_da.setCurrentText(cell_values[4])
        self.cb_calmaster_dy.setCurrentText(cell_values[5])
        self.entry_calmaster_val.setText(cell_values[6])
        self.cb_calmaster_inh.setCurrentText(cell_values[7])
        self.cb_calmaster_outh.setCurrentText(cell_values[8])
        self.cb_calmaster_nmh.setCurrentText(cell_values[9])
        self.entry_calmaster_remark.setText(cell_values[10])

    # Search data
    def search_data(self):
        yr = self.cb_calmaster_yr.currentText()
        mth = self.cb_calmaster_mth.currentText()
        da = self.cb_calmaster_da.currentText()
        dy = self.cb_calmaster_dy.currentText()
        inh = self.cb_calmaster_inh.currentText()
        outh = self.cb_calmaster_outh.currentText()
        nmh = self.cb_calmaster_nmh.currentText()
        
        conditions = {'v01': (yr, "yr = {}"), 'v02': (mth, "mth = {}"), 'v03': (da, "da = {}"),
                    'v04': (dy, "dy = '{}'"), 'v05': (inh, "inh = '{}'"), 'v06': (outh, "outh = '{}'"),
                    'v07': (nmh, "nmh like '%{}%'"),}
        
        selected_conditions = []

        for key, (value, condition_format) in conditions.items():
            if len(value) > 0:
                selected_conditions.append(condition_format.format(value))

        if not selected_conditions:
            QMessageBox.about(self, "검색 조건 확인", "검색 조건이 비어 있습니다!")
            return

        # Join the selected conditions to form the SQL query
        query = f"SELECT * FROM calmaster WHERE {' AND '.join(selected_conditions)} ORDER BY id"

        QMessageBox.about(self, "검색 조건 확인", f"연도: {yr} \n월: {mth} \n일: {da}\n요일: {dy}\n내근직OT: {inh}\n외근직OT: {outh}\n휴일설명: {nmh} \n\n위 조건으로 검색을 수행합니다!")

        sql_query, tv_widget, column_info, column_names, column_widths = self.common_query_statement() 
        self.populate_dialog(self.cursor, query, tv_widget, column_info, column_names, column_widths)

    # Calculate the sum of values in a specific column
    def calculate_days(self):
        self.entry_calmaster_days.clear()

        COLUMN = 6 # index 기준으로 count
        r = 0

        model = self.tv_calmaster.model()  # Get the model associated with the table view

        for i in range(model.rowCount()):
            item = model.item(i, COLUMN)
            if item and item.data(Qt.DisplayRole):
                value = float(item.data(Qt.DisplayRole))
                r += value
       
        self.entry_calmaster_days.setText(str(r))

    # Export data to Excel sheet                
    def export_table(self):
        output_subfolder = "data_list"          # set the output subfoler name
        tv_widget = self.tv_calmaster           # set the name of table widget
        sheet_name = "calmaster"                # set the excel sheet name
        filename = get_file_name(output_subfolder, sheet_name)    # get the file name / export_to_excel 밑에 두면 get_file_name을 두 번 실행시켜서 파일의 번호가 달라지므로 미리 만들어서 filename에 담아둠.
        numeric_columns=[0, 2, 3, 4, 6]         # set the numerical column index
        export_to_excel(output_subfolder, tv_widget, sheet_name, numeric_columns)
               
        if filename:
            full_file_path = os.path.join(output_subfolder, filename)
            self.excel_formatting(sheet_name, full_file_path)
        else:
            QMessageBox.about(self, "코드 확인", "파일을 찾지 못했습니다!")    

    # Excel sheet formatting
    def excel_formatting(self, sheet_name, full_file_path):
        wb = load_workbook(full_file_path)
        sheet_name = sheet_name
        ws = wb[sheet_name]
        last_row = ws.max_row + 1

        column_widths = [8, 10, 8, 6, 6, 6, 6, 10, 10, 15, 25]    # set the width of each column
        title_font = Font(bold=True, name="Arial", size=10)     # set the font for the first row
        arial_font = Font(name="Arial", size=10)                # set the forn from the second row to max row

        set_column_widths(ws, column_widths)    # reset column widths

        ws.freeze_panes = 'D2'              # freeze panes D2 cell
        ws.auto_filter.ref = ws.dimensions  # apply auto filter
        ws.sheet_view.showGridLines = False # remove gridlines
            
        set_font(ws, [1], range(1, len(column_widths) + 1), title_font)                 # first row font setting
        set_font(ws, range(2, last_row), range(1, len(column_widths) + 1), arial_font)  # from second row to last row font setting

        wb.save(full_file_path)
        QMessageBox.about(self, "파일 생성 완료", f"data_list folder에 \n엑셀 파일이 {full_file_path}로 \n생성 되었습니다!")  


    # Refresh data
    def refresh_data(self):
        self.clear_data()
        self.make_data()       

if __name__ == "__main__":
    
    log_subfolder = "logs"
    os.makedirs(log_subfolder, exist_ok=True)
    log_file_path = os.path.join(log_subfolder, "access_calmaster.log")

    logging.basicConfig(
        filename=log_file_path,  
        level=logging.INFO,    
        format="%(asctime)s [%(levelname)s] - %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
    app = QtWidgets.QApplication(sys.argv)
    dialog = CalMasterDialog()
    dialog.show()
    sys.exit(app.exec())
