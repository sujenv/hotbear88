import sys
import logging
from functools import partial
from openpyxl.styles import Font
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from PyQt5 import uic, QtWidgets, QtCore
from PyQt5.QtGui import QStandardItemModel, QKeySequence
from PyQt5.QtWidgets import QDialog, QMessageBox, QShortcut, QMenu
from PyQt5.QtCore import Qt
from datetime import datetime
from cal import CalendarView
from commonmd import *
#for non_ui version-------------------------
#from payment_product_ui import Ui_PaymentProductDialog

# Payment Product table contents -----------------------------------------------------
class PaymentProductDialog(QDialog, SubWindowBase):
#for non_ui version-------------------------
#class PaymentProductDialog(QDialog, Ui_PaymentProductDialog, SubWindowBase):
     
    def __init__(self, current_username = None, current_datetime = None):
        super().__init__()    
  
        self.conn, self.cursor = connect_to_database1()

        # load ui file
        uic.loadUi("payment_product.ui", self)
        #for non_ui version-------------------------
        #self.setupUi(self)

        # Add window flags
        self.setWindowFlags(self.windowFlags() | Qt.WindowMinimizeButtonHint | Qt.WindowMaximizeButtonHint | Qt.WindowCloseButtonHint)  

        # Initialize current_username and current_datetime directly
        self.current_username, self.current_datetime = initialize_username_and_datetime(current_username, current_datetime)

        # Enable automatic deletion on close
        self.setAttribute(QtCore.Qt.WA_DeleteOnClose)  

        # Create tv_paymentproduct and QSortFilterProxyModel
        self.model = QStandardItemModel()
        self.proxy_model = NumericStringSortModel(self.model)
        self.proxy_model.setSourceModel(self.model)

        # Define the column types
        column_types = ["", "numeric", "", "numeric"]

        # Set the custom delegate for the specific column
        delegate = NumericDelegate(column_types,self.tv_paymentproduct)
        self.tv_paymentproduct.setItemDelegate(delegate)
        self.tv_paymentproduct.setModel(self.proxy_model)

        # Enable sorting
        self.tv_paymentproduct.setSortingEnabled(True)

        # Enable alternating row colors
        self.tv_paymentproduct.setAlternatingRowColors(True)  
        
        # Hide the first index column
        self.tv_paymentproduct.verticalHeader().setVisible(False)

        # While selecting row in tv_paymentproduct, each cell values to displayed to designated widgets
        self.tv_paymentproduct.clicked.connect(self.show_selected_data)

        # Fill combobox items when the application starts
        self.get_combobox_contents()

        # Initiate display of data
        self.make_data()
        self.conn_button_to_method()
        self.connect_signal_slot()

        # Set tab order for input widgets
        self.set_tab_order()

        # Initiate CTRL+C, CTRL+V and ENTER
        self.setup_shortcuts()

        # Create context menus ----------------------------------------------------------------------------------
        self.context_menu1 = self.create_context_menu(self.entry_paymentproduct_sdt)
        self.context_menu2 = self.create_context_menu(self.entry_paymentproduct_edt)

        self.entry_paymentproduct_sdt.setContextMenuPolicy(Qt.CustomContextMenu)
        self.entry_paymentproduct_sdt.customContextMenuRequested.connect(self.show_context_menu1)

        self.entry_paymentproduct_edt.setContextMenuPolicy(Qt.CustomContextMenu)
        self.entry_paymentproduct_edt.customContextMenuRequested.connect(self.show_context_menu2)

        # Make log file
        self.make_logfiles("access_PaymentProductDialog.log")

    # Create a shortcut for CTRL+C/CTRL+V/ Return key
    def setup_shortcuts(self):
        self.copy_shortcut = QShortcut(QKeySequence.Copy, self.tv_paymentproduct, partial(self.copy_cells, self.tv_paymentproduct))
        self.paste_shortcut = QShortcut(QKeySequence.Paste, self.tv_paymentproduct, partial(self.paste_cells, self.tv_paymentproduct))
        self.return_shortcut = QShortcut(Qt.Key_Return, self.tv_paymentproduct, partial(self.handle_return_key, self.tv_paymentproduct))

    # Mouse Right click, show "달력보기" menu ---------------------------------------------------------------------
    def create_context_menu(self, target_lineedit):
        context_menu = QMenu()
        custom_action = context_menu.addAction("달력보기")
        custom_action.triggered.connect(lambda: self.show_calendar(target_lineedit))
        return context_menu

    def show_context_menu1(self, pos):
        self.context_menu1.exec_(self.entry_paymentproduct_sdt.mapToGlobal(pos))
    def show_context_menu2(self, pos):
        self.context_menu2.exec_(self.entry_paymentproduct_edt.mapToGlobal(pos))

    # Show Calendar
    def show_calendar(self, target_lineedit):
        calendar_dialog = CalendarView()
        calendar_dialog.selected_date_changed.connect(lambda date: self.set_selected_date(date, target_lineedit))
        calendar_dialog.exec()

    # Show selected date to the select Qlineedit
    def set_selected_date(self, date, target_lineedit):
        if target_lineedit == self.entry_paymentproduct_sdt:
            target_lineedit.setText(date)
        elif target_lineedit == self.entry_paymentproduct_edt:
            target_lineedit.setText(date)

    # Call process_key_event and pass the event and your QTableWidget instance
    def keyPressEvent(self, event):
        tv_widget = self.tv_paymentproduct
        self.process_key_event(event, tv_widget)

    # Pass combobox info and sql to next method
    def get_combobox_contents(self):
        self.insert_combobox_initiate(self.cb_paymentproduct_ename, "SELECT DISTINCT ename FROM employee ORDER BY ename")

    # Initiate Combo_Box 
    def insert_combobox_initiate(self, combo_box, sql_query):
        self.combobox_initializing(combo_box, sql_query) 
        self.entry_paymentproduct_ecode.setText("")
        self.cb_paymentproduct_ename.setCurrentIndex(0) 

    # Connect button to method
    def conn_button_to_method(self):
        self.pb_paymentproduct_show.clicked.connect(self.make_data)
        self.pb_paymentproduct_search.clicked.connect(self.search_data)        
        self.pb_paymentproduct_close.clicked.connect(self.close_dialog)
        self.pb_paymentproduct_clear.clicked.connect(self.clear_data)

        self.pb_paymentproduct_excel_export.clicked.connect(self.export_table)
        self.pb_paymentproduct_access_export.clicked.connect(self.export_data_to_access)

    # Connect Signal to Slot
    def connect_signal_slot(self):
        self.cb_paymentproduct_ename.activated.connect(self.ename_changed)
        self.entry_paymentproduct_sdt.editingFinished.connect(self.sdt_changed)

    # tab order for employee window
    def set_tab_order(self):
       
        widgets = [self.pb_paymentproduct_show, self.entry_paymentproduct_ddate, self.entry_paymentproduct_ecode, 
            self.entry_paymentproduct_remark, self.pb_paymentproduct_search, self.pb_paymentproduct_clear, 
            self.pb_paymentproduct_close]
        
        for i in range(len(widgets) - 1):
            self.setTabOrder(widgets[i], widgets[i + 1])

    # To reduce duplications
    def common_query_statement(self):
        tv_widget = self.tv_paymentproduct
        self.cursor.execute("SELECT * FROM vw_ap_list_daily WHERE 1=0")
        column_info = self.cursor.description
        column_names = [col[0] for col in column_info]

        sql_query = "Select * From vw_ap_list_daily"
        column_widths = [80, 100, 100, 100]

        return sql_query, tv_widget, column_info, column_names, column_widths 

    # show employee table data inside of the MDI
    def make_data(self):
        sql_query, tv_widget, column_info, column_names, column_widths = self.common_query_statement() 
        self.populate_dialog(self.cursor, sql_query, tv_widget, column_info, column_names,column_widths)


    # Search data
    def search_data(self):
        ename = self.cb_paymentproduct_ename.currentText()
        sdt = self.entry_paymentproduct_sdt.text()
        edt = self.entry_paymentproduct_edt.text()
        
        conditions = {'v01': (sdt, "actualdt >= #{}#"),
                      'v02': (edt, "actualdt <= #{}#"),
                      'v03': (ename, "ename like '%{}%'")
                      }
        
        selected_conditions = []

        for key, (value, condition_format) in conditions.items():
            if len(value) > 0:
                selected_conditions.append(condition_format.format(value))

        if not selected_conditions:
            QMessageBox.about(self, "검색 조건 확인", "검색 조건이 비어 있습니다!")
            return

        # Join the selected conditions to form the SQL query
        query = f"SELECT * FROM vw_ap_list_daily WHERE {' AND '.join(selected_conditions)} ORDER BY actualdt, ename"

        QMessageBox.about(self, "검색 조건 확인", f"근무자: {ename} \n시작일: {sdt} \n종료일:{edt} \n\n위 조건으로 검색을 수행합니다!")
        
        sql_query, tv_widget, column_info, column_names, column_widths = self.common_query_statement()
        self.populate_dialog(self.cursor, query, tv_widget, column_info, column_names,column_widths)

    def ename_changed(self):
        self.entry_paymentproduct_ecode.clear()
        selected_item = self.cb_paymentproduct_ename.currentText()

        if selected_item:
            query = f"SELECT DISTINCT ecode From employee WHERE ename ='{selected_item}'"
            line_edit_widgets = [self.entry_paymentproduct_ecode]
        
            # Check if any line edit widgets are provided
            if line_edit_widgets:
                self.lineEdit_contents(line_edit_widgets, query)
            else:
                pass

    def sdt_changed(self):
        # inputed string type date
        date_string = self.entry_paymentproduct_sdt.text()
        # convert string type date to date format
        startdt = datetime.strptime(date_string, "%Y/%m/%d")

        # Find the last day of the month for the given date
        _, last_day = calendar.monthrange(startdt.year, startdt.month)
        last_day_of_month = datetime(startdt.year, startdt.month, last_day)
        
        # Calculate the end date, which is one day before the last day of the month
        enddt = last_day_of_month - timedelta(days=0)
        
        # Format the end date as a string and set it to the desired widget
        enddt = enddt.strftime("%Y/%m/%d")        
        self.entry_paymentproduct_edt.setText(enddt)

    # Export data to MS Access Table
    def export_data_to_access(self):

        startdt = self.entry_paymentproduct_sdt.text()
        enddt = self.entry_paymentproduct_edt.text()

        result = QMessageBox.question(
            self,
            "새로운 데이터 추가 확인",
            f"시작일: {startdt} \n종료일: {enddt}\n\n"
            "위 기준으로 물품대 데이터를 추가하시겠습니까?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        
        # User clicked Yes, continue with the code
        if result == QMessageBox.Yes:
    
            dt1 = len(self.entry_paymentproduct_sdt.text())
            dt2 = len(self.entry_paymentproduct_edt.text())

            if dt1>0 and dt2>0:

                # set the name of tableview widget
                table_widget = self.tv_paymentproduct       

                # Get the model associated with the QTableView
                model = table_widget.model()
                
                # Check if there is data to insert
                if model.rowCount() == 0:
                    QMessageBox.about(self, "데이터 확인", "테이블뷰에 추가할 데이터가 없습니다!")
                    return

                # Extract data from the model and prepare for insertion
                data_to_insert = []
                
                for row in range(model.rowCount()):
                    # Extracting data for each column in a row
                    ddate = model.item(row, 0).text()
                    ecode = model.item(row, 1).text()
                    ename = model.item(row, 2).text()
                    payval = model.item(row, 3).text()
                    
                    # Create a tuple with the extracted values
                    data = (ddate, ecode, payval)
                    
                    # Append the tuple to the data_to_insert list
                    data_to_insert.append(data)

                try:
                    # Set up the connection to the MS Access database
                    relative_dbs_folder = 'dbs'
                    db_driver = '{Microsoft Access Driver (*.mdb, *.accdb)}'
                    filename = 'payroll.accdb'
                    db_path = os.path.join(relative_dbs_folder, filename)
                    conn = pyodbc.connect(rf'DRIVER={db_driver};' rf'DBQ={db_path};')
                    cursor = conn.cursor()        

                    # Insert data into the MonthlySalary table using executemany
                    cursor.executemany("""
                        INSERT INTO paymentproductval (ddate, ecode, payval)
                        VALUES (?, ?, ?)
                    """, data_to_insert)

                
                    # Commit the changes and close the connection
                    conn.commit()
                    conn.close()

                    QMessageBox.about(self, "데이터 추가", "데이터가 성공적으로 추가되었습니다!")

                except Exception as e:
                    QMessageBox.about(self, "에러 발생", f"데이터 추가 중 에러가 발생했습니다: {str(e)}")
            else:
                QMessageBox.critical(self, "필수 데이터 입력 안됨!", "시작일자, 종료일자, 지급예정일을 확인해 주세요!")
        
        else:
            # User clicked No, do nothing or handle as needed
            return

    # Export data to Excel sheet                
    def export_table(self):
        output_subfolder = "data_list"              # set the output subfoler name
        table_widget = self.tv_paymentproduct       # set the name of table widget
        sheet_name = "payment_product"              # set the excel sheet name
        filename = get_file_name(output_subfolder, sheet_name)    # get the file name / export_to_excel 밑에 두면 get_file_name을 두 번 실행시켜서 파일의 번호가 달라지므로 미리 만들어서 filename에 담아둠.
        numeric_columns=[0, 1]                      # set the numerical column index
        export_to_excel(output_subfolder, table_widget, sheet_name, numeric_columns)
               
        if filename:
            full_file_path = os.path.join(output_subfolder, filename)
            self.excel_formatting(sheet_name, full_file_path)
        else:
            QMessageBox.about(self, "코드 확인", "파일을 찾지 못했습니다!")    

    # Excel sheet formatting
    def excel_formatting(self, sheet_name, full_file_path):
        wb = load_workbook(full_file_path)
        sheet_name = sheet_name
        ws = wb[sheet_name]
        last_row = ws.max_row + 1

        column_widths = [8, 14, 10, 10]                                     # set the width of each column
        title_font = Font(bold=True, name="Arial", size=10)                 # set the font for the first row
        arial_font = Font(name="Arial", size=10)                            # set the forn from the second row to max row

        set_column_widths(ws, column_widths)        # reset column widths

        ws.freeze_panes = 'D2'                      # freeze panes D2 cell
        ws.auto_filter.ref = ws.dimensions          # apply auto filter
        ws.sheet_view.showGridLines = False         # remove gridlines
            
        set_font(ws, [1], range(1, len(column_widths) + 1), title_font)                 # first row font setting
        set_font(ws, range(2, last_row), range(1, len(column_widths) + 1), arial_font)  # from second row to last row font setting
        
        wb.save(full_file_path)
        QMessageBox.about(self, "파일 생성 완료", f"data_list folder에 \n엑셀 파일이 {full_file_path}로 \n생성 되었습니다!")   

    # clear input field entry
    def clear_data(self):
        for line_edit in self.findChildren(QtWidgets.QLineEdit):
            line_edit.clear()
        self.cb_paymentproduct_ename.setCurrentIndex(0) 

    # table widget cell double click
    def show_selected_data(self, item):
        # Get the row index of the clicked item
        row_index = item.row()

        # Initialize a list to store the cell values
        cell_values = []

        # Loop through the columns and retrieve the text from each cell
        for column_index in range(4):  # 4 columns
            cell_text = self.tv_paymentproduct.model().item(row_index, column_index).text()
            cell_values.append(cell_text)

        # Populate the input widgets with the data from the selected row
        self.entry_paymentproduct_ddate.setText(cell_values[0])
        self.entry_paymentproduct_ecode.setText(cell_values[1])
        self.cb_paymentproduct_ename.setCurrentText(cell_values[2])
        self.entry_paymentproduct_val.setText(cell_values[3])
    
    def refresh_data(self):
        self.clear_data()
        self.make_data()


if __name__ == "__main__":

    log_subfolder = "logs"
    os.makedirs(log_subfolder, exist_ok=True)
    log_file_path = os.path.join(log_subfolder, "access_PaymentProductDialog.log")

    logging.basicConfig(
        filename=log_file_path,  
        level=logging.INFO,    
        format="%(asctime)s [%(levelname)s] - %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
        
    app = QtWidgets.QApplication(sys.argv)
    dialog = PaymentProductDialog()
    dialog.show()
    sys.exit(app.exec())