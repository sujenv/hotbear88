import sys
import logging
from functools import partial
from openpyxl.styles import Font
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from PyQt5 import uic, QtWidgets, QtCore
from PyQt5.QtGui import QStandardItemModel, QKeySequence
from PyQt5.QtWidgets import QDialog, QMessageBox, QShortcut, QMenu, QInputDialog
from PyQt5.QtCore import Qt
from datetime import datetime
from cal import CalendarView
from commonmd import *
#for non_ui version-------------------------
#from bankacc_product_info_ui import Ui_BankAccProductInfoDialog

# Bank Account Product table contents -----------------------------------------------------
class BankAccProductInfoDialog(QDialog, SubWindowBase):
#for non_ui version-------------------------
#class BankAccProductInfoDialog(QDialog, Ui_BankAccProductInfoDialog, SubWindowBase):
     
    def __init__(self, current_username = None, current_datetime = None):
        super().__init__()    
  
        self.conn, self.cursor = connect_to_database1()

        # load ui file
        uic.loadUi("bankacc_product_info.ui", self)
        #for non_ui version-------------------------
        #self.setupUi(self)

        # Add window flags
        self.setWindowFlags(self.windowFlags() | Qt.WindowMinimizeButtonHint | Qt.WindowMaximizeButtonHint | Qt.WindowCloseButtonHint)  

        # Initialize current_username and current_datetime directly
        self.current_username, self.current_datetime = initialize_username_and_datetime(current_username, current_datetime)

        # Enable automatic deletion on close
        self.setAttribute(QtCore.Qt.WA_DeleteOnClose)  

        # Create tv_bankaccproduct and QSortFilterProxyModel
        self.model = QStandardItemModel()
        self.proxy_model = NumericStringSortModel(self.model)
        self.proxy_model.setSourceModel(self.model)

        # Define the column types
        column_types = ["numeric", "numeric", "", "numeric", "", "", "", "", "", "", "", "", ""]

        # Set the custom delegate for the specific column
        delegate = NumericDelegate(column_types,self.tv_bankaccproduct)
        self.tv_bankaccproduct.setItemDelegate(delegate)
        self.tv_bankaccproduct.setModel(self.proxy_model)

        # Enable sorting
        self.tv_bankaccproduct.setSortingEnabled(True)

        # Enable alternating row colors
        self.tv_bankaccproduct.setAlternatingRowColors(True)  
        
        # Hide the first index column
        self.tv_bankaccproduct.verticalHeader().setVisible(False)

        # While selecting row in tv_bankaccproduct, each cell values to displayed to designated widgets
        self.tv_bankaccproduct.clicked.connect(self.show_selected_data)

        # Fill combobox items when the application starts
        self.get_combobox_contents()

        # Initiate display of data
        self.make_data()
        self.conn_button_to_method()
        self.connect_signal_slot()

        # Set tab order for input widgets
        self.set_tab_order()

        # Initiate CTRL+C, CTRL+V and ENTER
        self.setup_shortcuts()

        # Create context menus ----------------------------------------------------------------------------------
        self.context_menu1 = self.create_context_menu(self.entry_baproduct_efffrom)
        self.context_menu2 = self.create_context_menu(self.entry_baproduct_effthru)

        self.entry_baproduct_efffrom.setContextMenuPolicy(Qt.CustomContextMenu)
        self.entry_baproduct_efffrom.customContextMenuRequested.connect(self.show_context_menu1)

        self.entry_baproduct_effthru.setContextMenuPolicy(Qt.CustomContextMenu)
        self.entry_baproduct_effthru.customContextMenuRequested.connect(self.show_context_menu2)

        self.entry_stylesheet_as_is()
        self.hide_bkaccno_change_widget()

        # Make log file
        self.make_logfiles("access_BankAccProductInfoDialog.log")

    # Create a shortcut for CTRL+C/CTRL+V/ Return key
    def setup_shortcuts(self):
        self.copy_shortcut = QShortcut(QKeySequence.Copy, self.tv_bankaccproduct, partial(self.copy_cells, self.tv_bankaccproduct))
        self.paste_shortcut = QShortcut(QKeySequence.Paste, self.tv_bankaccproduct, partial(self.paste_cells, self.tv_bankaccproduct))
        self.return_shortcut = QShortcut(Qt.Key_Return, self.tv_bankaccproduct, partial(self.handle_return_key, self.tv_bankaccproduct))

    # Change styles for the selected widgets 1
    def entry_stylesheet_as_is(self):
        self.entry_baproduct_changeccode.setStyleSheet('color:black;background:rgb(255,255,255)')
        self.cb_baproduct_changecname.setStyleSheet('color:black;background:rgb(255,255,255)')
        self.entry_baproduct_changedbaowner.setStyleSheet('color:black;background:rgb(255,255,255)')
        self.entry_baproduct_changedbankid.setStyleSheet('color:black;background:rgb(255,255,255)')
        self.cb_baproduct_changedbkname.setStyleSheet('color:black;background:rgb(255,255,255)')
        self.entry_baproduct_changedbankaccno.setStyleSheet('color:black;background:rgb(255,255,255)')
        self.entry_baproduct_changedefffrom.setStyleSheet('color:black;background:rgb(255,255,255)')
        self.entry_baproduct_changedeffthru.setStyleSheet('color:black;background:rgb(255,255,255)')
        self.entry_baproduct_changedremark.setStyleSheet('color:black;background:rgb(255,255,255)')

    # Change styles for the selected widgets 2
    def entry_stylesheet_to_be(self):
        self.entry_baproduct_changeccode.setStyleSheet('color:black;background:rgb(255,0,0)')
        self.cb_baproduct_changecname.setStyleSheet('color:black;background:rgb(255,0,0)')
        self.entry_baproduct_changedbaowner.setStyleSheet('color:white;background:rgb(255,0,0)')
        self.entry_baproduct_changedbankid.setStyleSheet('color:white;background:rgb(255,0,0)')
        self.cb_baproduct_changedbkname.setStyleSheet('color:white;background:rgb(255,0,0)')
        self.entry_baproduct_changedbankaccno.setStyleSheet('color:white;background:rgb(255,0,0)')
        self.entry_baproduct_changedefffrom.setStyleSheet('color:white;background:rgb(255,0,0)')
        self.entry_baproduct_changedeffthru.setStyleSheet('color:white;background:rgb(255,0,0)')
        self.entry_baproduct_changedremark.setStyleSheet('color:white;background:rgb(255,0,0)')

    # Show widgets for the cost change parts 
    def show_bkaccno_change_widget(self):
        self.pb_baproduct_changeinsert.setVisible(True)
        self.entry_baproduct_changeccode.setReadOnly(False)
        self.cb_baproduct_changecname.setEnabled(True)
        self.entry_baproduct_changedbaowner.setReadOnly(False)
        self.entry_baproduct_changedbankid.setReadOnly(False)
        self.cb_baproduct_changedbkname.setEnabled(True)
        self.entry_baproduct_changedbankaccno.setReadOnly(False)
        self.entry_baproduct_changedefffrom.setReadOnly(False)
        self.entry_baproduct_changedeffthru.setReadOnly(False)
        self.entry_baproduct_changedremark.setReadOnly(False)
        self.entry_stylesheet_to_be()

    # Hide widgets for the cost change parts 
    def hide_bkaccno_change_widget(self):
        self.pb_baproduct_changeinsert.setVisible(False)
        self.entry_baproduct_changeccode.setReadOnly(True)
        self.cb_baproduct_changecname.setEnabled(False)        
        self.entry_baproduct_changedbaowner.setReadOnly(True)
        self.entry_baproduct_changedbankid.setReadOnly(True)
        self.cb_baproduct_changedbkname.setEnabled(False)
        self.entry_baproduct_changedbankaccno.setReadOnly(True)      
        self.entry_baproduct_changedefffrom.setReadOnly(True)
        self.entry_baproduct_changedeffthru.setReadOnly(True)
        self.entry_baproduct_changedremark.setReadOnly(True)
        self.entry_stylesheet_as_is()

    # Mouse Right click, show "달력보기" menu ---------------------------------------------------------------------
    def create_context_menu(self, target_lineedit):
        context_menu = QMenu()
        custom_action = context_menu.addAction("달력보기")
        custom_action.triggered.connect(lambda: self.show_calendar(target_lineedit))
        return context_menu

    def show_context_menu1(self, pos):
        self.context_menu1.exec_(self.entry_baproduct_efffrom.mapToGlobal(pos))
    def show_context_menu2(self, pos):
        self.context_menu2.exec_(self.entry_baproduct_effthru.mapToGlobal(pos))

    # Show Calendar
    def show_calendar(self, target_lineedit):
        calendar_dialog = CalendarView()
        calendar_dialog.selected_date_changed.connect(lambda date: self.set_selected_date(date, target_lineedit))
        calendar_dialog.exec()

    # Show selected date to the select Qlineedit
    def set_selected_date(self, date, target_lineedit):
        if target_lineedit == self.entry_baproduct_efffrom:
            target_lineedit.setText(date)
        elif target_lineedit == self.entry_baproduct_effthru:
            target_lineedit.setText(date)

    # Call process_key_event and pass the event and your QTableWidget instance
    def keyPressEvent(self, event):
        tv_widget = self.tv_bankaccproduct
        self.process_key_event(event, tv_widget)

    # Display end of date only
    def display_eff_date(self):
        endofdate = "2050/12/31"

        return endofdate
    
    # Pass combobox info and sql to next method
    def get_combobox_contents(self):
        self.insert_combobox_initiate(self.cb_baproduct_cname, "SELECT DISTINCT cname FROM customer Where type01 = 's' ORDER BY cname")
        self.insert_combobox_initiate(self.cb_baproduct_ename, "SELECT DISTINCT ename FROM employee ORDER BY ename")
        self.insert_combobox_initiate(self.cb_baproduct_class2, "SELECT DISTINCT class2 FROM employee ORDER BY class2")
        self.insert_combobox_initiate(self.cb_baproduct_bkname, "SELECT DISTINCT bname FROM bankid ORDER BY bname")
        self.insert_combobox_initiate(self.cb_baproduct_changecname, "SELECT DISTINCT cname FROM customer Where type01 = 's' ORDER BY cname")
        self.insert_combobox_initiate(self.cb_baproduct_changedbkname, "SELECT DISTINCT bname FROM bankid ORDER BY bname")

    # Initiate Combo_Box 
    def insert_combobox_initiate(self, combo_box, sql_query):
        self.combobox_initializing(combo_box, sql_query) 
        self.lbl_baproduct_id.setText("")
        self.entry_baproduct_ccode.setText("")
        self.entry_baproduct_ecode.setText("")
        self.cb_baproduct_cname.setCurrentIndex(0) 
        self.cb_baproduct_ename.setCurrentIndex(0) 
        self.cb_baproduct_class2.setCurrentIndex(0) 
        self.cb_baproduct_changecname.setCurrentIndex(0) 
        self.cb_baproduct_changedbkname.setCurrentIndex(0) 

    # Connect button to method
    def conn_button_to_method(self):
        self.pb_baproduct_show.clicked.connect(self.make_data)
        self.pb_baproduct_search.clicked.connect(self.search_data)        
        self.pb_baproduct_close.clicked.connect(self.close_dialog)
        self.pb_baproduct_clear.clicked.connect(self.clear_data)

        self.pb_baproduct_insert.clicked.connect(self.tb_insert)
        self.pb_baproduct_update.clicked.connect(self.SelectionMessageBox)
        self.pb_baproduct_delete.clicked.connect(self.tb_delete)
        self.pb_baproduct_excel_export.clicked.connect(self.export_table)
        self.pb_baproduct_changeinsert.clicked.connect(self.reflect_bkaccno_change)
        
    # Connect Signal to Slot
    def connect_signal_slot(self):
        self.cb_baproduct_cname.activated.connect(self.cb_baproduct_cname_changed)
        self.cb_baproduct_ename.activated.connect(self.cb_baproduct_ename_changed)
        self.cb_baproduct_bkname.activated.connect(self.cb_baproduct_bkname_changed)
        self.entry_baproduct_efffrom.editingFinished.connect(self.sdt_changed)
        
        self.entry_baproduct_changedefffrom.editingFinished.connect(self.chgeffrom_changed)
        self.cb_baproduct_changecname.activated.connect(self.cb_ccname_changed)
        self.cb_baproduct_changedbkname.activated.connect(self.cb_cbkname_changed)

    # tab order for employee window
    def set_tab_order(self):
       
        widgets = [self.pb_baproduct_show, self.entry_baproduct_ccode, self.cb_baproduct_cname,
            self.entry_baproduct_ecode, self.cb_baproduct_ename, self.cb_baproduct_class2, 
            self.entry_baproduct_baowner, self.entry_baproduct_remark, 
            self.pb_baproduct_search, self.pb_baproduct_clear, self.pb_baproduct_insert, 
            self.pb_baproduct_update, self.pb_baproduct_delete, self.pb_baproduct_close]
        
        for i in range(len(widgets) - 1):
            self.setTabOrder(widgets[i], widgets[i + 1])

    # To reduce duplications
    def common_query_statement(self):
        tv_widget = self.tv_bankaccproduct
        self.cursor.execute("SELECT * FROM vw_bankacc_product WHERE 1=0")
        column_info = self.cursor.description
        column_names = [col[0] for col in column_info]

        sql_query = "Select * From vw_bankacc_product Order By id"
        column_widths = [80, 100, 100, 80, 80, 50, 100, 50, 150, 150, 100, 100, 150]

        return sql_query, tv_widget, column_info, column_names, column_widths 

    # show employee table data inside of the MDI
    def make_data(self):
        sql_query, tv_widget, column_info, column_names, column_widths = self.common_query_statement() 
        self.populate_dialog(self.cursor, sql_query, tv_widget, column_info, column_names,column_widths)

    # Get the value of other variables
    def get_basic_input(self):
        ccode = int(self.entry_baproduct_ccode.text())
        ecode = int(self.entry_baproduct_ecode.text())
        baowner = str(self.entry_baproduct_baowner.text())
        bankid = str(self.entry_baproduct_bankid.text())
        bankaccno = str(self.entry_baproduct_bankaccno.text())
        efffrom = str(self.entry_baproduct_efffrom.text())
        effthru = str(self.entry_baproduct_effthru.text())
        remark = str(self.entry_baproduct_remark.text())

        return ccode, ecode, baowner, bankid, bankaccno, efffrom, effthru, remark

    # Make Common values set
    def common_values_set(self):
        username = self.current_username
        user_id = self.userID_gen(username)
        formatted_datetime = self.dt_time_info()

        return username, user_id, formatted_datetime    

    # insert new employee data to MySQL table
    def tb_insert(self):
        currentid = self.lbl_baproduct_id.text()
        if not currentid:
            confirm_dialog = self.show_insert_confirmation_dialog()
            
            if confirm_dialog == QMessageBox.Yes:

                idx = self.max_row_id("bankacc_product")
                username, user_id, formatted_datetime = self.common_values_set()
                ccode, ecode, baowner, bankid, bankaccno, efffrom, effthru, remark = self.get_basic_input()  
                
                if (idx>0 and ccode>0 and ecode) and all(len(var) > 0 for var in (baowner, bankid, bankaccno)):
                
                    self.cursor.execute('''INSERT INTO bankacc_product (id, paycompany, ecode, baowner, bankid, bankaccno, efffrom, effthru, trxdate, userid, remark) 
                                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)'''
                                , (idx, ccode, ecode, baowner, bankid, bankaccno, efffrom, effthru, formatted_datetime, user_id, remark))
                    self.conn.commit()
                    self.show_insert_success_message()
                    self.refresh_data() 
                    logging.info(f"User {username} inserted id number {idx}, at the bankacc_product table.")
                else:
                    self.show_missing_message("입력 이상")
                    return
            else:
                self.show_cancel_message("데이터 추가 취소")
                return
        else:
            QMessageBox.information(self, "Input Error", "기존 id가 선택된 상태에서는 신규 입력이 불가합니다!")
            return

    # update 조건에 따라 분기 할 것
    def SelectionMessageBox(self):

        # If there's no selection
        if len(self.lbl_baproduct_id.text()) == 0:
            self.show_missing_message_update("입력 확인")

        # In case of row selection

        conA = '''물품대 계좌 내용 중 오류 수정 - 현재 행을 수정, 추가 행을 만들지 않음!'''
        conB = '''물품대 계좌 내용의 변경 또는 갱신 - 현재 행은 종요일 변경, 변경된 내용으로 추가 행을 만듦!'''
        
        conditions = [conA, conB]
        condition, okPressed = QInputDialog.getItem(self, "입력 조건 선택", "어떤 작업을 진행하시겠습니까?:", conditions, 0, False)

        if okPressed:
            if condition == conA:
                self.fix_typo()     
            elif condition == conB:
                self.show_bkaccno_change_widget()
            else:
                return

    # revise the values in the selected row
    def fix_typo(self):

        confirm_dialog = self.show_update_confirmation_dialog()

        if confirm_dialog == QMessageBox.Yes:
            idx = int(self.lbl_baproduct_id.text())
            username, user_id, formatted_datetime = self.common_values_set()         
            ccode, ecode, baowner, bankid, bankaccno, efffrom, effthru, remark = self.get_basic_input()  
            
            if (idx>0 and ccode>0):
                self.cursor.execute('''UPDATE bankacc_product SET paycompany=?, ecode=?, baowner=?, bankid=?, bankaccno=?, efffrom=?, effthru=?, trxdate=?, userid=?, remark=? WHERE id=?'''
                            , (ccode, ecode, baowner, bankid, bankaccno, efffrom, effthru, formatted_datetime, user_id, remark, idx))
                self.conn.commit()
                self.show_update_success_message()
                self.refresh_data()  
                logging.info(f"User {username} updated row number {idx} in the bankacc_product table.")            
            else:
                self.show_missing_message("입력 이상")
                return
        else:
            self.show_cancel_message("데이터 변경 취소")
            return

    # Get the Changed Infos
    def get_changed_info(self):
        cccode = int(self.entry_baproduct_changeccode.text())
        cecode = int(self.entry_baproduct_ecode.text())
        cbaowner = str(self.entry_baproduct_changedbaowner.text())
        cbankid = str(self.entry_baproduct_changedbankid.text())
        cbankaccno = str(self.entry_baproduct_changedbankaccno.text())
        cefffrom = str(self.entry_baproduct_changedefffrom.text())
        ceffthru = str(self.entry_baproduct_changedeffthru.text())
        cremark = str(self.entry_baproduct_changedremark.text())
        
        return cccode, cecode, cbaowner, cbankid, cbankaccno, cefffrom, ceffthru, cremark
    
    # Bank Account Change and Insert
    def reflect_bkaccno_change(self):

        username, user_id, formatted_datetime = self.common_values_set()
        cccode, cecode, cbaowner, cbankid, cbankaccno, cefffrom, ceffthru, cremark  = self.get_changed_info()

        idx = int(self.max_row_id("bankacc_product"))        
                
        if (idx>0 and cccode>0 and cecode>0) and all(len(var) > 0 for var in (cbaowner, cbankid, cbankaccno, cefffrom, ceffthru)):

            org_id = str(self.lbl_baproduct_id.text())
            effthru1 = str(self.entry_baproduct_effthru.text()) # 다시 불러와야 함..
            # 기존 id의 유효종료일을 변경유효시작일 -1 일로 수정
            self.cursor.execute('''UPDATE bankacc_product SET effthru=? WHERE id=?''', (effthru1, org_id))
            
            #변경된 내용을 신규로 추가
            self.cursor.execute('''INSERT INTO bankacc_product (id, paycompany, ecode, baowner, bankid, bankaccno, efffrom, effthru, trxdate, userid, remark) 
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)'''
                        , (idx, cccode, cecode, cbaowner, cbankid, cbankaccno, cefffrom, ceffthru, formatted_datetime, user_id, cremark))
            
            self.conn.commit()
            self.show_update_success_message()
            self.refresh_data()
            logging.info(f"User {username} updated row number {idx} in the apt contract table.")
            
        else:
            self.show_missing_message("입력 이상")
            return

        self.entry_stylesheet_as_is()
        self.hide_bkaccno_change_widget()

    # delete row according to id selected
    def tb_delete(self):
        confirm_dialog = self.show_delete_confirmation_dialog()

        if confirm_dialog == QMessageBox.Yes:
            idx = self.lbl_baproduct_id.text()
            username, user_id, formatted_datetime = self.common_values_set()

            self.cursor.execute("DELETE FROM bankacc_product WHERE id=?", (idx,))
            self.conn.commit()
            self.show_delete_success_message()
            self.refresh_data()  
            logging.info(f"User {username} deleted id number {idx}, at the bankacc_product table.")       
        else:
            self.show_cancel_message("데이터 삭제 취소")

    # Search data
    def search_data(self):
        
        cname = self.cb_baproduct_cname.currentText()
        ename = self.cb_baproduct_ename.currentText()
        class2 = self.cb_baproduct_class2.currentText()
        bname = self.cb_baproduct_bkname.currentText()
        
        conditions = {'v01': (cname, "cname like '%{}%'"),
                      'v02': (ename, "ename like '%{}%'"),
                      'v03': (class2, "class2 like '%{}%'"),
                      'v04': (bname, "bname like '%{}%'"),
                      }
        
        selected_conditions = []

        for key, (value, condition_format) in conditions.items():
            if len(value) > 0:
                selected_conditions.append(condition_format.format(value))

        if not selected_conditions:
            QMessageBox.about(self, "검색 조건 확인", "검색 조건이 비어 있습니다!")
            return

        # Join the selected conditions to form the SQL query
        query = f"SELECT * FROM vw_bankacc_product WHERE {' AND '.join(selected_conditions)} ORDER BY cname, ename"

        QMessageBox.about(self, "검색 조건 확인", f"업체명: {cname} \n직원명:{ename} \n현재근무자:{class2} \n은행명:{bname} \n\n위 조건으로 검색을 수행합니다!")
        
        sql_query, tv_widget, column_info, column_names, column_widths = self.common_query_statement()
        self.populate_dialog(self.cursor, query, tv_widget, column_info, column_names,column_widths)

    # Combobox Payment Company index changed
    def cb_baproduct_cname_changed(self):
        self.entry_baproduct_ccode.clear()
        selected_item = self.cb_baproduct_cname.currentText()

        if selected_item:
            query = f"SELECT DISTINCT ccode From customer WHERE cname ='{selected_item}'"
            line_edit_widgets = [self.entry_baproduct_ccode]
        
            # Check if any line edit widgets are provided
            if line_edit_widgets:
                self.lineEdit_contents(line_edit_widgets, query)
            else:
                pass

    # CEfffrom Changed
    def chgeffrom_changed(self):
        chg_date_str = self.entry_baproduct_changedefffrom.text()
        try:
            chg_date = parse_date(chg_date_str)                             # 날짜 문자열을 날짜 객체로 변환
            org_date = chg_date - timedelta(days=1)                         # 날짜에서 1일을 빼줍니다.
            org_date_str = org_date.strftime('%Y-%m-%d')                    # 결과를 문자열로 변환
            self.entry_baproduct_effthru.setText(org_date_str)              # 변경된 effthru 날짜를 표시
        
            endofdate = self.display_eff_date()
            self.entry_baproduct_changedeffthru.setText(endofdate)

        except ValueError as e:
            QMessageBox.critical(self, "날짜 형식 오류", str(e))             # 날짜 형식이 잘못된 경우 사용자에게 알림


    # Employee Name Index Changed
    def cb_baproduct_ename_changed(self):
        self.entry_baproduct_ecode.clear()
        selected_item = self.cb_baproduct_ename.currentText()

        if selected_item:
            query = f"SELECT DISTINCT ecode From employee WHERE ename ='{selected_item}'"
            line_edit_widgets = [self.entry_baproduct_ecode]
        
            # Check if any line edit widgets are provided
            if line_edit_widgets:
                self.lineEdit_contents(line_edit_widgets, query)
            else:
                pass

    # Bank Name Index Changed
    def cb_baproduct_bkname_changed(self):
        self.entry_baproduct_bankid.clear()
        selected_item = self.cb_baproduct_bkname.currentText()

        if selected_item:
            query = f"SELECT DISTINCT bcode From bankid WHERE bname ='{selected_item}'"
            line_edit_widgets = [self.entry_baproduct_bankid]
        
            # Check if any line edit widgets are provided
            if line_edit_widgets:
                self.lineEdit_contents(line_edit_widgets, query)
            else:
                pass

    # Effective Date Index Changed
    def sdt_changed(self):
        # inputed string type date
        date_string = self.entry_baproduct_efffrom.text()
        # convert string type date to date format
        startdt = datetime.strptime(date_string, "%Y/%m/%d")

        # Find the last day of the month for the given date
        _, last_day = calendar.monthrange(startdt.year, startdt.month)
        last_day_of_month = datetime(startdt.year, startdt.month, last_day)
        
        # Calculate the end date, which is one day before the last day of the month
        enddt = last_day_of_month - timedelta(days=0)
        effthru = "2050/12/31"
        
        # Format the end date as a string and set it to the desired widget
        enddt = enddt.strftime("%Y/%m/%d")
        
        self.entry_baproduct_effthru.setText(effthru)
        
    # Combobox Paycompany Name index changed
    def cb_ccname_changed(self):
        self.entry_baproduct_changeccode.clear()
        selected_item = self.cb_baproduct_changecname.currentText()

        if selected_item:
            query = f"SELECT DISTINCT ccode From customer WHERE cname ='{selected_item}'"
            line_edit_widgets = [self.entry_baproduct_changeccode]
        
            # Check if any line edit widgets are provided
            if line_edit_widgets:
                self.lineEdit_contents(line_edit_widgets, query)
            else:
                pass

    # Combox BK Name Index Changed
    def cb_cbkname_changed(self):
        self.entry_baproduct_changedbankid.clear()
        selected_item = self.cb_baproduct_changedbkname.currentText()

        if selected_item:
            query = f"SELECT DISTINCT bcode From bankid WHERE bname ='{selected_item}'"
            line_edit_widgets = [self.entry_baproduct_changedbankid]
        
            # Check if any line edit widgets are provided
            if line_edit_widgets:
                self.lineEdit_contents(line_edit_widgets, query)
            else:
                pass        

    # Export data to Excel sheet                
    def export_table(self):
        output_subfolder = "data_list"              # set the output subfoler name
        table_widget = self.tv_bankaccproduct       # set the name of table widget
        sheet_name = "bankacc_product"              # set the excel sheet name
        filename = get_file_name(output_subfolder, sheet_name)    # get the file name / export_to_excel 밑에 두면 get_file_name을 두 번 실행시켜서 파일의 번호가 달라지므로 미리 만들어서 filename에 담아둠.
        numeric_columns=[0, 1]                      # set the numerical column index
        export_to_excel(output_subfolder, table_widget, sheet_name, numeric_columns)
               
        if filename:
            full_file_path = os.path.join(output_subfolder, filename)
            self.excel_formatting(sheet_name, full_file_path)
        else:
            QMessageBox.about(self, "코드 확인", "파일을 찾지 못했습니다!")    

    # Excel sheet formatting
    def excel_formatting(self, sheet_name, full_file_path):
        wb = load_workbook(full_file_path)
        sheet_name = sheet_name
        ws = wb[sheet_name]
        last_row = ws.max_row + 1

        column_widths = [8, 14, 10, 10, 10, 8, 10, 10, 16, 20, 10, 10, 20]  # set the width of each column
        title_font = Font(bold=True, name="Arial", size=10)                 # set the font for the first row
        arial_font = Font(name="Arial", size=10)                            # set the forn from the second row to max row

        set_column_widths(ws, column_widths)        # reset column widths

        ws.freeze_panes = 'D2'                      # freeze panes D2 cell
        ws.auto_filter.ref = ws.dimensions          # apply auto filter
        ws.sheet_view.showGridLines = False         # remove gridlines
            
        set_font(ws, [1], range(1, len(column_widths) + 1), title_font)                 # first row font setting
        set_font(ws, range(2, last_row), range(1, len(column_widths) + 1), arial_font)  # from second row to last row font setting
        
        wb.save(full_file_path)
        QMessageBox.about(self, "파일 생성 완료", f"data_list folder에 \n엑셀 파일이 {full_file_path}로 \n생성 되었습니다!")   

    # clear input field entry
    def clear_data(self):
        self.lbl_baproduct_id.setText("")
        for line_edit in self.findChildren(QtWidgets.QLineEdit):
            line_edit.clear()

    # table widget cell double click
    def show_selected_data(self, item):
        # Get the row index of the clicked item
        row_index = item.row()

        # Initialize a list to store the cell values
        cell_values = []

        # Loop through the columns and retrieve the text from each cell
        for column_index in range(13):  # 13 columns
            cell_text = self.tv_bankaccproduct.model().item(row_index, column_index).text()
            cell_values.append(cell_text)

        # Populate the input widgets with the data from the selected row
        self.lbl_baproduct_id.setText(cell_values[0])
        self.entry_baproduct_ccode.setText(cell_values[1])
        self.cb_baproduct_cname.setCurrentText(cell_values[2])
        self.entry_baproduct_ecode.setText(cell_values[3])
        self.cb_baproduct_ename.setCurrentText(cell_values[4])
        self.cb_baproduct_class2.setCurrentText(cell_values[5])
        self.entry_baproduct_baowner.setText(cell_values[6])
        self.entry_baproduct_bankid.setText(cell_values[7])
        self.cb_baproduct_bkname.setCurrentText(cell_values[8])
        self.entry_baproduct_bankaccno.setText(cell_values[9])
        self.entry_baproduct_efffrom.setText(cell_values[10])
        self.entry_baproduct_effthru.setText(cell_values[11])
        self.entry_baproduct_remark.setText(cell_values[12])
    
    def refresh_data(self):
        self.clear_data()
        self.make_data()


if __name__ == "__main__":

    log_subfolder = "logs"
    os.makedirs(log_subfolder, exist_ok=True)
    log_file_path = os.path.join(log_subfolder, "access_BankAccProductInfoDialog.log")

    logging.basicConfig(
        filename=log_file_path,  
        level=logging.INFO,    
        format="%(asctime)s [%(levelname)s] - %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
        
    app = QtWidgets.QApplication(sys.argv)
    dialog = BankAccProductInfoDialog()
    dialog.show()
    sys.exit(app.exec())