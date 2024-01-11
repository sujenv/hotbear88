import sys
import logging
import math
from functools import partial
from openpyxl.styles import Font
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from PyQt5 import uic, QtWidgets, QtCore
from PyQt5.QtGui import QStandardItemModel, QKeySequence
from PyQt5.QtWidgets import QMessageBox, QDialog, QShortcut, QMenu
from PyQt5.QtCore import Qt
from commonmd import *
from cal import CalendarView

# Calendar Master table contents -----------------------------------------------------
class CalAbsenteeismDialog(QDialog, SubWindowBase):

    def __init__(self, current_username = None, current_datetime = None):
        super().__init__()

        self.conn, self.cursor = connect_to_database3()    
        uic.loadUi("calc_absenteeism.ui", self)

        # Add window flags
        self.setWindowFlags(self.windowFlags() | Qt.WindowMinimizeButtonHint | Qt.WindowMaximizeButtonHint | Qt.WindowCloseButtonHint)  

        # Initialize current_username and current_datetime directly
        self.current_username, self.current_datetime = initialize_username_and_datetime(current_username, current_datetime)

        # Enable automatic deletion on close
        self.setAttribute(QtCore.Qt.WA_DeleteOnClose)  

        # Create tv_employeelist and QSortFilterProxyModel
        self.model = QStandardItemModel()
        self.proxy_model = NumericStringSortModel(self.model)
        self.proxy_model.setSourceModel(self.model)

        # Define the column types
        column_types = ["numeric", "", "", ]  
        column_types1 = ["numeric", "", "numeric", "", "numeric", "numeric", "numeric", "" ] 

        # Set the custom delegate for the specific column
        delegate = NumericDelegate(column_types, self.tv_employeelist)
        self.tv_employeelist.setItemDelegate(delegate)
        self.tv_employeelist.setModel(self.proxy_model)

        delegate1 = NumericDelegate(column_types1, self.tv_abslist)
        self.tv_abslist.setItemDelegate(delegate1)
        self.tv_abslist.setModel(self.proxy_model)

        # Enable Sorting
        self.tv_employeelist.setSortingEnabled(True)
        self.tv_abslist.setSortingEnabled(True)

        # Enable alternating row colors
        self.tv_employeelist.setAlternatingRowColors(True)  
        self.tv_abslist.setAlternatingRowColors(True)  

        # Hide the first index column
        self.tv_employeelist.verticalHeader().setVisible(False)
        self.tv_abslist.verticalHeader().setVisible(False)

        # While selecting row in tv_widget, each cell values to displayed to designated widgets
        self.tv_employeelist.clicked.connect(self.show_selected_emplist_data)
        self.tv_abslist.clicked.connect(self.show_selected_abslist_data)
        
        # Initialize some widgets
        self.lbl_abs_id.setText("")
        self.entry_abs_reghr.setText("8")
        self.entry_abs_othr.setText("0")
        self.radio_reg.setChecked(True)
        
        # Initialize current date
        now = datetime.now()
        curr_date = now.strftime("%Y/%m/%d")
        dt = f"{curr_date}"        
        self.entry_abs_datein.setText(dt)

        # Fill combobox items when the application starts
        self.get_combobox_contents()

        # Populate the data
        self.make_data() 
        self.make_data1() 
        self.connect_btn_method()
        self.conn_signal_to_slot()

        # Set tab order for input widgets
        self.set_tab_order()

        # Initiate CTRL+C, CTRL+V and ENTER
        self.setup_shortcuts()

        # Create context menus ----------------------------------------------------------------------------------
        self.context_menu1 = self.create_context_menu(self.entry_abs_datein)
        self.entry_abs_datein.setContextMenuPolicy(Qt.CustomContextMenu)
        self.entry_abs_datein.customContextMenuRequested.connect(self.show_context_menu1)

        # Make log file
        self.make_logfiles("access_CalcAbsenteeism.log")        

    # Create a shortcut for CTRL+C/CTRL+V/ Return key
    def setup_shortcuts(self):
        self.copy_shortcut = QShortcut(QKeySequence.Copy, self.tv_employeelist, partial(self.copy_cells, self.tv_employeelist))
        self.paste_shortcut = QShortcut(QKeySequence.Paste, self.tv_employeelist, partial(self.paste_cells, self.tv_employeelist))
        self.return_shortcut = QShortcut(Qt.Key_Return, self.tv_employeelist, partial(self.handle_return_key, self.tv_employeelist))
        
        self.copy_shortcut = QShortcut(QKeySequence.Copy, self.tv_abslist, partial(self.copy_cells, self.tv_abslist))
        self.paste_shortcut = QShortcut(QKeySequence.Paste, self.tv_abslist, partial(self.paste_cells, self.tv_abslist))
        self.return_shortcut = QShortcut(Qt.Key_Return, self.tv_abslist, partial(self.handle_return_key, self.tv_abslist))

    # Mouse Right click, show "달력보기" menu ---------------------------------------------------------------------
    def create_context_menu(self, target_lineedit):
        context_menu = QMenu()
        custom_action = context_menu.addAction("달력보기")
        custom_action.triggered.connect(lambda: self.show_calendar(target_lineedit))
        return context_menu

    def show_context_menu1(self, pos):
        self.context_menu1.exec_(self.entry_abs_datein.mapToGlobal(pos))

    # Show selected date to the select Qlineedit
    def set_selected_date(self, date, target_lineedit):
        if target_lineedit == self.entry_abs_datein:
            target_lineedit.setText(date)

    # Show Calendar
    def show_calendar(self, target_lineedit):
        calendar_dialog = CalendarView()
        calendar_dialog.selected_date_changed.connect(lambda date: self.set_selected_date(date, target_lineedit))
        calendar_dialog.exec()

    # Call process_key_event and pass the event and your QTableWidget instance
    def keyPressEvent(self, event):
        tv_widget = self.tv_employeelist    
        self.process_key_event(event, tv_widget)

    def keyPressEvent1(self, event):
        tv_widget = self.tv_abslist
        self.process_key_event(event, tv_widget)

    # Pass combobox info and sql to next method
    def get_combobox_contents(self):
        pass
        #self.insert_combobox_initiate(self.cb_out_year, "SELECT DISTINCT fy FROM basicdata ORDER BY fy")

    # Initiate Combo_Box 
    def insert_combobox_initiate(self, combo_box, sql_query):
        self.combobox_initializing(combo_box, sql_query) 
        #self.cb_out_year.setCurrentIndex(0) 

    # Connect button to method
    def connect_btn_method(self):
        self.pb_abs_clear.clicked.connect(self.clear_data)
        self.pb_abs_close.clicked.connect(self.close_dialog)
        self.pb_abs_insert.clicked.connect(self.tv_insert)
        self.pb_abs_update.clicked.connect(self.tv_update)
        self.pb_abs_delete.clicked.connect(self.tv_delete)
      
    # Connect signal to method    
    def conn_signal_to_slot(self):
        self.radio_reg.toggled.connect(self.onRadioButtonClicked)
        self.radio_ot.toggled.connect(self.onRadioButtonClicked)
        self.radio_off.toggled.connect(self.onRadioButtonClicked)

    #def onRadioButtonClicked(self):
    #    is_reg_checked = self.radio_reg.isChecked()
    #    is_ot_checked = self.radio_ot.isChecked()
    #    is_off_checked = self.radio_off.isChecked()
        
    #    if is_reg_checked:
    #        self.entry_abs_reghr.setText("8")
    #        self.entry_abs_othr.setText("0")
    #        self.entry_abs_off.setText("0")
    #        self.setColors(self.entry_abs_reghr, 'yellow')
    #        self.setColors(self.entry_abs_othr, 'light gray')
    #        self.setColors(self.entry_abs_off, 'light gray')            
    #    elif is_ot_checked:
    #        self.entry_abs_reghr.setText("0")
    #        self.entry_abs_othr.setText("8")
    #        self.entry_abs_off.setText("0")
    #        self.setColors(self.entry_abs_reghr, 'light gray')
    #        self.setColors(self.entry_abs_othr, 'yellow')
    #        self.setColors(self.entry_abs_off, 'light gray')            
    #    elif is_off_checked:
    #        self.entry_abs_reghr.setText("0")
    #        self.entry_abs_othr.setText("0")
    #        self.entry_abs_off.setText("8")
    #        self.setColors(self.entry_abs_reghr, 'light gray')
    #        self.setColors(self.entry_abs_othr, 'light gray')
    #        self.setColors(self.entry_abs_off, 'yellow')            
    #    else:
    #        return

    def onRadioButtonClicked(self):
        is_reg_checked = self.radio_reg.isChecked()
        is_ot_checked = self.radio_ot.isChecked()
        is_off_checked = self.radio_off.isChecked()

        self.entry_abs_reghr.setText("8" if is_reg_checked else "0")
        self.entry_abs_othr.setText("8" if is_ot_checked else "0")
        self.entry_abs_off.setText("8" if is_off_checked else "0")

        self.setColors(self.entry_abs_reghr, 'yellow' if is_reg_checked else 'light gray')
        self.setColors(self.entry_abs_othr, 'yellow' if is_ot_checked else 'light gray')
        self.setColors(self.entry_abs_off, 'yellow' if is_off_checked else 'light gray')

    def setColors(self, widget, color):
        widget.setStyleSheet(f'background-color: {color}; color: black')

    # tab order for calmaster window
    def set_tab_order(self):
        widgets = [self.pb_abs_insert , self.pb_abs_delete ,]
        
        for i in range(len(widgets) - 1):
            self.setTabOrder(widgets[i], widgets[i + 1])


    def query_statement(self):
        tv_widget = self.tv_employeelist
        
        self.cursor.execute("SELECT * FROM vw_employee_abs WHERE 1=0")
        column_info = self.cursor.description
        column_names = [col[0] for col in column_info]

        sql_query = "Select * from vw_employee_abs order by ename"
        column_widths = [80, 80, 80]

        return sql_query, tv_widget, column_info, column_names, column_widths

    # To reduce duplications
    def make_data(self):
        sql_query, tv_widget, column_info, column_names, column_widths = self.query_statement() 
        self.populate_dialog(self.cursor, sql_query, tv_widget, column_info, column_names,column_widths)

    def query_statement1(self):
        tv_widget = self.tv_abslist
        
        self.cursor.execute("SELECT * FROM vw_absenteeism WHERE 1=0")
        column_info = self.cursor.description
        column_names = [col[0] for col in column_info]

        sql_query = "Select * from vw_absenteeism order By id"
        column_widths = [60, 80, 80, 80, 60, 60, 60, 100]

        return sql_query, tv_widget, column_info, column_names, column_widths

    # To reduce duplications
    def make_data1(self):
        sql_query, tv_widget, column_info, column_names, column_widths = self.query_statement1() 
        self.populate_dialog(self.cursor, sql_query, tv_widget, column_info, column_names,column_widths)

    # Get the value of other variables
    def get_abs_input(self):
        datein = self.entry_abs_datein.text()
        ecode = self.entry_employee_id.text()
        reghr = self.entry_abs_reghr.text()
        othr = self.entry_abs_othr.text()
        offhr = self.entry_abs_off.text()
        remark = self.entry_employee_remark.text()

        return datein, ecode, reghr, othr, offhr, remark
        
    # Make Common values set
    def common_values_set(self):
        username = self.current_username
        user_id = self.userID_gen(username)
        formatted_datetime = self.dt_time_info()
        return username, user_id, formatted_datetime

    # insert new calc_salary data to MySQL table
    def tv_insert(self):
        confirm_dialog = self.show_insert_confirmation_dialog()

        if confirm_dialog == QMessageBox.Yes:
            
            idx = self.max_row_id("absenteeism")                                  # Get the max id 
            username, user_id, formatted_datetime = self.common_values_set()
            datein, ecode, reghr, othr, offhr, remark = self.get_abs_input()         # Get the value of other variables

            if (idx>0):

                self.cursor.execute('''INSERT INTO absenteeism 
                            (id, datein, ecode, reg, ot, off, trxdate, userid, remark) 
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)'''
                            , (idx, datein, ecode, reghr, othr, offhr, formatted_datetime, user_id, remark))
                self.conn.commit()
                self.show_insert_success_message()
                self.refresh_data()
                self.make_data1()
                logging.info(f"User {username} inserted {idx} row to the absenteeism table.")
            else:
                self.show_missing_message("입력 이상")
                return
        else:
            self.show_cancel_message("데이터 추가 취소")
            return

    def tv_update(self):
        confirm_dialog = self.show_insert_confirmation_dialog()
        
        if confirm_dialog == QMessageBox.Yes:
            idx = int(self.lbl_abs_id.text())                                  # Get the max id 
            username, user_id, formatted_datetime = self.common_values_set()
            datein, ecode, reghr, othr, offhr, remark = self.get_abs_input()         # Get the value of other variables

            if (idx>0):
                self.cursor.execute('''UPDATE absenteeism SET datein=?, ecode=?, reg=?, ot=?, off=?, trxdate=?, userid=?, remark=? WHERE id=?'''
                            , (datein, ecode, reghr, othr, offhr, formatted_datetime, user_id, remark, idx,))
                self.conn.commit()
                self.show_update_success_message()
                self.refresh_data()
                self.make_data1()
                logging.info(f"User {username} Updated {idx} row to the absenteeism table.")
            else:
                self.show_missing_message("입력 이상")
                return
        else:
            self.show_cancel_message("데이터 추가 취소")
            return

    def tv_delete(self):
        confirm_dialog = self.show_delete_confirmation_dialog()

        if confirm_dialog == QMessageBox.Yes:

            idx = self.lbl_abs_id.text()
            username, user_id, formatted_datetime = self.common_values_set()

            self.cursor.execute("DELETE FROM absenteeism WHERE id=?", (idx,))
            self.conn.commit()
            self.show_delete_success_message()
            self.refresh_data()  
            self.make_data1()            
            logging.info(f"User {username} deleted {idx} row to the absenteeism table.")                
            
        else:
            self.show_cancel_message("데이터 삭제 취소")

    # clear input field entry
    def clear_data(self):
        for line_edit in self.findChildren(QtWidgets.QLineEdit):
            line_edit.clear()

        self.lbl_abs_id.setText("")
  
    # table view click
    def show_selected_emplist_data(self, item):
        # Get the row index of the clicked item
        row_index = item.row()

        # Initialize a list to store the cell values
        cell_values = []

        # Loop through the columns and retrieve the text from each cell
        for column_index in range(3): 
            cell_text = self.tv_employeelist.model().item(row_index, column_index).text()
            cell_values.append(cell_text)
            
        # Populate the input widgets with the data from the selected row
        self.entry_employee_id.setText(cell_values[0])
        self.entry_employee_name.setText(cell_values[1])
        self.entry_employee_remark.setText(cell_values[2])
    
    # table view 2 click
    def show_selected_abslist_data(self, item):
        # Get the row index of the clicked item
        row_index = item.row()

        # Initialize a list to store the cell values
        cell_values = []

        # Loop through the columns and retrieve the text from each cell
        for column_index in range(8): 
            cell_text = self.tv_abslist.model().item(row_index, column_index).text()
            cell_values.append(cell_text)

        # Populate the input widgets with the data from the selected row
        self.lbl_abs_id.setText(cell_values[0])
        self.entry_abs_datein.setText(cell_values[1])
        self.entry_employee_id.setText(cell_values[2])
        self.entry_employee_name.setText(cell_values[3])
        self.entry_abs_reghr.setText(cell_values[4])
        self.entry_abs_othr.setText(cell_values[5])
        self.entry_abs_off.setText(cell_values[6])
        self.entry_employee_remark.setText(cell_values[7])


    # 선택된 각 위젯의 내용을 엑셀로 내보내기
    def export_data(self):
        output_subfolder = "data_list"                              # set the output subfoler name
        sheet_name = "Calc_Absenteeism"                             # set the excel sheet name
        filename = get_file_name(output_subfolder, sheet_name)      # get the file name

        # Ensure the subfolder exists; create it if it doesn't
        os.makedirs(output_subfolder, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        if filename:
            full_file_path = os.path.join(output_subfolder, filename)
    
            # 각 위젯에서 내용 가져와서 엑셀에 쓰기
            data_to_export = [
                (self.label_40.text(), self.lbl_abs_id.text()),
                #(self..text(), self..text()),
                #(self..text(), self..text()),
                #(self..text(), self..text()),
                #(self..text(), self..text()),
                #(self..text(), self..text()),
                #(self..text(), self..text()),
                #(self..text(), self..text()),
                #(self..text(), self..text()),
                #(self..text(), self..text()),
                #(self..text(), self..text()),
                #(self..text(), self..text()),
                #(self..text(), self..text()),
                #(self..text() , self..text()),
                #(self..text() , self..text()),
                #(self..text() , self..text() ),
                #(self..text() , self..text() ),
                #(self..text() , self..text() ),
                #(self..text() , self..text() ),
                #(self..text() , self..text() ),
                ]

            #for i, value in enumerate(data_to_export, start=1):
            #    ws.cell(row=i, column=1, value=value)
            
            for i, (label, value) in enumerate(data_to_export, start=1):
                ws.cell(row=i, column=1, value=label)
                ws.cell(row=i, column=2, value=value)
            
            # 엑셀 파일 저장
            wb.save(full_file_path)
        
            self.excel_formatting(sheet_name, full_file_path)

    # Excel sheet formatting
    def excel_formatting(self, sheet_name, full_file_path):
        wb = load_workbook(full_file_path)
        sheet_name = sheet_name
        ws = wb[sheet_name]
        last_row = ws.max_row + 2

        # Insert headers at A1 and B1
        ws.insert_rows(1)
        ws['A1'] = '구분'
        ws['B1'] = '내용'

        column_widths = [20, 15]                                # set the width of each column
        title_font = Font(bold=True, name="Arial", size=10)     # set the font for the first row
        arial_font = Font(name="Arial", size=10)                # set the forn from the second row to max row

        set_column_widths(ws, column_widths)    # reset column widths

        ws.freeze_panes = 'B2'              # freeze panes D2 cell
        ws.auto_filter.ref = ws.dimensions  # apply auto filter
        ws.sheet_view.showGridLines = False # remove gridlines
            
        set_font(ws, [1], range(1, len(column_widths) + 1), title_font)                 # first row font setting
        set_font(ws, range(2, last_row), range(1, len(column_widths) + 1), arial_font)  # from second row to last row font setting

        wb.save(full_file_path)
        QMessageBox.about(self, "파일 생성 완료", f"data_list folder에 \n엑셀 파일이 {full_file_path}로 \n생성 되었습니다!")  

    # Refresh data
    def refresh_data(self):
        self.clear_data()
        self.make_data()       

if __name__ == "__main__":
    
    log_subfolder = "logs"
    os.makedirs(log_subfolder, exist_ok=True)
    log_file_path = os.path.join(log_subfolder, "access_CalcAbsenteeism.log")

    logging.basicConfig(
        filename=log_file_path,  
        level=logging.INFO,    
        format="%(asctime)s [%(levelname)s] - %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
    app = QtWidgets.QApplication(sys.argv)
    dialog = CalAbsenteeismDialog()
    dialog.show()
    sys.exit(app.exec())
