import sys
import logging
from functools import partial
from openpyxl.styles import Font
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from PyQt5 import uic, QtWidgets, QtCore
from PyQt5.QtGui import QStandardItemModel, QKeySequence
from PyQt5.QtWidgets import QDialog, QMessageBox, QShortcut, QMenu, QInputDialog
from PyQt5.QtCore import Qt
from datetime import datetime
from cal import CalendarView
from commonmd import *
#for non_ui version-------------------------
#from oil_payment_info_ui import Ui_OilUsageInfoDialog

# Oil Usage table contents -----------------------------------------------------
class OilUsageInfoDialog(QDialog, SubWindowBase):
#for non_ui version-------------------------
#class OilUsageInfoDialog(QDialog, Ui_OilUsageInfoDialog, SubWindowBase):
     
    def __init__(self, current_username = None, current_datetime = None):
        super().__init__()    
  
        self.conn, self.cursor = connect_to_database3()

        # load ui file
        uic.loadUi("oil_payment_info.ui", self)
        #for non_ui version-------------------------
        #self.setupUi(self)

        # Add window flags
        self.setWindowFlags(self.windowFlags() | Qt.WindowMinimizeButtonHint | Qt.WindowMaximizeButtonHint | Qt.WindowCloseButtonHint)  

        # Initialize current_username and current_datetime directly
        self.current_username, self.current_datetime = initialize_username_and_datetime(current_username, current_datetime)

        # Enable automatic deletion on close
        self.setAttribute(QtCore.Qt.WA_DeleteOnClose)  

        # Create tv_oilusageinfo and QSortFilterProxyModel
        self.model = QStandardItemModel()
        self.proxy_model = NumericStringSortModel(self.model)
        self.proxy_model.setSourceModel(self.model)

        # Define the column types
        column_types = ["numeric", "numeric", "", "", "", "", "", "numeric", "",]

        # Set the custom delegate for the specific column
        delegate = NumericDelegate(column_types,self.tv_oilusageinfo)
        self.tv_oilusageinfo.setItemDelegate(delegate)
        self.tv_oilusageinfo.setModel(self.proxy_model)

        # Enable sorting
        self.tv_oilusageinfo.setSortingEnabled(True)

        # Enable alternating row colors
        self.tv_oilusageinfo.setAlternatingRowColors(True)  
        
        # Hide the first index column
        self.tv_oilusageinfo.verticalHeader().setVisible(False)

        # While selecting row in tv_oilusageinfo, each cell values to displayed to designated widgets
        self.tv_oilusageinfo.clicked.connect(self.show_selected_data)

        # Fill combobox items when the application starts
        self.get_combobox_contents()

        # Initiate display of data
        self.make_data()
        self.conn_button_to_method()
        self.connect_signal_slot()

        # Set tab order for input widgets
        self.set_tab_order()

        # Initiate CTRL+C, CTRL+V and ENTER
        self.setup_shortcuts()

        # Create context menus ----------------------------------------------------------------------------------
        self.context_menu1 = self.create_context_menu(self.entry_oilusage_getdt)

        self.entry_oilusage_getdt.setContextMenuPolicy(Qt.CustomContextMenu)
        self.entry_oilusage_getdt.customContextMenuRequested.connect(self.show_context_menu1)

        # Make log file
        self.make_logfiles("access_OilUsageInfoDialog.log")

    # Create a shortcut for CTRL+C/CTRL+V/ Return key
    def setup_shortcuts(self):
        self.copy_shortcut = QShortcut(QKeySequence.Copy, self.tv_oilusageinfo, partial(self.copy_cells, self.tv_oilusageinfo))
        self.paste_shortcut = QShortcut(QKeySequence.Paste, self.tv_oilusageinfo, partial(self.paste_cells, self.tv_oilusageinfo))
        self.return_shortcut = QShortcut(Qt.Key_Return, self.tv_oilusageinfo, partial(self.handle_return_key, self.tv_oilusageinfo))


    # Mouse Right click, show "달력보기" menu ---------------------------------------------------------------------
    def create_context_menu(self, target_lineedit):
        context_menu = QMenu()
        custom_action = context_menu.addAction("달력보기")
        custom_action.triggered.connect(lambda: self.show_calendar(target_lineedit))
        return context_menu

    def show_context_menu1(self, pos):
        self.context_menu1.exec_(self.entry_oilusage_getdt.mapToGlobal(pos))

    # Show Calendar
    def show_calendar(self, target_lineedit):
        calendar_dialog = CalendarView()
        calendar_dialog.selected_date_changed.connect(lambda date: self.set_selected_date(date, target_lineedit))
        calendar_dialog.exec()

    # Show selected date to the select Qlineedit
    def set_selected_date(self, date, target_lineedit):
        if target_lineedit == self.entry_oilusage_getdt:
            target_lineedit.setText(date)


    # Call process_key_event and pass the event and your QTableWidget instance
    def keyPressEvent(self, event):
        tv_widget = self.tv_oilusageinfo
        self.process_key_event(event, tv_widget)

    # Display end of date only
    def display_eff_date(self):
        endofdate = "2050/12/31"

        return endofdate
    
    # Pass combobox info and sql to next method
    def get_combobox_contents(self):
        self.insert_combobox_initiate(self.cb_oilusage_ename, "SELECT DISTINCT ename FROM employee ORDER BY ename")

    # Initiate Combo_Box 
    def insert_combobox_initiate(self, combo_box, sql_query):
        self.combobox_initializing(combo_box, sql_query) 
        self.lbl_oilusage_id.setText("")
        self.entry_oilusage_carno.setText("")
        self.entry_oilusage_ecode.setText("")
        self.cb_oilusage_ename.setCurrentIndex(0) 

    # Connect button to method
    def conn_button_to_method(self):
        self.pb_oilusage_show.clicked.connect(self.make_data)
        self.pb_oilusage_search.clicked.connect(self.search_data)        
        self.pb_oilusage_close.clicked.connect(self.close_dialog)
        self.pb_oilusage_clear.clicked.connect(self.clear_data)

        self.pb_oilusage_insert.clicked.connect(self.tb_insert)
        self.pb_oilusage_update.clicked.connect(self.tb_update)
        self.pb_oilusage_delete.clicked.connect(self.tb_delete)
        self.pb_oilusage_xlexport.clicked.connect(self.export_table)
        
    # Connect Signal to Slot
    def connect_signal_slot(self):
        self.cb_oilusage_ename.activated.connect(self.cb_oilusage_ename_changed)

    # tab order for employee window
    def set_tab_order(self):
       
        widgets = [self.pb_oilusage_show, self.entry_oilusage_ecode, self.cb_oilusage_ename,
            self.entry_oilusage_carno, self.entry_oilusage_getdt, self.entry_oilusage_otype,
            self.entry_oilusage_qty, self.entry_oilusage_amt, self.entry_oilusage_remark, 
            
            self.pb_oilusage_search, self.pb_oilusage_clear, self.pb_oilusage_insert, 
            self.pb_oilusage_update, self.pb_oilusage_delete, self.pb_oilusage_close]
        
        for i in range(len(widgets) - 1):
            self.setTabOrder(widgets[i], widgets[i + 1])

    # To reduce duplications
    def common_query_statement(self):
        tv_widget = self.tv_oilusageinfo
        self.cursor.execute("SELECT * FROM vw_oilusage WHERE 1=0")
        column_info = self.cursor.description
        column_names = [col[0] for col in column_info]

        sql_query = "Select * From vw_oilusage"
        column_widths = [80, 100, 100, 100, 100, 100, 100, 100, 150]

        return sql_query, tv_widget, column_info, column_names, column_widths 

    # show employee table data inside of the MDI
    def make_data(self):
        sql_query, tv_widget, column_info, column_names, column_widths = self.common_query_statement() 
        self.populate_dialog(self.cursor, sql_query, tv_widget, column_info, column_names,column_widths)

    # Get the value of other variables
    def get_basic_input(self):
        getdt = str(self.entry_oilusage_getdt.text())
        otype = str(self.entry_oilusage_otype.text())
        carno = str(self.entry_oilusage_carno.text())
        qty = float(self.entry_oilusage_qty.text())
        amt = float(self.entry_oilusage_amt.text())
        remark = str(self.entry_oilusage_remark.text())

        return getdt, otype, carno, qty, amt, remark

    # Make Common values set
    def common_values_set(self):
        username = self.current_username
        user_id = self.userID_gen(username)
        formatted_datetime = self.dt_time_info()

        return username, user_id, formatted_datetime    

    # insert new employee data to MySQL table
    def tb_insert(self):
        currentid = self.lbl_oilusage_id.text()
        if not currentid:
            confirm_dialog = self.show_insert_confirmation_dialog()
            
            if confirm_dialog == QMessageBox.Yes:

                idx = self.max_row_id("oilusage")
                username, user_id, formatted_datetime = self.common_values_set()
                getdt, otype, carno, qty, amt, remark = self.get_basic_input()  
                
                if (idx>0 and qty and amt) and all(len(var) > 0 for var in (otype, carno, getdt)):
                
                    self.cursor.execute('''INSERT INTO oilusage (id, getdt, otype, carno, qty, amt, trxdate, userid, remark) 
                                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)'''
                                , (idx, getdt, otype, carno, qty, amt, formatted_datetime, user_id, remark))
                    self.conn.commit()
                    self.show_insert_success_message()
                    self.refresh_data() 
                    logging.info(f"User {username} inserted id number {idx}, at the oilusage table.")
                else:
                    self.show_missing_message("입력 이상")
                    return
            else:
                self.show_cancel_message("데이터 추가 취소")
                return
        else:
            QMessageBox.information(self, "Input Error", "기존 id가 선택된 상태에서는 신규 입력이 불가합니다!")
            return

    # revise the values in the selected row
    def tb_update(self):

        confirm_dialog = self.show_update_confirmation_dialog()

        if confirm_dialog == QMessageBox.Yes:
            idx = int(self.lbl_oilusage_id.text())
            username, user_id, formatted_datetime = self.common_values_set()         
            getdt, otype, carno, qty, amt, remark = self.get_basic_input()  
                
            if (idx>0 and qty and amt) and all(len(var) > 0 for var in (otype, carno, getdt)):
                self.cursor.execute('''UPDATE oilusage SET getdt=?, otype=?, carno=?, qty=?, amt=?, trxdate=?, userid=?, remark=? WHERE id=?'''
                            , (getdt, otype, carno, qty, amt, formatted_datetime, user_id, remark, idx))
                self.conn.commit()
                self.show_update_success_message()
                self.refresh_data()  
                logging.info(f"User {username} updated row number {idx} in the oilusage table.")            
            else:
                self.show_missing_message("입력 이상")
                return
        else:
            self.show_cancel_message("데이터 변경 취소")
            return


    # delete row according to id selected
    def tb_delete(self):
        confirm_dialog = self.show_delete_confirmation_dialog()

        if confirm_dialog == QMessageBox.Yes:
            idx = self.lbl_oilusage_id.text()
            username, user_id, formatted_datetime = self.common_values_set()

            self.cursor.execute("DELETE FROM oilusage WHERE id=?", (idx,))
            self.conn.commit()
            self.show_delete_success_message()
            self.refresh_data()  
            logging.info(f"User {username} deleted id number {idx}, at the oilusage table.")       
        else:
            self.show_cancel_message("데이터 삭제 취소")

    # Search data
    def search_data(self):
      
        ename = self.cb_oilusage_ename.currentText()
        
        conditions = {'v01': (ename, "ename like '%{}%'"),
                      }
        
        selected_conditions = []

        for key, (value, condition_format) in conditions.items():
            if len(value) > 0:
                selected_conditions.append(condition_format.format(value))

        if not selected_conditions:
            QMessageBox.about(self, "검색 조건 확인", "검색 조건이 비어 있습니다!")
            return

        # Join the selected conditions to form the SQL query
        query = f"SELECT * FROM vw_oilusage WHERE {' AND '.join(selected_conditions)} ORDER BY  ename"

        QMessageBox.about(self, "검색 조건 확인", f"직원명:{ename}  \n\n위 조건으로 검색을 수행합니다!")
        
        sql_query, tv_widget, column_info, column_names, column_widths = self.common_query_statement()
        self.populate_dialog(self.cursor, query, tv_widget, column_info, column_names,column_widths)


    # Employee Name Index Changed
    def cb_oilusage_ename_changed(self):
        self.entry_oilusage_ecode.clear()
        selected_item = self.cb_oilusage_ename.currentText()

        if selected_item:
            query = f"SELECT DISTINCT ecode From employee WHERE ename ='{selected_item}'"
            line_edit_widgets = [self.entry_oilusage_ecode]
        
            # Check if any line edit widgets are provided
            if line_edit_widgets:
                self.lineEdit_contents(line_edit_widgets, query)
            else:
                pass
 
    # Export data to Excel sheet                
    def export_table(self):
        output_subfolder = "data_list"              # set the output subfoler name
        table_widget = self.tv_oilusageinfo             # set the name of table widget
        sheet_name = "oil_payment"                 # set the excel sheet name
        filename = get_file_name(output_subfolder, sheet_name)    # get the file name / export_to_excel 밑에 두면 get_file_name을 두 번 실행시켜서 파일의 번호가 달라지므로 미리 만들어서 filename에 담아둠.
        numeric_columns=[0, 1]                      # set the numerical column index
        export_to_excel(output_subfolder, table_widget, sheet_name, numeric_columns)
               
        if filename:
            full_file_path = os.path.join(output_subfolder, filename)
            self.excel_formatting(sheet_name, full_file_path)
        else:
            QMessageBox.about(self, "코드 확인", "파일을 찾지 못했습니다!")    

    # Excel sheet formatting
    def excel_formatting(self, sheet_name, full_file_path):
        wb = load_workbook(full_file_path)
        sheet_name = sheet_name
        ws = wb[sheet_name]
        last_row = ws.max_row + 1

        column_widths = [8, 14, 10, 10, 10, 8, 10, 10, 16, 20, 10, 10, 20]  # set the width of each column
        title_font = Font(bold=True, name="Arial", size=10)                 # set the font for the first row
        arial_font = Font(name="Arial", size=10)                            # set the forn from the second row to max row

        set_column_widths(ws, column_widths)        # reset column widths

        ws.freeze_panes = 'D2'                      # freeze panes D2 cell
        ws.auto_filter.ref = ws.dimensions          # apply auto filter
        ws.sheet_view.showGridLines = False         # remove gridlines
            
        set_font(ws, [1], range(1, len(column_widths) + 1), title_font)                 # first row font setting
        set_font(ws, range(2, last_row), range(1, len(column_widths) + 1), arial_font)  # from second row to last row font setting
        
        wb.save(full_file_path)
        QMessageBox.about(self, "파일 생성 완료", f"data_list folder에 \n엑셀 파일이 {full_file_path}로 \n생성 되었습니다!")   

    # clear input field entry
    def clear_data(self):
        self.lbl_oilusage_id.setText("")
        for line_edit in self.findChildren(QtWidgets.QLineEdit):
            line_edit.clear()

    # table widget cell double click
    def show_selected_data(self, item):
        # Get the row index of the clicked item
        row_index = item.row()

        # Initialize a list to store the cell values
        cell_values = []

        # Loop through the columns and retrieve the text from each cell
        for column_index in range(9):  # 9 columns
            cell_text = self.tv_oilusageinfo.model().item(row_index, column_index).text()
            cell_values.append(cell_text)

        # Populate the input widgets with the data from the selected row
        self.lbl_oilusage_id.setText(cell_values[0])
        self.entry_oilusage_ecode.setText(cell_values[1])
        self.cb_oilusage_ename.setCurrentText(cell_values[2])
        self.entry_oilusage_carno.setText(cell_values[3])
        self.entry_oilusage_getdt.setText(cell_values[4])
        self.entry_oilusage_otype.setText(cell_values[5])
        self.entry_oilusage_qty.setText(cell_values[6])
        self.entry_oilusage_amt.setText(cell_values[7])
        self.entry_oilusage_remark.setText(cell_values[8])
    
    def refresh_data(self):
        self.clear_data()
        self.make_data()


if __name__ == "__main__":

    log_subfolder = "logs"
    os.makedirs(log_subfolder, exist_ok=True)
    log_file_path = os.path.join(log_subfolder, "access_OilUsageInfoDialog.log")

    logging.basicConfig(
        filename=log_file_path,  
        level=logging.INFO,    
        format="%(asctime)s [%(levelname)s] - %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
        
    app = QtWidgets.QApplication(sys.argv)
    dialog = OilUsageInfoDialog()
    dialog.show()
    sys.exit(app.exec())