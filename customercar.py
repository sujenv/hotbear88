import sys
import logging
from functools import partial
from openpyxl.styles import Font
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from PyQt5 import uic, QtWidgets, QtCore
from PyQt5.QtGui import QStandardItemModel, QKeySequence
from PyQt5.QtWidgets import QDialog, QMessageBox, QMenu, QShortcut
from PyQt5.QtCore import Qt
from datetime import datetime
from commonmd import *
from cal import CalendarView
#for non_ui version-------------------------
#from customercar_ui import Ui_CustomerCarDialog

# customer car table contents -----------------------------------------------------
class CustomerCarDialog(QDialog, SubWindowBase):
#for non_ui version-------------------------
#class CustomerCarDialog(QDialog, Ui_CustomerCarDialog, SubWindowBase):

    def __init__(self, current_username = None, current_datetime = None):
        super().__init__()   
        
        self.conn, self.cursor = connect_to_database1()
        
        # Load ui file
        uic.loadUi("customercar.ui", self)
        #for non_ui version-------------------------
        #self.setupUi(self)

        # Add window flags
        self.setWindowFlags(self.windowFlags() | Qt.WindowMinimizeButtonHint | Qt.WindowMaximizeButtonHint | Qt.WindowCloseButtonHint)  

        # Initialize current_username and current_datetime directly
        self.current_username, self.current_datetime = initialize_username_and_datetime(current_username, current_datetime)

        # Enable automatic deletion on close
        self.setAttribute(QtCore.Qt.WA_DeleteOnClose)  

        # Create tv_customercar and QSortFilterProxyModel
        self.model = QStandardItemModel()
        self.proxy_model = NumericStringSortModel(self.model)
        self.proxy_model.setSourceModel(self.model)

        # Define the column types
        column_types = ["numeric", "", "", "", "numeric", "", "", "", "", ""]
        
        # Set the custom delegate for the specific column
        delegate = NumericDelegate(column_types, self.tv_customercar)
        self.tv_customercar.setItemDelegate(delegate)
        self.tv_customercar.setModel(self.proxy_model)
       
        # Enable sorting
        self.tv_customercar.setSortingEnabled(True)

        # Enable alternating row colors
        self.tv_customercar.setAlternatingRowColors(True)  
        
        # Hide the first index column
        self.tv_customercar.verticalHeader().setVisible(False)

        # While selecting row in tv_customercar, each cell values to displayed to designated widgets
        self.tv_customercar.clicked.connect(self.show_selected_data)

        # Fill combobox items when the application starts
        self.get_combobox_contents()

        # Initiate display of data
        self.make_data()
        self.conn_button_to_method()
        self.connect_signal_slot()

        # Set tab order for input widgets
        self.set_tab_order()

        # Initiate CTRL+C, CTRL+V and ENTER
        self.setup_shortcuts()

        # Create context menus ----------------------------------------------------------------------------------
        self.context_menu1 = self.create_context_menu(self.entry_customercar_efffrom)
        self.context_menu2 = self.create_context_menu(self.entry_customercar_effthru)

        self.entry_customercar_efffrom.setContextMenuPolicy(Qt.CustomContextMenu)
        self.entry_customercar_efffrom.customContextMenuRequested.connect(self.show_context_menu1)

        self.entry_customercar_effthru.setContextMenuPolicy(Qt.CustomContextMenu)
        self.entry_customercar_effthru.customContextMenuRequested.connect(self.show_context_menu2)

        # Make log file
        self.make_logfiles("access_customercar.log")

    # Create a shortcut for CTRL+C/CTRL+V/ Return key
    def setup_shortcuts(self):
        self.copy_shortcut = QShortcut(QKeySequence.Copy, self.tv_customercar, partial(self.copy_cells, self.tv_customercar))
        self.paste_shortcut = QShortcut(QKeySequence.Paste, self.tv_customercar, partial(self.paste_cells, self.tv_customercar))
        self.return_shortcut = QShortcut(Qt.Key_Return, self.tv_customercar, partial(self.handle_return_key, self.tv_customercar))

    # Mouse Right click, show "달력보기" menu ---------------------------------------------------------------------
    def create_context_menu(self, target_lineedit):
        context_menu = QMenu()
        custom_action = context_menu.addAction("달력보기")
        custom_action.triggered.connect(lambda: self.show_calendar(target_lineedit))
        return context_menu

    def show_context_menu1(self, pos):
        self.context_menu1.exec_(self.entry_customercar_efffrom.mapToGlobal(pos))

    def show_context_menu2(self, pos):
        self.context_menu2.exec_(self.entry_customercar_effthru.mapToGlobal(pos))

    # Show Calendar
    def show_calendar(self, target_lineedit):
        calendar_dialog = CalendarView()
        calendar_dialog.selected_date_changed.connect(lambda date: self.set_selected_date(date, target_lineedit))
        calendar_dialog.exec()

    # Show selected date to the select Qlineedit
    def set_selected_date(self, date, target_lineedit):
        if target_lineedit == self.entry_customercar_efffrom:
            target_lineedit.setText(date)
        elif target_lineedit == self.entry_customercar_effthru:
            target_lineedit.setText(date)
            
    # Call process_key_event and pass the event and your QTableWidget instance
    def keyPressEvent(self, event):
        tv_widget = self.tv_customercar
        self.process_key_event(event, tv_widget)

    # Pass combobox info and sql to next method
    def get_combobox_contents(self):
        self.initialize_comboboxes(self.cb_customercar_cname,)
                                      
    # Initiate Combo_Box 
    def initialize_comboboxes(self, *combo_boxes):
        for combo_box in combo_boxes:
            field_name = combo_box.objectName().replace("cb_customercar_", "")
            sql_query = f"SELECT DISTINCT {field_name} FROM vw_cus_car ORDER BY {field_name}"
            self.combobox_initializing(combo_box, sql_query)
            self.clear_comboboxes(combo_box)                  

    # Clear Combo_Box contents
    def clear_comboboxes(self, combo_box):
        self.lbl_customercar_id.setText("")
        self.entry_customercar_code.setText("")
        self.clear_combobox_selections(combo_box)                         

    # Connect the button to the method
    def conn_button_to_method(self):
        # Connect the button to the method
        self.pb_customercar_show.clicked.connect(self.make_data)
        self.pb_customercar_search.clicked.connect(self.search_data)     
        self.pb_customercar_close.clicked.connect(self.close_dialog)
        self.pb_customercar_clearinput.clicked.connect(self.clear_data)

        self.pb_customercar_insert.clicked.connect(self.tb_insert)
        self.pb_customercar_update.clicked.connect(self.tb_update)
        self.pb_customercar_delete.clicked.connect(self.tb_delete)
        
        self.pb_customercar_excel_export.clicked.connect(self.export_table)

    # Connect Signal to Slot
    def connect_signal_slot(self):
        self.cb_customercar_cname.activated.connect(self.cb_customercar_cname_changed)
        self.entry_customercar_bno.editingFinished.connect(self.make_carfullno)      

    # Tab order for sub window
    def set_tab_order(self):
        widgets = [self.entry_customercar_code, self.cb_customercar_cname, self.entry_customercar_fno, 
                self.entry_customercar_bno, self.entry_customercar_fullno, self.entry_customercar_type, 
                self.entry_customercar_efffrom, self.entry_customercar_effthru, self.entry_customercar_remark, 
                self.pb_customercar_show, self.pb_customercar_search, self.pb_customercar_clearinput, 
                self.pb_customercar_insert, self.pb_customercar_update, self.pb_customercar_delete, 
                self.pb_customercar_close]

        for i in range(len(widgets) - 1):
            self.setTabOrder(widgets[i], widgets[i + 1])

    # To reduce duplications
    def common_query_statement(self):
        tv_widget = self.tv_customercar
        
        self.cursor.execute("Select id, cpno1, cpno2, cp, ccode, cname, type01, efffrom, effthru, remark From vw_cus_car WHERE 1=0")
        column_info = self.cursor.description
        column_names = [col[0] for col in column_info]    
        
        sql_query = "Select id, cpno1, cpno2, cp, ccode, cname, type01, efffrom, effthru, remark From vw_cus_car order by id"
        column_widths = [50, 50, 80, 130, 100, 150]

        return sql_query, tv_widget, column_info, column_names, column_widths

    # Make table data
    def make_data(self):
        sql_query, tv_widget, column_info, column_names, column_widths = self.common_query_statement() 
        self.populate_dialog(self.cursor, sql_query, tv_widget, column_info, column_names,column_widths)

    # Get the value of other variables
    def get_customercar_input(self):

        ccode = int(self.entry_customercar_code.text())
        fno = str(self.entry_customercar_fno.text())
        bno = str(self.entry_customercar_bno.text())
        efffrom = str(self.entry_customercar_efffrom.text())
        effthru = str(self.entry_customercar_effthru.text())
        remark = str(self.entry_customercar_remark.text())

        return ccode, fno, bno, efffrom, effthru, remark

    # Make Common values set
    def common_values_set(self):
        username = self.current_username
        user_id = self.userID_gen(username)
        formatted_datetime = self.dt_time_info()
        return username, user_id, formatted_datetime

    # insert new ar_list data to MySQL table
    def tb_insert(self):
        currentid = self.lbl_customercar_id.text()
        if not currentid:
            confirm_dialog = self.show_insert_confirmation_dialog()
            
            if confirm_dialog == QMessageBox.Yes:
                id = self.max_row_id("customercar")
                username, user_id, formatted_datetime = self.common_values_set()
                ccode, fno, bno, efffrom, effthru, remark = self.get_customercar_input() 

            self.cursor.execute('''INSERT INTO customercar (id, cpno1, cpno2, ccode, efffrom, effthru, trxdate, userid, remark)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)''', 
                        (id, fno, bno, ccode, efffrom, effthru, formatted_datetime, user_id, remark))
            
            self.conn.commit()
            QMessageBox.about(self, "새로운 데이터 등록 완료", "새로 추가한 내용의 데이터가 등록 완료되었습니다.")
            self.refresh_data()  
        
        else:
            QMessageBox.information(self, "Input Error", "기존 id가 선택된 상태에서는 신규 입력이 불가합니다!")
            return

    # revise the values in the selected row
    def tb_update(self):
        confirm_dialog = self.show_update_confirmation_dialog()

        if confirm_dialog == QMessageBox.Yes:
            id = int(self.lbl_customercar_id.text())
            username, user_id, formatted_datetime = self.common_values_set()
            ccode, fno, bno, efffrom, effthru, remark = self.get_customercar_input() 

            if (id>0 and ccode>0) and all(len(var) > 0 for var in (fno, bno, efffrom, effthru)):
                self.cursor.execute('''UPDATE customercar SET 
                            cpno1=?, cpno2=?, ccode=?, efffrom=?, effthru=?, trxdate=?, userid=?, remark=? WHERE id=?'''
                            , (fno, bno, ccode, efffrom, effthru, formatted_datetime, user_id, remark, id))
                self.conn.commit()
                self.show_update_success_message()
                self.refresh_data()
                logging.info(f"User {username} updated row number {id} in the customercar table.")
            else:
                self.show_missing_message("입력 이상")
                return
        else:
            self.show_cancel_message("데이터 변경 취소")
            return

    # delete row according to id selected
    def tb_delete(self):
        confirm_dialog = self.show_delete_confirmation_dialog()

        if confirm_dialog == QMessageBox.Yes:
            id = self.lbl_customer_id.text()
            username, user_id, formatted_datetime = self.common_values_set()
            self.cursor.execute("DELETE FROM customercar WHERE id=?", (id,))
            self.conn.commit()
            self.show_delete_success_message()
            self.refresh_data()
            logging.info(f"User {username} deleted {id} row to the customercar table.")
        else:
            self.show_cancel_message("데이터 삭제 취소")
            return
        
    # Search data
    def search_data(self):
        cname = self.cb_customercar_cname.currentText()
        bno = self.entry_customercar_bno.text()

        conditions = {
                    'v01': (cname, "cname like '%{}%'"),
                    'v02': (bno, "cpno2 like '%{}%'"),}
        
        selected_conditions = []

        for key, (value, condition_format) in conditions.items():
            if len(value) > 0:
                selected_conditions.append(condition_format.format(value))

        if not selected_conditions:
            QMessageBox.about(self, "검색 조건 확인", "검색 조건이 비어 있습니다!")
            return

        # Join the selected conditions to form the SQL query
        query = f"SELECT id, cpno1, cpno2, cp, ccode, cname, type01, efffrom, effthru, remark FROM vw_cus_car WHERE {' AND '.join(selected_conditions)} ORDER BY cname"

        QMessageBox.about(self, "검색 조건 확인", f"거래처명: {cname}\n 차량번호: {bno} \n\n위 조건으로 검색을 수행합니다!")
        
        sql_query, tv_widget, column_info, column_names, column_widths = self.common_query_statement()
        self.populate_dialog(self.cursor, query, tv_widget, column_info, column_names,column_widths)

    # Combobox index changed
    def cb_customercar_cname_changed(self):
        self.entry_customercar_code.clear()
        selected_item = self.cb_customercar_cname.currentText()

        if selected_item:
            query = f"SELECT DISTINCT ccode From vw_cus_car WHERE cname ='{selected_item}'"
            line_edit_widgets = [self.entry_customercar_code]
        
            # Check if any line edit widgets are provided
            if line_edit_widgets:
                self.lineEdit_contents(line_edit_widgets, query)
            else:
                pass

    # car plate number editing finished and connect
    def make_carfullno(self):
        fno = str(self.entry_customercar_fno.text())
        bno = str(self.entry_customercar_bno.text())
        self.entry_customercar_fullno.setText(fno + bno)

    # Export data to Excel sheet                
    def export_table(self):
        output_subfolder = "data_list"          # set the output subfoler name
        tv_widget = self.tv_customercar         # set the name of table widget
        sheet_name = "customercar"              # set the excel sheet name
        filename = get_file_name(output_subfolder, sheet_name)    # get the file name / export_to_excel 밑에 두면 get_file_name을 두 번 실행시켜서 파일의 번호가 달라지므로 미리 만들어서 filename에 담아둠.
        numeric_columns=[0, 4]                  # set the numerical column index
        export_to_excel(output_subfolder, tv_widget, sheet_name, numeric_columns)
               
        if filename:
            full_file_path = os.path.join(output_subfolder, filename)
            self.excel_formatting(sheet_name, full_file_path)
        else:
            QMessageBox.about(self, "코드 확인", "파일을 찾지 못했습니다!")    

    # Excel sheet formatting
    def excel_formatting(self, sheet_name, full_file_path):
        wb = load_workbook(full_file_path)
        sheet_name = sheet_name
        ws = wb[sheet_name]
        last_row = ws.max_row + 1

        column_widths = [8, 10, 10, 12, 12, 15, 10, 12, 12, 25]                 # set the width of each column
        title_font = Font(bold=True, name="Arial", size=10)     # set the font for the first row
        arial_font = Font(name="Arial", size=10)                # set the forn from the second row to max row

        set_column_widths(ws, column_widths)    # reset column widths

        ws.freeze_panes = 'G2'                  # freeze panes G2 cell
        ws.auto_filter.ref = ws.dimensions      # apply auto filter
        ws.sheet_view.showGridLines = False     # remove gridlines
            
        set_font(ws, [1], range(1, len(column_widths) + 1), title_font)                 # first row font setting
        set_font(ws, range(2, last_row), range(1, len(column_widths) + 1), arial_font)  # from second row to last row font setting
        
        wb.save(full_file_path)
        QMessageBox.about(self, "파일 생성 완료", f"data_list folder에 \n엑셀 파일이 {full_file_path}로 \n생성 되었습니다!") 

    # clear input field entry
    def clear_data(self):
        self.lbl_customercar_id.setText("")
        clear_widget_data(self)

    # table widget cell double click
    def show_selected_data(self, item):
        # Get the row index of the clicked item
        row_index = item.row()

        # Initialize a list to store the cell values
        cell_values = []

        # Loop through the columns and retrieve the text from each cell
        for column_index in range(10):  # 10 columns
            cell_text = self.tv_customercar.model().item(row_index, column_index).text()
            cell_values.append(cell_text)

        # Populate the input widgets with the data from the selected row
        self.lbl_customercar_id.setText(cell_values[0])
        self.entry_customercar_code.setText(cell_values[4])
        self.cb_customercar_cname.setCurrentText(cell_values[5])
        self.entry_customercar_fno.setText(cell_values[1])
        self.entry_customercar_bno.setText(cell_values[2])
        self.entry_customercar_fullno.setText(cell_values[3])
        self.entry_customercar_type.setText(cell_values[6])
        self.entry_customercar_efffrom.setText(cell_values[7])
        self.entry_customercar_effthru.setText(cell_values[8])
        self.entry_customercar_remark.setText(cell_values[9])

    # Refresh data table
    def refresh_data(self):
        self.clear_data()
        self.make_data()

if __name__ == "__main__":

    log_subfolder = "logs"
    os.makedirs(log_subfolder, exist_ok=True)
    log_file_path = os.path.join(log_subfolder, "access_customercar.log")

    logging.basicConfig(
        filename=log_file_path,  
        level=logging.INFO,    
        format="%(asctime)s [%(levelname)s] - %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    app = QtWidgets.QApplication(sys.argv)
    dialog = CustomerCarDialog()
    dialog.show()
    sys.exit(app.exec())