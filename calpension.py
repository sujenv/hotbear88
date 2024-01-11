import sys
import logging
import math
from functools import partial
from openpyxl.styles import Font
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from PyQt5 import uic, QtWidgets, QtCore
from PyQt5.QtGui import QStandardItemModel, QKeySequence
from PyQt5.QtWidgets import QMessageBox, QDialog, QShortcut, QMenu
from PyQt5.QtCore import Qt
from commonmd import *

# Calendar Master table contents -----------------------------------------------------
class CalPensionDialog(QDialog, SubWindowBase):

    def __init__(self, current_username = None, current_datetime = None):
        super().__init__()

        self.conn, self.cursor = connect_to_database3()    
        uic.loadUi("calpension.ui", self)

        # Add window flags
        self.setWindowFlags(self.windowFlags() | Qt.WindowMinimizeButtonHint | Qt.WindowMaximizeButtonHint | Qt.WindowCloseButtonHint)  

        # Initialize current_username and current_datetime directly
        self.current_username, self.current_datetime = initialize_username_and_datetime(current_username, current_datetime)

        # Enable automatic deletion on close
        self.setAttribute(QtCore.Qt.WA_DeleteOnClose)  

        # Create tv_pensionrate and QSortFilterProxyModel
        self.model = QStandardItemModel()
        self.proxy_model = NumericStringSortModel(self.model)
        self.proxy_model.setSourceModel(self.model)

        # Define the column types
        column_types = ["numeric", "", "numeric", "numeric", "numeric", "numeric", "numeric", ""] 
        column_types1 = ["numeric", "numeric", "", "", "numeric", "", "", ""] 

        # Set the custom delegate for the specific column
        delegate = NumericDelegate(column_types, self.tv_pensionrate)
        self.tv_pensionrate.setItemDelegate(delegate)
        self.tv_pensionrate.setModel(self.proxy_model)

        delegate1 = NumericDelegate(column_types1, self.tv_employeeinfo)
        self.tv_employeeinfo.setItemDelegate(delegate1)
        self.tv_employeeinfo.setModel(self.proxy_model)

        # Enable Sorting
        self.tv_pensionrate.setSortingEnabled(True)
        self.tv_employeeinfo.setSortingEnabled(True)

        # Enable alternating row colors
        self.tv_pensionrate.setAlternatingRowColors(True)  
        self.tv_employeeinfo.setAlternatingRowColors(True)  

        # Hide the first index column
        self.tv_pensionrate.verticalHeader().setVisible(False)
        self.tv_employeeinfo.verticalHeader().setVisible(False)

        # While selecting row in tv_widget, each cell values to displayed to designated widgets
        self.tv_employeeinfo.clicked.connect(self.show_selected_employeeinfo_data)
        self.tv_pensionrate.clicked.connect(self.show_selected_pensionrate_data)

        # Fill combobox items when the application starts
        self.get_combobox_contents()

        # Populate the data
        self.make_data() 
        self.make_data1() 
        self.connect_btn_method()
        self.conn_signal_to_slot()

        # Set tab order for input widgets
        self.set_tab_order()

        # Initiate CTRL+C, CTRL+V and ENTER
        self.setup_shortcuts()

        # Make log file
        self.make_logfiles("access_CalPension.log")        

    # Create a shortcut for CTRL+C/CTRL+V/ Return key
    def setup_shortcuts(self):
        self.copy_shortcut = QShortcut(QKeySequence.Copy, self.tv_pensionrate, partial(self.copy_cells, self.tv_pensionrate))
        self.paste_shortcut = QShortcut(QKeySequence.Paste, self.tv_pensionrate, partial(self.paste_cells, self.tv_pensionrate))
        self.return_shortcut = QShortcut(Qt.Key_Return, self.tv_pensionrate, partial(self.handle_return_key, self.tv_pensionrate))
        self.copy_shortcut = QShortcut(QKeySequence.Copy, self.tv_employeeinfo, partial(self.copy_cells, self.tv_employeeinfo))
        self.paste_shortcut = QShortcut(QKeySequence.Paste, self.tv_employeeinfo, partial(self.paste_cells, self.tv_employeeinfo))
        self.return_shortcut = QShortcut(Qt.Key_Return, self.tv_employeeinfo, partial(self.handle_return_key, self.tv_employeeinfo))

    # Call process_key_event and pass the event and your QTableWidget instance
    def keyPressEvent(self, event):
        tv_widget = self.tv_pensionrate    
        self.process_key_event(event, tv_widget)

    def keyPressEvent1(self, event):
        tv_widget = self.tv_employeeinfo
        self.process_key_event(event, tv_widget)

    # Pass combobox info and sql to next method
    def get_combobox_contents(self):
        self.insert_combobox_initiate(self.cb_pension_ename, "SELECT DISTINCT ename FROM vw_employeeinfo ORDER BY ename")
        self.insert_combobox_initiate(self.cb_pension_dedhc, "SELECT DISTINCT HC FROM taxtable ORDER BY HC")
        self.insert_combobox_initiate(self.cb_pension_year, "SELECT DISTINCT fy FROM basicdata ORDER BY fy")

    # Initiate Combo_Box 
    def insert_combobox_initiate(self, combo_box, sql_query):
        self.combobox_initializing(combo_box, sql_query) 
        self.cb_pension_ename.setCurrentIndex(0) 
        self.cb_pension_dedhc.setCurrentIndex(0) 
        self.cb_pension_year.setCurrentIndex(0) 

    # Connect button to method
    def connect_btn_method(self):
        self.pb_pension_show.clicked.connect(self.make_data)
        self.pb_pension_clear.clicked.connect(self.clear_data)
        self.pb_pension_excel_export.clicked.connect(self.export_data)
        self.pb_pension_close.clicked.connect(self.close_dialog)
        #self.pb_pension_calc.clicked.connect(self.calc_salary)
      
    # Connect signal to method    
    def conn_signal_to_slot(self):
        self.entry_pension_taxedsalary.editingFinished.connect(self.base_salary_changed)
        self.cb_pension_ename.activated.connect(self.employeeinfo_ename_changed)
        self.cb_pension_year.activated.connect(self.cb_pension_year_changed)

    # Base Salary change
    def base_salary_changed(self):
        self.entry_pension_itvalue.clear()
        self.entry_pension_rtvalue.clear()

        selected_text = self.entry_pension_taxedsalary.text()
        selected_item = self.cb_pension_dedhc.currentText()

        if selected_item and selected_text:
            query = f"SELECT * FROM taxtable WHERE GE <= {selected_text} AND HC = {selected_item} ORDER BY id DESC"
            try:
                self.cursor.execute(query)

                result = self.cursor.fetchone()
                if result:
                    item01 = str(result[4])
                    self.entry_pension_itvalue.setText(item01)
                    
            except Exception as e:
                print(f"Error executing SQL query: {e}")

        itv = math.floor(float(self.entry_pension_itvalue.text())/1)*1
        rtv = math.floor((float(self.entry_pension_itvalue.text())*0.1)/10) * 10
        self.entry_pension_rtvalue.setText(str(rtv))
        
        # Calculate National Pension-------------------------------------------------------------------
        # 1. Calculate the age
        mybirthday = self.entry_pension_birthdate.text()
        # Convert the birthdate string to a datetime object
        birthdate = datetime.strptime(mybirthday, "%Y/%m/%d")
        # Get the current date
        today = datetime.now()
        # Calculate the age
        age = today.year - birthdate.year - ((today.month, today.day) < (birthdate.month, birthdate.day))
        # Check if the person is 60 years old or older
        is_60_or_older = age >= 60

        # 2. Calculate National Pension Value and Display
        if is_60_or_older:
            self.entry_pension_npvalue.setText("0")
        else:
            np = math.floor(float(self.entry_pension_taxedsalary.text()) * float(self.entry_pension_nprate.text())/10)*10
            self.entry_pension_npvalue.setText(str(np))

        # Calculate Health Insurance-------------------------------------------------------------------
        hi = math.floor(float(self.entry_pension_taxedsalary.text()) * float(self.entry_pension_hcrate.text())/10)*10
        self.entry_pension_hcvalue.setText(str(hi))

        # 3. Calculate Long Term Care-------------------------------------------------------------------
        ltc = math.floor(hi * float(self.entry_pension_ltcrate.text()) / 10) * 10
        self.entry_pension_ltcvalue.setText(str(ltc))

        # 4. Calculate Long Term Care-------------------------------------------------------------------
        ei = math.floor(float(self.entry_pension_taxedsalary.text()) * float(self.entry_pension_eirate.text())/10)*10
        self.entry_pension_eivalue.setText(str(ei))

        # 5. Total Duduction 
        ttl_pen = itv + rtv + np + hi + ltc + ei
        self.entry_pension_total.setText(str(ttl_pen))

    # tab order for calmaster window
    def set_tab_order(self):
        widgets = [self.pb_pension_show, self.cb_pension_ename, self.entry_pension_birthdate, 
            self.cb_pension_dedhc, self.cb_pension_year, self.entry_pension_taxedsalary, ]
        
        for i in range(len(widgets) - 1):
            self.setTabOrder(widgets[i], widgets[i + 1])


    def query_statement(self):
        tv_widget = self.tv_pensionrate
        
        self.cursor.execute("SELECT * FROM basicdata WHERE 1=0")
        column_info = self.cursor.description
        column_names = [col[0] for col in column_info]

        sql_query = "Select * from basicdata order By id"
        column_widths = [80, 80, 80, 80, 80, 80, 80, 150]

        return sql_query, tv_widget, column_info, column_names, column_widths

    # To reduce duplications
    def make_data(self):
        sql_query, tv_widget, column_info, column_names, column_widths = self.query_statement() 
        self.populate_dialog(self.cursor, sql_query, tv_widget, column_info, column_names,column_widths)

    def query_statement1(self):
        tv_widget = self.tv_employeeinfo
        
        self.cursor.execute("SELECT id, ecode, ename, regid, nofamily, regpension, joindt, remark FROM vw_employeeinfo WHERE 1=0")
        column_info = self.cursor.description
        column_names = [col[0] for col in column_info]

        sql_query = "Select id, ecode, ename, regid, nofamily, regpension, joindt, remark from vw_employeeinfo order By ename"
        column_widths = [80, 100, 100, 150, 80, 100, 100, 150]

        return sql_query, tv_widget, column_info, column_names, column_widths

    # To reduce duplications
    def make_data1(self):
        sql_query, tv_widget, column_info, column_names, column_widths = self.query_statement1() 
        self.populate_dialog(self.cursor, sql_query, tv_widget, column_info, column_names,column_widths)

    # Make Common values set
    def common_values_set(self):
        username = self.current_username
        user_id = self.userID_gen(username)
        formatted_datetime = self.dt_time_info()
        return username, user_id, formatted_datetime

    # clear input field entry
    def clear_data(self):
        for line_edit in self.findChildren(QtWidgets.QLineEdit):
            line_edit.clear()
        self.lbl_pension_regid.setText("")
        self.cb_pension_ename.setCurrentIndex(0)
        self.cb_pension_dedhc.setCurrentIndex(0)
        self.cb_pension_year.setCurrentIndex(0)

    # Combobox apt type index changed
    def employeeinfo_ename_changed(self):
        self.entry_pension_eid.clear()
        self.entry_pension_regid.clear()
        selected_item = self.cb_pension_ename.currentText()

        if selected_item:
            query = f"SELECT DISTINCT ecode, regid From employeeinfo WHERE ename ='{selected_item}'"
            line_edit_widgets = [self.entry_pension_eid, self.entry_pension_regid]
        
            # Check if any line edit widgets are provided
            if line_edit_widgets:
                self.lineEdit_contents(line_edit_widgets, query)
                registration_number = self.entry_pension_regid.text()       
                # 주민번호에 대한 생년월일을 출력
                formatted_birthdate = format_birthdate(registration_number)
                self.entry_pension_birthdate.setText(formatted_birthdate)
            else:
                pass

    def cb_pension_year_changed(self):
        self.entry_pension_nprate.clear()
        self.entry_pension_hcrate.clear()
        self.entry_pension_ltcrate.clear()
        self.entry_pension_eirate.clear()

        selected_item = self.cb_pension_year.currentText()

        if selected_item:
            query = f"SELECT DISTINCT np, hi, ltc, ei From basicdata WHERE fy ='{selected_item}'"
            line_edit_widgets = [self.entry_pension_nprate, self.entry_pension_hcrate, self.entry_pension_ltcrate, self.entry_pension_eirate]
        
            # Check if any line edit widgets are provided
            if line_edit_widgets:
                self.lineEdit_contents(line_edit_widgets, query)
            else:
                pass

    # table view click
    def show_selected_pensionrate_data(self, item):
        # Get the row index of the clicked item
        row_index = item.row()

        # Initialize a list to store the cell values
        cell_values = []

        # Loop through the columns and retrieve the text from each cell
        for column_index in range(8): 
            cell_text = self.tv_pensionrate.model().item(row_index, column_index).text()
            cell_values.append(cell_text)
            
        # Populate the input widgets with the data from the selected row
        self.cb_pension_year.setCurrentText(cell_values[1])
        self.entry_pension_nprate.setText(cell_values[3])
        self.entry_pension_hcrate.setText(cell_values[4])
        self.entry_pension_ltcrate.setText(cell_values[5])
        self.entry_pension_eirate.setText(cell_values[6])


    # table view click
    def show_selected_employeeinfo_data(self, item):
        # Get the row index of the clicked item
        row_index = item.row()

        # Initialize a list to store the cell values
        cell_values = []

        # Loop through the columns and retrieve the text from each cell
        for column_index in range(6): 
            cell_text = self.tv_employeeinfo.model().item(row_index, column_index).text()
            cell_values.append(cell_text)
            
        # Populate the input widgets with the data from the selected row
        self.entry_pension_eid.setText(cell_values[1])
        self.cb_pension_ename.setCurrentText(cell_values[2])
        self.entry_pension_regid.setText(cell_values[3])
        self.cb_pension_dedhc.setCurrentText(cell_values[4])

        registration_number = self.entry_pension_regid.text()       
        # 주민번호에 대한 생년월일을 출력
        formatted_birthdate = format_birthdate(registration_number)
        self.entry_pension_birthdate.setText(formatted_birthdate)

    # 선택된 각 위젯의 내용을 엑셀로 내보내기
    def export_data(self):
        output_subfolder = "data_list"          # set the output subfoler name
        sheet_name = "pension"                   # set the excel sheet name
        filename = get_file_name(output_subfolder, sheet_name)    # get the file name

        # Ensure the subfolder exists; create it if it doesn't
        os.makedirs(output_subfolder, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        if filename:
            full_file_path = os.path.join(output_subfolder, filename)
    
            # 각 위젯에서 내용 가져와서 엑셀에 쓰기
            data_to_export = [
                (self.label_2.text(), self.entry_pension_eid.text()),
                (self.label_11.text(), self.cb_pension_ename.currentText()),
                (self.label_13.text(), self.entry_pension_birthdate.text()),
                (self.label_14.text(), self.cb_pension_dedhc.currentText()),
                (self.label_22.text(), self.cb_pension_year.currentText()),
                (self.label_15.text(), self.entry_pension_taxedsalary.text()),
                (self.label_6.text(), self.entry_pension_itvalue.text()),
                (self.label_20.text(), self.entry_pension_rtvalue.text()),
                (self.label_16.text(), self.entry_pension_nprate.text()),
                (self.label_17.text(), self.entry_pension_npvalue.text()),
                (self.label_27.text(), self.entry_pension_hcrate.text()),
                (self.label_28.text(), self.entry_pension_hcvalue.text()),
                (self.label_29.text(), self.entry_pension_ltcrate.text()),
                (self.label_23.text() , self.entry_pension_ltcvalue.currentText()),
                (self.label_19.text() , self.entry_pension_eirate.text()),
                (self.label_21.text() , self.entry_pension_eivalue.text() ),
                #(self..text() , self..text() ),
                #(self..text() , self..text() ),
                #(self..text() , self..text() ),
                #(self..text() , self..text() ),
                ]

            #for i, value in enumerate(data_to_export, start=1):
            #    ws.cell(row=i, column=1, value=value)
            
            for i, (label, value) in enumerate(data_to_export, start=1):
                ws.cell(row=i, column=1, value=label)
                ws.cell(row=i, column=2, value=value)
            
            # 엑셀 파일 저장
            wb.save(full_file_path)
        
            self.excel_formatting(sheet_name, full_file_path)

    # Excel sheet formatting
    def excel_formatting(self, sheet_name, full_file_path):
        wb = load_workbook(full_file_path)
        sheet_name = sheet_name
        ws = wb[sheet_name]
        last_row = ws.max_row + 2

        # Insert headers at A1 and B1
        ws.insert_rows(1)
        ws['A1'] = '구분'
        ws['B1'] = '내용'

        column_widths = [20, 15]                                # set the width of each column
        title_font = Font(bold=True, name="Arial", size=10)     # set the font for the first row
        arial_font = Font(name="Arial", size=10)                # set the forn from the second row to max row

        set_column_widths(ws, column_widths)    # reset column widths

        ws.freeze_panes = 'B2'              # freeze panes D2 cell
        ws.auto_filter.ref = ws.dimensions  # apply auto filter
        ws.sheet_view.showGridLines = False # remove gridlines
            
        set_font(ws, [1], range(1, len(column_widths) + 1), title_font)                 # first row font setting
        set_font(ws, range(2, last_row), range(1, len(column_widths) + 1), arial_font)  # from second row to last row font setting

        wb.save(full_file_path)
        QMessageBox.about(self, "파일 생성 완료", f"data_list folder에 \n엑셀 파일이 {full_file_path}로 \n생성 되었습니다!")  

    # Refresh data
    def refresh_data(self):
        self.clear_data()
        self.make_data()       

if __name__ == "__main__":
    
    log_subfolder = "logs"
    os.makedirs(log_subfolder, exist_ok=True)
    log_file_path = os.path.join(log_subfolder, "access_CalPension.log")

    logging.basicConfig(
        filename=log_file_path,  
        level=logging.INFO,    
        format="%(asctime)s [%(levelname)s] - %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
    app = QtWidgets.QApplication(sys.argv)
    dialog = CalPensionDialog()
    dialog.show()
    sys.exit(app.exec())
