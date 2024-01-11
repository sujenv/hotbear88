import sys
import logging
from functools import partial
from openpyxl.styles import Font
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from PyQt5 import uic, QtWidgets, QtCore
from PyQt5.QtGui import QStandardItemModel, QKeySequence
from PyQt5.QtWidgets import QDialog, QMessageBox, QShortcut, QMenu, QInputDialog
from PyQt5.QtCore import Qt
from datetime import datetime
from cal import CalendarView
from commonmd import *
#for non_ui version-------------------------
#from gpai_payment_info_ui import Ui_GPAIInfoDialog

# Bank Account Product table contents -----------------------------------------------------
class GPAIInfoDialog(QDialog, SubWindowBase):
#for non_ui version-------------------------
#class GPAIInfoDialog(QDialog, Ui_GPAIInfoDialog, SubWindowBase):
     
    def __init__(self, current_username = None, current_datetime = None):
        super().__init__()    
  
        self.conn, self.cursor = connect_to_database3()

        # load ui file
        uic.loadUi("gpai_payment_info.ui", self)
        #for non_ui version-------------------------
        #self.setupUi(self)

        # Add window flags
        self.setWindowFlags(self.windowFlags() | Qt.WindowMinimizeButtonHint | Qt.WindowMaximizeButtonHint | Qt.WindowCloseButtonHint)  

        # Initialize current_username and current_datetime directly
        self.current_username, self.current_datetime = initialize_username_and_datetime(current_username, current_datetime)

        # Enable automatic deletion on close
        self.setAttribute(QtCore.Qt.WA_DeleteOnClose)  

        # Create tv_gpaiinfo and QSortFilterProxyModel
        self.model = QStandardItemModel()
        self.proxy_model = NumericStringSortModel(self.model)
        self.proxy_model.setSourceModel(self.model)

        # Define the column types
        column_types = ["numeric", "numeric", "", "numeric", "numeric", "numeric", "", "", "", ]

        # Set the custom delegate for the specific column
        delegate = NumericDelegate(column_types,self.tv_gpaiinfo)
        self.tv_gpaiinfo.setItemDelegate(delegate)
        self.tv_gpaiinfo.setModel(self.proxy_model)

        # Enable sorting
        self.tv_gpaiinfo.setSortingEnabled(True)

        # Enable alternating row colors
        self.tv_gpaiinfo.setAlternatingRowColors(True)  
        
        # Hide the first index column
        self.tv_gpaiinfo.verticalHeader().setVisible(False)

        # While selecting row in tv_gpaiinfo, each cell values to displayed to designated widgets
        self.tv_gpaiinfo.clicked.connect(self.show_selected_data)

        # Fill combobox items when the application starts
        self.get_combobox_contents()

        # Initiate display of data
        self.make_data()
        self.conn_button_to_method()
        self.connect_signal_slot()

        # Set tab order for input widgets
        self.set_tab_order()

        # Initiate CTRL+C, CTRL+V and ENTER
        self.setup_shortcuts()

        # Create context menus ----------------------------------------------------------------------------------
        self.context_menu1 = self.create_context_menu(self.entry_gpai_efffrom)
        self.context_menu2 = self.create_context_menu(self.entry_gpai_effthru)
        self.context_menu3 = self.create_context_menu(self.entry_gpai_cefffrom)
        self.context_menu4 = self.create_context_menu(self.entry_gpai_ceffthru)

        self.entry_gpai_efffrom.setContextMenuPolicy(Qt.CustomContextMenu)
        self.entry_gpai_efffrom.customContextMenuRequested.connect(self.show_context_menu1)

        self.entry_gpai_effthru.setContextMenuPolicy(Qt.CustomContextMenu)
        self.entry_gpai_effthru.customContextMenuRequested.connect(self.show_context_menu2)

        self.entry_gpai_cefffrom.setContextMenuPolicy(Qt.CustomContextMenu)
        self.entry_gpai_cefffrom.customContextMenuRequested.connect(self.show_context_menu3)

        self.entry_gpai_ceffthru.setContextMenuPolicy(Qt.CustomContextMenu)
        self.entry_gpai_ceffthru.customContextMenuRequested.connect(self.show_context_menu4)

        self.entry_stylesheet_as_is()
        self.hide_bkaccno_change_widget()

        # Make log file
        self.make_logfiles("access_GPAIInfoDialog.log")

    # Create a shortcut for CTRL+C/CTRL+V/ Return key
    def setup_shortcuts(self):
        self.copy_shortcut = QShortcut(QKeySequence.Copy, self.tv_gpaiinfo, partial(self.copy_cells, self.tv_gpaiinfo))
        self.paste_shortcut = QShortcut(QKeySequence.Paste, self.tv_gpaiinfo, partial(self.paste_cells, self.tv_gpaiinfo))
        self.return_shortcut = QShortcut(Qt.Key_Return, self.tv_gpaiinfo, partial(self.handle_return_key, self.tv_gpaiinfo))

    # Change styles for the selected widgets 1
    def entry_stylesheet_as_is(self):
        self.entry_gpai_cttlval.setStyleSheet('color:white;background:rgb(0,0,0)')
        self.entry_gpai_ccomval.setStyleSheet('color:white;background:rgb(0,0,0)')
        self.entry_gpai_cindval.setStyleSheet('color:white;background:rgb(0,0,0)')
        self.entry_gpai_cefffrom.setStyleSheet('color:white;background:rgb(0,0,0)')
        self.entry_gpai_ceffthru.setStyleSheet('color:white;background:rgb(0,0,0)')
        self.entry_gpai_cremark.setStyleSheet('color:white;background:rgb(0,0,0)')

    # Change styles for the selected widgets 2
    def entry_stylesheet_to_be(self):
        self.entry_gpai_cttlval.setStyleSheet('color:black;background:rgb(255,255,0)')
        self.entry_gpai_ccomval.setStyleSheet('color:black;background:rgb(255,255,0)')
        self.entry_gpai_cindval.setStyleSheet('color:black;background:rgb(255,255,0)')
        self.entry_gpai_cefffrom.setStyleSheet('color:black;background:rgb(255,255,0)')
        self.entry_gpai_ceffthru.setStyleSheet('color:black;background:rgb(255,255,0)')
        self.entry_gpai_cremark.setStyleSheet('color:black;background:rgb(255,255,0)')

    # Show widgets for the cost change parts 
    def show_bkaccno_change_widget(self):
        self.pb_gpai_changeinsert.setVisible(True)
        self.entry_gpai_cttlval.setReadOnly(False)
        self.entry_gpai_ccomval.setReadOnly(False)
        self.entry_gpai_cindval.setReadOnly(False)
        self.entry_gpai_cefffrom.setReadOnly(False)
        self.entry_gpai_ceffthru.setReadOnly(False)
        self.entry_gpai_cremark.setReadOnly(False)
        self.entry_stylesheet_to_be()

    # Hide widgets for the cost change parts 
    def hide_bkaccno_change_widget(self):
        self.pb_gpai_changeinsert.setVisible(False)
        self.entry_gpai_cttlval.setReadOnly(True)
        self.entry_gpai_ccomval.setReadOnly(True)
        self.entry_gpai_cindval.setReadOnly(True)
        self.entry_gpai_cefffrom.setReadOnly(True)
        self.entry_gpai_ceffthru.setReadOnly(True)
        self.entry_gpai_cremark.setReadOnly(True)
        self.entry_stylesheet_as_is()

    # Mouse Right click, show "달력보기" menu ---------------------------------------------------------------------
    def create_context_menu(self, target_lineedit):
        context_menu = QMenu()
        custom_action = context_menu.addAction("달력보기")
        custom_action.triggered.connect(lambda: self.show_calendar(target_lineedit))
        return context_menu

    def show_context_menu1(self, pos):
        self.context_menu1.exec_(self.entry_gpai_efffrom.mapToGlobal(pos))
    def show_context_menu2(self, pos):
        self.context_menu2.exec_(self.entry_gpai_effthru.mapToGlobal(pos))
    def show_context_menu3(self, pos):
        self.context_menu3.exec_(self.entry_gpai_cefffrom.mapToGlobal(pos))
    def show_context_menu4(self, pos):
        self.context_menu4.exec_(self.entry_gpai_ceffthru.mapToGlobal(pos))

    # Show Calendar
    def show_calendar(self, target_lineedit):
        calendar_dialog = CalendarView()
        calendar_dialog.selected_date_changed.connect(lambda date: self.set_selected_date(date, target_lineedit))
        calendar_dialog.exec()

    # Show selected date to the select Qlineedit
    def set_selected_date(self, date, target_lineedit):
        if target_lineedit == self.entry_gpai_efffrom:
            target_lineedit.setText(date)
        elif target_lineedit == self.entry_gpai_effthru:
            target_lineedit.setText(date)
        elif target_lineedit == self.entry_gpai_cefffrom:
            target_lineedit.setText(date)
        elif target_lineedit == self.entry_gpai_ceffthru:
            target_lineedit.setText(date)

    # Call process_key_event and pass the event and your QTableWidget instance
    def keyPressEvent(self, event):
        tv_widget = self.tv_gpaiinfo
        self.process_key_event(event, tv_widget)

    # Display end of date only
    def display_eff_date(self):
        endofdate = "2050/12/31"

        return endofdate
    
    # Pass combobox info and sql to next method
    def get_combobox_contents(self):
        self.insert_combobox_initiate(self.cb_gpai_ename, "SELECT DISTINCT ename FROM employee ORDER BY ename")

    # Initiate Combo_Box 
    def insert_combobox_initiate(self, combo_box, sql_query):
        self.combobox_initializing(combo_box, sql_query) 
        self.lbl_gpai_id.setText("")
        self.entry_gpai_ttlval.setText("")
        self.entry_gpai_ecode.setText("")
        self.cb_gpai_ename.setCurrentIndex(0) 

    # Connect button to method
    def conn_button_to_method(self):
        self.pb_gpai_show.clicked.connect(self.make_data)
        self.pb_gpai_search.clicked.connect(self.search_data)        
        self.pb_gpai_close.clicked.connect(self.close_dialog)
        self.pb_gpai_clear.clicked.connect(self.clear_data)

        self.pb_gpai_insert.clicked.connect(self.tb_insert)
        self.pb_gpai_update.clicked.connect(self.SelectionMessageBox)
        self.pb_gpai_delete.clicked.connect(self.tb_delete)
        self.pb_gpai_excel_export.clicked.connect(self.export_table)
        self.pb_gpai_changeinsert.clicked.connect(self.reflect_bkaccno_change)
        
    # Connect Signal to Slot
    def connect_signal_slot(self):
        self.cb_gpai_ename.activated.connect(self.cb_gpai_ename_changed)
        self.entry_gpai_efffrom.editingFinished.connect(self.sdt_changed)     
        self.entry_gpai_cefffrom.editingFinished.connect(self.chgeffrom_changed)
        self.entry_gpai_cttlval.editingFinished.connect(self.cttlval_changed)

    # tab order for employee window
    def set_tab_order(self):
       
        widgets = [self.pb_gpai_show, self.entry_gpai_ecode, self.cb_gpai_ename,
            self.entry_gpai_ttlval, self.entry_gpai_comval, self.entry_gpai_indval, 
            self.entry_gpai_efffrom, self.entry_gpai_effthru, self.entry_gpai_remark, 
            
            self.entry_gpai_cttlval, self.entry_gpai_ccomval, self.entry_gpai_cindval, 
            self.entry_gpai_cefffrom, self.entry_gpai_ceffthru, self.entry_gpai_cremark,

            self.pb_gpai_search, self.pb_gpai_clear, self.pb_gpai_insert, 
            self.pb_gpai_update, self.pb_gpai_delete, self.pb_gpai_close]
        
        for i in range(len(widgets) - 1):
            self.setTabOrder(widgets[i], widgets[i + 1])

    # To reduce duplications
    def common_query_statement(self):
        tv_widget = self.tv_gpaiinfo
        self.cursor.execute("SELECT * FROM vw_gpai WHERE 1=0")
        column_info = self.cursor.description
        column_names = [col[0] for col in column_info]

        sql_query = "Select * From vw_gpai Order By ename"
        column_widths = [80, 100, 100, 100, 100, 100, 100, 100, 150]

        return sql_query, tv_widget, column_info, column_names, column_widths 

    # show employee table data inside of the MDI
    def make_data(self):
        sql_query, tv_widget, column_info, column_names, column_widths = self.common_query_statement() 
        self.populate_dialog(self.cursor, sql_query, tv_widget, column_info, column_names,column_widths)

    # Get the value of other variables
    def get_basic_input(self):
        ecode = int(self.entry_gpai_ecode.text())
        ttlval = int(self.entry_gpai_ttlval.text())
        comval = str(self.entry_gpai_comval.text())
        efffrom = str(self.entry_gpai_efffrom.text())
        effthru = str(self.entry_gpai_effthru.text())
        remark = str(self.entry_gpai_remark.text())

        return ecode, ttlval, comval, efffrom, effthru, remark

    # Make Common values set
    def common_values_set(self):
        username = self.current_username
        user_id = self.userID_gen(username)
        formatted_datetime = self.dt_time_info()

        return username, user_id, formatted_datetime    

    # insert new employee data to MySQL table
    def tb_insert(self):
        currentid = self.lbl_gpai_id.text()
        if not currentid:
            confirm_dialog = self.show_insert_confirmation_dialog()
            
            if confirm_dialog == QMessageBox.Yes:

                idx = self.max_row_id("gpai")
                username, user_id, formatted_datetime = self.common_values_set()
                ecode, ttlval, comval, efffrom, effthru, remark = self.get_basic_input()  
                
                if (idx>0 and ecode and ttlval and comval) and all(len(var) > 0 for var in (efffrom, effthru)):
                
                    self.cursor.execute('''INSERT INTO gpai (id, ecode, ttlval, comval, efffrom, effthru, trxdate, userid, remark) 
                                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)'''
                                , (idx, ecode, ttlval, comval, efffrom, effthru, formatted_datetime, user_id, remark))
                    self.conn.commit()
                    self.show_insert_success_message()
                    self.refresh_data() 
                    logging.info(f"User {username} inserted id number {idx}, at the group personal accident insurance table.")
                else:
                    self.show_missing_message("입력 이상")
                    return
            else:
                self.show_cancel_message("데이터 추가 취소")
                return
        else:
            QMessageBox.information(self, "Input Error", "기존 id가 선택된 상태에서는 신규 입력이 불가합니다!")
            return

    # update 조건에 따라 분기 할 것
    def SelectionMessageBox(self):

        # If there's no selection
        if len(self.lbl_gpai_id.text()) == 0:
            self.show_missing_message_update("입력 확인")

        # In case of row selection

        conA = '''물품대 계좌 내용 중 오류 수정 - 현재 행을 수정, 추가 행을 만들지 않음!'''
        conB = '''물품대 계좌 내용의 변경 또는 갱신 - 현재 행의 종료일 변경, 변경된 내용으로 추가 행을 만듦!'''
        
        conditions = [conA, conB]
        condition, okPressed = QInputDialog.getItem(self, "입력 조건 선택", "어떤 작업을 진행하시겠습니까?:", conditions, 0, False)

        if okPressed:
            if condition == conA:
                self.fix_typo()     
            elif condition == conB:
                self.show_bkaccno_change_widget()
            else:
                return

    # revise the values in the selected row
    def fix_typo(self):

        confirm_dialog = self.show_update_confirmation_dialog()

        if confirm_dialog == QMessageBox.Yes:
            idx = int(self.lbl_gpai_id.text())
            username, user_id, formatted_datetime = self.common_values_set()         
            ecode, ttlval, comval, efffrom, effthru, remark = self.get_basic_input()  
            
            if (idx>0 and ecode and ttlval and comval) and all(len(var) > 0 for var in (efffrom, effthru)):
                self.cursor.execute('''UPDATE gpai SET ecode=?, ttlval=?, comval=?, efffrom=?, effthru=?, trxdate=?, userid=?, remark=? WHERE id=?'''
                            , (ecode, ttlval, comval, efffrom, effthru, formatted_datetime, user_id, remark, idx))
                self.conn.commit()
                self.show_update_success_message()
                self.refresh_data()  
                logging.info(f"User {username} updated row number {idx} in the group personal accident insurance table.")            
            else:
                self.show_missing_message("입력 이상")
                return
        else:
            self.show_cancel_message("데이터 변경 취소")
            return

    # Get the Changed Infos
    def get_changed_info(self):
        cecode = int(self.entry_gpai_ecode.text())
        cttlval = str(self.entry_gpai_cttlval.text())
        ccomval = str(self.entry_gpai_ccomval.text())
        cefffrom = str(self.entry_gpai_cefffrom.text())
        ceffthru = str(self.entry_gpai_ceffthru.text())
        cremark = str(self.entry_gpai_cremark.text())
        
        return cecode, cttlval, ccomval, cefffrom, ceffthru, cremark
    
    # Bank Account Change and Insert
    def reflect_bkaccno_change(self):

        username, user_id, formatted_datetime = self.common_values_set()
        cecode, cttlval, ccomval, cefffrom, ceffthru, cremark  = self.get_changed_info()

        idx = int(self.max_row_id("gpai"))        
                
        if (idx>0 and cecode and cttlval and ccomval) and all(len(var) > 0 for var in (cefffrom, ceffthru)):

            org_id = str(self.lbl_gpai_id.text())
            effthru1 = str(self.entry_gpai_effthru.text()) # 다시 불러와야 함..
            # 기존 id의 유효종료일을 변경유효시작일 -1 일로 수정
            self.cursor.execute('''UPDATE gpai SET effthru=? WHERE id=?''', (effthru1, org_id))
            
            #변경된 내용을 신규로 추가
            self.cursor.execute('''INSERT INTO gpai (id, ecode, ttlval, comval, efffrom, effthru, trxdate, userid, remark) 
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)'''
                        , (idx, cecode, cttlval, ccomval, cefffrom, ceffthru, formatted_datetime, user_id, cremark))
            
            self.conn.commit()
            self.show_update_success_message()
            self.refresh_data()
            logging.info(f"User {username} updated row number {idx} in the group personal accident insurance table.")
            
        else:
            self.show_missing_message("입력 이상")
            return

        self.entry_stylesheet_as_is()
        self.hide_bkaccno_change_widget()

    # delete row according to id selected
    def tb_delete(self):
        confirm_dialog = self.show_delete_confirmation_dialog()

        if confirm_dialog == QMessageBox.Yes:
            idx = self.lbl_gpai_id.text()
            username, user_id, formatted_datetime = self.common_values_set()

            self.cursor.execute("DELETE FROM gpai WHERE id=?", (idx,))
            self.conn.commit()
            self.show_delete_success_message()
            self.refresh_data()  
            logging.info(f"User {username} deleted id number {idx}, at the group personal accident insurance table.")       
        else:
            self.show_cancel_message("데이터 삭제 취소")

    # Search data
    def search_data(self):
      
        ename = self.cb_gpai_ename.currentText()

        
        conditions = {'v01': (ename, "ename like '%{}%'"),
                      }
        
        selected_conditions = []

        for key, (value, condition_format) in conditions.items():
            if len(value) > 0:
                selected_conditions.append(condition_format.format(value))

        if not selected_conditions:
            QMessageBox.about(self, "검색 조건 확인", "검색 조건이 비어 있습니다!")
            return

        # Join the selected conditions to form the SQL query
        query = f"SELECT * FROM vw_gpai WHERE {' AND '.join(selected_conditions)} ORDER BY  ename"

        QMessageBox.about(self, "검색 조건 확인", f"직원명:{ename}  \n\n위 조건으로 검색을 수행합니다!")
        
        sql_query, tv_widget, column_info, column_names, column_widths = self.common_query_statement()
        self.populate_dialog(self.cursor, query, tv_widget, column_info, column_names,column_widths)


    # Employee Name Index Changed
    def cb_gpai_ename_changed(self):
        self.entry_gpai_ecode.clear()
        selected_item = self.cb_gpai_ename.currentText()

        if selected_item:
            query = f"SELECT DISTINCT ecode From employee WHERE ename ='{selected_item}'"
            line_edit_widgets = [self.entry_gpai_ecode]
        
            # Check if any line edit widgets are provided
            if line_edit_widgets:
                self.lineEdit_contents(line_edit_widgets, query)
            else:
                pass

    # Effective Date Index Changed
    def sdt_changed(self):
        # inputed string type date
        date_string = self.entry_gpai_efffrom.text()
        
        # convert string type date to date format
        try:
            startdt = datetime.strptime(date_string, "%Y-%m-%d")
        except ValueError:
            startdt = datetime.strptime(date_string, "%Y/%m/%d")

        # Find the last day of the month for the given date
        _, last_day = calendar.monthrange(startdt.year, startdt.month)
        last_day_of_month = datetime(startdt.year, startdt.month, last_day)
        
        # Calculate the end date, which is one day before the last day of the month
        enddt = last_day_of_month - timedelta(days=0)
        effthru = "2050/12/31"
        
        # Format the end date as a string and set it to the desired widget
        enddt = enddt.strftime("%Y/%m/%d")
        
        self.entry_gpai_effthru.setText(effthru)
          
    def cttlval_changed(self):
        self.entry_gpai_ccomval.setText("")
        self.entry_gpai_cindval.setText("")

        compay = 15000
        self.entry_gpai_ccomval.setText(str(compay))
        indpay = int(self.entry_gpai_cttlval.text()) - compay
        self.entry_gpai_cindval.setText(str(indpay))

    # CEfffrom Changed
    def chgeffrom_changed(self):
        chg_date_str = self.entry_gpai_cefffrom.text()
        try:
            chg_date = parse_date(chg_date_str)                             # 날짜 문자열을 날짜 객체로 변환
            org_date = chg_date - timedelta(days=1)                         # 날짜에서 1일을 빼줍니다.
            org_date_str = org_date.strftime('%Y-%m-%d')                    # 결과를 문자열로 변환
            self.entry_gpai_effthru.setText(org_date_str)                   # 변경된 effthru 날짜를 표시
        
            endofdate = self.display_eff_date()
            self.entry_gpai_ceffthru.setText(endofdate)

        except ValueError as e:
            QMessageBox.critical(self, "날짜 형식 오류", str(e))             # 날짜 형식이 잘못된 경우 사용자에게 알림


    # Export data to Excel sheet                
    def export_table(self):
        output_subfolder = "data_list"              # set the output subfoler name
        table_widget = self.tv_gpaiinfo             # set the name of table widget
        sheet_name = "GPAI_payment"                 # set the excel sheet name
        filename = get_file_name(output_subfolder, sheet_name)    # get the file name / export_to_excel 밑에 두면 get_file_name을 두 번 실행시켜서 파일의 번호가 달라지므로 미리 만들어서 filename에 담아둠.
        numeric_columns=[0, 1]                      # set the numerical column index
        export_to_excel(output_subfolder, table_widget, sheet_name, numeric_columns)
               
        if filename:
            full_file_path = os.path.join(output_subfolder, filename)
            self.excel_formatting(sheet_name, full_file_path)
        else:
            QMessageBox.about(self, "코드 확인", "파일을 찾지 못했습니다!")    

    # Excel sheet formatting
    def excel_formatting(self, sheet_name, full_file_path):
        wb = load_workbook(full_file_path)
        sheet_name = sheet_name
        ws = wb[sheet_name]
        last_row = ws.max_row + 1

        column_widths = [8, 14, 10, 10, 10, 8, 10, 10, 16, 20, 10, 10, 20]  # set the width of each column
        title_font = Font(bold=True, name="Arial", size=10)                 # set the font for the first row
        arial_font = Font(name="Arial", size=10)                            # set the forn from the second row to max row

        set_column_widths(ws, column_widths)        # reset column widths

        ws.freeze_panes = 'D2'                      # freeze panes D2 cell
        ws.auto_filter.ref = ws.dimensions          # apply auto filter
        ws.sheet_view.showGridLines = False         # remove gridlines
            
        set_font(ws, [1], range(1, len(column_widths) + 1), title_font)                 # first row font setting
        set_font(ws, range(2, last_row), range(1, len(column_widths) + 1), arial_font)  # from second row to last row font setting
        
        wb.save(full_file_path)
        QMessageBox.about(self, "파일 생성 완료", f"data_list folder에 \n엑셀 파일이 {full_file_path}로 \n생성 되었습니다!")   

    # clear input field entry
    def clear_data(self):
        self.lbl_gpai_id.setText("")
        for line_edit in self.findChildren(QtWidgets.QLineEdit):
            line_edit.clear()

    # table widget cell double click
    def show_selected_data(self, item):
        # Get the row index of the clicked item
        row_index = item.row()

        # Initialize a list to store the cell values
        cell_values = []

        # Loop through the columns and retrieve the text from each cell
        for column_index in range(9):  # 9 columns
            cell_text = self.tv_gpaiinfo.model().item(row_index, column_index).text()
            cell_values.append(cell_text)

        # Populate the input widgets with the data from the selected row
        self.lbl_gpai_id.setText(cell_values[0])
        self.entry_gpai_ecode.setText(cell_values[1])
        self.cb_gpai_ename.setCurrentText(cell_values[2])
        self.entry_gpai_ttlval.setText(cell_values[3])
        self.entry_gpai_comval.setText(cell_values[4])
        self.entry_gpai_indval.setText(cell_values[5])
        self.entry_gpai_efffrom.setText(cell_values[6])
        self.entry_gpai_effthru.setText(cell_values[7])
        self.entry_gpai_remark.setText(cell_values[8])
    
    def refresh_data(self):
        self.clear_data()
        self.make_data()


if __name__ == "__main__":

    log_subfolder = "logs"
    os.makedirs(log_subfolder, exist_ok=True)
    log_file_path = os.path.join(log_subfolder, "access_GPAIInfoDialog.log")

    logging.basicConfig(
        filename=log_file_path,  
        level=logging.INFO,    
        format="%(asctime)s [%(levelname)s] - %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
        
    app = QtWidgets.QApplication(sys.argv)
    dialog = GPAIInfoDialog()
    dialog.show()
    sys.exit(app.exec())