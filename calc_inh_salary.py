import sys
import logging
import math
from functools import partial
from openpyxl.styles import Font
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from PyQt5 import uic, QtWidgets, QtCore
from PyQt5.QtGui import QStandardItemModel, QKeySequence
from PyQt5.QtWidgets import QMessageBox, QDialog, QShortcut, QMenu
from PyQt5.QtCore import Qt
from commonmd import *

# Calendar Master table contents -----------------------------------------------------
class CalSalaryInhDialog(QDialog, SubWindowBase):

    def __init__(self, current_username = None, current_datetime = None):
        super().__init__()

        self.conn, self.cursor = connect_to_database3()    
        uic.loadUi("calc_inh_salary.ui", self)

        # Add window flags
        self.setWindowFlags(self.windowFlags() | Qt.WindowMinimizeButtonHint | Qt.WindowMaximizeButtonHint | Qt.WindowCloseButtonHint)  

        # Initialize current_username and current_datetime directly
        self.current_username, self.current_datetime = initialize_username_and_datetime(current_username, current_datetime)

        # Enable automatic deletion on close
        self.setAttribute(QtCore.Qt.WA_DeleteOnClose)  

        # Create tv_otdays and QSortFilterProxyModel
        self.model = QStandardItemModel()
        self.proxy_model = NumericStringSortModel(self.model)
        self.proxy_model.setSourceModel(self.model)

        # Define the column types
        column_types = ["numeric", "", "numeric", ]  
        column_types1 = ["numeric", "numeric", "numeric", "numeric", "numeric", "numeric", 
                        "numeric", "numeric", "numeric", "numeric", "numeric", "numeric",
                        "numeric", "numeric", "numeric", "numeric", "numeric", "numeric",
                        "numeric", "numeric", ] 

        # Set the custom delegate for the specific column
        delegate = NumericDelegate(column_types, self.tv_otdays)
        self.tv_otdays.setItemDelegate(delegate)
        self.tv_otdays.setModel(self.proxy_model)

        delegate1 = NumericDelegate(column_types1, self.tv_salarylist)
        self.tv_salarylist.setItemDelegate(delegate1)
        self.tv_salarylist.setModel(self.proxy_model)

        # Enable Sorting
        self.tv_otdays.setSortingEnabled(True)
        self.tv_salarylist.setSortingEnabled(True)

        # Enable alternating row colors
        self.tv_otdays.setAlternatingRowColors(True)  
        self.tv_salarylist.setAlternatingRowColors(True)  

        # Hide the first index column
        self.tv_otdays.verticalHeader().setVisible(False)
        self.tv_salarylist.verticalHeader().setVisible(False)

        # While selecting row in tv_widget, each cell values to displayed to designated widgets
        self.tv_otdays.clicked.connect(self.show_selected_otdays_data)
        self.tv_salarylist.clicked.connect(self.show_selected_salarylist_data)
        
        self.entry_in_workinghrs.setText("8")
        self.entry_in_working_days.setText("5")
        self.entry_in_saturdayworkhrs.setText("8")
        self.entry_in_weekdayholidaywkhrs.setText("8")
        self.lbl_in_id.setText("")
        self.ckbox_ot.setChecked(True)
        self.ckbox_al.setChecked(True)        

        # Fill combobox items when the application starts
        self.get_combobox_contents()

        # Populate the data
        self.make_data() 
        self.make_data1() 
        self.connect_btn_method()
        self.conn_signal_to_slot()

        # Set tab order for input widgets
        self.set_tab_order()

        # Initiate CTRL+C, CTRL+V and ENTER
        self.setup_shortcuts()

        # Make log file
        self.make_logfiles("access_CalcSalaryInh.log")        

    # Create a shortcut for CTRL+C/CTRL+V/ Return key
    def setup_shortcuts(self):
        self.copy_shortcut = QShortcut(QKeySequence.Copy, self.tv_otdays, partial(self.copy_cells, self.tv_otdays))
        self.paste_shortcut = QShortcut(QKeySequence.Paste, self.tv_otdays, partial(self.paste_cells, self.tv_otdays))
        self.return_shortcut = QShortcut(Qt.Key_Return, self.tv_otdays, partial(self.handle_return_key, self.tv_otdays))
        
        self.copy_shortcut = QShortcut(QKeySequence.Copy, self.tv_salarylist, partial(self.copy_cells, self.tv_salarylist))
        self.paste_shortcut = QShortcut(QKeySequence.Paste, self.tv_salarylist, partial(self.paste_cells, self.tv_salarylist))
        self.return_shortcut = QShortcut(Qt.Key_Return, self.tv_salarylist, partial(self.handle_return_key, self.tv_salarylist))

    # Call process_key_event and pass the event and your QTableWidget instance
    def keyPressEvent(self, event):
        tv_widget = self.tv_otdays    
        self.process_key_event(event, tv_widget)

    def keyPressEvent1(self, event):
        tv_widget = self.tv_salarylist
        self.process_key_event(event, tv_widget)

    # Pass combobox info and sql to next method
    def get_combobox_contents(self):
        self.insert_combobox_initiate(self.cb_in_year, "SELECT DISTINCT fy FROM basicdata ORDER BY fy")

    # Initiate Combo_Box 
    def insert_combobox_initiate(self, combo_box, sql_query):
        self.combobox_initializing(combo_box, sql_query) 
        self.cb_in_year.setCurrentIndex(0) 

    # Connect button to method
    def connect_btn_method(self):
        self.pb_in_calc.clicked.connect(self.calc_change)
        self.pb_in_clear.clicked.connect(self.clear_data)
        self.pb_in_close.clicked.connect(self.close_dialog)
        self.pb_in_insert.clicked.connect(self.tv_insert)
        self.pb_in_delete.clicked.connect(self.tv_delete)
      
    # Connect signal to method    
    def conn_signal_to_slot(self):
        self.cb_in_year.activated.connect(self.cb_in_year_changed)
        self.entry_in_wage.editingFinished.connect(self.mw_changed)
        self.entry_in_saturdayworkhrs.editingFinished.connect(self.sat_changed)
        self.entry_in_weekdayothrsperday.editingFinished.connect(self.wkdyot_changed)
        self.entry_in_weekdayotdaysperweek.editingFinished.connect(self.wkdyot_changed)

        self.ckbox_ot.stateChanged.connect(self.ckbox_ot_click)
        self.ckbox_al.stateChanged.connect(self.ckbox_al_click)

    # Execute Calculation
    def calc_change(self):

        lwkhr = len(self.entry_in_workinghrs.text())
        lwkday = len(self.entry_in_working_days.text())
        lmw = len(self.entry_in_wage.text())

        if abs(lwkhr) > 0 and abs(lwkday) > 0 and abs(lmw) > 0:
            self.mw_changed()
        else:
            QMessageBox.about(self, "입력 확인", f"3개의 조건값 \n일일근무시간: {lwkhr} \n주간근무일수: {lwkday} \n시급: {lmw} \n중 하나 이상이 입력되지 않았습니다!")

    # 주 40시간 기준급여 계산
    def base_cal(self):
        mw = self.entry_in_wage.text()
        self.clear_data
        self.entry_in_wage.setText(mw)

        # 최저임금
        mw = float(mw)
        
        # 주간 근로 시간 계산
        dailyworkinghours = float(self.entry_in_workinghrs.text())
        workingdaysperweek = float(self.entry_in_working_days.text())
        workinghoursperweek = math.floor(workingdaysperweek * dailyworkinghours/1)*1
        self.entry_in_workinghrsperweek.setText(str(workinghoursperweek))

        # 월간 근로 시간 계산
        avgworkinghourspermonth = math.floor(round(workinghoursperweek * 4.345,0)/1)*1
        self.entry_in_workinghrspermonth.setText(str(avgworkinghourspermonth))

        # 월 기본급여 계산
        salary_a = math.floor(mw * avgworkinghourspermonth/1)*1
        self.entry_in_salary_a.setText(str(salary_a))

    # 주휴 수당 계산
    def sl_cal(self):
        # 최저 시급
        mw = float(self.entry_in_wage.text())
        
        # 주간 근로 시간 계산
        dailyworkinghours = float(self.entry_in_workinghrs.text())
        workingdaysperweek = float(self.entry_in_working_days.text())
        workinghoursperweek = math.floor(workingdaysperweek * dailyworkinghours/1)*1

        # 주휴시간 계산
        slhrperweek = workinghoursperweek / workingdaysperweek
        self.entry_in_slpayweekhrs.setText(str(slhrperweek))

        # 주휴수당 계산
        slhrperweekvalue = mw * slhrperweek
        self.entry_in_slpayweekvalue.setText(str(slhrperweekvalue))

        # 월평균 주휴시간 계산
        slhrpermonth = round(slhrperweek * 4.345,0)
        self.entry_in_slpaymonthhrs.setText(str(slhrpermonth))

        # 월평균 주휴수당 계산
        slhrpermonthvalue = math.floor(mw * slhrpermonth/1)*1
        self.entry_in_slpaymonthvalue.setText(str(slhrpermonthvalue))

        # 7. 주간 Paid hour 계산
        paidhrsperweek = workinghoursperweek + slhrperweek
        self.entry_in_weeklpaidhrs.setText(str(paidhrsperweek))

        # 8. 월소정근로 시간 계산
        contractedworkinghrspermonth = math.floor(round(paidhrsperweek * 365/12/7, 0)/1)*1
        self.entry_in_conmonthworkhrs.setText(str(contractedworkinghrspermonth))

        # 9. 월기준급여 계산
        salary_c = math.floor(mw * contractedworkinghrspermonth/1)*1
        self.entry_in_salary_c.setText(str(salary_c))

    # 토요일 근무 수당 계산
    def sat_cal(self):
        mw = float(self.entry_in_wage.text())
        wkhrsperday = float(self.entry_in_workinghrs.text())   
        salary_c = float(self.entry_in_salary_c.text())
        
        # 연간 토요 근무일수
        saturdays = round(365/7, 3)
        self.entry_in_saturdayyearworkdays.setText(str(saturdays))

        # 월간 토요 근무일수
        saturdayspermonth = 4.345
        self.entry_in_saturdaymonthwkdays.setText(str(saturdayspermonth))

        # 월간 토요 근무시간
        saturdayspermonthwkhrs = round(saturdayspermonth * wkhrsperday,0)
        self.entry_in_saturdaymonthwkhrs.setText(str(saturdayspermonthwkhrs))

        # 월간 토요 근무수당
        saturdayspermonthval = math.floor(mw * saturdayspermonthwkhrs * 1.5/1)*1
        self.entry_in_saturdaymonthwkvalue.setText(str(saturdayspermonthval))

        # 월기준급여 + 토요근무 수당
        salary_e = salary_c + saturdayspermonthval
        self.entry_in_salary_e.setText(str(salary_e))        

    # 주중 공휴일 근무 수당 계산
    def wkdyhol_cal(self):
        mw = float(self.entry_in_wage.text())
        wkhrsperday = float(self.entry_in_workinghrs.text())
        saturdayspermonth = 4.345
        saturdayspermonthval = math.floor(mw * round(saturdayspermonth * wkhrsperday,0) * 1.5/1)*1
        salary_c = float(self.entry_in_salary_c.text())
        
        # 일일 근무시간
        wkhrsperday = float(self.entry_in_workinghrs.text())

        # 연간 평일 특근 일수
        holidaysinweekdaysinyear = 15
        self.entry_in_weekdayholidaysperyear.setText(str(holidaysinweekdaysinyear))

        # 월간 평일특근일수
        holidaysinweekdayspermonth = holidaysinweekdaysinyear / 12
        self.entry_in_weekdayholidayspermonth.setText(str(holidaysinweekdayspermonth))

        # 월간 평일특근시간
        wkhrholidaysinweekdayspermonth = holidaysinweekdayspermonth * wkhrsperday
        self.entry_in_weekdayholidayspermonthwkhrs.setText(str(wkhrholidaysinweekdayspermonth))

        # 월간 평일특근수당
        holidaysinweekdayspermonthvalue = math.floor(mw * wkhrholidaysinweekdayspermonth * 1.5/1)*1
        self.entry_in_weekdayholidayspermonthvalue.setText(str(holidaysinweekdayspermonthvalue))

        # 월기준급여 + 토요근무 수당 + 평일 공휴일 계속 근로
        salary_e = salary_c + saturdayspermonthval
        salary_g = math.floor(salary_e + holidaysinweekdayspermonthvalue/1)*1
        self.entry_in_salary_g.setText(str(salary_g))        

    # 주중 고정 OT시간에 대한 수당 계산
    def wkot_cal(self):
        mw = float(self.entry_in_wage.text())

        # 평일 + 토요일 OT 2시간 고정시
        raw_input = self.entry_in_weekdayothrsperday.text()
        fixedOT = int(raw_input) if raw_input else 0

        raw_input1 = self.entry_in_weekdayotdaysperweek.text()
        daysforfixedOT = int(raw_input1) if raw_input1 else 0
        
        self.entry_in_weekdayothrsperday.setText(str(fixedOT))
        self.entry_in_weekdayotdaysperweek.setText(str(daysforfixedOT))

        # 평일 + 토요일 OT 2시간 고정시 주간 OT 시간
        fixedOThrsperweek = fixedOT * daysforfixedOT
        self.entry_in_weekdayothrsperweek.setText(str(fixedOThrsperweek))

        # 평일 + 토요일 OT 2시간 고정시 월평균 OT 시간
        fixedOThrspermonth = fixedOThrsperweek * 4.345
        self.entry_in_weekdayothrspermonth.setText(str(fixedOThrspermonth))

        # 평일 + 토요일 OT 2시간 고정시 월평균 OT 수당
        fixedOTvaluepermonth = math.floor(round(mw * fixedOThrspermonth * 1.5,-1)/1)*1
        self.entry_in_weekdayotvaluepermonth.setText(str(fixedOTvaluepermonth))

        # 월기준급여 + 토요근무 수당 + 평일 공휴일 계속 근로 + 추가 OT 6일간 2시간
        salary_g = float(self.entry_in_salary_g.text())
        salary_i = math.floor((salary_g + fixedOTvaluepermonth)/1)*1
        self.entry_in_salary_i.setText(str(salary_i))

    # 연차 수당 포함 전체 지급액 계산
    def al_cal(self):
        mw = float(self.entry_in_wage.text())        
        dailyworkinghours = float(self.entry_in_workinghrs.text())
        workingdaysperweek = float(self.entry_in_working_days.text())
        workinghoursperweek = math.floor(workingdaysperweek * dailyworkinghours/1)*1

        # 연차일수
        annualleave = 15
        self.entry_in_annualleavedays.setText(str(annualleave))

        # 연차수당
        annualleavevalue = math.floor(mw * annualleave * workinghoursperweek / workingdaysperweek /1)*1
        self.entry_in_annualleavevalue.setText(str(annualleavevalue))

        # 월평균연차수당
        monthlyalval = math.floor(annualleavevalue / 12 / 1 ) * 1
        self.entry_in_annualleavevaluepermonth.setText(str(monthlyalval))

        # 월기준급여 + 토요근무 수당 + 평일 공휴일 계속 근로 + 추가 OT 6일간 2시간 + 연차 월별 균등 지급 조건 
        salary_i = float(self.entry_in_salary_i.text())
        salary_k = math.floor((salary_i + monthlyalval)/1)*1
        self.entry_in_salary_k.setText(str(salary_k))        

    # Weekday OT hour changed
    def wkdyot_changed(self):
        mw = float(self.entry_in_wage.text())

        # 평일 + 토요일 OT 2시간 고정시
        fixedOT = int(self.entry_in_weekdayothrsperday.text())
        daysforfixedOT = int(self.entry_in_weekdayotdaysperweek.text())
        
        # 일일 추가 OT 시간 X 근무일수 = 주간 추가 OT 시간
        fixedOThrsperweek = fixedOT * daysforfixedOT
        self.entry_in_weekdayothrsperweek.setText(str(fixedOThrsperweek))

        # 월평균 추가 OT 시간
        fixedOThrspermonth = fixedOThrsperweek * 4.345
        fixedOThrspermonth = round(fixedOThrspermonth,3)
        self.entry_in_weekdayothrspermonth.setText(str(fixedOThrspermonth))

        # 월평균 추가 OT 수당
        fixedOTvaluepermonth = math.floor(round(mw * fixedOThrspermonth * 1.5,-1)/1)*1
        self.entry_in_weekdayotvaluepermonth.setText(str(fixedOTvaluepermonth))

        # 월기준급여 + 토요근무 수당 + 평일 공휴일 계속 근로 + 추가 OT 6일간 2시간
        salary_g = float(self.entry_in_salary_g.text())
        salary_i = math.floor((salary_g + fixedOTvaluepermonth)/1)*1
        self.entry_in_salary_i.setText(str(salary_i))

        # Salary_K
        sal_k = salary_i + float(self.entry_in_annualleavevaluepermonth.text())
        self.entry_in_salary_k.setText(str(sal_k))

    # Saturday working hour changed
    def sat_changed(self):
        mw = float(self.entry_in_wage.text())
        satwkhrsperday = float(self.entry_in_saturdayworkhrs.text())   
        salary_c = float(self.entry_in_salary_c.text())
        
        # 연간 토요 근무일수
        saturdays = round(365/7, 3)
        self.entry_in_saturdayyearworkdays.setText(str(saturdays))

        # 월간 토요 근무일수
        saturdayspermonth = 4.345
        self.entry_in_saturdaymonthwkdays.setText(str(saturdayspermonth))

        # 월간 토요 근무시간
        if satwkhrsperday == 4:
            saturdayspermonthwkhrs = 17.5
        elif satwkhrsperday == 10:
            saturdayspermonthwkhrs = 43.5
        else:
            saturdayspermonthwkhrs = round(saturdayspermonth * satwkhrsperday,0)
        
        self.entry_in_saturdaymonthwkhrs.setText(str(saturdayspermonthwkhrs))

        # 월간 토요 근무수당
        saturdayspermonthval = math.floor(mw * saturdayspermonthwkhrs * 1.5/1)*1
        self.entry_in_saturdaymonthwkvalue.setText(str(saturdayspermonthval))

        # 월기준급여 + 토요근무 수당
        salary_e = salary_c + saturdayspermonthval
        self.entry_in_salary_e.setText(str(salary_e))    
        
        # Salary_G
        satotval = salary_e + float(self.entry_in_weekdayholidayspermonthvalue.text())
        self.entry_in_salary_g.setText(str(satotval))

        # Salary_I
        sal_i = satotval + float(self.entry_in_weekdayotvaluepermonth.text())
        self.entry_in_salary_i.setText(str(sal_i))
        
        # Salary_K
        sal_k = sal_i + float(self.entry_in_annualleavevaluepermonth.text())
        self.entry_in_salary_k.setText(str(sal_k))

    # Working Days changed
    def mw_changed(self):

        self.base_cal()
        self.sl_cal()
        self.sat_cal()
        self.wkdyhol_cal()
        self.wkot_cal()
        self.al_cal()



    def ckbox_ot_click(self):
        checkbox_state = self.ckbox_ot.isChecked()
        
        if checkbox_state:
        #20. 평일 + 토요일 OT 2시간 고정시
            mw = float(self.entry_in_wage.text())
            fixedOT = 2
            daysforfixedOT = 6
            self.entry_in_weekdayothrsperday.setText(str(fixedOT))
            self.entry_in_weekdayotdaysperweek.setText(str(daysforfixedOT))
            salary_g = float(self.entry_in_salary_g.text())

            #21. 평일 + 토요일 OT 2시간 고정시 주간 OT 시간
            fixedOThrsperweek = fixedOT * daysforfixedOT
            self.entry_in_weekdayothrsperweek.setText(str(fixedOThrsperweek))

            #22. 평일 + 토요일 OT 2시간 고정시 월평균 OT 시간
            fixedOThrspermonth = fixedOThrsperweek * 4.345
            self.entry_in_weekdayothrspermonth.setText(str(fixedOThrspermonth))

            #23. 평일 + 토요일 OT 2시간 고정시 월평균 OT 수당
            fixedOTvaluepermonth = math.floor(round(mw * fixedOThrspermonth * 1.5,-1)/1)*1
            self.entry_in_weekdayotvaluepermonth.setText(str(fixedOTvaluepermonth))

            #24. 월기준급여 + 토요근무 수당 + 평일 공휴일 계속 근로 + 추가 OT 6일간 2시간
            salary_i = math.floor((salary_g + fixedOTvaluepermonth)/1)*1
            self.entry_in_salary_i.setText(str(salary_i))
            self.ckbox_al_click()
        
        else:
            self.entry_in_weekdayothrsperday.setText("0")
            self.entry_in_weekdayotdaysperweek.setText("0")
            self.entry_in_weekdayothrsperweek.setText("0")
            self.entry_in_weekdayothrspermonth.setText("0")
            self.entry_in_weekdayotvaluepermonth.setText("0")
            self.entry_in_salary_i.setText("0")
            self.ckbox_al_click()

    def ckbox_al_click(self):
        checkbox_state = self.ckbox_al.isChecked()
        
        if checkbox_state:
            mw = 0 if self.entry_in_wage.text() == '' else float(self.entry_in_wage.text())
            workinghoursperweek = 0 if self.entry_in_workinghrsperweek.text() == '' else float(self.entry_in_workinghrsperweek.text())
            workingdaysperweek = 0 if self.entry_in_working_days.text() == '' else float(self.entry_in_working_days.text())
            salary_g = 0 if self.entry_in_salary_g.text() == '' else float(self.entry_in_salary_g.text())
            #salary_i = float(self.entry_in_salary_i.text())
            salary_i = 0 if self.entry_in_salary_i.text() == '' else float(self.entry_in_salary_i.text())

            #25. 연차일수
            annualleave = 15
            self.entry_in_annualleavedays.setText(str(annualleave))

            #26. 연차수당
            annualleavevalue = math.floor(mw * annualleave * workinghoursperweek / workingdaysperweek /1)*1
            self.entry_in_annualleavevalue.setText(str(annualleavevalue))

            #27. 월평균연차수당
            monthlyalval = math.floor(annualleavevalue / 12 / 1 ) * 1
            self.entry_in_annualleavevaluepermonth.setText(str(monthlyalval))

            #28. 월기준급여 + 토요근무 수당 + 평일 공휴일 계속 근로 + 추가 OT 6일간 2시간 + 연차 월별 균등 지급 조건 
            if salary_i > 0:
                salary_k = math.floor((salary_i + monthlyalval)/1)*1
            else:
                salary_k = math.floor((salary_g + monthlyalval)/1)*1
            self.entry_in_salary_k.setText(str(salary_k))
        else:
            self.entry_in_annualleavedays.setText("0")
            self.entry_in_annualleavevalue.setText("0")
            self.entry_in_annualleavevaluepermonth.setText("0")
            self.entry_in_salary_k.setText("0")


    # tab order for calmaster window
    def set_tab_order(self):
        widgets = [self.cb_in_year, self.entry_in_workinghrs, self.entry_in_working_days, 
            self.entry_in_wage, self.entry_in_workinghrsperweek , self.entry_in_workinghrspermonth ,
            self.entry_in_salary_a , self.entry_in_slpayweekhrs , self.entry_in_slpayweekvalue ,
            self.entry_in_slpaymonthhrs , self.entry_in_slpaymonthvalue , self.entry_in_weeklpaidhrs ,
            self.entry_in_conmonthworkhrs , self.entry_in_salary_c , self.entry_in_saturdayyearworkdays ,
            self.entry_in_saturdaymonthwkdays , self.entry_in_saturdaymonthwkhrs , self.entry_in_saturdaymonthwkvalue ,
            self.entry_in_salary_e , self.entry_in_weekdayholidaysperyear , self.entry_in_weekdayholidayspermonth ,
            self.entry_in_weekdayholidayspermonthwkhrs , self.entry_in_weekdayholidayspermonthvalue , self.entry_in_salary_g ,
            self.entry_in_weekdayothrsperday , self.entry_in_weekdayotdaysperweek , self.entry_in_weekdayothrsperweek ,
            self.entry_in_weekdayothrspermonth , self.entry_in_weekdayotvaluepermonth , self.entry_in_salary_i ,
            self.entry_in_annualleavedays , self.entry_in_annualleavevalue , self.entry_in_annualleavevaluepermonth ,
            self.entry_in_salary_k , self.pb_in_calc , self.pb_in_insert , self.pb_in_delete ,
            ]
        
        for i in range(len(widgets) - 1):
            self.setTabOrder(widgets[i], widgets[i + 1])


    def query_statement(self):
        tv_widget = self.tv_otdays
        
        self.cursor.execute("SELECT id, fy, mw FROM basicdata WHERE 1=0")
        column_info = self.cursor.description
        column_names = [col[0] for col in column_info]

        sql_query = "Select id, fy, mw from basicdata order By id"
        column_widths = [80, 80, 80]

        return sql_query, tv_widget, column_info, column_names, column_widths

    # To reduce duplications
    def make_data(self):
        sql_query, tv_widget, column_info, column_names, column_widths = self.query_statement() 
        self.populate_dialog(self.cursor, sql_query, tv_widget, column_info, column_names,column_widths)

    def query_statement1(self):
        tv_widget = self.tv_salarylist
        
        self.cursor.execute("SELECT * FROM calc_salary_inh WHERE 1=0")
        column_info = self.cursor.description
        column_names = [col[0] for col in column_info]

        sql_query = "Select * from calc_salary_inh order By id"
        column_widths = [80, 80, 80, 80, 80, 80, 80, 80, 80, 80, 80, 80, 80, 80, 80, 80, 80, 80, 80, 80]

        return sql_query, tv_widget, column_info, column_names, column_widths

    # To reduce duplications
    def make_data1(self):
        sql_query, tv_widget, column_info, column_names, column_widths = self.query_statement1() 
        self.populate_dialog(self.cursor, sql_query, tv_widget, column_info, column_names,column_widths)

    # Get the value of other variables
    def get_calsalary_input(self):
        dwkhr = float(self.entry_in_workinghrs.text())
        wkwkday = float(self.entry_in_working_days.text())
        wage = float(self.entry_in_wage.text())
        wkwkhr = float(self.entry_in_workinghrsperweek.text())
        mthwkhr = float(self.entry_in_workinghrspermonth.text())
        mthbssal1 = float(self.entry_in_salary_a.text())
        wkslhr = float(self.entry_in_slpayweekhrs.text())
        mthslhr = float(self.entry_in_slpaymonthhrs.text())
        mthslval = float(self.entry_in_slpaymonthvalue.text())
        mthconwkhr = float(self.entry_in_conmonthworkhrs.text())
        mthbssal2 = float(self.entry_in_salary_c.text())
        mthsatval = float(self.entry_in_saturdaymonthwkvalue.text())
        mthbssal3  = float(self.entry_in_salary_e.text())
        mthwkhdval = float(self.entry_in_weekdayholidayspermonthvalue.text())
        mthbssal4 = float(self.entry_in_salary_g.text())
        mthwkdotval = float(self.entry_in_weekdayotvaluepermonth.text())
        mthbssal5 = float(self.entry_in_salary_i.text())
        mthalval = float(self.entry_in_annualleavevaluepermonth.text())
        mthbssal6 = float(self.entry_in_salary_k.text())

        return dwkhr, wkwkday, wage, wkwkhr, mthwkhr, mthbssal1, wkslhr, mthslhr, mthslval, mthconwkhr, mthbssal2, mthsatval, mthbssal3, mthwkhdval, mthbssal4, mthwkdotval, mthbssal5, mthalval, mthbssal6
    
    # Make Common values set
    def common_values_set(self):
        username = self.current_username
        user_id = self.userID_gen(username)
        formatted_datetime = self.dt_time_info()
        return username, user_id, formatted_datetime

    # insert new calc_salary data to MySQL table
    def tv_insert(self):

        confirm_dialog = self.show_insert_confirmation_dialog()
        
        if confirm_dialog == QMessageBox.Yes:
            
            idx = self.max_row_id("calc_salary_inh")                                  # Get the max id 
            username, user_id, formatted_datetime = self.common_values_set()
            dwkhr, wkwkday, wage, wkwkhr, mthwkhr, mthbssal1, wkslhr, mthslhr, mthslval, mthconwkhr, mthbssal2, mthsatval, mthbssal3, mthwkhdval, mthbssal4, mthwkdotval, mthbssal5, mthalval, mthbssal6 = self.get_calsalary_input()         # Get the value of other variables

            if (idx>0):

                self.cursor.execute('''INSERT INTO calc_salary_inh 
                            (id, dwkhr, wkwkday, wage, wkwkhr, mthwkhr, mthbssal1, wkslhr, mthslhr, mthslval, mthconwkhr, mthbssal2, mthsatval, mthbssal3, mthwkhdval, mthbssal4, mthwkdotval, mthbssal5, mthalval, mthbssal6) 
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)'''
                            , (idx, dwkhr, wkwkday, wage, wkwkhr, mthwkhr, mthbssal1, wkslhr, mthslhr, mthslval, mthconwkhr, mthbssal2, mthsatval, mthbssal3, mthwkhdval, mthbssal4, mthwkdotval, mthbssal5, mthalval, mthbssal6))
                self.conn.commit()
                self.show_insert_success_message()
                self.refresh_data()
                self.make_data1()
                logging.info(f"User {username} inserted {idx} row to the calc_salary_inh table.")
            else:
                self.show_missing_message("입력 이상")
                return
        else:
            self.show_cancel_message("데이터 추가 취소")
            return

    def tv_delete(self):
        confirm_dialog = self.show_delete_confirmation_dialog()

        if confirm_dialog == QMessageBox.Yes:

            idx = self.lbl_in_id.text()
            username, user_id, formatted_datetime = self.common_values_set()

            self.cursor.execute("DELETE FROM calc_salary_inh WHERE id=?", (idx,))
            self.conn.commit()
            self.show_delete_success_message()
            self.refresh_data()  
            self.make_data1()            
            logging.info(f"User {username} deleted {idx} row to the calc_salary_inh table.")                
            
        else:
            self.show_cancel_message("데이터 삭제 취소")

    # clear input field entry
    def clear_data(self):
        for line_edit in self.findChildren(QtWidgets.QLineEdit):
            line_edit.clear()

        self.lbl_in_id.setText("")
        self.cb_in_year.setCurrentIndex(0)
        self.entry_in_workinghrs.setText("8")
        self.entry_in_working_days.setText("5")
        self.entry_in_saturdayworkhrs.setText("8")
        self.entry_in_weekdayholidaywkhrs.setText("8")
        self.ckbox_ot.setChecked(True)
        self.ckbox_al.setChecked(True)

    def cb_in_year_changed(self):
        self.entry_in_wage.clear()

        selected_item = self.cb_in_year.currentText()

        if selected_item:
            query = f"SELECT DISTINCT mw From basicdata WHERE fy ='{selected_item}'"
            line_edit_widgets = [self.entry_in_wage, ]
        
            # Check if any line edit widgets are provided
            if line_edit_widgets:
                self.lineEdit_contents(line_edit_widgets, query)
            else:
                pass

    # table view click
    def show_selected_otdays_data(self, item):
        # Get the row index of the clicked item
        row_index = item.row()

        # Initialize a list to store the cell values
        cell_values = []

        # Loop through the columns and retrieve the text from each cell
        for column_index in range(3): 
            cell_text = self.tv_otdays.model().item(row_index, column_index).text()
            cell_values.append(cell_text)
            
        # Populate the input widgets with the data from the selected row
        self.cb_in_year.setCurrentText(cell_values[1])
        self.entry_in_wage.setText(cell_values[2])
    
    # table view 2 click
    def show_selected_salarylist_data(self, item):
        # Get the row index of the clicked item
        row_index = item.row()

        # Initialize a list to store the cell values
        cell_values = []

        # Loop through the columns and retrieve the text from each cell
        for column_index in range(20): 
            cell_text = self.tv_salarylist.model().item(row_index, column_index).text()
            cell_values.append(cell_text)

        # Populate the input widgets with the data from the selected row
        self.lbl_in_id.setText(cell_values[0])
        #self.cb_in_year.setCurrentText(cell_values[1])
        self.entry_in_workinghrs.setText(cell_values[1])
        self.entry_in_working_days.setText(cell_values[2])
        self.entry_in_wage.setText(cell_values[3])
        self.entry_in_workinghrsperweek.setText(cell_values[4])
        self.entry_in_workinghrspermonth.setText(cell_values[5])
        self.entry_in_salary_a.setText(cell_values[6])
        self.entry_in_slpayweekhrs.setText(cell_values[7])
        self.entry_in_slpaymonthhrs.setText(cell_values[8])
        self.entry_in_slpaymonthvalue.setText(cell_values[9])
        self.entry_in_conmonthworkhrs.setText(cell_values[10])
        self.entry_in_salary_c.setText(cell_values[11])
        self.entry_in_saturdaymonthwkvalue.setText(cell_values[12])
        self.entry_in_salary_e.setText(cell_values[13])
        self.entry_in_weekdayholidayspermonthvalue.setText(cell_values[14])
        self.entry_in_salary_g.setText(cell_values[15])
        self.entry_in_weekdayotvaluepermonth.setText(cell_values[16])
        self.entry_in_salary_i.setText(cell_values[17])
        self.entry_in_annualleavevaluepermonth.setText(cell_values[18])
        self.entry_in_salary_k.setText(cell_values[19])                    


    # 선택된 각 위젯의 내용을 엑셀로 내보내기
    def export_data(self):
        output_subfolder = "data_list"          # set the output subfoler name
        sheet_name = "Calc_Salary_Inh"                   # set the excel sheet name
        filename = get_file_name(output_subfolder, sheet_name)    # get the file name

        # Ensure the subfolder exists; create it if it doesn't
        os.makedirs(output_subfolder, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        if filename:
            full_file_path = os.path.join(output_subfolder, filename)
    
            # 각 위젯에서 내용 가져와서 엑셀에 쓰기
            data_to_export = [
                (self.label.text(), self.cb_in_year.currentText()),
                #(self..text(), self..text()),
                #(self..text(), self..text()),
                #(self..text(), self..text()),
                #(self..text(), self..text()),
                #(self..text(), self..text()),
                #(self..text(), self..text()),
                #(self..text(), self..text()),
                #(self..text(), self..text()),
                #(self..text(), self..text()),
                #(self..text(), self..text()),
                #(self..text(), self..text()),
                #(self..text(), self..text()),
                #(self..text() , self..text()),
                #(self..text() , self..text()),
                #(self..text() , self..text() ),
                #(self..text() , self..text() ),
                #(self..text() , self..text() ),
                #(self..text() , self..text() ),
                #(self..text() , self..text() ),
                ]

            #for i, value in enumerate(data_to_export, start=1):
            #    ws.cell(row=i, column=1, value=value)
            
            for i, (label, value) in enumerate(data_to_export, start=1):
                ws.cell(row=i, column=1, value=label)
                ws.cell(row=i, column=2, value=value)
            
            # 엑셀 파일 저장
            wb.save(full_file_path)
        
            self.excel_formatting(sheet_name, full_file_path)

    # Excel sheet formatting
    def excel_formatting(self, sheet_name, full_file_path):
        wb = load_workbook(full_file_path)
        sheet_name = sheet_name
        ws = wb[sheet_name]
        last_row = ws.max_row + 2

        # Insert headers at A1 and B1
        ws.insert_rows(1)
        ws['A1'] = '구분'
        ws['B1'] = '내용'

        column_widths = [20, 15]                                # set the width of each column
        title_font = Font(bold=True, name="Arial", size=10)     # set the font for the first row
        arial_font = Font(name="Arial", size=10)                # set the forn from the second row to max row

        set_column_widths(ws, column_widths)    # reset column widths

        ws.freeze_panes = 'B2'              # freeze panes D2 cell
        ws.auto_filter.ref = ws.dimensions  # apply auto filter
        ws.sheet_view.showGridLines = False # remove gridlines
            
        set_font(ws, [1], range(1, len(column_widths) + 1), title_font)                 # first row font setting
        set_font(ws, range(2, last_row), range(1, len(column_widths) + 1), arial_font)  # from second row to last row font setting

        wb.save(full_file_path)
        QMessageBox.about(self, "파일 생성 완료", f"data_list folder에 \n엑셀 파일이 {full_file_path}로 \n생성 되었습니다!")  

    # Refresh data
    def refresh_data(self):
        self.clear_data()
        self.make_data()       

if __name__ == "__main__":
    
    log_subfolder = "logs"
    os.makedirs(log_subfolder, exist_ok=True)
    log_file_path = os.path.join(log_subfolder, "access_CalcSalaryInh.log")

    logging.basicConfig(
        filename=log_file_path,  
        level=logging.INFO,    
        format="%(asctime)s [%(levelname)s] - %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
    app = QtWidgets.QApplication(sys.argv)
    dialog = CalSalaryInhDialog()
    dialog.show()
    sys.exit(app.exec())
