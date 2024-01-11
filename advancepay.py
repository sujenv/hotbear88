import sys
import logging
from functools import partial
from openpyxl.styles import Font
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from PyQt5 import uic, QtWidgets, QtCore
from PyQt5.QtGui import QStandardItemModel, QKeySequence
from PyQt5.QtWidgets import QDialog, QMessageBox, QMenu, QShortcut
from PyQt5.QtCore import Qt
from datetime import datetime
from commonmd import *
from cal import CalendarView
#for non_ui version-------------------------
#from advancepay_ui import Ui_AdvancePayDialog

# advancepay table contents -----------------------------------------------------
class AdvancePayDialog(QDialog, SubWindowBase):
#for non_ui version-------------------------
#class AdvancePayDialog(QDialog, Ui_AdvancePayDialog, SubWindowBase): 
    
    def __init__(self, current_username = None, current_datetime = None):
        super().__init__()

        self.conn, self.cursor = connect_to_database1()

        uic.loadUi("advancepay.ui", self)
        #for non_ui version-------------------------
        #self.setupUi(self)

        # Add window flags
        self.setWindowFlags(self.windowFlags() | Qt.WindowMinimizeButtonHint | Qt.WindowMaximizeButtonHint | Qt.WindowCloseButtonHint)  

        # Initialize current_username and current_datetime directly
        self.current_username, self.current_datetime = initialize_username_and_datetime(current_username, current_datetime)

        # Enable automatic deletion on close
        self.setAttribute(QtCore.Qt.WA_DeleteOnClose)  

        # Create tv_advancepay and QSortFilterProxyModel
        self.model = QStandardItemModel()
        self.proxy_model = NumericStringSortModel(self.model)
        self.proxy_model.setSourceModel(self.model)

        # Define the column types
        column_types = ["numeric", "numeric", "", "", "", "numeric", "", "", ""]
        
        # Set the custom delegate for the specific column
        delegate = NumericDelegate(column_types, self.tv_advancepay)
        self.tv_advancepay.setItemDelegate(delegate)
        self.tv_advancepay.setModel(self.proxy_model)
       
        # Enable sorting
        self.tv_advancepay.setSortingEnabled(True)

        # Enable alternating row colors
        self.tv_advancepay.setAlternatingRowColors(True)  
        
        # Hide the first index column
        self.tv_advancepay.verticalHeader().setVisible(False)

        # While selecting row in tv_advancepay, each cell values to displayed to designated widgets
        self.tv_advancepay.clicked.connect(self.show_selected_data)

        # Fill combobox items when the application starts
        self.get_combobox_contents()

        # Initial Display of data
        self.make_data()
        self.connect_btn_method()
        self.conn_signal_to_slot()

        # Set tab order for input widgets
        self.set_tab_order()

        # Initiate CTRL+C, CTRL+V and ENTER
        self.setup_shortcuts()
        
        # Create context menus ----------------------------------------------------------------------------------
        self.context_menu1 = self.create_context_menu(self.entry_advancepay_transactiondate)
        self.entry_advancepay_transactiondate.setContextMenuPolicy(Qt.CustomContextMenu)
        self.entry_advancepay_transactiondate.customContextMenuRequested.connect(self.show_context_menu1)

        ddt = self.display_trx_date()
        self.entry_advancepay_transactiondate.setText(ddt)
        
        # Make log file
        self.make_logfiles("access_advancepay.log")

    # Create a shortcut for CTRL+C/CTRL+V/ Return key
    def setup_shortcuts(self):
        self.copy_shortcut = QShortcut(QKeySequence.Copy, self.tv_advancepay, partial(self.copy_cells, self.tv_advancepay))
        self.paste_shortcut = QShortcut(QKeySequence.Paste, self.tv_advancepay, partial(self.paste_cells, self.tv_advancepay))
        self.return_shortcut = QShortcut(Qt.Key_Return, self.tv_advancepay, partial(self.handle_return_key, self.tv_advancepay))

    # Mouse Right click, show "달력보기" menu ---------------------------------------------------------------------
    def create_context_menu(self, target_lineedit):
        context_menu = QMenu()
        custom_action = context_menu.addAction("달력보기")
        custom_action.triggered.connect(lambda: self.show_calendar(target_lineedit))
        return context_menu

    def show_context_menu1(self, pos):
        self.context_menu1.exec_(self.entry_advancepay_transactiondate.mapToGlobal(pos))

    # Show Calendar
    def show_calendar(self, target_lineedit):
        calendar_dialog = CalendarView()
        calendar_dialog.selected_date_changed.connect(lambda date: self.set_selected_date(date, target_lineedit))
        calendar_dialog.exec()

    # Show selected date to the select Qlineedit
    def set_selected_date(self, date, target_lineedit):
        if target_lineedit == self.entry_advancepay_transactiondate:
            target_lineedit.setText(date)

    # Call process_key_event and pass the event and your QTableWidget instance
    def keyPressEvent(self, event):
        tv_widget = self.tv_advancepay
        self.process_key_event(event, tv_widget)

    # Pass combobox info and sql to next method
    def get_combobox_contents(self):
        self.insert_combobox_initiate(self.cb_advancepay_ename, "SELECT DISTINCT ename FROM employee ORDER BY ename")
        self.insert_combobox_initiate(self.cb_advancepay_class1, "SELECT DISTINCT class1 FROM employee ORDER BY class1")
        self.insert_combobox_initiate(self.cb_advancepay_active, "SELECT DISTINCT active FROM advancepay ORDER BY active")

    # Initiate Combo_Box 
    def insert_combobox_initiate(self, combo_box, sql_query):    
        self.combobox_initializing(combo_box, sql_query)
        self.lbl_advancepay_id.setText("")
        self.entry_advancepay_ecode.setText("")
        self.cb_advancepay_ename.setCurrentIndex(0) 
        self.cb_advancepay_class1.setCurrentIndex(0)
        self.cb_advancepay_active.setCurrentIndex(0)


    # Connect button to method
    def connect_btn_method(self):
        self.pb_advancepay_show.clicked.connect(self.make_data)
        self.pb_advancepay_cancel.clicked.connect(self.close_dialog)
        self.pb_advancepay_search.clicked.connect(self.search_data)
        self.pb_advancepay_clear.clicked.connect(self.clear_data)
        self.pb_advancepay_insert.clicked.connect(self.tb_insert)
        self.pb_advancepay_update.clicked.connect(self.tb_update)
        self.pb_advancepay_delete.clicked.connect(self.tb_delete)
        self.pb_advancepay_check_amount.clicked.connect(self.calculate_sum)
        self.pb_advancepay_excel_export.clicked.connect(self.export_table)

    # Connect signal to method    
    def conn_signal_to_slot(self):
        self.cb_advancepay_ename.activated.connect(self.advancepay_ename_changed)

    # tab order for advancepay window
    def set_tab_order(self):
        widgets = [self.pb_advancepay_show, self.entry_advancepay_ecode, self.cb_advancepay_ename,
            self.cb_advancepay_class1, self.entry_advancepay_class2, self.entry_advancepay_amount,
            self.entry_advancepay_transactiondate, self.cb_advancepay_active, self.entry_advancepay_remark,
            self.entry_advancepay_ttl, self.pb_advancepay_search, self.pb_advancepay_clear, 
            self.pb_advancepay_insert, self.pb_advancepay_update, self.pb_advancepay_delete, 
            self.pb_advancepay_cancel, self.pb_advancepay_check_amount, self.pb_advancepay_excel_export]
        
        for i in range(len(widgets) - 1):
            self.setTabOrder(widgets[i], widgets[i + 1])

    # To reduce duplications
    def common_query_statement(self):
        tv_widget = self.tv_advancepay
        self.cursor.execute("SELECT id, ecode, ename, class1, class2, amount, transactiondate, active, remark FROM vw_advancepay WHERE 1=0")
        column_info = self.cursor.description
        column_names = [col[0] for col in column_info]

        sql_query = "Select id, ecode, ename, class1, class2, amount, transactiondate, active, remark from vw_advancepay order By id DESC"
        column_widths = [80, 100, 100, 50, 50, 100, 100, 50, 150]

        return sql_query, tv_widget, column_info, column_names, column_widths 

    # show advancepay table data inside of the MDI
    def make_data(self):
        sql_query, tv_widget, column_info, column_names, column_widths = self.common_query_statement() 
        self.populate_dialog(self.cursor, sql_query, tv_widget, column_info, column_names,column_widths)

        ddt = self.display_trx_date()
        self.entry_advancepay_transactiondate.setText(ddt)
   
    # Get the value of other variables
    def get_advancepay_input(self):
        ecode = int(self.entry_advancepay_ecode.text())
        ename = str(self.cb_advancepay_ename.currentText())
        class1 = str(self.cb_advancepay_class1.currentText())
        class2 = str(self.entry_advancepay_class2.text())
                
        input_val = self.entry_advancepay_amount.text()
        if input_val.replace(".", "").replace("-", "").isdigit():
            amount = float(input_val)
        else:
            amount = 0

        #amount = float(self.entry_advancepay_amount.text())
        inputdate = str(self.entry_advancepay_transactiondate.text())
        active = str(self.cb_advancepay_active.currentText())
        remark = str(self.entry_advancepay_remark.text())

        return ecode, ename, class1, class2, amount, inputdate, active, remark

    # Make Common values set
    def common_values_set(self):
        username = self.current_username
        user_id = self.userID_gen(username)
        formatted_datetime = self.dt_time_info()

        return username, user_id, formatted_datetime

    # insert new advancepay data to MySQL table
    def tb_insert(self):
        confirm_dialog = self.show_insert_confirmation_dialog()
        
        if confirm_dialog == QMessageBox.Yes:
            
            idx = self.max_row_id("advancepay")
            username, user_id, formatted_datetime = self.common_values_set()
            ecode, ename, class1, class2, amount, inputdate, active, remark = self.get_advancepay_input() 

            if (idx>0 and ecode>0 and abs(amount)>=0) and all(len(var) > 0 for var in (ename, class1, inputdate, active)):

                self.cursor.execute('''INSERT INTO advancepay (id, ecode, amount, active, transactiondate, trxdate, userid, remark) 
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?)'''
                            , (idx, ecode, amount, active, inputdate, formatted_datetime, user_id, remark))
                self.conn.commit()
                self.show_insert_success_message()
                self.refresh_data() 
                logging.info(f"User {username} inserted {idx} row to the advancepay table.")
            else:
                self.show_missing_message("입력 이상")
                return
        else:
            self.show_cancel_message("데이터 추가 취소")
            return

    # revise the values in the selected row
    def tb_update(self):
        confirm_dialog = self.show_update_confirmation_dialog()

        if confirm_dialog == QMessageBox.Yes:
            
            idx = int(self.lbl_advancepay_id.text())
            username, user_id, formatted_datetime = self.common_values_set() 
            ecode, ename, class1, class2, amount, inputdate, active, remark = self.get_advancepay_input() 

            if (idx>0 and ecode>0 and abs(amount)>=0) and all(len(var) > 0 for var in (ename, class1, inputdate, active)):

                self.cursor.execute('''UPDATE advancepay SET 
                            ecode=?, amount=?, active=?, transactiondate=?, trxdate=?, userid=?, remark=? WHERE id=?'''
                            , (ecode, amount, active, inputdate, formatted_datetime, user_id, remark, idx))
                self.conn.commit()
                self.show_update_success_message()
                self.refresh_data()  
                logging.info(f"User {username} updated row number {idx} in the advancepay table.")
            else:
                self.show_missing_message("입력 이상")
                return
        else:
            self.show_cancel_message("데이터 변경 취소")
            return

    # delete row according to id selected
    def tb_delete(self):
        confirm_dialog = self.show_delete_confirmation_dialog()

        if confirm_dialog == QMessageBox.Yes:
            idx = self.lbl_advancepay_id.text()
            username, user_id, formatted_datetime = self.common_values_set()
            self.cursor.execute("DELETE FROM advancepay WHERE id=?", (idx,))
            self.conn.commit()
            self.show_delete_success_message()
            self.refresh_data()  
            logging.info(f"User {username} deleted {idx} row to the advancepay table.")                
            
        else:
            self.show_cancel_message("데이터 삭제 취소")

    # Search data
    def search_data(self):
        ename = self.cb_advancepay_ename.currentText()
        class1 = self.cb_advancepay_class1.currentText()
        active = self.cb_advancepay_active.currentText()
        
        conditions = {'v01': (ename, "ename like '%{}%'"),
                    'v02': (class1, "class1 like '%{}%'"),
                    'v03': (active, "active like '%{}%'"),
                    }
        
        selected_conditions = []

        for key, (value, condition_format) in conditions.items():
            if len(value) > 0:
                selected_conditions.append(condition_format.format(value))

        if not selected_conditions:
            QMessageBox.about(self, "검색 조건 확인", "검색 조건이 비어 있습니다!")
            return

        # Join the selected conditions to form the SQL query
        query = f"SELECT id, ecode, ename, class1, class2, amount, transactiondate, active, remark FROM vw_advancepay WHERE {' AND '.join(selected_conditions)} ORDER BY id desc"

        QMessageBox.about(self, "검색 조건 확인", f"이름: {ename} \n구분: {class1} \n현행: {active} \n\n위 조건으로 검색을 수행합니다!")
        
        sql_query, tv_widget, column_info, column_names, column_widths = self.common_query_statement() 
        self.populate_dialog(self.cursor, query, tv_widget, column_info, column_names,column_widths)

    # 콤보박스 내용 변경에 따라 lineedit 값 변경 + 옆 콤보박스 내용 함께 변경하기
    def advancepay_ename_changed(self):
        self.entry_advancepay_ecode.clear()
        selected_item = self.cb_advancepay_ename.currentText()

        if selected_item:
            query = f"SELECT DISTINCT ecode, class2 From employee WHERE ename = '{selected_item}'"
            line_edit_widgets = [self.entry_advancepay_ecode, self.entry_advancepay_class2]
        
            # Check if any line edit widgets are provided
            if line_edit_widgets:
                self.lineEdit_contents(line_edit_widgets, query)

                sql_query = f"SELECT DISTINCT class1 FROM employee WHERE ename = '{selected_item}'"
                combo_box = self.cb_advancepay_class1
                self.cursor.execute(sql_query)
                
                combo_box.clear() # Clear existing items
                items = self.cursor.fetchall()
                combo_box.addItems([item[0] for item in items])

            else:
                pass

    # Calculate the sum of values in a specific column
    def calculate_sum(self):
        self.entry_advancepay_ttl.clear()

        COLUMN = 5
        r = 0
        
        # Assuming you have a QStandardItemModel set for your QTableView
        model = self.tv_advancepay.model()

        if model is not None:
            for i in range(model.rowCount()):
                item = model.item(i, COLUMN)
                if item and item.data(Qt.DisplayRole):
                    value = float(item.data(Qt.DisplayRole))
                    r += value
       
        self.entry_advancepay_ttl.setText(str(r))

    # Export data to Excel sheet                
    def export_table(self):
        output_subfolder = "data_list"          # set the output subfoler name
        table_widget = self.tv_advancepay       # set the name of table widget
        sheet_name = "advancepay"               # set the excel sheet name
        filename = get_file_name(output_subfolder, sheet_name)    # get the file name / export_to_excel 밑에 두면 get_file_name을 두 번 실행시켜서 파일의 번호가 달라지므로 미리 만들어서 filename에 담아둠.
        numeric_columns=[0, 1, 5]                     # set the numerical column index
        export_to_excel(output_subfolder, table_widget, sheet_name, numeric_columns)
               
        if filename:
            full_file_path = os.path.join(output_subfolder, filename)
            self.excel_formatting(sheet_name, full_file_path)
        else:
            QMessageBox.about(self, "코드 확인", "파일을 찾지 못했습니다!")    

    # Excel sheet formatting
    def excel_formatting(self, sheet_name, full_file_path):
        wb = load_workbook(full_file_path)
        sheet_name = sheet_name
        ws = wb[sheet_name]
        last_row = ws.max_row + 1

        column_widths = [6, 10, 10, 8, 8, 15, 15, 8, 25]        # set the width of each column
        title_font = Font(bold=True, name="Arial", size=10)     # set the font for the first row
        arial_font = Font(name="Arial", size=10)                # set the forn from the second row to max row

        set_column_widths(ws, column_widths)    # reset column widths

        ws.freeze_panes = 'D2'              # freeze panes D2 cell
        ws.auto_filter.ref = ws.dimensions  # apply auto filter
        ws.sheet_view.showGridLines = False # remove gridlines
            
        set_font(ws, [1], range(1, len(column_widths) + 1), title_font)                 # first row font setting
        set_font(ws, range(2, last_row), range(1, len(column_widths) + 1), arial_font)  # from second row to last row font setting
        
        wb.save(full_file_path)
        QMessageBox.about(self, "파일 생성 완료", f"data_list folder에 \n엑셀 파일이 {full_file_path}로 \n생성 되었습니다!")  


    # clear input field entry
    def clear_data(self):
        self.lbl_advancepay_id.setText("")
        for line_edit in self.findChildren(QtWidgets.QLineEdit):
            line_edit.clear()

    # table widget cell double click
    def show_selected_data(self, item):
        # Get the row index of the clicked item
        row_index = item.row()

        # Initialize a list to store the cell values
        cell_values = []

        # Loop through the columns and retrieve the text from each cell
        for column_index in range(9):  # 9 columns
            cell_text = self.tv_advancepay.model().item(row_index, column_index).text()
            cell_values.append(cell_text)

        # Populate the input widgets with the data from the selected row
        self.lbl_advancepay_id.setText(cell_values[0])
        self.entry_advancepay_ecode.setText(cell_values[1])
        self.cb_advancepay_ename.setCurrentText(cell_values[2])
        self.cb_advancepay_class1.setCurrentText(cell_values[3])
        self.entry_advancepay_class2.setText(cell_values[4])
        self.entry_advancepay_amount.setText(cell_values[5])
        self.entry_advancepay_transactiondate.setText(cell_values[6])
        self.cb_advancepay_active.setCurrentText(cell_values[7])
        self.entry_advancepay_remark.setText(cell_values[8])
    
    def refresh_data(self):
        self.clear_data()
        self.make_data()

if __name__ == "__main__":

    log_subfolder = "logs"
    os.makedirs(log_subfolder, exist_ok=True)
    log_file_path = os.path.join(log_subfolder, "access_advancepay.log")

    logging.basicConfig(
        filename=log_file_path,  
        level=logging.INFO,    
        format="%(asctime)s [%(levelname)s] - %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    app = QtWidgets.QApplication(sys.argv)
    dialog = AdvancePayDialog()
    dialog.show()
    sys.exit(app.exec())